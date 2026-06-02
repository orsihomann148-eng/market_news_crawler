#!/usr/bin/env python3
from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import hmac
import json
import io
import os
import re
import shutil
import signal
import tempfile
import threading
import time
import traceback
import uuid
from contextlib import redirect_stdout, redirect_stderr
from copy import copy
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
import sys
from typing import Any
from urllib.parse import urlparse

import requests
from flask import Flask, Response, jsonify, redirect, render_template, request, session, url_for
from openpyxl import load_workbook

import article_browser as article_browser_utils
import briefing_table as briefing_table_utils
import news_summary as news_summary_utils
import web_settings as web_settings_utils
import dedupe
import db_store
from country_config import (
    COUNTRY_CONFIGS,
    DEFAULT_COUNTRY_CODE,
    country_dict_setting,
    default_country_file_paths,
    country_options,
    delete_custom_country_config,
    get_country_config,
    legacy_country_file_names,
    normalize_country_code,
    normalize_new_country_code,
    normalize_project_relative_path,
    resolve_project_path,
    save_country_config_patch,
    save_custom_country_config,
)
import news_crawler
import source_manager
import xlsx_source_test

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / 'outputs'
RUNTIME_OUTPUT_DIR = Path.cwd() / 'outputs'
LEGACY_WORKSPACE_OUTPUT_DIR = BASE_DIR.parents[2] / 'outputs' if len(BASE_DIR.parents) > 2 else RUNTIME_OUTPUT_DIR
DEFAULT_COUNTRY_FILE_PATHS = default_country_file_paths(DEFAULT_COUNTRY_CODE)
DEFAULT_EXTRA_SOURCES_PATH = DEFAULT_COUNTRY_FILE_PATHS['extra_sources_path']
DEFAULT_ARTICLES_CSV_PATH = BASE_DIR / 'articles.csv'
APP_SETTINGS_PATH = BASE_DIR / 'web_app_settings.json'
ARTICLE_STAR_STORE_PATH = BASE_DIR / 'article_star_store.json'
SOURCE_CAPABILITY_CACHE_PATH = BASE_DIR / DEFAULT_COUNTRY_FILE_PATHS['source_capability_cache_path']
JOB_TIMING_HISTORY_PATH = BASE_DIR / 'job_timing_history.json'
BRIEFING_TABLE_TEMPLATE_FILENAME = '生成资讯表.xlsx'
BRIEFING_TABLE_OUTPUT_DIR = OUTPUT_DIR / 'generated_briefing_tables'
NEWS_SUMMARY_OUTPUT_DIR = OUTPUT_DIR / 'generated_news_summaries'
SETTINGS_SECRET_PREFIX = 'enc-v1:'
SURVEY_API_SETTING_FIELDS = ('survey_api_url', 'survey_api_key', 'survey_api_model')
CONFIG_PREVIOUS_FIELD_MAP = {
    'related_news_search_keywords': 'related_news_search_keywords_previous',
    'report_search_keywords': 'report_search_keywords_previous',
    'survey_system_prompt': 'survey_system_prompt_previous',
}
PLATFORM_ALIASES_PREVIOUS_FIELD = 'platform_aliases_text_previous'

TIMING_HISTORY_LIMIT = 80
NEWS_STAGE_ORDER = [
    'setup',
    'crawl_site',
    'promo_search',
    'translation',
    'survey_filter',
    'ai_dedupe',
    'finalize_output',
    'capability_refresh',
    'done',
]
SOURCE_STAGE_ORDER = ['setup', 'run', 'done']

app = Flask(__name__)
app.secret_key = web_settings_utils.load_or_create_app_secret()

SIDES = [('media', '\u5a92\u4f53\u4fa7'), ('buyer', '\u4e70\u5bb6\u4fa7'), ('seller', '\u5356\u5bb6\u4fa7')]
DEFAULT_NEWS_SIDES = ['media', 'buyer', 'seller']
WEB_VIEW_STATE: dict[str, Any] = {
    'result': None,
    'active_tab': 'home',
    'country_code': DEFAULT_COUNTRY_CODE,
    'news_form_state': None,
    'news_platforms_text': None,
    'news_sides': DEFAULT_NEWS_SIDES.copy(),
    'news_filter_state': None,
    'source_state': None,
}
JOB_STORE: dict[str, dict[str, Any]] = {}
JOB_STORE_LOCK = threading.Lock()
ARTICLE_STAR_STORE_LOCK = threading.Lock()


APP_SECRET_PATH = web_settings_utils.APP_SECRET_PATH


def _sync_web_settings_paths() -> None:
    web_settings_utils.APP_SETTINGS_PATH = APP_SETTINGS_PATH
    web_settings_utils.APP_SECRET_PATH = APP_SECRET_PATH


AUTH_SESSION_KEY = 'market_news_authenticated'
AUTH_ALLOWED_ENDPOINTS = {
    'login',
    'logout',
    'setup_admin',
    'static',
}
PASSWORD_HASH_ITERATIONS = 240_000


def auth_settings() -> dict[str, Any]:
    payload = read_app_settings()
    auth_payload = payload.get('auth')
    return auth_payload if isinstance(auth_payload, dict) else {}


def admin_password_configured() -> bool:
    auth_payload = auth_settings()
    return bool(auth_payload.get('admin_password_hash') and auth_payload.get('admin_password_salt'))


def hash_admin_password(password: str, salt_hex: str | None = None) -> tuple[str, str]:
    normalized = str(password or '')
    salt = bytes.fromhex(salt_hex) if salt_hex else os.urandom(16)
    digest = hashlib.pbkdf2_hmac('sha256', normalized.encode('utf-8'), salt, PASSWORD_HASH_ITERATIONS)
    return salt.hex(), digest.hex()


def save_admin_password(password: str) -> None:
    payload = read_app_settings()
    salt_hex, password_hash = hash_admin_password(password)
    payload['auth'] = {
        'admin_password_salt': salt_hex,
        'admin_password_hash': password_hash,
        'password_updated_at': datetime.now().isoformat(timespec='seconds'),
    }
    write_app_settings(payload)


def verify_admin_password(password: str) -> bool:
    auth_payload = auth_settings()
    salt_hex = str(auth_payload.get('admin_password_salt') or '')
    expected_hash = str(auth_payload.get('admin_password_hash') or '')
    if not salt_hex or not expected_hash:
        return False
    try:
        _, actual_hash = hash_admin_password(password, salt_hex)
    except ValueError:
        return False
    return hmac.compare_digest(actual_hash, expected_hash)


def mark_logged_in() -> None:
    session[AUTH_SESSION_KEY] = True


def logged_in() -> bool:
    return bool(session.get(AUTH_SESSION_KEY))


@app.before_request
def require_login():
    endpoint = request.endpoint or ''
    if endpoint in AUTH_ALLOWED_ENDPOINTS:
        return None
    if endpoint == 'setup_admin' or request.path.startswith('/static/'):
        return None
    if not admin_password_configured():
        if request.path.startswith('/api/'):
            return jsonify({'ok': False, 'error': 'auth_not_configured'}), 401
        return redirect(url_for('setup_admin', next=request.full_path if request.query_string else request.path))
    if logged_in():
        return None
    if request.path.startswith('/api/'):
        return jsonify({'ok': False, 'error': 'login_required'}), 401
    return redirect(url_for('login', next=request.full_path if request.query_string else request.path))


app_secret_dir = web_settings_utils.app_secret_dir
normalize_country_request = web_settings_utils.normalize_country_request
country_path = web_settings_utils.country_path
output_dir_matches_country = web_settings_utils.output_dir_matches_country
install_terminal_signal_guards = web_settings_utils.install_terminal_signal_guards
ensure_secret_file_permissions = web_settings_utils.ensure_secret_file_permissions
derive_secret_keys = web_settings_utils.derive_secret_keys
build_keystream = web_settings_utils.build_keystream
iter_news_api_setting_states = web_settings_utils.iter_news_api_setting_states
settings_payload_needs_secret_migration = web_settings_utils.settings_payload_needs_secret_migration
text_config_changed = web_settings_utils.text_config_changed
backup_news_filter_previous_values = web_settings_utils.backup_news_filter_previous_values
extract_survey_api_settings = web_settings_utils.extract_survey_api_settings
extract_survey_api_settings_from_form = web_settings_utils.extract_survey_api_settings_from_form


def load_or_create_app_secret() -> bytes:
    _sync_web_settings_paths()
    return web_settings_utils.load_or_create_app_secret()


def encrypt_secret_value(value: str) -> str:
    _sync_web_settings_paths()
    return web_settings_utils.encrypt_secret_value(value)


def decrypt_secret_value(value: str) -> str:
    _sync_web_settings_paths()
    return web_settings_utils.decrypt_secret_value(value)


def encrypt_settings_payload(payload: dict[str, Any]) -> dict[str, Any]:
    _sync_web_settings_paths()
    return web_settings_utils.encrypt_settings_payload(payload)


def decrypt_settings_payload(payload: dict[str, Any]) -> dict[str, Any]:
    _sync_web_settings_paths()
    return web_settings_utils.decrypt_settings_payload(payload)


def read_app_settings() -> dict[str, Any]:
    _sync_web_settings_paths()
    return web_settings_utils.read_app_settings()


def write_app_settings(payload: dict[str, Any]) -> None:
    _sync_web_settings_paths()
    web_settings_utils.write_app_settings(payload)


def read_country_settings(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, Any]:
    _sync_web_settings_paths()
    return web_settings_utils.read_country_settings(country_code)


def set_country_previous_config_value(country_code: str, field: str, value: Any) -> None:
    _sync_web_settings_paths()
    web_settings_utils.set_country_previous_config_value(country_code, field, value)


def read_country_previous_config_value(country_code: str, field: str) -> str:
    _sync_web_settings_paths()
    return web_settings_utils.read_country_previous_config_value(country_code, field)


def read_global_survey_api_settings() -> dict[str, str]:
    _sync_web_settings_paths()
    return web_settings_utils.read_global_survey_api_settings()


def persist_global_survey_api_settings(settings: dict[str, str]) -> None:
    _sync_web_settings_paths()
    web_settings_utils.persist_global_survey_api_settings(settings)


def read_country_ai_prompt_setting() -> str:
    _sync_web_settings_paths()
    return web_settings_utils.read_country_ai_prompt_setting()


def persist_country_ai_prompt_setting(prompt: str) -> None:
    _sync_web_settings_paths()
    web_settings_utils.persist_country_ai_prompt_setting(prompt)


def remove_country_settings(country_code: str) -> None:
    _sync_web_settings_paths()
    web_settings_utils.remove_country_settings(country_code)


def read_last_country_code() -> str:
    _sync_web_settings_paths()
    return web_settings_utils.read_last_country_code()


def read_article_star_store() -> dict[str, dict[str, Any]]:
    try:
        db_store.migrate_star_json_if_needed(ARTICLE_STAR_STORE_PATH)
        return db_store.load_star_store()
    except Exception:
        if not ARTICLE_STAR_STORE_PATH.exists():
            return {}
        try:
            payload = json.loads(ARTICLE_STAR_STORE_PATH.read_text(encoding='utf-8'))
        except Exception:
            return {}
        if not isinstance(payload, dict):
            return {}
        return {
            str(key): value
            for key, value in payload.items()
            if isinstance(key, str) and isinstance(value, dict)
        }


def write_article_star_store(payload: dict[str, dict[str, Any]]) -> None:
    for article_id, value in payload.items():
        if isinstance(value, dict):
            db_store.set_article_star(str(article_id), value, bool(value.get('starred', True)))


def build_article_star_key(
    *,
    platform_label: str = '',
    title: str = '',
    title_original: str = '',
    source_name: str = '',
    article_url: str = '',
    source_url: str = '',
    published_at: str = '',
) -> str:
    raw = '||'.join(
        item.strip()
        for item in [
            platform_label,
            title,
            title_original,
            source_name,
            article_url,
            source_url,
            published_at,
        ]
    )
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()


def sanitize_news_filter_state_for_storage(
    state: dict[str, Any] | None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, Any]:
    normalized = build_news_filter_state(state, country_code=country_code)
    return {
        'survey_filter_mode': normalized.get('survey_filter_mode') or xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE,
        'survey_api_url': normalized.get('survey_api_url') or '',
        'survey_api_key': normalized.get('survey_api_key') or '',
        'survey_api_model': normalized.get('survey_api_model') or '',
        'survey_system_prompt': normalized.get('survey_system_prompt') or xlsx_source_test.default_survey_ai_system_prompt(country_code),
        'promo_search_enabled': bool(normalized.get('promo_search_enabled')),
        'related_news_search_enabled': bool(normalized.get('related_news_search_enabled')),
        'report_ranking_search_enabled': bool(normalized.get('report_ranking_search_enabled')),
        'promo_search_engine': normalized.get('promo_search_engine') or 'both',
        'promo_search_keywords': normalized.get('promo_search_keywords') or xlsx_source_test.default_promo_search_keywords_text(country_code),
        'related_news_search_keywords': normalized.get('related_news_search_keywords') or xlsx_source_test.default_related_news_search_keywords_text(country_code),
        'report_search_keywords': normalized.get('report_search_keywords') or xlsx_source_test.default_report_search_keywords_text(country_code),
    }


def sanitize_news_form_state_for_storage(
    state: dict[str, Any] | None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, Any]:
    normalized = build_news_form_state(state, country_code=country_code)
    return {
        'date_mode': normalized.get('date_mode') or 'days',
        'days': str(normalized.get('days') or '7'),
        'start_date': normalized.get('start_date') or '',
        'end_date': normalized.get('end_date') or '',
        'translate_to': normalized.get('translate_to') or 'zh-CN',
        'output_dir': normalized.get('output_dir') or 'outputs',
        'recall_mode': normalized.get('recall_mode') or xlsx_source_test.DEFAULT_RECALL_MODE,
        'news_platforms_text': normalized.get('news_platforms_text') if normalized.get('news_platforms_text') is not None else default_news_platforms_text(country_code),
        'news_sides': list(normalized.get('news_sides') or DEFAULT_NEWS_SIDES.copy()),
    }


def sanitize_source_state_for_storage(state: dict[str, Any] | None) -> dict[str, Any]:
    normalized = build_source_manager_state(state)
    return {
        'skip_api': bool(normalized.get('skip_api')),
        'force_api': bool(normalized.get('force_api')),
        'api_url': normalized.get('api_url') or '',
        'api_key': normalized.get('api_key') or '',
        'api_model': normalized.get('api_model') or '',
    }


def persist_app_settings(
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    news_form_state: dict[str, Any] | None = None,
    news_filter_state: dict[str, Any] | None = None,
    source_state: dict[str, Any] | None = None,
) -> None:
    payload = read_app_settings()
    normalized_country_code = normalize_country_request(country_code)
    countries = payload.get('countries')
    if not isinstance(countries, dict):
        countries = {}
    country_payload = countries.get(normalized_country_code)
    if not isinstance(country_payload, dict):
        country_payload = {}
    if news_form_state is not None:
        country_payload['news_form_state'] = sanitize_news_form_state_for_storage(news_form_state, country_code=normalized_country_code)
    if news_filter_state is not None:
        existing_news_filter_state = country_payload.get('news_filter_state')
        current_news_filter_state = build_news_filter_state(
            existing_news_filter_state if isinstance(existing_news_filter_state, dict) else None,
            country_code=normalized_country_code,
        )
        sanitized_news_filter_state = sanitize_news_filter_state_for_storage(news_filter_state, country_code=normalized_country_code)
        backup_news_filter_previous_values(country_payload, current_news_filter_state, sanitized_news_filter_state)
        country_payload['news_filter_state'] = sanitized_news_filter_state
        ai_settings = payload.get('ai_settings')
        if not isinstance(ai_settings, dict):
            ai_settings = {}
        for key, value in extract_survey_api_settings(sanitized_news_filter_state).items():
            ai_settings[key] = value
        payload['ai_settings'] = ai_settings
    countries[normalized_country_code] = country_payload
    payload['countries'] = countries
    payload['last_country_code'] = normalized_country_code
    if source_state is not None:
        payload['source_state'] = sanitize_source_state_for_storage(source_state)
    write_app_settings(payload)


def apply_overrides_preserving_blank_api_fields(
    base_state: dict[str, Any],
    overrides: dict[str, Any] | None,
    *,
    preserved_fields: list[str],
) -> dict[str, Any]:
    if not overrides:
        return base_state

    for key, value in overrides.items():
        if key in preserved_fields and isinstance(value, str) and not value.strip():
            continue
        base_state[key] = value
    return base_state


def run_cli(func, argv: list[str], **kwargs: Any) -> dict[str, Any]:
    buffer = io.StringIO()
    exit_code = 0
    try:
        with redirect_stdout(buffer), redirect_stderr(buffer):
            result = func(argv, **kwargs)
            exit_code = int(result or 0)
    except SystemExit as exc:
        exit_code = int(exc.code) if isinstance(exc.code, int) else 1
    except Exception:
        exit_code = 1
        buffer.write('\n[\u5185\u90e8\u5f02\u5e38]')
        buffer.write(traceback.format_exc())
    return {
        'argv': redact_cli_args(argv),
        'exit_code': exit_code,
        'output': buffer.getvalue().strip(),
    }


SENSITIVE_CLI_FLAGS = {
    '--survey-api-key',
    '--api-key',
}


def redact_cli_args(argv: list[str] | None) -> list[str]:
    if not argv:
        return []
    redacted: list[str] = []
    hide_next = False
    for item in argv:
        text = str(item)
        if hide_next:
            redacted.append('***')
            hide_next = False
            continue
        matched_flag = next((flag for flag in SENSITIVE_CLI_FLAGS if text == flag), None)
        if matched_flag:
            redacted.append(text)
            hide_next = True
            continue
        matched_prefix = next((flag for flag in SENSITIVE_CLI_FLAGS if text.startswith(f'{flag}=')), None)
        if matched_prefix:
            redacted.append(f'{matched_prefix}=***')
            continue
        redacted.append(text)
    return redacted


def extract_api_error_detail(exc: Exception) -> str:
    response = getattr(exc, 'response', None)
    if response is None:
        message_parts: list[str] = []
        exception_chain = [exc]
        visited_ids = {id(exc)}
        current = exc
        while True:
            next_exc = getattr(current, '__cause__', None) or getattr(current, '__context__', None)
            if not isinstance(next_exc, BaseException) or id(next_exc) in visited_ids:
                break
            visited_ids.add(id(next_exc))
            exception_chain.append(next_exc)
            current = next_exc

        normalized_chain = ' | '.join(
            f'{type(item).__name__}: {str(item or "").strip()}'
            for item in exception_chain
            if str(item or '').strip()
        ).lower()
        if isinstance(exc, requests.exceptions.ProxyError) or 'proxyerror' in normalized_chain:
            message_parts.append('\u8bf7\u6c42\u770b\u8d77\u6765\u53d7\u5230\u4e86\u4ee3\u7406\u6216 VPN \u5f71\u54cd\u3002\u8bf7\u5148\u5173\u95ed\u4ee3\u7406\u3001VPN\uff0c\u6216\u6539\u7528\u53ef\u76f4\u63a5\u8bbf\u95ee HTTPS \u7684\u7f51\u7edc\u540e\u91cd\u8bd5\u3002')
        elif isinstance(exc, requests.exceptions.SSLError) or 'certificate verify failed' in normalized_chain or 'ssl' in normalized_chain:
            message_parts.append('\u8bf7\u6c42\u9047\u5230\u4e86 SSL \u8bc1\u4e66\u6821\u9a8c\u95ee\u9898\u3002\u8bf7\u68c0\u67e5\u7cfb\u7edf\u65f6\u95f4\u3001\u7f51\u7edc\u73af\u5883\uff0c\u6216\u786e\u8ba4\u76ee\u6807 HTTPS \u7ad9\u70b9\u8bc1\u4e66\u662f\u5426\u6b63\u5e38\u3002')
        elif isinstance(exc, requests.exceptions.ConnectTimeout) or 'connecttimeout' in normalized_chain:
            message_parts.append('\u8fde\u63a5\u76ee\u6807\u7ad9\u70b9\u8d85\u65f6\u3002\u8bf7\u68c0\u67e5\u5f53\u524d\u7f51\u7edc\u3001\u4ee3\u7406\u8bbe\u7f6e\uff0c\u6216\u7a0d\u540e\u91cd\u8bd5\u3002')
        elif isinstance(exc, requests.exceptions.ReadTimeout) or 'readtimeout' in normalized_chain:
            message_parts.append('\u76ee\u6807\u7ad9\u70b9\u54cd\u5e94\u8d85\u65f6\u3002\u53ef\u80fd\u662f\u7f51\u7edc\u8f83\u6162\u6216\u5bf9\u65b9\u7ad9\u70b9\u6682\u65f6\u8f83\u5fd9\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5\u3002')
        elif isinstance(exc, requests.exceptions.ConnectionError) or 'name resolution' in normalized_chain or 'nodename nor servname provided' in normalized_chain:
            message_parts.append('\u7f51\u7edc\u8fde\u63a5\u6216 DNS \u89e3\u6790\u5931\u8d25\u3002\u8bf7\u68c0\u67e5\u672c\u673a\u7f51\u7edc\u3001\u4ee3\u7406\u8bbe\u7f6e\uff0c\u6216\u786e\u8ba4\u57df\u540d\u53ef\u4ee5\u6b63\u5e38\u8bbf\u95ee\u3002')

        raw_message = ' | '.join(
            str(item or '').strip()
            for item in exception_chain
            if str(item or '').strip()
        )
        if raw_message:
            message_parts.append(raw_message[:500])
        return '\uff1b'.join(part for part in message_parts if part)
    try:
        body_text = str(response.text or '').strip()
    except Exception:
        body_text = ''
    if not body_text:
        return ''
    try:
        payload = json.loads(body_text)
    except Exception:
        return body_text[:500]
    if isinstance(payload, dict):
        error_payload = payload.get('error')
        if isinstance(error_payload, dict):
            message = str(error_payload.get('message') or '').strip()
            if message:
                return message
        message = str(payload.get('message') or '').strip()
        if message:
            return message
    return body_text[:500]


def test_survey_api_connection(
    *,
    api_url: str,
    api_key: str,
    api_model: str,
) -> dict[str, Any]:
    normalized_url = xlsx_source_test.normalize_chat_completions_url(api_url)
    normalized_key = str(api_key or '').strip()
    normalized_model = str(api_model or '').strip()
    if not normalized_url or not normalized_key or not normalized_model:
        raise RuntimeError('\u8c03\u7528 AI API \u524d\u9700\u8981\u5b8c\u6574\u914d\u7f6e API URL\u3001Key \u548c\u6a21\u578b\u540d\u3002')

    started_at = time.time()
    payload = xlsx_source_test.call_survey_filter_api(
        [
            {'role': 'system', 'content': 'Return a JSON object only.'},
            {'role': 'user', 'content': 'Please reply with {"ok":true,"message":"api test passed"} exactly.'},
        ],
        normalized_url,
        normalized_key,
        normalized_model,
    )
    elapsed_ms = int(round((time.time() - started_at) * 1000))
    return {
        'api_url': normalized_url,
        'api_model': normalized_model,
        'elapsed_ms': elapsed_ms,
        'payload': payload,
    }


read_output_metadata_for_file = article_browser_utils.read_output_metadata_for_file
article_csv_translation_status_label = article_browser_utils.article_csv_translation_status_label
classify_article_csv_label = article_browser_utils.classify_article_csv_label
build_article_csv_display_label = article_browser_utils.build_article_csv_display_label
is_article_browser_csv = article_browser_utils.is_article_browser_csv
count_csv_data_rows = article_browser_utils.count_csv_data_rows
iter_output_dirs = article_browser_utils.iter_output_dirs
article_source_value = article_browser_utils.article_source_value
db_article_source_value = article_browser_utils.db_article_source_value
parse_db_article_source = article_browser_utils.parse_db_article_source
is_db_article_source = article_browser_utils.is_db_article_source
resolve_output_dir_for_run = article_browser_utils.resolve_output_dir_for_run
list_recent_output_dirs = article_browser_utils.list_recent_output_dirs
list_article_csv_paths = article_browser_utils.list_article_csv_paths
list_article_csv_files = article_browser_utils.list_article_csv_files
list_user_article_sources = article_browser_utils.list_user_article_sources
resolve_article_csv_path = article_browser_utils.resolve_article_csv_path
translate_article_summary_fallback = article_browser_utils.translate_article_summary_fallback
translate_article_title_for_display = article_browser_utils.translate_article_title_for_display
normalize_article_card_summary = article_browser_utils.normalize_article_card_summary
format_article_published_date = article_browser_utils.format_article_published_date
fetch_article_summary_fallback = article_browser_utils.fetch_article_summary_fallback
load_article_rows = article_browser_utils.load_article_rows
load_article_rows_from_db_source = article_browser_utils.load_article_rows_from_db_source
article_export_row_score = article_browser_utils.article_export_row_score
format_starred_updated_at = article_browser_utils.format_starred_updated_at
build_starred_export_rows = article_browser_utils.build_starred_export_rows

BRIEFING_TABLE_HEADERS = briefing_table_utils.BRIEFING_TABLE_HEADERS
BRIEFING_AI_HEADERS = briefing_table_utils.BRIEFING_AI_HEADERS

find_briefing_table_template_path = briefing_table_utils.find_briefing_table_template_path
parse_article_publish_dt = briefing_table_utils.parse_article_publish_dt
format_tracking_week = briefing_table_utils.format_tracking_week
format_briefing_publish_date = briefing_table_utils.format_briefing_publish_date
briefing_country_label = briefing_table_utils.briefing_country_label
briefing_country_zh_label = briefing_table_utils.briefing_country_zh_label
briefing_segment_options = briefing_table_utils.briefing_segment_options
compact_cell_text = briefing_table_utils.compact_cell_text
infer_briefing_news_type = briefing_table_utils.infer_briefing_news_type
source_name_for_briefing = briefing_table_utils.source_name_for_briefing
nps_metric_from_dimensions = briefing_table_utils.nps_metric_from_dimensions
direction_symbol = briefing_table_utils.direction_symbol
normalize_briefing_sentiment = briefing_table_utils.normalize_briefing_sentiment
normalize_briefing_segment_text = briefing_table_utils.normalize_briefing_segment_text
is_invalid_briefing_segment = briefing_table_utils.is_invalid_briefing_segment
infer_briefing_segment = briefing_table_utils.infer_briefing_segment
normalize_briefing_segment = briefing_table_utils.normalize_briefing_segment
direct_briefing_fields = briefing_table_utils.direct_briefing_fields
default_briefing_ai_fields = briefing_table_utils.default_briefing_ai_fields
build_briefing_ai_messages = briefing_table_utils.build_briefing_ai_messages
normalize_enum = briefing_table_utils.normalize_enum
get_ai_item_value = briefing_table_utils.get_ai_item_value
normalize_briefing_ai_item = briefing_table_utils.normalize_briefing_ai_item
briefing_payload_items = briefing_table_utils.briefing_payload_items
generate_briefing_ai_fields = briefing_table_utils.generate_briefing_ai_fields


def build_article_briefing_table(
    rows: list[dict[str, Any]],
    *,
    country_code: str,
    api_settings: dict[str, str],
    ai_stats: dict[str, int] | None = None,
) -> Path:
    briefing_table_utils.BRIEFING_TABLE_OUTPUT_DIR = BRIEFING_TABLE_OUTPUT_DIR
    return briefing_table_utils.build_article_briefing_table(
        rows,
        country_code=country_code,
        api_settings=api_settings,
        ai_stats=ai_stats,
    )


def build_article_browser_state(
    args=None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    user_friendly_sources: bool = False,
) -> dict[str, Any]:
    article_sources = (
        list_user_article_sources(country_code=country_code)
        if user_friendly_sources
        else list_article_csv_files(country_code=country_code)
    )
    selected_source = (args.get('article_source') if args else '') or ''
    if user_friendly_sources and selected_source:
        available_source_values = {str(source.get('value') or '') for source in article_sources}
        if selected_source not in available_source_values:
            selected_source = ''
    if not selected_source and article_sources:
        selected_source = article_sources[0]['value']
    selected_platform = (args.get('article_platform') if args else '') or ''
    keyword = (args.get('article_keyword') if args else '') or ''
    raw_star_filter = ((args.get('article_starred_only') if args else '') or '').strip().lower()
    star_filter = raw_star_filter if raw_star_filter in {'starred', 'unstarred'} else ''

    raw_limit = (args.get('article_limit') if args else '') or 'all'
    if str(raw_limit).strip().lower() == 'all':
        limit = None
    else:
        try:
            limit = int(raw_limit)
        except ValueError:
            limit = 20
        limit = max(5, min(limit, 200))

    if is_db_article_source(selected_source):
        article_rows, available_platforms, file_total_rows, matching_rows, file_starred_rows = load_article_rows_from_db_source(
            selected_source,
            selected_platform=selected_platform,
            keyword=keyword,
            star_filter=star_filter,
            limit=limit,
            hydrate_remote_summary=False,
        )
        selected_source_value = selected_source
        resolved_path = selected_source
        file_exists = file_total_rows > 0
    else:
        csv_path = resolve_article_csv_path(selected_source, country_code=country_code)
        article_rows, available_platforms, file_total_rows, matching_rows, file_starred_rows = load_article_rows(
            csv_path,
            selected_platform=selected_platform,
            keyword=keyword,
            star_filter=star_filter,
            limit=limit,
            hydrate_remote_summary=False,
        )
        selected_source_value = article_source_value(csv_path) if csv_path.exists() else selected_source
        resolved_path = str(csv_path)
        file_exists = csv_path.exists()

    return enrich_article_browser_display_state({
        'sources': article_sources,
        'selected_source': selected_source_value,
        'selected_platform': selected_platform,
        'available_platforms': available_platforms,
        'available_platform_options': platform_display_options(available_platforms),
        'keyword': keyword,
        'star_filter': star_filter,
        'starred_only': star_filter == 'starred',
        'limit': 'all' if limit is None else limit,
        'rows': article_rows,
        'total_rows': len(article_rows),
        'file_total_rows': file_total_rows,
        'matching_rows': matching_rows,
        'file_starred_rows': file_starred_rows,
        'resolved_path': resolved_path,
        'file_exists': file_exists,
    })


def build_article_browser_stub_state(
    args=None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    user_friendly_sources: bool = False,
) -> dict[str, Any]:
    article_sources = (
        list_user_article_sources(country_code=country_code)
        if user_friendly_sources
        else list_article_csv_files(country_code=country_code)
    )
    selected_source = (args.get('article_source') if args else '') or ''
    if user_friendly_sources and selected_source:
        available_source_values = {str(source.get('value') or '') for source in article_sources}
        if selected_source not in available_source_values:
            selected_source = ''
    if not selected_source and article_sources:
        selected_source = article_sources[0]['value']
    selected_platform = (args.get('article_platform') if args else '') or ''
    keyword = (args.get('article_keyword') if args else '') or ''
    raw_star_filter = ((args.get('article_starred_only') if args else '') or '').strip().lower()
    star_filter = raw_star_filter if raw_star_filter in {'starred', 'unstarred'} else ''

    raw_limit = (args.get('article_limit') if args else '') or 'all'
    if str(raw_limit).strip().lower() == 'all':
        limit = None
    else:
        try:
            limit = int(raw_limit)
        except ValueError:
            limit = 20
        limit = max(5, min(limit, 200))

    selected_source_value = selected_source
    resolved_path = selected_source
    file_exists = False
    if not selected_source and article_sources:
        selected_source_value = article_sources[0]['value']
        resolved_path = selected_source_value
        file_exists = True
    elif is_db_article_source(selected_source):
        run_id, stage = parse_db_article_source(selected_source) or ('', 'after')
        selected_source_value = selected_source
        resolved_path = selected_source
        file_exists = bool(run_id and db_store.load_articles(run_id, stage))
    else:
        csv_path = resolve_article_csv_path(selected_source, country_code=country_code)
        resolved_path = str(csv_path)
        file_exists = csv_path.exists()
        if csv_path.exists():
            selected_source_value = article_source_value(csv_path)

    return enrich_article_browser_display_state({
        'sources': article_sources,
        'selected_source': selected_source_value,
        'selected_platform': selected_platform,
        'available_platforms': [],
        'available_platform_options': [],
        'keyword': keyword,
        'star_filter': star_filter,
        'starred_only': star_filter == 'starred',
        'limit': 'all' if limit is None else limit,
        'rows': [],
        'total_rows': 0,
        'file_total_rows': 0,
        'matching_rows': 0,
        'file_starred_rows': 0,
        'resolved_path': resolved_path,
        'file_exists': file_exists,
    })


def manual_article_payload_from_form(form) -> tuple[str, dict[str, Any]]:
    country_code = normalize_country_request(form.get('country_code'))
    raw_article_source = (form.get('article_source') or '').strip()
    parsed = parse_db_article_source(raw_article_source)
    if parsed is None:
        raise ValueError('\u8bf7\u9009\u62e9 SQLite \u7684 after \u6570\u636e\u6e90\uff1bCSV \u5386\u53f2\u6587\u4ef6\u4e0d\u80fd\u76f4\u63a5\u6dfb\u52a0\u624b\u52a8\u65b0\u95fb\u3002')
    run_id, stage = parsed
    if stage != 'after':
        raise ValueError('\u624b\u52a8\u65b0\u95fb\u53ea\u80fd\u6dfb\u52a0\u5230\u7b5b\u9009\u540e\u7684 after \u65b0\u95fb\u5e93\u3002')

    title = (form.get('title') or '').strip()
    title_translated = (form.get('title_translated') or '').strip()
    article_url = (form.get('article_url') or '').strip()
    platform_label = (form.get('platform_label') or '').strip()
    published_at = (form.get('published_at') or '').strip()
    if not (title or article_url):
        raise ValueError('\u8bf7\u81f3\u5c11\u586b\u5199\u65b0\u95fb\u6807\u9898\u6216 URL\u3002')
    if not platform_label:
        raise ValueError('\u8bf7\u586b\u5199\u5e73\u53f0\u540d\u79f0\u3002')
    if not published_at:
        raise ValueError('\u8bf7\u586b\u5199\u53d1\u5e03\u65f6\u95f4\u3002')

    briefing_sentiment = normalize_briefing_sentiment(form.get('briefing_sentiment') or '')
    if not briefing_sentiment:
        briefing_sentiment = 'Neutral'
    industry_trend_flag = str(form.get('industry_trend_flag') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    source_name = (form.get('source_name') or '').strip()
    payload = {
        'country_code': country_code,
        'platform': platform_label,
        'platform_label': platform_label,
        'matched_brands': [item.strip() for item in re.split(r'[,|/]+', platform_label) if item.strip()] or [platform_label],
        'title': title,
        'title_translated': title_translated,
        'summary': (form.get('summary') or '').strip(),
        'summary_translated': (form.get('summary_translated') or '').strip(),
        'article_url': article_url,
        'verification_final_url': article_url,
        'source_name': source_name,
        'source_site': source_name,
        'source_url': (form.get('source_url') or '').strip(),
        'published_at': published_at,
        'category': (form.get('category') or '').strip(),
        'survey_dimensions': (form.get('survey_dimensions') or '').strip(),
        'survey_question_ids': (form.get('survey_question_ids') or '').strip(),
        'survey_indicator_examples': (form.get('survey_indicator_examples') or '').strip(),
        'briefing_sentiment': briefing_sentiment,
        'briefing_sentiment_reason': (form.get('briefing_sentiment_reason') or '').strip(),
        'industry_trend_flag': industry_trend_flag,
        'industry_trend_category': (form.get('industry_trend_category') or '').strip() if industry_trend_flag else '',
        'industry_trend_impact': xlsx_source_test.normalize_industry_trend_impact(form.get('industry_trend_impact')) if industry_trend_flag else '',
        'industry_trend_reason': (form.get('industry_trend_reason') or '').strip() if industry_trend_flag else '',
        'source_discovery': 'manual_added',
        'manual_added': True,
    }
    return run_id, payload


def saved_survey_api_settings_for_country(country_code: str) -> dict[str, str]:
    country_settings = read_country_settings(country_code)
    api_settings = extract_survey_api_settings(
        country_settings.get('news_filter_state') if isinstance(country_settings.get('news_filter_state'), dict) else None
    )
    if not all(api_settings.get(key) for key in SURVEY_API_SETTING_FIELDS):
        global_settings = read_global_survey_api_settings()
        api_settings = {**global_settings, **api_settings}
    api_settings['survey_api_url'] = xlsx_source_test.normalize_chat_completions_url(api_settings.get('survey_api_url') or '')
    return api_settings


def validate_manual_article_source(raw_article_source: str) -> tuple[str, str]:
    parsed = parse_db_article_source((raw_article_source or '').strip())
    if parsed is None:
        raise ValueError('\u8bf7\u9009\u62e9 SQLite \u7684 after \u6570\u636e\u6e90\uff1bCSV \u5386\u53f2\u6587\u4ef6\u4e0d\u80fd\u76f4\u63a5\u6dfb\u52a0\u624b\u52a8\u65b0\u95fb\u3002')
    run_id, stage = parsed
    if stage != 'after':
        raise ValueError('\u624b\u52a8\u65b0\u95fb\u53ea\u80fd\u6dfb\u52a0\u5230\u7b5b\u9009\u540e\u7684 after \u65b0\u95fb\u5e93\u3002')
    return run_id, stage


def filter_rows_by_selected_article_ids(rows: list[dict[str, Any]], form: Any) -> list[dict[str, Any]]:
    selection_mode = str(form.get('article_selection_mode') or '').strip().lower()
    if selection_mode != 'selected':
        return rows
    selected_ids = [
        str(item).strip()
        for item in form.getlist('selected_article_ids')
        if str(item).strip()
    ]
    if not selected_ids:
        raise ValueError('\u8bf7\u81f3\u5c11\u9009\u62e9 1 \u6761\u65b0\u95fb\u518d\u751f\u6210\u6750\u6599\u3002')
    selected_id_set = set(selected_ids)
    filtered_rows = [
        row
        for row in rows
        if str(row.get('article_id') or '').strip() in selected_id_set
    ]
    if not filtered_rows:
        raise ValueError('\u5f53\u524d\u52fe\u9009\u7684\u65b0\u95fb\u5df2\u66f4\u65b0\u6216\u4e0d\u5b58\u5728\uff0c\u8bf7\u5237\u65b0\u5217\u8868\u540e\u91cd\u8bd5\u3002')
    return filtered_rows

def fetch_manual_article_url_context(article_url: str) -> dict[str, Any]:
    normalized_url = (article_url or '').strip()
    context: dict[str, Any] = {'article_url': normalized_url}
    if not normalized_url:
        return context
    session = requests.Session()
    start = datetime(2000, 1, 1)
    end = datetime.now() + timedelta(days=3650)
    meta = xlsx_source_test.extract_article_metadata(
        normalized_url,
        session,
        start,
        end,
        source_discovery='manual_ai_suggest',
    )
    if isinstance(meta, dict):
        context.update({key: value for key, value in meta.items() if value})
        return context
    try:
        response = session.get(
            normalized_url,
            timeout=15,
            headers={'User-Agent': 'Mozilla/5.0 (compatible; MarketNewsCrawler/1.0)'},
        )
        context['article_url'] = response.url or normalized_url
        context['http_status'] = response.status_code
        html = response.text[:200000]
        title_match = re.search(r'<title[^>]*>(.*?)</title>', html, flags=re.I | re.S)
        if title_match:
            context['title'] = xlsx_source_test.clean_text(re.sub(r'<[^>]+>', ' ', title_match.group(1)))
        description_match = re.search(
            r'<meta[^>]+(?:name|property)=["\'](?:description|og:description)["\'][^>]+content=["\']([^"\']+)["\']',
            html,
            flags=re.I | re.S,
        )
        if description_match:
            context['summary'] = xlsx_source_test.clean_text(description_match.group(1))
    except Exception as exc:
        context['fetch_error'] = str(exc)[:200]
    return context


def manual_article_suggestion_messages(
    *,
    country_code: str,
    article_url: str,
    article_context: dict[str, Any],
    platform_hint: str = '',
) -> list[dict[str, str]]:
    country_label = get_country_config(country_code)['label']
    platforms = xlsx_source_test.list_available_platform_labels(country_code)
    try:
        indicators = xlsx_source_test.load_survey_indicators_from_xlsx(str(country_path(country_code, 'xlsx_path')))
    except Exception:
        indicators = []
    indicator_lines = [
        f'{item.question_id}: {item.dimension} - {item.prompt_zh or item.prompt_en}'
        for item in indicators[:80]
    ]
    return [
        {
            'role': 'system',
            'content': (
                '\u4f60\u662f\u4e00\u4e2a\u7528\u4e8e\u65b0\u95fb\u5f55\u5165\u7684\u7ed3\u6784\u5316\u52a9\u624b\u3002'
                '\u8bf7\u6839\u636e URL \u3001\u9875\u9762\u4e0a\u4e0b\u6587\u3001\u56fd\u5bb6\u548c\u53ef\u7528\u5e73\u53f0\u5217\u8868\uff0c'
                '\u63a8\u6d4b\u5e76\u586b\u5199\u65b0\u95fb\u624b\u52a8\u5f55\u5165\u6240\u9700\u5b57\u6bb5\u3002'
                '\u8f93\u51fa\u5fc5\u987b\u662f JSON object\uff0c\u4e0d\u8981\u8f93\u51fa markdown \u6216\u989d\u5916\u89e3\u91ca\u3002'
                '\u6807\u9898\u3001\u6458\u8981\u3001\u5173\u8054\u8bf4\u660e\u7528\u4e2d\u6587\u3002'
                '\u5e73\u53f0\u540d\u53ea\u80fd\u4ece available_platforms \u4e2d\u9009\uff0c\u4e0d\u5141\u8bb8\u81ea\u9020\u65b0\u503c\u3002'
                '\u82e5\u4fe1\u606f\u4e0d\u8db3\uff0c\u53ef\u4ee5\u5408\u7406\u4fdd\u5b88\u63a8\u65ad\uff0c\u4f46\u4e0d\u8981\u7f16\u9020\u7ec6\u8282\u3002'
                '\u8bf7\u4f18\u5148\u8fd4\u56de\u8fd9\u4e9b\u5b57\u6bb5\uff1atitle, title_translated, article_url, published_at, source_name, '
                'platform_label, summary, survey_dimensions, survey_question_ids, '
                'survey_indicator_examples, briefing_sentiment, industry_trend_flag, '
                'industry_trend_category, industry_trend_impact, industry_trend_reason.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_label': country_label,
                    'article_url': article_url,
                    'platform_hint': platform_hint,
                    'available_platforms': platforms,
                    'article_context': article_context,
                    'survey_indicators': indicator_lines,
                },
                ensure_ascii=False,
            ),
        },
    ]

def normalize_manual_article_suggestion(payload: dict[str, Any], *, article_url: str, country_code: str) -> dict[str, Any]:
    raw = payload.get('article') if isinstance(payload.get('article'), dict) else payload
    platform = xlsx_source_test.clean_text(raw.get('platform_label') or raw.get('platform') or '')
    known_platforms = set(xlsx_source_test.list_available_platform_labels(country_code))
    platform_lookup = {item.lower(): item for item in known_platforms}
    platform = platform_lookup.get(platform.lower(), platform if platform in known_platforms else '')
    published_at = xlsx_source_test.clean_text(raw.get('published_at') or raw.get('publish_date') or '')
    if published_at:
        parsed_dt = xlsx_source_test.parse_dt(published_at)
        if parsed_dt is not None:
            published_at = parsed_dt.date().isoformat()
    fields = {
        'title': xlsx_source_test.clean_text(raw.get('title') or ''),
        'title_translated': xlsx_source_test.clean_text(raw.get('title_translated') or raw.get('title_zh') or ''),
        'article_url': xlsx_source_test.clean_text(raw.get('article_url') or article_url),
        'published_at': published_at,
        'source_name': xlsx_source_test.clean_text(raw.get('source_name') or ''),
        'platform_label': platform,
        'summary': xlsx_source_test.clean_text(raw.get('summary') or ''),
        'survey_dimensions': xlsx_source_test.clean_text(raw.get('survey_dimensions') or ''),
        'survey_question_ids': xlsx_source_test.clean_text(raw.get('survey_question_ids') or ''),
        'survey_indicator_examples': xlsx_source_test.clean_text(raw.get('survey_indicator_examples') or ''),
        'briefing_sentiment': xlsx_source_test.normalize_briefing_sentiment(raw.get('briefing_sentiment')),
        'industry_trend_flag': bool(raw.get('industry_trend_flag')),
        'industry_trend_category': xlsx_source_test.clean_text(raw.get('industry_trend_category') or ''),
        'industry_trend_impact': xlsx_source_test.normalize_industry_trend_impact(raw.get('industry_trend_impact')),
        'industry_trend_reason': xlsx_source_test.clean_text(raw.get('industry_trend_reason') or ''),
    }
    dataset = {
        'manualArticleId': '',
        'platformLabel': fields['platform_label'],
        'title': fields['title'],
        'titleTranslated': fields['title_translated'],
        'articleUrl': fields['article_url'],
        'sourceName': fields['source_name'],
        'publishedAt': fields['published_at'],
        'summary': fields['summary'],
        'surveyDimensions': fields['survey_dimensions'],
        'surveyQuestionIds': fields['survey_question_ids'],
        'surveyIndicatorExamples': fields['survey_indicator_examples'],
        'briefingSentiment': fields['briefing_sentiment'],
        'industryTrendFlag': '1' if fields['industry_trend_flag'] else '0',
        'industryTrendCategory': fields['industry_trend_category'],
        'industryTrendImpact': fields['industry_trend_impact'],
        'industryTrendReason': fields['industry_trend_reason'],
    }
    return {'fields': fields, 'dataset': dataset}


def build_article_source_payload(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, Any]:
    sources = list_article_csv_files(country_code=country_code)
    latest_source = sources[0]['value'] if sources else ''
    return {
        'sources': sources,
        'latest_source': latest_source,
    }


def default_news_platforms_text(country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    return '\n'.join(xlsx_source_test.list_available_platform_labels(country_code))


def parse_platform_text(raw_text: str) -> list[str]:
    return [
        item.strip()
        for line in raw_text.replace('\u3001', ',').splitlines()
        for item in line.split(',')
        if item.strip()
    ]


def split_news_platforms_for_country(
    platforms_text: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> tuple[list[str], list[str]]:
    known_platforms = xlsx_source_test.list_available_platform_labels(country_code)
    known_lookup = {platform.lower(): platform for platform in known_platforms}
    alias_lookup = {
        alias.lower(): platform
        for alias, platform in xlsx_source_test.requested_display_platform_lookup(country_code).items()
        if platform in known_platforms
    }
    selected_known: list[str] = []
    custom_platforms: list[str] = []

    for item in parse_platform_text(platforms_text or ''):
        matched = known_lookup.get(item.lower()) or alias_lookup.get(item.lower())
        if matched:
            if matched not in selected_known:
                selected_known.append(matched)
        elif item not in custom_platforms:
            custom_platforms.append(item)
    return selected_known, custom_platforms


def normalize_news_platforms_text_for_country(
    platforms_text: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    selected_known, custom_platforms = split_news_platforms_for_country(platforms_text, country_code)
    if not selected_known and not custom_platforms and platforms_text is None:
        return default_news_platforms_text(country_code)
    return '\n'.join([*selected_known, *custom_platforms])


def parse_multiline_items(raw_text: str) -> list[str]:
    items: list[str] = []
    for line in str(raw_text or '').replace('\u3001', ',').splitlines():
        for item in line.split(','):
            normalized = item.strip()
            if normalized and normalized not in items:
                items.append(normalized)
    return items


def unique_texts(values: list[str]) -> list[str]:
    items: list[str] = []
    for value in values:
        normalized = str(value or '').strip()
        if normalized and normalized not in items:
            items.append(normalized)
    return items


def contains_cjk(value: str) -> bool:
    return bool(re.search(r'[\u3400-\u9fff]', str(value or '')))


def alias_country_terms(market_terms: list[str], country_code: str) -> list[str]:
    terms = [
        str(term or '').strip()
        for term in market_terms
        if str(term or '').strip() and not contains_cjk(str(term or ''))
    ]
    terms.append(country_code.title())
    return unique_texts(terms)


def normalize_platform_alias_token(value: str) -> str:
    normalized = str(value or '').strip()
    if not normalized:
        return ''
    normalized = normalized.replace('\uff08', '(').replace('\uff09', ')')
    normalized = re.sub(r'\([^)]*\)', ' ', normalized)
    normalized = re.sub(r'[^A-Za-z0-9]+', ' ', normalized).strip().lower()
    return re.sub(r'\s+', ' ', normalized)


def platform_alias_concept(platform: str) -> str:
    normalized = normalize_platform_alias_token(platform)
    collapsed = normalized.replace(' ', '')
    if not normalized:
        return ''
    if (
        'instagram' in normalized
        or collapsed in {'ig', 'ins', 'insdtc', 'insshop', 'instagramshop', 'instagramshopping'}
        or ('ins' in normalized and 'dtc' in normalized)
    ):
        return 'instagram_shopping'
    if 'tiktok' in normalized or collapsed in {'tts', 'ttsop', 'tiktokshop', 'ttshop'}:
        return 'tiktok_shop'
    if 'amazon' in normalized:
        return 'amazon'
    if 'zalando' in normalized:
        return 'zalando'
    if 'shein' in normalized:
        return 'shein'
    if 'temu' in normalized:
        return 'temu'
    if 'rakuten' in normalized:
        return 'rakuten'
    return ''


def platform_alias_seed_terms(platform: str) -> list[str]:
    platform_name = str(platform or '').strip()
    concept = platform_alias_concept(platform_name)
    concept_terms = {
        'instagram_shopping': [
            platform_name,
            'Instagram Shopping',
            'Instagram Shop',
            'Instagram',
            'Instagram DTC',
            'IG',
            'INS',
            'INS-DTC',
        ],
        'tiktok_shop': [
            platform_name,
            'TikTok Shop',
            'TikTok',
            'TTS',
            'TT Shop',
        ],
        'amazon': [platform_name, 'Amazon'],
        'zalando': [platform_name, 'Zalando'],
        'shein': [platform_name, 'SHEIN', 'Shein'],
        'temu': [platform_name, 'Temu', 'TEMU'],
        'rakuten': [platform_name, 'Rakuten', 'Rakuten Ichiba'],
    }
    return unique_texts(concept_terms.get(concept, [platform_name]))


def build_semantic_platform_aliases(
    platform: str,
    *,
    country_code: str,
    google_gl: str,
    market_terms: list[str],
) -> list[str]:
    platform_name = str(platform or '').strip()
    if not platform_name:
        return []
    country_terms = alias_country_terms(market_terms, country_code)
    primary_country = next((term for term in country_terms if str(term or '').isascii()), country_terms[0] if country_terms else country_code.title())
    seed_terms = platform_alias_seed_terms(platform_name)
    candidates: list[str] = []

    def add_candidate(value: str) -> None:
        normalized = str(value or '').strip()
        if normalized:
            candidates.append(normalized)

    for seed in seed_terms:
        add_candidate(seed)
        if google_gl:
            add_candidate(f'{seed} {google_gl}')
        if primary_country:
            add_candidate(f'{seed} {primary_country}')
        for term in country_terms[:4]:
            add_candidate(f'{seed} {term}')

    concept = platform_alias_concept(platform_name)
    if concept == 'amazon' and google_gl:
        add_candidate(f'Amazon {google_gl}')
        if len(google_gl) == 2:
            add_candidate(f'Amazon.{google_gl.lower()}')
    elif concept == 'instagram_shopping':
        for term in country_terms[:4]:
            add_candidate(f'Instagram Shopping {term}')
            add_candidate(f'Instagram {term}')
            add_candidate(f'Instagram DTC {term}')
        if google_gl:
            add_candidate(f'Instagram Shopping {google_gl}')
            add_candidate(f'Instagram {google_gl}')
    elif concept == 'tiktok_shop':
        for term in country_terms[:4]:
            add_candidate(f'TikTok Shop {term}')
            add_candidate(f'TikTok {term}')
        if google_gl:
            add_candidate(f'TikTok Shop {google_gl}')

    return [
        alias
        for alias in unique_texts(candidates)
        if alias and not contains_cjk(alias)
    ]


def parse_platform_alias_text(raw_text: str) -> dict[str, list[str]]:
    aliases: dict[str, list[str]] = {}
    for line in str(raw_text or '').splitlines():
        normalized = line.replace('\uff5c', '|').strip()
        if not normalized:
            continue
        if '|' in normalized:
            parts = [part.strip() for part in normalized.split('|') if part.strip()]
        elif '=' in normalized:
            platform, raw_aliases = normalized.split('=', 1)
            parts = [platform.strip()] + parse_multiline_items(raw_aliases)
        else:
            parts = parse_multiline_items(normalized)
        platform = parts[0]
        aliases[platform] = unique_texts(parts[1:] or [platform])
    return aliases


def format_platform_alias_text(alias_map: dict[str, list[str]], platform_order: list[str] | None = None) -> str:
    lines: list[str] = []
    emitted: set[str] = set()
    order = platform_order or []
    for platform in order + sorted(alias_map):
        if platform in emitted:
            continue
        emitted.add(platform)
        aliases = [
            alias
            for alias in unique_texts(list(alias_map.get(platform, [])))
            if alias != platform
        ]
        lines.append(' | '.join([platform, *aliases]))
    return '\n'.join(lines)


def generate_platform_aliases(
    platform: str,
    *,
    country_code: str,
    country_label: str,
    google_gl: str,
    market_terms: list[str],
) -> list[str]:
    platform_name = str(platform or '').strip()
    if not platform_name:
        return []
    return build_semantic_platform_aliases(
        platform_name,
        country_code=country_code,
        google_gl=google_gl,
        market_terms=market_terms,
    )


def quote_search_term(term: str) -> str:
    normalized = str(term or '').strip()
    if not normalized:
        return ''
    if re.search(r'\s', normalized):
        return f'"{normalized}"'
    return normalized


def build_market_search_block(terms: list[str]) -> str:
    tokens = [quote_search_term(term) for term in terms if quote_search_term(term)]
    return f"({' OR '.join(tokens)})" if tokens else ''


def parse_official_sources(raw_text: str) -> dict[str, dict[str, str]]:
    sources: dict[str, dict[str, str]] = {}
    for line in str(raw_text or '').splitlines():
        normalized = line.strip()
        if not normalized:
            continue
        parts = [part.strip() for part in normalized.split('|')]
        if len(parts) < 3:
            continue
        platform, label, url = parts[0], parts[1], parts[2]
        if platform and label and url:
            sources[platform] = {'label': label, 'url': url}
    return sources


def generic_promo_query_blocks() -> list[list[str]]:
    return [
        [
            'promo_core',
            '(promotion OR promo OR discount OR coupon OR sale OR deal OR offer OR campaign OR cashback)',
        ],
        [
            'points_rewards',
            '(points OR loyalty OR reward OR rewards OR cashback)',
        ],
        [
            'price_event',
            '("flash sale" OR voucher OR bundle OR "special offer" OR "limited time" OR "price cut" OR markdown)',
        ],
        [
            'shopping_festivals',
            '("black friday" OR "cyber monday" OR "prime day" OR "shopping festival" OR "shopping season")',
        ],
        [
            'seasonal_holidays',
            '(christmas OR valentine OR "mothers day" OR "father\'s day" OR holiday OR seasonal OR gift)',
        ],
    ]


def generic_report_query_blocks() -> list[list[str]]:
    return [
        [
            'ranking_report',
            '(ranking OR rankings OR benchmark OR report OR survey OR study OR index OR data OR stats OR "market share" OR comparison)',
        ],
    ]


def query_blocks_from_text(raw_text: str, fallback: list[list[str]], prefix: str) -> list[list[str]]:
    blocks: list[list[str]] = []
    for index, line in enumerate(str(raw_text or '').splitlines(), start=1):
        normalized = line.strip()
        if normalized:
            blocks.append([f'{prefix}_{index}', normalized])
    return blocks or fallback


def create_minimal_country_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'survey_indicators'
    worksheet.cell(1, 1, '\u95ee\u5377\u6307\u6807\u7ef4\u62a4\u8868')
    worksheet.cell(3, 1, '\u7ef4\u5ea6')
    worksheet.cell(3, 2, '\u6807\u7b7e')
    worksheet.cell(3, 3, '\u82f1\u6587\u63cf\u8ff0')
    worksheet.cell(3, 4, '\u4e2d\u6587\u63cf\u8ff0')
    worksheet.cell(4, 1, 'Price')
    worksheet.cell(4, 2, 'Price')
    worksheet.cell(4, 3, 'Price perception and promotion relevance')
    worksheet.cell(4, 4, '\u4ef7\u683c\u611f\u77e5\u4e0e\u4fc3\u9500\u76f8\u5173\u6027')
    workbook.save(path)


def ensure_country_xlsx_file(config: dict[str, Any]) -> None:
    target_path = resolve_project_path(str(config['xlsx_path']), base_dir=BASE_DIR)
    if target_path.exists():
        return

    template_candidates: list[Path] = []
    for candidate in [BASE_DIR / 'source_survey.xlsx', BASE_DIR / 'source_survey_france.xlsx']:
        if candidate not in template_candidates:
            template_candidates.append(candidate)
    for existing_xlsx in BASE_DIR.glob('*.xlsx'):
        if existing_xlsx not in template_candidates:
            template_candidates.append(existing_xlsx)
    for existing_xlsx in BASE_DIR.glob('country_data/*/*.xlsx'):
        if existing_xlsx not in template_candidates:
            template_candidates.append(existing_xlsx)

    for template_path in template_candidates:
        if template_path == target_path or not template_path.exists():
            continue
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, target_path)
        return

    target_path.parent.mkdir(parents=True, exist_ok=True)
    create_minimal_country_xlsx(target_path)


def ensure_country_support_files(config: dict[str, Any]) -> None:
    initial_files = {
        'extra_sources_path': [],
        'adapter_configs_path': {},
        'site_credentials_path': {},
        'source_capability_cache_path': {'version': 1, 'entries': {}},
    }
    for key, default_payload in initial_files.items():
        path = resolve_project_path(str(config[key]), base_dir=BASE_DIR)
        if path.exists():
            continue
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(default_payload, ensure_ascii=False, indent=2), encoding='utf-8')
    ensure_country_xlsx_file(config)


def cleanup_custom_country_generated_files(country_code: str, config: dict[str, Any]) -> None:
    expected_name_options = {
        key: {
            str(default_country_file_paths(country_code).get(key) or '').strip(),
            str(legacy_country_file_names(country_code).get(key) or '').strip(),
        }
        for key in ['xlsx_path', 'extra_sources_path', 'adapter_configs_path', 'site_credentials_path', 'source_capability_cache_path']
    }
    removed_dirs: set[Path] = set()
    for key, expected_names in expected_name_options.items():
        configured_name = str(config.get(key) or '').strip()
        if not configured_name or configured_name not in expected_names:
            continue
        target_path = resolve_project_path(configured_name, base_dir=BASE_DIR)
        if target_path.is_file():
            try:
                target_path.unlink()
                removed_dirs.add(target_path.parent)
            except OSError:
                continue
    for directory in sorted(removed_dirs, key=lambda item: len(item.parts), reverse=True):
        if directory == BASE_DIR:
            continue
        try:
            directory.rmdir()
        except OSError:
            continue


def build_country_config_from_form(form) -> tuple[str, dict[str, Any], list[str]]:
    errors: list[str] = []
    raw_code = (form.get('country_code') or '').strip()
    country_code = normalize_new_country_code(raw_code)
    if not country_code:
        errors.append('\u56fd\u5bb6\u4ee3\u7801\u4e0d\u80fd\u4e3a\u7a7a\uff1b\u5efa\u8bae\u4f7f\u7528\u5c0f\u5199\u82f1\u6587\uff0c\u4f8b\u5982 germany\u3001italy\u3001spain\u3002')
    if country_code in COUNTRY_CONFIGS:
        errors.append('\u8fd9\u4e2a\u56fd\u5bb6\u4ee3\u7801\u5df2\u7ecf\u662f\u5185\u7f6e\u914d\u7f6e\uff0c\u8bf7\u6362\u4e00\u4e2a\u65b0\u4ee3\u7801\u3002')

    label = (form.get('country_label') or '').strip()
    if not label:
        errors.append('\u8bf7\u586b\u5199\u56fd\u5bb6\u663e\u793a\u540d\u3002')

    google_hl = (form.get('google_news_hl') or '').strip().lower() or 'en'
    google_gl = (form.get('google_news_gl') or '').strip().upper()
    if not google_gl:
        google_gl = country_code[:2].upper() if len(country_code) >= 2 else country_code.upper()
    google_ceid = (form.get('google_news_ceid') or '').strip() or f'{google_gl}:{google_hl}'
    bing_market = (form.get('bing_news_market') or '').strip() or f'{google_hl}-{google_gl}'
    timezone_name = (form.get('timezone') or '').strip() or 'UTC'

    market_terms = parse_multiline_items(form.get('market_terms') or '')
    if not market_terms:
        market_terms = [country_code, label]
    market_search_block = (form.get('market_search_block') or '').strip() or build_market_search_block(market_terms)
    platform_labels = parse_multiline_items(form.get('platform_labels') or '')
    platform_aliases = parse_platform_alias_text(form.get('platform_aliases') or '')

    output_slug = (form.get('output_slug') or '').strip() or country_code
    output_slug = normalize_new_country_code(output_slug) or country_code
    default_paths = default_country_file_paths(country_code)
    xlsx_path = normalize_project_relative_path((form.get('xlsx_path') or '').strip() or default_paths['xlsx_path'])
    include_xlsx_sources = str(form.get('include_xlsx_sources') or '').strip().lower() in {'1', 'true', 'on', 'yes'}
    promo_query_blocks = query_blocks_from_text(
        form.get('promo_search_keywords') or '',
        generic_promo_query_blocks(),
        'promo',
    )
    report_query_blocks = query_blocks_from_text(
        form.get('report_search_keywords') or '',
        generic_report_query_blocks(),
        'report',
    )
    platform_search_term_overrides: dict[str, list[str]] = {}
    for platform in platform_labels:
        platform_search_term_overrides[platform] = platform_aliases.get(platform) or generate_platform_aliases(
            platform,
            country_code=country_code,
            country_label=label,
            google_gl=google_gl,
            market_terms=market_terms,
        )
    for platform, aliases in platform_aliases.items():
        if platform not in platform_search_term_overrides:
            platform_search_term_overrides[platform] = unique_texts([platform, *aliases])

    config = {
        'code': country_code,
        'label': label,
        'consumer_label': (form.get('consumer_label') or '').strip() or f'{label}\u6d88\u8d39\u8005',
        'market_label': (form.get('market_label') or '').strip() or f'{label}\u5e02\u573a',
        'timezone': timezone_name,
        'google_news_hl': google_hl,
        'google_news_gl': google_gl,
        'google_news_ceid': google_ceid,
        'bing_news_market': bing_market,
        'market_terms': market_terms,
        'market_search_block': market_search_block,
        'xlsx_path': xlsx_path,
        'include_xlsx_sources': include_xlsx_sources,
        'extra_sources_path': default_paths['extra_sources_path'],
        'adapter_configs_path': default_paths['adapter_configs_path'],
        'site_credentials_path': default_paths['site_credentials_path'],
        'source_capability_cache_path': default_paths['source_capability_cache_path'],
        'output_slug': output_slug,
        'legacy_output_prefixes': [],
        'app_title': '\u65b0\u95fb\u8d44\u8baf\u6293\u53d6\u5de5\u5177',
        'promo_search_query_blocks': promo_query_blocks,
        'related_news_query_blocks': promo_query_blocks,
        'report_query_blocks': report_query_blocks,
        'dimension_search_term_overrides': {},
        'platform_display_overrides': {},
        'platform_search_term_overrides': platform_search_term_overrides,
        'default_platform_search_term_overrides': platform_search_term_overrides,
        'available_platform_labels': platform_labels,
        'platform_alias_exclude_tokens': [],
        'official_sources': parse_official_sources(form.get('official_sources') or ''),
        'rakuten_press_path_date_pattern': r'/(\d{4})/(\d{2})/(\d{2})/',
        'rakuten_press_date_group_mode': 'ymd',
        'tiktok_newsroom_lang': (form.get('tiktok_newsroom_lang') or '').strip() or google_hl,
    }
    return country_code, config, errors

def normalize_ai_string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [
            str(item).strip()
            for item in value
            if str(item or '').strip()
        ]
    if isinstance(value, str):
        return parse_multiline_items(value)
    return []


def normalize_ai_query_lines(value: Any) -> list[str]:
    if isinstance(value, list):
        return [
            str(item).strip()
            for item in value
            if str(item or '').strip()
        ]
    if isinstance(value, str):
        return [
            line.strip()
            for line in value.splitlines()
            if line.strip()
        ]
    return []


def ensure_country_ai_query_lines(
    value: Any,
    fallback_blocks: list[list[str]],
    *,
    minimum_lines: int | None = None,
) -> list[str]:
    lines = unique_texts(normalize_ai_query_lines(value))
    fallback_lines = unique_texts([
        str(block[1]).strip()
        for block in fallback_blocks
        if isinstance(block, (list, tuple)) and len(block) >= 2 and str(block[1]).strip()
    ])
    target_count = max(0, minimum_lines if minimum_lines is not None else len(fallback_lines))
    if not lines:
        return fallback_lines[:target_count] if target_count else fallback_lines
    if target_count <= 0:
        return lines
    for line in fallback_lines:
        if len(lines) >= target_count:
            break
        if line not in lines:
            lines.append(line)
    return lines


def normalize_ai_official_sources(value: Any) -> str:
    lines: list[str] = []
    if isinstance(value, dict):
        iterable = [
            {
                'platform': platform,
                **source,
            }
            for platform, source in value.items()
            if isinstance(source, dict)
        ]
    elif isinstance(value, list):
        iterable = value
    else:
        iterable = []

    for item in iterable:
        if not isinstance(item, dict):
            continue
        platform = str(item.get('platform') or item.get('name') or '').strip()
        label = str(item.get('label') or item.get('source_label') or '').strip()
        url = str(item.get('url') or item.get('source_url') or '').strip()
        if platform and label and url:
            line = f'{platform} | {label} | {url}'
            if line not in lines:
                lines.append(line)
    return '\n'.join(lines)


def normalize_ai_platform_aliases(
    value: Any,
    platform_labels: list[str],
    *,
    country_code: str = '',
    google_gl: str = '',
    market_terms: list[str] | None = None,
) -> str:
    alias_map: dict[str, list[str]] = {}
    if isinstance(value, dict):
        for platform, aliases in value.items():
            normalized_platform = str(platform or '').strip()
            if not normalized_platform:
                continue
            alias_map[normalized_platform] = normalize_ai_string_list(aliases)
    elif isinstance(value, list):
        for item in value:
            if not isinstance(item, dict):
                continue
            platform = str(item.get('platform') or item.get('label') or item.get('name') or '').strip()
            if not platform:
                continue
            alias_map[platform] = normalize_ai_string_list(item.get('aliases') or item.get('terms'))
    for platform in platform_labels:
        semantic_aliases = build_semantic_platform_aliases(
            platform,
            country_code=country_code or 'global',
            google_gl=google_gl,
            market_terms=market_terms or [],
        )
        alias_map[platform] = unique_texts([platform, *(alias_map.get(platform) or []), *semantic_aliases])
    return format_platform_alias_text(alias_map, platform_labels)


def build_platform_alias_suggestion_messages(
    *,
    country_code: str,
    platforms: list[str],
    current_alias_text: str,
) -> list[dict[str, str]]:
    country = get_country_config(country_code)
    return [
        {
            'role': 'system',
            'content': (
                'You generate search aliases for ecommerce platforms. Return JSON only. '
                'Keep platform keys unchanged and provide practical local-language aliases. '
                'Do not add platforms that are not requested.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_label': country.get('label', country_code),
                    'platforms': platforms,
                    'current_alias_text': current_alias_text,
                    'expected_schema': {'aliases': {'Platform': ['alias1', 'alias2']}},
                },
                ensure_ascii=False,
            ),
        },
    ]

def suggest_platform_aliases_with_ai(form) -> dict[str, Any]:
    country_code = normalize_country_request(form.get('country_code'))
    api_settings = extract_survey_api_settings_from_form(form)
    api_url = api_settings.get('survey_api_url') or ''
    api_key = api_settings.get('survey_api_key') or ''
    api_model = api_settings.get('survey_api_model') or ''
    if not api_url or not api_key or not api_model:
        raise RuntimeError('Please fill AI API URL, Key and model first.')

    existing_alias_map = parse_platform_alias_text(form.get('platform_aliases_text') or '')
    selected_known_platforms = [
        item.strip()
        for item in form.getlist('news_builtin_platforms')
        if item.strip()
    ]
    custom_platforms = parse_platform_text(form.get('news_custom_platforms') or '')
    target_platforms = unique_texts(selected_known_platforms + custom_platforms)
    if not target_platforms:
        target_platforms = xlsx_source_test.list_available_platform_labels(country_code)
    for platform in target_platforms:
        existing_alias_map.setdefault(platform, [platform])

    payload = xlsx_source_test.call_survey_filter_api(
        build_platform_alias_suggestion_messages(
            country_code=country_code,
            target_platforms=target_platforms,
            existing_alias_map=existing_alias_map,
        ),
        api_url,
        api_key,
        api_model,
    )
    persist_global_survey_api_settings(api_settings)
    if not isinstance(payload, dict):
        raise RuntimeError('AI returned an invalid JSON object.')

    suggested_alias_text = normalize_ai_platform_aliases(
        payload.get('platform_aliases') or payload.get('aliases') or payload,
        target_platforms,
        country_code=country_code,
        google_gl=str(get_country_config(country_code).get('google_news_gl') or '').upper(),
        market_terms=[str(item) for item in get_country_config(country_code).get('market_terms', []) if str(item).strip()],
    )
    suggested_alias_map = parse_platform_alias_text(suggested_alias_text)
    merged_alias_map: dict[str, list[str]] = {
        platform: unique_texts([platform, *aliases])
        for platform, aliases in existing_alias_map.items()
        if platform.strip()
    }
    for platform, aliases in suggested_alias_map.items():
        merged_alias_map[platform] = unique_texts([platform, *merged_alias_map.get(platform, []), *aliases])
    merged_alias_map = canonicalize_platform_alias_map(merged_alias_map, country_code)
    known_platforms = xlsx_source_test.list_available_platform_labels(country_code)
    next_text = format_platform_alias_text(merged_alias_map, unique_texts([*known_platforms, *target_platforms]))
    backup_platform_aliases_previous_value(country_code, next_text)
    save_country_config_patch(country_code, {'platform_search_term_overrides': merged_alias_map})

    return {
        'country_code': country_code,
        'platform_count': len(merged_alias_map),
        'optimized_platforms': target_platforms,
        'text': next_text,
        'api_url': api_url,
        'api_model': api_model,
    }


def require_survey_api_settings_from_form(form) -> dict[str, str]:
    api_settings = extract_survey_api_settings_from_form(form)
    api_url = api_settings.get('survey_api_url') or ''
    api_key = api_settings.get('survey_api_key') or ''
    api_model = api_settings.get('survey_api_model') or ''
    if not api_url or not api_key or not api_model:
        raise RuntimeError('Please fill AI API URL, Key and model first.')
    return api_settings


def normalize_ai_text_block(value: Any) -> str:
    if isinstance(value, list):
        return '\n'.join(str(item or '').strip() for item in value if str(item or '').strip())
    return str(value or '').strip()


def survey_indicator_prompt_brief(country_code: str, limit: int = 80) -> list[dict[str, str]]:
    try:
        indicators = xlsx_source_test.load_survey_indicators_from_xlsx(str(country_path(country_code, 'xlsx_path')))
    except Exception:
        return []
    rows: list[dict[str, str]] = []
    for indicator in indicators[:limit]:
        rows.append(
            {
                'dimension': indicator.dimension,
                'question_id': indicator.question_id,
                'prompt_en': indicator.prompt_en,
                'prompt_zh': indicator.prompt_zh,
            }
        )
    return rows


def build_keyword_suggestion_messages(
    *,
    country_code: str,
    related_keywords_text: str,
    report_keywords_text: str,
    mode: str,
) -> list[dict[str, str]]:
    country = get_country_config(country_code)
    return [
        {
            'role': 'system',
            'content': (
                'You optimize ecommerce news search keyword blocks. Return JSON only. '
                'Each keyword block should be one line and should improve recall while avoiding generic noise.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_label': country.get('label', country_code),
                    'mode': mode,
                    'related_news_search_keywords': related_keywords_text,
                    'report_search_keywords': report_keywords_text,
                    'expected_fields': ['related_news_search_keywords', 'report_search_keywords'],
                },
                ensure_ascii=False,
            ),
        },
    ]

def suggest_news_filter_keywords_with_ai(form) -> dict[str, Any]:
    country_code = normalize_country_request(form.get('country_code'))
    api_settings = require_survey_api_settings_from_form(form)
    related_keywords = (form.get('related_news_search_keywords') or '').strip()
    report_keywords = (form.get('report_search_keywords') or '').strip()
    if not related_keywords:
        related_keywords = xlsx_source_test.default_related_news_search_keywords_text(country_code)
    if not report_keywords:
        report_keywords = xlsx_source_test.default_report_search_keywords_text(country_code)

    payload = xlsx_source_test.call_survey_filter_api(
        build_keyword_suggestion_messages(
            country_code=country_code,
            related_news_keywords=related_keywords,
            report_keywords=report_keywords,
        ),
        api_settings['survey_api_url'],
        api_settings['survey_api_key'],
        api_settings['survey_api_model'],
    )
    persist_global_survey_api_settings(api_settings)
    if not isinstance(payload, dict):
        raise RuntimeError('AI returned an invalid JSON object.')

    suggested_related = normalize_ai_text_block(
        payload.get('related_news_search_keywords')
        or payload.get('promo_search_keywords')
        or payload.get('related_keywords')
    )
    suggested_report = normalize_ai_text_block(
        payload.get('report_search_keywords')
        or payload.get('report_keywords')
    )
    if not suggested_related or not suggested_report:
        raise RuntimeError('AI did not return related_news_search_keywords and report_search_keywords.')

    related_auto_split_count = xlsx_source_test.keyword_auto_split_count(suggested_related)
    report_auto_split_count = xlsx_source_test.keyword_auto_split_count(suggested_report)
    suggested_related = xlsx_source_test.normalize_keyword_blocks_for_storage(suggested_related)
    suggested_report = xlsx_source_test.normalize_keyword_blocks_for_storage(suggested_report)

    existing_state = read_country_settings(country_code).get('news_filter_state')
    news_filter_state = build_news_filter_state(existing_state if isinstance(existing_state, dict) else None, country_code=country_code)
    news_filter_state['related_news_search_keywords'] = suggested_related
    news_filter_state['report_search_keywords'] = suggested_report
    news_filter_state = build_news_filter_state(news_filter_state, country_code=country_code)
    persist_app_settings(country_code=country_code, news_filter_state=news_filter_state)
    related_block_count = len(xlsx_source_test.normalize_related_news_search_keyword_blocks(
        news_filter_state['related_news_search_keywords'],
        country_code,
        xlsx_source_test.DEFAULT_RECALL_MODE,
    ))
    default_related_block_count = len(xlsx_source_test.normalize_related_news_search_keyword_blocks('', country_code, xlsx_source_test.DEFAULT_RECALL_MODE))
    keyword_warning = ''
    if related_block_count < max(1, default_related_block_count // 2):
        keyword_warning = 'Keyword block count is low and may reduce recall.'
    return {
        'country_code': country_code,
        'related_news_search_keywords': news_filter_state['related_news_search_keywords'],
        'report_search_keywords': news_filter_state['report_search_keywords'],
        'related_news_keyword_auto_split_count': related_auto_split_count,
        'report_keyword_auto_split_count': report_auto_split_count,
        'related_news_search_keyword_block_count': related_block_count,
        'keyword_warning': keyword_warning,
        'api_url': api_settings['survey_api_url'],
        'api_model': api_settings['survey_api_model'],
    }


def build_survey_prompt_suggestion_messages(
    *,
    country_code: str,
    current_prompt: str,
) -> list[dict[str, str]]:
    country = get_country_config(country_code)
    return [
        {
            'role': 'system',
            'content': (
                'You improve an AI survey-filter prompt for ecommerce market news. Return JSON only. '
                'The prompt should keep relevant platform, NPS, consumer perception, regulation, service, trend, and industry-impact news.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_label': country.get('label', country_code),
                    'current_prompt': current_prompt,
                    'expected_field': 'survey_system_prompt',
                },
                ensure_ascii=False,
            ),
        },
    ]

def suggest_news_filter_survey_prompt_with_ai(form) -> dict[str, Any]:
    country_code = normalize_country_request(form.get('country_code'))
    api_settings = require_survey_api_settings_from_form(form)
    current_prompt = (form.get('survey_system_prompt') or '').strip() or xlsx_source_test.default_survey_ai_system_prompt(country_code)

    payload = xlsx_source_test.call_survey_filter_api(
        build_survey_prompt_suggestion_messages(
            country_code=country_code,
            current_prompt=current_prompt,
        ),
        api_settings['survey_api_url'],
        api_settings['survey_api_key'],
        api_settings['survey_api_model'],
    )
    persist_global_survey_api_settings(api_settings)
    if not isinstance(payload, dict):
        raise RuntimeError('AI returned an invalid JSON object.')
    suggested_prompt = str(
        payload.get('survey_system_prompt')
        or payload.get('system_prompt')
        or payload.get('prompt')
        or ''
    ).strip()
    if not suggested_prompt:
        raise RuntimeError('AI did not return survey_system_prompt.')

    existing_state = read_country_settings(country_code).get('news_filter_state')
    news_filter_state = build_news_filter_state(existing_state if isinstance(existing_state, dict) else None, country_code=country_code)
    news_filter_state['survey_system_prompt'] = suggested_prompt
    news_filter_state = build_news_filter_state(news_filter_state, country_code=country_code)
    persist_app_settings(country_code=country_code, news_filter_state=news_filter_state)
    return {
        'country_code': country_code,
        'survey_system_prompt': news_filter_state['survey_system_prompt'],
        'survey_system_prompt_source': news_filter_state['survey_system_prompt_source'],
        'survey_system_prompt_source_label': news_filter_state['survey_system_prompt_source_label'],
        'api_url': api_settings['survey_api_url'],
        'api_model': api_settings['survey_api_model'],
    }


def default_country_ai_prompt_text() -> str:
    return (
        'You are helping configure a market news crawler country profile. '
        'Return JSON only. Include country_code, country_label, google_news_gl, '
        'market_terms, default_platforms, platform_aliases, related_news_keywords, '
        'report_keywords, official_sources, and survey_prompt when useful.'
    )

def build_country_suggestion_messages(country_name: str, form) -> list[dict[str, str]]:
    return [
        {
            'role': 'system',
            'content': (
                'You help configure a country for an ecommerce news crawler. Return JSON only. '
                'Use concise country settings, market terms, platform labels, aliases, and search keyword blocks.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_name': country_name,
                    'existing_form': {key: form.get(key) for key in form.keys()},
                    'expected_fields': [
                        'country_code', 'country_label', 'consumer_label', 'market_label',
                        'google_news_hl', 'google_news_gl', 'google_news_ceid', 'bing_news_market',
                        'market_terms', 'platform_labels', 'platform_aliases',
                        'promo_search_keywords', 'report_search_keywords', 'official_sources',
                    ],
                },
                ensure_ascii=False,
            ),
        },
    ]

def suggest_country_form_fields_with_ai(form) -> dict[str, Any]:
    country_name = (form.get('country_label') or form.get('country_name') or '').strip()
    if not country_name:
        raise RuntimeError('Please enter a country name first.')

    api_settings = extract_survey_api_settings_from_form(form)
    saved_api_settings = read_global_survey_api_settings()
    api_url = api_settings.get('survey_api_url') or saved_api_settings.get('survey_api_url') or ''
    api_key = api_settings.get('survey_api_key') or saved_api_settings.get('survey_api_key') or ''
    api_model = api_settings.get('survey_api_model') or saved_api_settings.get('survey_api_model') or ''
    if not api_url or not api_key or not api_model:
        raise RuntimeError('Please fill AI API URL, Key and model first.')

    payload = xlsx_source_test.call_survey_filter_api(
        build_country_suggestion_messages(country_name, form),
        api_url,
        api_key,
        api_model,
    )
    persist_global_survey_api_settings({
        'survey_api_url': api_url,
        'survey_api_key': api_key,
        'survey_api_model': api_model,
    })
    if not isinstance(payload, dict):
        raise RuntimeError('AI returned an invalid JSON object.')

    return {
        'country_code': str(payload.get('country_code') or form.get('country_code') or '').strip(),
        'country_label': str(payload.get('country_label') or country_name).strip(),
        'consumer_label': str(payload.get('consumer_label') or '').strip(),
        'market_label': str(payload.get('market_label') or country_name).strip(),
        'google_news_hl': str(payload.get('google_news_hl') or '').strip(),
        'google_news_gl': str(payload.get('google_news_gl') or '').strip(),
        'google_news_ceid': str(payload.get('google_news_ceid') or '').strip(),
        'bing_news_market': str(payload.get('bing_news_market') or '').strip(),
        'market_terms': normalize_ai_string_list(payload.get('market_terms')),
        'platform_labels': normalize_ai_string_list(payload.get('platform_labels') or payload.get('default_platforms')),
        'platform_aliases': normalize_ai_text_block(payload.get('platform_aliases') or payload.get('aliases')),
        'promo_search_keywords': normalize_ai_text_block(payload.get('promo_search_keywords') or payload.get('related_news_keywords')),
        'report_search_keywords': normalize_ai_text_block(payload.get('report_search_keywords') or payload.get('report_keywords')),
        'official_sources': normalize_ai_official_sources(payload.get('official_sources')),
        'survey_prompt': str(payload.get('survey_prompt') or payload.get('survey_system_prompt') or '').strip(),
        'api_url': api_url,
        'api_model': api_model,
    }

def build_news_platform_ui_state(
    platforms_text: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, Any]:
    known_platforms = xlsx_source_test.list_available_platform_labels(country_code)
    if platforms_text is None:
        return {
            'known_platforms': known_platforms,
            'selected_known_platforms': known_platforms,
            'custom_platforms_text': '',
        }

    selected_known, custom_platforms = split_news_platforms_for_country(platforms_text, country_code)

    return {
        'known_platforms': known_platforms,
        'selected_known_platforms': selected_known,
        'custom_platforms_text': '\n'.join(custom_platforms),
    }


UI_PLATFORM_DISPLAY_OVERRIDES = {
    'amazon': 'Amazon',
    'ebay': 'eBay',
    'ig': 'Instagram',
    'instagram': 'Instagram',
    'shein': 'SHEIN',
    'temu': 'TEMU',
    'tts': 'TikTok Shop',
    'tiktok shop': 'TikTok Shop',
}


def user_platform_display_label(platform: str) -> str:
    text = str(platform or '').strip()
    if not text:
        return ''
    pieces = re.split(r'(\s*(?:\||,|/|、|，)\s*)', text)
    display_parts: list[str] = []
    for piece in pieces:
        if not piece:
            continue
        if re.fullmatch(r'\s*(?:\||,|/|、|，)\s*', piece):
            display_parts.append(piece)
            continue
        token = piece.strip()
        mapped = UI_PLATFORM_DISPLAY_OVERRIDES.get(token.lower())
        display_parts.append(piece.replace(token, mapped) if mapped else piece)
    return ''.join(display_parts).strip()


def user_platform_display_labels(platforms: list[str]) -> dict[str, str]:
    return {platform: user_platform_display_label(platform) for platform in platforms}


def platform_display_options(platforms: list[str]) -> list[dict[str, str]]:
    return [
        {'value': str(platform), 'label': user_platform_display_label(str(platform))}
        for platform in platforms
    ]


def enrich_article_browser_display_state(state: dict[str, Any]) -> dict[str, Any]:
    state['available_platform_options'] = platform_display_options(state.get('available_platforms') or [])
    for row in state.get('rows') or []:
        if isinstance(row, dict):
            row['platform_label_display'] = user_platform_display_label(
                str(row.get('platform_label') or row.get('platform') or '')
            )
    return state


def build_news_form_state(
    overrides: dict[str, Any] | None = None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, Any]:
    state = {
        'date_mode': 'days',
        'days': '7',
        'start_date': '',
        'end_date': '',
        'translate_to': 'zh-CN',
        'output_dir': 'outputs',
        'recall_mode': xlsx_source_test.DEFAULT_RECALL_MODE,
        'news_platforms_text': default_news_platforms_text(country_code),
        'news_sides': DEFAULT_NEWS_SIDES.copy(),
    }
    persisted = read_country_settings(country_code).get('news_form_state')
    if isinstance(persisted, dict):
        state.update(persisted)
    if overrides:
        state.update(overrides)

    if state.get('date_mode') not in {'days', 'range'}:
        state['date_mode'] = 'days'
    try:
        state['days'] = str(max(1, int(str(state.get('days') or '7').strip() or '7')))
    except (TypeError, ValueError):
        state['days'] = '7'
    state['start_date'] = str(state.get('start_date') or '').strip()
    state['end_date'] = str(state.get('end_date') or '').strip()
    if state.get('translate_to') not in {'zh-CN', 'en'}:
        state['translate_to'] = 'zh-CN'
    state['output_dir'] = str(state.get('output_dir') or 'outputs').strip() or 'outputs'
    state['recall_mode'] = str(state.get('recall_mode') or xlsx_source_test.DEFAULT_RECALL_MODE).strip().lower()
    if state['recall_mode'] not in xlsx_source_test.RECALL_MODE_CHOICES:
        state['recall_mode'] = xlsx_source_test.DEFAULT_RECALL_MODE
    if state.get('news_platforms_text') is None:
        state['news_platforms_text'] = default_news_platforms_text(country_code)
    else:
        state['news_platforms_text'] = normalize_news_platforms_text_for_country(
            str(state.get('news_platforms_text') or ''),
            country_code,
        )

    normalized_sides = [
        item.strip()
        for item in (state.get('news_sides') or [])
        if str(item).strip() in {'media', 'buyer', 'seller'}
    ]
    state['news_sides'] = normalized_sides or DEFAULT_NEWS_SIDES.copy()
    return state


def build_news_filter_state(
    overrides: dict[str, Any] | None = None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, Any]:
    state = {
        'survey_filter_mode': xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE,
        'survey_api_url': '',
        'survey_api_key': '',
        'survey_api_model': '',
        'survey_system_prompt': xlsx_source_test.default_survey_ai_system_prompt(country_code),
        'promo_search_enabled': False,
        'related_news_search_enabled': False,
        'report_ranking_search_enabled': False,
        'promo_search_engine': 'both',
        'promo_search_keywords': xlsx_source_test.default_promo_search_keywords_text(country_code),
        'related_news_search_keywords': xlsx_source_test.default_related_news_search_keywords_text(country_code),
        'report_search_keywords': xlsx_source_test.default_report_search_keywords_text(country_code),
    }
    persisted = read_country_settings(country_code).get('news_filter_state')
    if isinstance(persisted, dict):
        state.update(persisted)
    state.update(read_global_survey_api_settings())
    state = apply_overrides_preserving_blank_api_fields(
        state,
        overrides,
        preserved_fields=['survey_api_url', 'survey_api_key', 'survey_api_model'],
    )
    legacy_promo_enabled = bool(state.get('promo_search_enabled'))
    if legacy_promo_enabled and not bool(state.get('related_news_search_enabled')) and not bool(state.get('report_ranking_search_enabled')):
        state['related_news_search_enabled'] = True
        state['report_ranking_search_enabled'] = True
    if state.get('survey_filter_mode') not in {'keyword', 'ai'}:
        state['survey_filter_mode'] = xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE
    state['survey_system_prompt'] = xlsx_source_test.normalize_survey_system_prompt(state.get('survey_system_prompt'), country_code)
    state['survey_system_prompt_source'] = xlsx_source_test.survey_system_prompt_source(
        state.get('survey_system_prompt'),
        country_code,
    )
    state['survey_system_prompt_source_label'] = xlsx_source_test.survey_system_prompt_source_label(
        state.get('survey_system_prompt'),
        country_code,
    )
    state['related_news_search_enabled'] = bool(state.get('related_news_search_enabled'))
    state['report_ranking_search_enabled'] = bool(state.get('report_ranking_search_enabled'))
    state['promo_search_enabled'] = bool(state.get('related_news_search_enabled') or state.get('report_ranking_search_enabled'))
    if state.get('promo_search_engine') not in xlsx_source_test.PROMO_SEARCH_ENGINE_CHOICES:
        state['promo_search_engine'] = 'both'
    state['promo_search_keywords'] = xlsx_source_test.normalize_promo_search_keywords_text(
        state.get('promo_search_keywords'),
        country_code,
    )
    legacy_related_keywords = state.get('related_news_search_keywords')
    legacy_promo_keywords = state.get('promo_search_keywords')
    if (
        not xlsx_source_test.clean_text(legacy_related_keywords)
        and xlsx_source_test.clean_text(legacy_promo_keywords)
        and xlsx_source_test.clean_text(legacy_promo_keywords) != xlsx_source_test.clean_text(xlsx_source_test.default_promo_search_keywords_text(country_code))
    ):
        legacy_related_keywords = legacy_promo_keywords
    state['related_news_search_keywords'] = xlsx_source_test.normalize_related_news_search_keywords_text(
        legacy_related_keywords,
        country_code,
    )
    state['related_news_search_keywords'] = xlsx_source_test.normalize_keyword_blocks_for_storage(
        state['related_news_search_keywords']
    )
    state['report_search_keywords'] = xlsx_source_test.normalize_report_search_keywords_text(
        state.get('report_search_keywords'),
        country_code,
    )
    state['report_search_keywords'] = xlsx_source_test.normalize_keyword_blocks_for_storage(
        state['report_search_keywords']
    )
    return state


def load_source_manager_records(country_code: str = DEFAULT_COUNTRY_CODE) -> list[dict[str, Any]]:
    source_platform_to_known_label: dict[str, str] = {}
    for known_platform in xlsx_source_test.list_available_platform_labels(country_code):
        _, source_platforms, _ = xlsx_source_test.resolve_requested_platforms([known_platform], country_code)
        for source_platform in source_platforms:
            source_platform_to_known_label.setdefault(str(source_platform), known_platform)

    records = source_manager.load_source_registry(str(country_path(country_code, 'extra_sources_path')))
    for row in records:
        raw_selector_label = xlsx_source_test.display_platform_label(str(row.get('platform') or ''), country_code)
        row['platform_selector_label'] = raw_selector_label
        row['platform_selector_display_label'] = user_platform_display_label(raw_selector_label)
        row['platform_filter_label'] = source_platform_to_known_label.get(str(row.get('platform') or ''), '')
        row['platform_filter_display_label'] = user_platform_display_label(row['platform_filter_label'])
        if str(row.get('side') or '').strip() == 'media':
            row['platform_display'] = user_platform_display_label(
                source_manager.display_platform(row.get('platform', ''), row.get('side', ''))
            )
        else:
            row['platform_display'] = user_platform_display_label(
                xlsx_source_test.display_platform_label(str(row.get('platform') or ''), country_code)
            )
    records.sort(key=lambda row: (row['platform'], row['side'], row['source_url']))
    return records


def parse_bool_arg(value: Any) -> bool:
    normalized = str(value or '').strip().lower()
    return normalized in {'1', 'true', 'yes', 'on'}


def clamp_int(value: float, minimum: int, maximum: int) -> int:
    return max(minimum, min(maximum, int(round(value))))


def format_seconds(total_seconds: int | float | None) -> str:
    if total_seconds is None:
        return '00:00'
    seconds = max(0, int(round(float(total_seconds))))
    hours, remainder = divmod(seconds, 3600)
    minutes, remain = divmod(remainder, 60)
    if hours:
        return f'{hours:02d}:{minutes:02d}:{remain:02d}'
    return f'{minutes:02d}:{remain:02d}'

def read_job_timing_history() -> dict[str, list[dict[str, Any]]]:
    if not JOB_TIMING_HISTORY_PATH.exists():
        return {'news_runs': [], 'source_runs': []}
    try:
        payload = json.loads(JOB_TIMING_HISTORY_PATH.read_text(encoding='utf-8'))
    except Exception:
        return {'news_runs': [], 'source_runs': []}
    if not isinstance(payload, dict):
        return {'news_runs': [], 'source_runs': []}
    return {
        'news_runs': payload.get('news_runs') if isinstance(payload.get('news_runs'), list) else [],
        'source_runs': payload.get('source_runs') if isinstance(payload.get('source_runs'), list) else [],
    }


def write_job_timing_history(payload: dict[str, list[dict[str, Any]]]) -> None:
    JOB_TIMING_HISTORY_PATH.write_text(
        json.dumps(
            {
                'version': 1,
                'updated_at': int(time.time()),
                'news_runs': payload.get('news_runs', [])[-TIMING_HISTORY_LIMIT:],
                'source_runs': payload.get('source_runs', [])[-TIMING_HISTORY_LIMIT:],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding='utf-8',
    )


def weighted_average(values: list[tuple[float, float]]) -> float | None:
    total_weight = 0.0
    weighted_total = 0.0
    for value, weight in values:
        if value <= 0 or weight <= 0:
            continue
        total_weight += weight
        weighted_total += value * weight
    if total_weight <= 0:
        return None
    return weighted_total / total_weight


def news_history_weight(record: dict[str, Any], *, total_sites: int, promo_enabled: bool, ai_enabled: bool) -> float:
    weight = 1.0
    if bool(record.get('promo_search_enabled')) == promo_enabled:
        weight += 1.2
    if bool(record.get('ai_filter_enabled')) == ai_enabled:
        weight += 1.2
    record_sites = max(1, int(record.get('total_sites') or 0))
    if total_sites > 0:
        site_ratio = min(record_sites, total_sites) / max(record_sites, total_sites)
        weight += site_ratio
    return weight


def source_history_weight(record: dict[str, Any], *, action: str, skip_api: bool) -> float:
    weight = 1.0
    if str(record.get('action') or '') == action:
        weight += 1.5
    if bool(record.get('skip_api')) == skip_api:
        weight += 0.8
    return weight


def average_news_stage_seconds(
    records: list[dict[str, Any]],
    stage_name: str,
    *,
    total_sites: int,
    promo_enabled: bool,
    ai_enabled: bool,
) -> float | None:
    return weighted_average(
        [
            (
                float((record.get('stage_timings') or {}).get(stage_name) or 0),
                news_history_weight(record, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled),
            )
            for record in records
        ]
    )


def average_news_crawl_per_site_seconds(
    records: list[dict[str, Any]],
    *,
    total_sites: int,
    promo_enabled: bool,
    ai_enabled: bool,
) -> float | None:
    values: list[tuple[float, float]] = []
    for record in records:
        stage_timings = record.get('stage_timings') or {}
        record_sites = max(1, int(record.get('total_sites') or 0))
        crawl_seconds = float(stage_timings.get('crawl_site') or 0)
        if crawl_seconds <= 0:
            continue
        values.append(
            (
                crawl_seconds / record_sites,
                news_history_weight(record, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled),
            )
        )
    return weighted_average(values)


def average_news_promo_per_task_seconds(
    records: list[dict[str, Any]],
    *,
    total_sites: int,
    promo_enabled: bool,
    ai_enabled: bool,
) -> float | None:
    values: list[tuple[float, float]] = []
    for record in records:
        task_count = max(1, int(record.get('promo_task_count') or 0))
        promo_seconds = float((record.get('stage_timings') or {}).get('promo_search') or 0)
        if promo_seconds <= 0:
            continue
        values.append(
            (
                promo_seconds / task_count,
                news_history_weight(record, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled),
            )
        )
    return weighted_average(values)


def average_news_capability_refresh_per_site_seconds(
    records: list[dict[str, Any]],
    *,
    total_sites: int,
    promo_enabled: bool,
    ai_enabled: bool,
) -> float | None:
    values: list[tuple[float, float]] = []
    for record in records:
        record_sites = max(1, int(record.get('total_sites') or 0))
        refresh_seconds = float((record.get('stage_timings') or {}).get('capability_refresh') or 0)
        if refresh_seconds <= 0:
            continue
        values.append(
            (
                refresh_seconds / record_sites,
                news_history_weight(record, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled),
            )
        )
    return weighted_average(values)


def average_source_stage_seconds(records: list[dict[str, Any]], stage_name: str, *, action: str, skip_api: bool) -> float | None:
    return weighted_average(
        [
            (
                float((record.get('stage_timings') or {}).get(stage_name) or 0),
                source_history_weight(record, action=action, skip_api=skip_api),
            )
            for record in records
        ]
    )


def build_news_timing_profile(argv: list[str]) -> dict[str, Any]:
    args = xlsx_source_test.parse_args(argv)
    total_sites, _, _, selected_platform_count = estimate_news_site_counts(argv)
    promo_enabled = bool(getattr(args, 'promo_search', False) or getattr(args, 'search_related_news', False) or getattr(args, 'search_report_ranking', False))
    ai_enabled = getattr(args, 'survey_filter_mode', xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE) == 'ai'
    promo_task_count = len(
        xlsx_source_test.build_promo_search_tasks(
            args.platforms,
            args.promo_search_engine,
            getattr(args, 'promo_search_keywords', ''),
            survey_indicators=xlsx_source_test.load_survey_indicators_from_xlsx(args.xlsx),
            related_news_enabled=bool(getattr(args, 'search_related_news', False)),
            related_news_keywords_text=getattr(args, 'related_news_search_keywords', ''),
            report_ranking_enabled=bool(getattr(args, 'search_report_ranking', False)),
            report_keywords_text=getattr(args, 'report_search_keywords', ''),
            country_code=args.country,
        )
    ) if promo_enabled else 0

    history = read_job_timing_history().get('news_runs', [])
    fallback_setup = 5
    fallback_crawl = max(10, total_sites * 6)
    fallback_promo = 0 if not promo_enabled else max(10, 8 + promo_task_count * 2)
    fallback_finalize = max(
        28,
        24 + total_sites + (52 if ai_enabled else 18) + (12 if promo_enabled else 0) + max(0, promo_task_count // 3),
    )
    fallback_refresh = max(8, min(60, total_sites * 2))

    setup_avg = average_news_stage_seconds(history, 'setup', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    crawl_per_site_avg = average_news_crawl_per_site_seconds(history, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    promo_per_task_avg = average_news_promo_per_task_seconds(history, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    finalize_avg = average_news_stage_seconds(history, 'finalize', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    translation_avg = average_news_stage_seconds(history, 'translation', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    survey_filter_avg = average_news_stage_seconds(history, 'survey_filter', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    ai_dedupe_avg = average_news_stage_seconds(history, 'ai_dedupe', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    finalize_output_avg = average_news_stage_seconds(history, 'finalize_output', total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)
    refresh_per_site_avg = average_news_capability_refresh_per_site_seconds(history, total_sites=total_sites, promo_enabled=promo_enabled, ai_enabled=ai_enabled)

    setup_seconds = int(round(((setup_avg or fallback_setup) * 0.7) + (fallback_setup * 0.3)))
    crawl_seconds = int(round((((crawl_per_site_avg or (fallback_crawl / max(total_sites, 1))) * total_sites) * 0.75) + (fallback_crawl * 0.25)))
    promo_seconds = 0 if not promo_enabled else int(round((((promo_per_task_avg or (fallback_promo / max(promo_task_count, 1))) * promo_task_count) * 0.75) + (fallback_promo * 0.25)))
    finalize_seconds = int(round(((finalize_avg or fallback_finalize) * 0.7) + (fallback_finalize * 0.3)))
    refresh_seconds = int(round((((refresh_per_site_avg or (fallback_refresh / max(total_sites, 1))) * total_sites) * 0.65) + (fallback_refresh * 0.35)))

    if crawl_seconds > 0:
        crawl_seconds = min(crawl_seconds, max(10, 4 + total_sites * 5))
    if promo_enabled and promo_seconds > 0:
        promo_seconds = min(promo_seconds, max(10, 6 + promo_task_count * 2))
        finalize_seconds = max(finalize_seconds, promo_seconds + (36 if ai_enabled else 18))

    postprocess_total_seconds = max(12, finalize_seconds)
    if ai_enabled:
        fallback_translation_seconds = max(8, int(round(postprocess_total_seconds * 0.08)))
        fallback_survey_filter_seconds = max(18, int(round(postprocess_total_seconds * 0.84)))
        fallback_ai_dedupe_seconds = max(8, int(round(postprocess_total_seconds * 0.06)))
    else:
        fallback_translation_seconds = max(8, int(round(postprocess_total_seconds * 0.18)))
        fallback_survey_filter_seconds = max(12, int(round(postprocess_total_seconds * 0.58)))
        fallback_ai_dedupe_seconds = max(6, int(round(postprocess_total_seconds * 0.08)))
    fallback_output_seconds = max(2, postprocess_total_seconds - fallback_translation_seconds - fallback_survey_filter_seconds - fallback_ai_dedupe_seconds)

    def blend_stage_seconds(stage_avg: float | None, fallback_seconds: int, *, minimum: int) -> int:
        if stage_avg and stage_avg > 0:
            return max(minimum, int(round((stage_avg * 0.8) + (fallback_seconds * 0.2))))
        return max(minimum, fallback_seconds)

    translation_seconds = blend_stage_seconds(translation_avg, fallback_translation_seconds, minimum=8)
    survey_filter_seconds = blend_stage_seconds(survey_filter_avg, fallback_survey_filter_seconds, minimum=12 if not ai_enabled else 18)
    ai_dedupe_seconds = blend_stage_seconds(ai_dedupe_avg, fallback_ai_dedupe_seconds, minimum=4 if not ai_enabled else 6)
    finalize_output_seconds = blend_stage_seconds(finalize_output_avg, fallback_output_seconds, minimum=2)

    stages = {
        'setup': max(3, setup_seconds),
        'crawl_site': max(8, crawl_seconds),
        'promo_search': 0 if not promo_enabled else max(12, promo_seconds),
        'translation': translation_seconds,
        'survey_filter': survey_filter_seconds,
        'ai_dedupe': ai_dedupe_seconds,
        'finalize_output': finalize_output_seconds,
        'capability_refresh': max(4, refresh_seconds),
    }
    total_seconds = sum(seconds for seconds in stages.values() if seconds > 0)
    return {
        'task_kind': 'news',
        'order': NEWS_STAGE_ORDER,
        'stages': stages,
        'total_seconds': max(30, total_seconds),
        'total_sites': total_sites,
        'selected_platform_count': selected_platform_count,
        'promo_enabled': promo_enabled,
        'promo_task_count': promo_task_count,
        'ai_enabled': ai_enabled,
    }


def build_source_manager_timing_profile(source_state: dict[str, Any]) -> dict[str, Any]:
    action = str(source_state.get('source_action') or 'list')
    skip_api = bool(source_state.get('skip_api'))
    history = read_job_timing_history().get('source_runs', [])

    fallback_setup = 3
    if action == 'add':
        fallback_run = 8 if skip_api else 28
    elif action in {'remove', 'reactivate'}:
        fallback_run = 5
    else:
        fallback_run = 4

    setup_avg = average_source_stage_seconds(history, 'setup', action=action, skip_api=skip_api)
    run_avg = average_source_stage_seconds(history, 'run', action=action, skip_api=skip_api)
    setup_seconds = int(round(((setup_avg or fallback_setup) * 0.7) + (fallback_setup * 0.3)))
    run_seconds = int(round(((run_avg or fallback_run) * 0.7) + (fallback_run * 0.3)))

    stages = {
        'setup': max(2, setup_seconds),
        'run': max(2, run_seconds),
    }
    return {
        'task_kind': 'sources',
        'order': SOURCE_STAGE_ORDER,
        'stages': stages,
        'total_seconds': max(4, sum(stages.values())),
        'action': action,
        'skip_api': skip_api,
    }


def stage_expected_seconds(job: dict[str, Any], stage_name: str) -> int:
    timing_profile = job.get('timing_profile') or {}
    stage_map = timing_profile.get('stages') if isinstance(timing_profile, dict) else {}
    if isinstance(stage_map, dict):
        try:
            return max(1, int(stage_map.get(stage_name) or 0))
        except (TypeError, ValueError):
            return 1
    return 1


def stage_sequence(job: dict[str, Any]) -> list[str]:
    timing_profile = job.get('timing_profile') or {}
    order = timing_profile.get('order') if isinstance(timing_profile, dict) else None
    if isinstance(order, list) and order:
        return [str(item) for item in order]
    return NEWS_STAGE_ORDER if job.get('task_kind') == 'news' else SOURCE_STAGE_ORDER


def extract_progress_counts(text: str) -> tuple[int, int]:
    if not text:
        return 0, 0
    match = re.search(r'(\d+)\s*/\s*(\d+)', text)
    if not match:
        return 0, 0
    return int(match.group(1)), int(match.group(2))


def progress_signature(job: dict[str, Any]) -> str:
    stage = str(job.get('stage') or '')
    total_sites = int(job.get('total_sites') or 0)
    completed_sites = int(job.get('completed_sites') or 0)
    progress_message = str(job.get('progress_message') or '')
    current_site = str(job.get('current_site') or '')
    completed_units, total_units = extract_progress_counts(' '.join(part for part in [current_site, progress_message] if part))
    return '||'.join(
        [
            stage,
            str(completed_sites),
            str(total_sites),
            str(completed_units),
            str(total_units),
        ]
    )


def current_stage_elapsed_seconds(job: dict[str, Any], *, now_ts: float | None = None) -> int:
    current_time = now_ts or time.time()
    stage_started_at = float(job.get('stage_started_at') or job.get('started_at') or current_time)
    return max(0, int(current_time - stage_started_at))


def close_current_stage_timing(job: dict[str, Any], *, now_ts: float | None = None) -> None:
    stage_name = str(job.get('stage') or '')
    if not stage_name or stage_name in {'queued', 'done'}:
        return
    current_time = now_ts or time.time()
    stage_started_at = float(job.get('stage_started_at') or job.get('started_at') or current_time)
    elapsed = max(0, int(round(current_time - stage_started_at)))
    if elapsed <= 0:
        return
    stage_timings = job.setdefault('stage_timings', {})
    stage_timings[stage_name] = int(stage_timings.get(stage_name) or 0) + elapsed
    job['stage_started_at'] = current_time


def build_timing_summary_lines(summary: dict[str, Any]) -> list[str]:
    lines = ['[Timing]']
    total_seconds = int(summary.get('total_seconds') or 0)
    if total_seconds > 0:
        lines.append(f'Total elapsed: {format_seconds(total_seconds)}')
    stage_timings = summary.get('stage_timings') or {}
    order = NEWS_STAGE_ORDER if summary.get('task_kind') == 'news' else SOURCE_STAGE_ORDER
    stage_labels = {
        'setup': 'Setup',
        'crawl_site': 'Site crawl',
        'run': 'Run',
        'promo': 'Supplemental search',
        'translate': 'Translate',
        'survey_filter': 'Survey filter',
        'ai_dedupe': 'AI dedupe',
        'write_outputs': 'Write outputs',
        'capability_refresh': 'Capability refresh',
        'finalize': 'Finalize',
    }
    for stage_name in order:
        seconds = int(stage_timings.get(stage_name) or 0)
        if seconds <= 0:
            continue
        lines.append(f'- {stage_labels.get(stage_name, stage_name)}: {format_seconds(seconds)}')
    return lines

def append_job_timing_history(summary: dict[str, Any]) -> None:
    history = read_job_timing_history()
    if summary.get('task_kind') == 'news':
        history['news_runs'].append(summary)
    else:
        history['source_runs'].append(summary)
    write_job_timing_history(history)


def summarize_completed_job(job: dict[str, Any]) -> dict[str, Any]:
    total_seconds = max(1, int(round(float(job.get('finished_at') or time.time()) - float(job.get('started_at') or time.time()))))
    stage_timings = {
        str(key): int(value)
        for key, value in (job.get('stage_timings') or {}).items()
        if value
    }
    if job.get('task_kind') == 'news':
        args = xlsx_source_test.parse_args(job.get('argv') or [])
        timing_profile = job.get('timing_profile') or {}
        return {
            'task_kind': 'news',
            'completed_at': int(job.get('finished_at') or time.time()),
            'total_seconds': total_seconds,
            'total_sites': int(job.get('total_sites') or 0),
            'selected_platform_count': int(job.get('selected_platform_count') or 0),
            'promo_search_enabled': bool(getattr(args, 'promo_search', False)),
            'promo_task_count': int(timing_profile.get('promo_task_count') or 0),
            'ai_filter_enabled': getattr(args, 'survey_filter_mode', xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE) == 'ai',
            'stage_timings': stage_timings,
        }
    form_state = job.get('form_state') or {}
    return {
        'task_kind': 'sources',
        'completed_at': int(job.get('finished_at') or time.time()),
        'total_seconds': total_seconds,
        'action': str(form_state.get('source_action') or 'list'),
        'skip_api': bool(form_state.get('skip_api')),
        'stage_timings': stage_timings,
    }


def parse_float_value(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def parse_int_value(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default


def latest_source_yield_lookup(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, dict[str, Any]]:
    for row in list_recent_output_dirs(limit=12, country_code=country_code):
        output_path = Path(str(row.get('path') or ''))
        sources_csv = output_path / 'sources.csv'
        if not sources_csv.exists() or not sources_csv.is_file():
            continue
        lookup: dict[str, dict[str, Any]] = {}
        try:
            with sources_csv.open('r', encoding='utf-8-sig', newline='') as handle:
                reader = csv.DictReader(handle)
                for source_row in reader:
                    source_url = str(source_row.get('source_url') or '').strip()
                    if not source_url:
                        continue
                    cache_key = source_manager.capability_cache_key(source_url)
                    lookup[cache_key] = {
                        'last_run_id': output_path.name,
                        'last_source_status': str(source_row.get('status') or '').strip(),
                        'last_recent_article_count': parse_int_value(source_row.get('recent_article_count')),
                        'last_candidate_count': parse_int_value(source_row.get('candidate_count')),
                        'last_matched_brand_count': parse_int_value(source_row.get('matched_brand_count')),
                        'last_elapsed_seconds': parse_float_value(source_row.get('elapsed_seconds')),
                        'last_recent_yield': str(source_row.get('recent_yield') or '').strip(),
                        'last_zero_yield_reason': str(source_row.get('zero_yield_reason') or '').strip(),
                        'last_source_recommendation': str(source_row.get('source_recommendation') or '').strip(),
                        'last_search_endpoint_status': str(source_row.get('search_endpoint_status') or '').strip(),
                        'last_selector_match_count': parse_int_value(source_row.get('selector_match_count')),
                        'last_parsed_date_count': parse_int_value(source_row.get('parsed_date_count')),
                    }
        except Exception:
            continue
        if lookup:
            return lookup
    return {}


def build_source_record_cards(
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    show_inactive: bool = False,
    platform: str = '',
    side: str = '',
    source_url: str = '',
) -> list[dict[str, Any]]:
    records = source_manager.list_sources(
        str(country_path(country_code, 'extra_sources_path')),
        show_inactive,
        source_url=source_url or None,
        platform=platform or None,
        side=side or None,
    )
    capability_cache = source_manager.read_capability_cache(str(country_path(country_code, 'source_capability_cache_path')))
    source_yield_lookup = latest_source_yield_lookup(country_code)
    cards: list[dict[str, Any]] = []
    for record in records:
        card = dict(record)
        if str(record.get('side') or '').strip() == 'media':
            card['platform_display'] = user_platform_display_label(
                source_manager.display_platform(record.get('platform', ''), record.get('side', ''))
            )
        else:
            card['platform_display'] = user_platform_display_label(
                xlsx_source_test.display_platform_label(str(record.get('platform') or ''), country_code)
            )
        cache_key = source_manager.capability_cache_key(record['source_url'])
        merged_card = source_manager.merge_capability_snapshot(
            card,
            capability_cache.get(cache_key),
            adapter_configs_path=str(country_path(country_code, 'adapter_configs_path')),
        )
        merged_card.update(source_yield_lookup.get(cache_key, {}))
        cards.append(merged_card)
    return cards


def refresh_source_capability_cache_for_news(argv: list[str]) -> None:
    args = xlsx_source_test.parse_args(argv)
    country_code = args.country
    records = source_manager.list_sources(
        str(country_path(country_code, 'extra_sources_path')),
        False,
        platform=None,
        side=None,
    )
    if args.sides:
        allowed_sides = {item.strip() for item in args.sides if item.strip()}
        records = [record for record in records if str(record.get('side') or '').strip() in allowed_sides]
    if args.platforms:
        _, selected_source_platforms, custom_platforms = xlsx_source_test.resolve_requested_platforms(args.platforms, country_code)
        selected_source_platform_set = {item.strip().lower() for item in selected_source_platforms if item.strip()}
        selected_custom_platform_set = {item.strip().lower() for item in custom_platforms if item.strip()}
        platform_lookup = xlsx_source_test.requested_platform_lookup(country_code)

        filtered_records: list[dict[str, Any]] = []
        for record in records:
            record_side = str(record.get('side') or '').strip()
            if record_side == 'media':
                filtered_records.append(record)
                continue

            record_platform = str(record.get('platform') or '').strip()
            normalized_platform = str(platform_lookup.get(record_platform.lower(), record_platform.lower())).strip().lower()
            if normalized_platform in selected_source_platform_set or record_platform.lower() in selected_custom_platform_set:
                filtered_records.append(record)
        records = filtered_records
    if not records:
        return
    source_manager.refresh_capability_cache_for_records(
        records,
        adapter_configs_path=str(country_path(country_code, 'adapter_configs_path')),
        cache_path=str(country_path(country_code, 'source_capability_cache_path')),
    )


def build_source_record_payload(
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    show_inactive: bool = False,
    platform: str = '',
    side: str = '',
    source_url: str = '',
) -> dict[str, Any]:
    return {
        'records': load_source_manager_records(country_code),
        'cards': build_source_record_cards(
            country_code=country_code,
            show_inactive=show_inactive,
            platform=platform,
            side=side,
            source_url=source_url,
        ),
    }


def build_platform_alias_state(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, Any]:
    platforms = xlsx_source_test.list_available_platform_labels(country_code)
    alias_map: dict[str, list[str]] = {}
    for platform in platforms:
        aliases = xlsx_source_test.platform_search_terms(platform, country_code)
        alias_map[platform] = aliases or [platform]
    return {
        'text': format_platform_alias_text(alias_map, platforms),
        'platform_count': len(alias_map),
    }


def backup_platform_aliases_previous_value(country_code: str, next_text: str) -> None:
    current_text = str(build_platform_alias_state(country_code).get('text') or '').strip()
    if current_text and text_config_changed(current_text, next_text):
        set_country_previous_config_value(country_code, PLATFORM_ALIASES_PREVIOUS_FIELD, current_text)


def clean_alias_map_for_platforms(alias_map: dict[str, Any], platforms: list[str]) -> dict[str, list[str]]:
    cleaned: dict[str, list[str]] = {}
    for platform in platforms:
        aliases = alias_map.get(platform)
        if not isinstance(aliases, list):
            aliases = []
        cleaned[platform] = [
            alias
            for alias in unique_texts([str(item or '').strip() for item in aliases])
            if alias and not contains_cjk(alias)
        ] or [platform]
    return cleaned


def build_generated_platform_alias_map(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, list[str]]:
    country_config = get_country_config(country_code)
    market_terms = [str(item) for item in country_config.get('market_terms', []) if str(item).strip()]
    return {
        platform: generate_platform_aliases(
            platform,
            country_code=country_code,
            country_label=str(country_config.get('label') or ''),
            google_gl=str(country_config.get('google_news_gl') or '').upper(),
            market_terms=market_terms,
        )
        for platform in xlsx_source_test.list_available_platform_labels(country_code)
    }


def build_default_platform_alias_map(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, list[str]]:
    platforms = xlsx_source_test.list_available_platform_labels(country_code)
    default_overrides = country_dict_setting(country_code, 'default_platform_search_term_overrides')
    if default_overrides:
        return clean_alias_map_for_platforms(default_overrides, platforms)

    configured_overrides = country_dict_setting(country_code, 'platform_search_term_overrides')
    if configured_overrides:
        alias_map = {platform: xlsx_source_test.platform_search_terms(platform, country_code) for platform in platforms}
        return clean_alias_map_for_platforms(alias_map, platforms)

    return build_generated_platform_alias_map(country_code)


def canonicalize_platform_alias_map(alias_map: dict[str, list[str]], country_code: str) -> dict[str, list[str]]:
    known_platforms = set(xlsx_source_test.list_available_platform_labels(country_code))
    canonical: dict[str, list[str]] = {}
    for raw_platform, raw_aliases in alias_map.items():
        platform = str(raw_platform or '').strip()
        aliases = [
            alias
            for alias in unique_texts([str(alias or '').strip() for alias in raw_aliases])
            if not contains_cjk(alias)
        ]
        if platform not in known_platforms:
            matched_known = next((alias for alias in aliases if alias in known_platforms), '')
            if matched_known:
                aliases = unique_texts([platform, *[alias for alias in aliases if alias != matched_known]])
                platform = matched_known
        canonical[platform] = unique_texts([platform, *canonical.get(platform, []), *aliases])
    return canonical


def build_source_manager_state(overrides: dict[str, Any] | None = None) -> dict[str, Any]:
    state = {
        'source_action': 'list',
        'sm_side': '',
        'sm_platform_text': '',
        'sm_url_text': '',
        'sm_platform_select': '',
        'sm_url_select': '',
        'show_inactive': False,
        'skip_api': False,
        'force_api': False,
        'api_url': '',
        'api_key': '',
        'api_model': '',
    }
    persisted = read_app_settings().get('source_state')
    if isinstance(persisted, dict):
        state.update(persisted)
    global_api_settings = read_global_survey_api_settings()
    source_api_fallbacks = {
        'api_url': global_api_settings.get('survey_api_url') or '',
        'api_key': global_api_settings.get('survey_api_key') or '',
        'api_model': global_api_settings.get('survey_api_model') or '',
    }
    for key, fallback_value in source_api_fallbacks.items():
        if fallback_value and not str(state.get(key) or '').strip():
            state[key] = fallback_value
    state = apply_overrides_preserving_blank_api_fields(
        state,
        overrides,
        preserved_fields=['api_url', 'api_key', 'api_model'],
    )
    return state


def build_news_crawler_argv(form) -> tuple[list[str], dict[str, Any], dict[str, Any]]:
    country_code = normalize_country_request(form.get('country_code'))
    country_config = get_country_config(country_code)
    ensure_country_support_files(country_config)
    mode = (form.get('date_mode') or 'days').strip()
    days = (form.get('days') or '7').strip()
    start_date = (form.get('start_date') or '').strip()
    end_date = (form.get('end_date') or '').strip()
    translate_to = (form.get('translate_to') or 'zh-CN').strip()
    output_dir = (form.get('output_dir') or 'outputs').strip()
    recall_values = form.getlist('recall_mode') if hasattr(form, 'getlist') else [form.get('recall_mode')]
    normalized_recall_values = [str(item).strip().lower() for item in recall_values if item is not None and str(item).strip()]
    recall_mode = (
        'balanced'
        if 'balanced' in normalized_recall_values
        else ('strict' if normalized_recall_values else xlsx_source_test.DEFAULT_RECALL_MODE)
    )
    if recall_mode not in xlsx_source_test.RECALL_MODE_CHOICES:
        recall_mode = xlsx_source_test.DEFAULT_RECALL_MODE
    argv: list[str] = [
        '--country', country_code,
        '--xlsx', str(country_path(country_code, 'xlsx_path')),
        '--extra-sources', str(country_path(country_code, 'extra_sources_path')),
        '--adapter-configs', str(country_path(country_code, 'adapter_configs_path')),
        '--site-credentials', str(country_path(country_code, 'site_credentials_path')),
    ]
    if mode == 'range':
        if start_date:
            argv += ['--start-date', start_date]
        if end_date:
            argv += ['--end-date', end_date]
    else:
        argv += ['--days', days]

    selected_known_platforms = [
        item.strip()
        for item in form.getlist('news_builtin_platforms')
        if item.strip()
    ]
    custom_platforms_text = (form.get('news_custom_platforms') or '').strip()
    normalized_known, normalized_custom = split_news_platforms_for_country(
        '\n'.join([*selected_known_platforms, *parse_platform_text(custom_platforms_text)]),
        country_code,
    )
    selected_platforms = []
    for platform in normalized_known + normalized_custom:
        if platform not in selected_platforms:
            selected_platforms.append(platform)
    for platform in selected_platforms:
        argv += ['--platform', platform]

    selected_sides = [
        item.strip()
        for item in form.getlist('news_sides')
        if item.strip() in {'media', 'buyer', 'seller'}
    ] or DEFAULT_NEWS_SIDES.copy()
    argv += ['--sides', *selected_sides]
    news_form_state = build_news_form_state(
        {
            'date_mode': mode,
            'days': days,
            'start_date': start_date,
            'end_date': end_date,
            'translate_to': translate_to,
            'output_dir': output_dir,
            'recall_mode': recall_mode,
            'news_platforms_text': '\n'.join(selected_platforms),
            'news_sides': selected_sides,
        },
        country_code=country_code,
    )

    survey_filter_mode = (form.get('survey_filter_mode') or xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE).strip().lower()
    survey_api_url = (form.get('survey_api_url') or '').strip()
    survey_api_key = (form.get('survey_api_key') or '').strip()
    survey_api_model = (form.get('survey_api_model') or '').strip()
    survey_system_prompt = (form.get('survey_system_prompt') or '').strip()
    related_news_search_enabled = bool(form.get('related_news_search_enabled'))
    report_ranking_search_enabled = bool(form.get('report_ranking_search_enabled'))
    promo_search_enabled = related_news_search_enabled or report_ranking_search_enabled
    promo_search_engine = (form.get('promo_search_engine') or 'both').strip().lower()
    promo_search_keywords = (form.get('promo_search_keywords') or '').strip()
    related_news_search_keywords = (form.get('related_news_search_keywords') or '').strip()
    report_search_keywords = (form.get('report_search_keywords') or '').strip()
    news_filter_state = build_news_filter_state(
        {
            'survey_filter_mode': survey_filter_mode,
            'survey_api_url': survey_api_url,
            'survey_api_key': survey_api_key,
            'survey_api_model': survey_api_model,
            'survey_system_prompt': survey_system_prompt,
            'promo_search_enabled': promo_search_enabled,
            'related_news_search_enabled': related_news_search_enabled,
            'report_ranking_search_enabled': report_ranking_search_enabled,
            'promo_search_engine': promo_search_engine,
            'promo_search_keywords': promo_search_keywords,
            'related_news_search_keywords': related_news_search_keywords,
            'report_search_keywords': report_search_keywords,
        },
        country_code=country_code,
    )

    argv += ['--translate-to', news_form_state['translate_to'], '--output-dir', resolve_output_dir_for_run(news_form_state['output_dir'])]
    argv += ['--recall-mode', news_form_state['recall_mode']]
    argv += ['--survey-filter-mode', news_filter_state['survey_filter_mode']]
    if news_filter_state['survey_filter_mode'] == 'ai':
        if news_filter_state.get('survey_api_url'):
            argv += ['--survey-api-url', news_filter_state['survey_api_url']]
        if news_filter_state.get('survey_api_key'):
            argv += ['--survey-api-key', news_filter_state['survey_api_key']]
        if news_filter_state.get('survey_api_model'):
            argv += ['--survey-api-model', news_filter_state['survey_api_model']]
        if news_filter_state['survey_system_prompt'] and news_filter_state['survey_system_prompt'] != xlsx_source_test.default_survey_ai_system_prompt(country_code):
            argv += ['--survey-system-prompt', news_filter_state['survey_system_prompt']]

    if news_filter_state['promo_search_enabled']:
        argv += ['--promo-search-engine', news_filter_state['promo_search_engine']]
        if news_filter_state['related_news_search_enabled']:
            argv.append('--search-related-news')
        if news_filter_state['report_ranking_search_enabled']:
            argv.append('--search-report-ranking')
        if news_filter_state['related_news_search_keywords'] and news_filter_state['related_news_search_keywords'] != xlsx_source_test.default_related_news_search_keywords_text(country_code):
            argv += ['--related-news-search-keywords', news_filter_state['related_news_search_keywords']]
        if news_filter_state['report_search_keywords'] and news_filter_state['report_search_keywords'] != xlsx_source_test.default_report_search_keywords_text(country_code):
            argv += ['--report-search-keywords', news_filter_state['report_search_keywords']]

    return argv, news_form_state, news_filter_state


def build_source_manager_argv(form) -> tuple[list[str], dict[str, Any]]:
    action = (form.get('source_action') or 'list').strip()
    side = (form.get('sm_side') or '').strip()
    if action == 'add' and not side:
        side = 'media'
    if action == 'add' and side == 'all':
        side = 'media'

    source_state = {
        'source_action': action,
        'sm_side': side,
        'sm_platform_text': (form.get('sm_platform_text') or '').strip(),
        'sm_url_text': (form.get('sm_url_text') or '').strip(),
        'sm_platform_select': (form.get('sm_platform_select') or '').strip(),
        'sm_url_select': (form.get('sm_url_select') or '').strip(),
        'show_inactive': bool(form.get('show_inactive')),
        'skip_api': bool(form.get('skip_api')),
        'force_api': bool(form.get('force_api')),
        'api_url': (form.get('api_url') or '').strip(),
        'api_key': (form.get('api_key') or '').strip(),
        'api_model': (form.get('api_model') or '').strip(),
    }

    argv: list[str] = []
    platform_select = source_state['sm_platform_select']
    url_select = source_state['sm_url_select']
    platform_text = source_state['sm_platform_text']
    url_text = source_state['sm_url_text']

    if action == 'list':
        argv.append('--list')
        if url_select:
            argv.insert(0, url_select)
        if platform_select:
            argv += ['--platform', platform_select]
        if side:
            argv += ['--side', side]
        if source_state['show_inactive']:
            argv.append('--show-inactive')
    elif action == 'add':
        if not url_text:
            raise ValueError('Source URL is required.')
        platform = platform_text
        if side == 'media':
            platform = source_manager.GENERAL_MEDIA_PLATFORM
        elif not platform:
            raise ValueError('Platform is required for buyer/seller sources.')
        argv.insert(0, url_text)
        argv += ['--platform', platform]
        if side:
            argv += ['--side', side]
        if source_state['skip_api']:
            argv.append('--skip-api')
        else:
            api_url = str(source_state.get('api_url') or '').strip()
            api_key = str(source_state.get('api_key') or '').strip()
            api_model = str(source_state.get('api_model') or '').strip()
            if api_url:
                argv += ['--api-url', api_url]
            if api_key:
                argv += ['--api-key', api_key]
            if api_model:
                argv += ['--api-model', api_model]
            if source_state['force_api']:
                argv.append('--force-api')
    elif action == 'reactivate':
        argv.append('--reactivate')
        if url_select:
            argv.insert(0, url_select)
        if platform_select:
            argv += ['--platform', platform_select]
        if side:
            argv += ['--side', side]
    elif action == 'remove':
        argv.append('--remove')
        if url_select:
            argv.insert(0, url_select)
        if platform_select:
            argv += ['--platform', platform_select]
        if side:
            argv += ['--side', side]
    return argv, source_state

def estimate_news_seconds(argv: list[str]) -> int:
    return int(build_news_timing_profile(argv).get('total_seconds') or 30)


def estimate_news_site_counts(argv: list[str]) -> tuple[int, int, int, int]:
    args = xlsx_source_test.parse_args(argv)
    country_config = get_country_config(args.country)
    source_groups = [xlsx_source_test.load_extra_sources(args.extra_sources)]
    if bool(country_config.get('include_xlsx_sources')):
        source_groups.insert(0, xlsx_source_test.load_sources_from_xlsx(args.xlsx))
    sources = xlsx_source_test.select_sources_for_run(
        xlsx_source_test.merge_source_entries(*source_groups),
        args.platforms,
        args.sides,
        args.country,
    )
    selected_display_platforms, _, custom_platforms = xlsx_source_test.resolve_requested_platforms(args.platforms, args.country)
    total_sites = len(sources)
    selected_platform_count = len(selected_display_platforms) + len(custom_platforms)
    return total_sites, 0, 0, selected_platform_count


def estimate_source_manager_seconds(source_state: dict[str, Any]) -> int:
    return int(build_source_manager_timing_profile(source_state).get('total_seconds') or 8)


def estimate_source_manager_site_counts(source_state: dict[str, Any]) -> tuple[int, int, int, int]:
    action = source_state.get('source_action')
    if action == 'list':
        return 0, 0, 0, 0
    return 1, 0, 1, 1


def create_job(
    *,
    active_tab: str,
    argv: list[str],
    estimate_seconds: int,
    country_code: str = DEFAULT_COUNTRY_CODE,
    form_state: dict[str, Any] | None = None,
    task_kind: str | None = None,
    timing_profile: dict[str, Any] | None = None,
) -> str:
    job_id = uuid.uuid4().hex
    now_ts = time.time()
    with JOB_STORE_LOCK:
        JOB_STORE[job_id] = {
            'job_id': job_id,
            'status': 'queued',
            'stage': 'queued',
            'active_tab': active_tab,
            'country_code': country_code,
            'task_kind': task_kind or active_tab,
            'argv': argv,
            'output': '',
            'exit_code': None,
            'started_at': now_ts,
            'stage_started_at': now_ts,
            'finished_at': None,
            'initial_estimate_seconds': estimate_seconds,
            'estimate_seconds': estimate_seconds,
            'timing_profile': timing_profile or {},
            'form_state': form_state or {},
            'total_sites': 0,
            'completed_sites': 0,
            'active_sites': 0,
            'current_site': '',
            'last_completed_site': '',
            'progress_message': '',
            'progress_percent': None,
            'selected_platform_count': 0,
            'stage_timings': {},
        }
    return job_id


def update_job(job_id: str, **fields: Any) -> None:
    with JOB_STORE_LOCK:
        job = JOB_STORE.get(job_id)
        if not job:
            return
        next_progress_percent = fields.get('progress_percent')
        if next_progress_percent is not None:
            try:
                next_progress_percent = int(next_progress_percent)
            except (TypeError, ValueError):
                next_progress_percent = None
            current_progress_percent = job.get('progress_percent')
            if next_progress_percent is not None and current_progress_percent is not None:
                try:
                    fields['progress_percent'] = max(int(current_progress_percent), next_progress_percent)
                except (TypeError, ValueError):
                    fields['progress_percent'] = next_progress_percent
            else:
                fields['progress_percent'] = next_progress_percent
        next_stage = fields.get('stage')
        if next_stage and next_stage != job.get('stage'):
            transition_time = time.time()
            close_current_stage_timing(job, now_ts=transition_time)
            fields['stage_started_at'] = transition_time
        job.update(fields)


def run_job_async(
    *,
    job_id: str,
    func,
    argv: list[str],
    active_tab: str,
    country_code: str = DEFAULT_COUNTRY_CODE,
    news_form_state: dict[str, Any] | None = None,
    news_platforms_text: str | None = None,
    news_sides: list[str] | None = None,
    news_filter_state: dict[str, Any] | None = None,
    source_state: dict[str, Any] | None = None,
) -> None:
    def worker() -> None:
        try:
            update_job(job_id, status='running')
            if func in {news_crawler.main, xlsx_source_test.main}:
                def progress_callback(payload: dict[str, Any]) -> None:
                    update_job(
                        job_id,
                        stage=payload.get('stage', ''),
                        total_sites=payload.get('total_sites', 0),
                        completed_sites=payload.get('completed_sites', 0),
                        active_sites=payload.get('active_sites', 0),
                        current_site=payload.get('current_site', ''),
                        last_completed_site=payload.get('last_completed_site', ''),
                        progress_message=payload.get('message', ''),
                        progress_percent=payload.get('progress_percent'),
                    )

                result = run_cli(func, argv, progress_callback=progress_callback)
            else:
                update_job(job_id, stage='run')
                result = run_cli(func, argv)

            if func in {news_crawler.main, xlsx_source_test.main} and int(result.get('exit_code') or 0) == 0:
                update_job(
                    job_id,
                    stage='capability_refresh',
                    progress_message='Refreshing source capability cache...',
                    current_site='Source capability cache',
                    active_sites=0,
                    progress_percent=None,
                )
                refresh_source_capability_cache_for_news(argv)

            finished_at = time.time()
            update_job(
                job_id,
                status='completed',
                stage='done',
                output=result['output'],
                exit_code=result['exit_code'],
                argv=result['argv'],
                finished_at=finished_at,
                result=result,
                news_platforms_text=(news_form_state or {}).get('news_platforms_text', news_platforms_text),
                news_sides=(news_form_state or {}).get('news_sides', news_sides),
                news_filter_state=build_news_filter_state(news_filter_state, country_code=country_code),
                source_state=build_source_manager_state(source_state),
                active_tab=active_tab,
                country_code=country_code,
                active_sites=0,
                current_site='',
                last_completed_site='',
                progress_percent=100,
            )
            with JOB_STORE_LOCK:
                job_snapshot = dict(JOB_STORE.get(job_id) or {})
                job_snapshot['stage_timings'] = dict((JOB_STORE.get(job_id) or {}).get('stage_timings') or {})
                job_snapshot['timing_profile'] = dict((JOB_STORE.get(job_id) or {}).get('timing_profile') or {})
            timing_summary = summarize_completed_job(job_snapshot)
            timing_lines = build_timing_summary_lines(timing_summary)
            if int(result.get('exit_code') or 0) == 0:
                append_job_timing_history(timing_summary)
            final_output = str(result.get('output') or '').rstrip()
            if timing_lines:
                final_output = f"{final_output}\n\n" + "\n".join(timing_lines) if final_output else "\n".join(timing_lines)
                update_job(job_id, output=final_output)
                result['output'] = final_output
            remember_view_state(
                result=result,
                active_tab=active_tab,
                country_code=country_code,
                news_form_state=news_form_state,
                news_platforms_text=news_platforms_text,
                news_sides=news_sides,
                news_filter_state=news_filter_state,
                source_state=source_state,
            )
        except Exception as exc:
            finished_at = time.time()
            update_job(
                job_id,
                status='failed',
                stage='done',
                output=f'Job failed: {exc}',
                exit_code=1,
                finished_at=finished_at,
                active_sites=0,
                current_site='',
                progress_percent=100,
            )

    threading.Thread(target=worker, daemon=True).start()

def estimate_finalization_seconds(job: dict[str, Any], baseline_estimate_seconds: int) -> int:
    expected_finalize = sum(
        stage_expected_seconds(job, stage_name)
        for stage_name in ['translation', 'survey_filter', 'ai_dedupe', 'finalize_output']
    )
    if expected_finalize > 1:
        return expected_finalize
    total_sites = int(job.get('total_sites') or 0)
    return max(18, min(180, baseline_estimate_seconds // 3 + total_sites))


def estimate_current_stage_remaining(job: dict[str, Any], *, elapsed_seconds: int) -> int:
    stage = str(job.get('stage') or '')
    stage_elapsed = current_stage_elapsed_seconds(job)
    expected_stage_seconds = stage_expected_seconds(job, stage)
    total_sites = int(job.get('total_sites') or 0)
    completed_sites = int(job.get('completed_sites') or 0)
    active_sites = int(job.get('active_sites') or 0)
    current_site = str(job.get('current_site') or '')
    progress_message = str(job.get('progress_message') or '')
    combined_text = ' '.join(part for part in [current_site, progress_message] if part)
    baseline_remaining = max(0, expected_stage_seconds - stage_elapsed)

    if stage == 'crawl_site' and total_sites > 0:
        remaining_sites = max(0, total_sites - completed_sites)
        if remaining_sites <= 0:
            return 0
        progress_remaining = expected_stage_seconds * (remaining_sites / max(total_sites, 1))
        if completed_sites > 0 and stage_elapsed > 0:
            observed_per_site = stage_elapsed / max(completed_sites, 1)
            observed_remaining = observed_per_site * remaining_sites
            progress_remaining = min(observed_remaining, progress_remaining * 1.25)
        remaining = (baseline_remaining * 0.6) + (progress_remaining * 0.4)
        if active_sites > 0:
            remaining *= 0.92
        return max(1, int(round(remaining)))

    if stage == 'promo_search':
        completed_units, total_units = extract_progress_counts(combined_text)
        if total_units > 0 and completed_units > 0 and total_units >= completed_units:
            remaining_units = max(0, total_units - completed_units)
            progress_remaining = expected_stage_seconds * (remaining_units / max(total_units, 1))
            observed_per_unit = stage_elapsed / max(completed_units, 1)
            observed_remaining = observed_per_unit * remaining_units
            progress_remaining = min(observed_remaining, progress_remaining * 1.1)
            remaining = (baseline_remaining * 0.3) + (progress_remaining * 0.7)
            return max(1, int(round(remaining)))

    if stage in {'translation', 'survey_filter', 'ai_dedupe'}:
        completed_units, total_units = extract_progress_counts(combined_text)
        if total_units > 0 and completed_units > 0 and total_units >= completed_units:
            remaining_units = max(0, total_units - completed_units)
            progress_remaining = expected_stage_seconds * (remaining_units / max(total_units, 1))
            observed_per_unit = stage_elapsed / max(completed_units, 1)
            observed_remaining = observed_per_unit * remaining_units
            progress_remaining = min(observed_remaining, progress_remaining * 1.2)
            remaining = (baseline_remaining * 0.65) + (progress_remaining * 0.35)
            return max(1, int(round(remaining)))

    if stage == 'finalize_output':
        return max(1, baseline_remaining)

    if stage == 'capability_refresh' and total_sites > 0:
        return max(1, baseline_remaining)

    return max(1, baseline_remaining)


def stabilize_remaining_seconds(job: dict[str, Any], remaining_seconds: int, *, now_ts: float) -> int:
    signature = progress_signature(job)
    last_signature = str(job.get('eta_signature') or '')
    last_remaining = job.get('eta_remaining_seconds')
    last_recorded_at = float(job.get('eta_recorded_at') or now_ts)
    countdown_remaining: int | None = None

    if last_remaining is not None:
        countdown_elapsed = max(0, int(now_ts - last_recorded_at))
        countdown_remaining = max(0, int(last_remaining) - countdown_elapsed)
    if countdown_remaining is not None:
        if last_signature == signature:
            remaining_seconds = min(int(remaining_seconds), countdown_remaining)
        else:
            remaining_seconds = min(int(remaining_seconds), countdown_remaining + 3)

    job['eta_signature'] = signature
    job['eta_remaining_seconds'] = max(0, int(remaining_seconds))
    job['eta_recorded_at'] = now_ts
    return max(0, int(remaining_seconds))


def estimate_dynamic_timing(job: dict[str, Any], elapsed_seconds: int) -> tuple[int, int | None, bool]:
    if job.get('status') == 'completed':
        return max(1, elapsed_seconds), 0, False
    baseline_estimate_seconds = max(1, int(job.get('initial_estimate_seconds') or job.get('estimate_seconds') or 1))
    current_stage = str(job.get('stage') or '')
    order = stage_sequence(job)

    try:
        stage_index = order.index(current_stage)
    except ValueError:
        stage_index = 0

    future_stage_seconds = 0
    for next_stage in order[stage_index + 1:]:
        if next_stage == 'done':
            continue
        future_stage_seconds += stage_expected_seconds(job, next_stage)

    current_remaining = estimate_current_stage_remaining(job, elapsed_seconds=elapsed_seconds)
    dynamic_remaining_seconds = max(0, current_remaining + future_stage_seconds)
    dynamic_remaining_seconds = stabilize_remaining_seconds(job, dynamic_remaining_seconds, now_ts=time.time())
    dynamic_estimate_seconds = max(elapsed_seconds + 1, elapsed_seconds + dynamic_remaining_seconds)
    dynamic_estimate_seconds = max(dynamic_estimate_seconds, min(baseline_estimate_seconds, elapsed_seconds + dynamic_remaining_seconds + 1))
    remaining_seconds = max(0, dynamic_estimate_seconds - elapsed_seconds)
    is_overdue = elapsed_seconds > baseline_estimate_seconds and remaining_seconds > 0
    return dynamic_estimate_seconds, remaining_seconds, is_overdue


def calculate_progress_percent(job: dict[str, Any], elapsed_seconds: int, estimate_seconds: int) -> int:
    if job.get('status') == 'completed':
        return 100

    stage = str(job.get('stage') or '')
    order = stage_sequence(job)
    stage_names = [name for name in order if name != 'done']
    total_expected_seconds = sum(stage_expected_seconds(job, name) for name in stage_names)
    if total_expected_seconds <= 0:
        raw_ratio = elapsed_seconds / max(estimate_seconds, 1)
        return clamp_int(raw_ratio * 100, 2, 99)

    if stage not in stage_names:
        raw_ratio = elapsed_seconds / max(estimate_seconds, 1)
        return clamp_int(raw_ratio * 100, 2, 99)

    completed_stage_seconds = 0
    for name in stage_names:
        if name == stage:
            break
        completed_stage_seconds += stage_expected_seconds(job, name)

    current_stage_expected_seconds = stage_expected_seconds(job, stage)
    current_stage_remaining_seconds = estimate_current_stage_remaining(job, elapsed_seconds=elapsed_seconds)
    current_stage_remaining_seconds = max(0, min(current_stage_expected_seconds, current_stage_remaining_seconds))
    current_stage_completed_seconds = max(0, current_stage_expected_seconds - current_stage_remaining_seconds)

    percent = ((completed_stage_seconds + current_stage_completed_seconds) / max(total_expected_seconds, 1)) * 100
    return clamp_int(percent, 2, 99)


def stabilize_progress_percent(job: dict[str, Any], progress_percent: int) -> int:
    stabilized = clamp_int(progress_percent, 0, 100)
    progress_floor = job.get('progress_percent')
    use_backend_floor = job.get('task_kind') != 'news' or job.get('status') == 'completed'
    if progress_floor is not None and not use_backend_floor:
        try:
            use_backend_floor = int(progress_floor) >= 97
        except (TypeError, ValueError):
            use_backend_floor = False
    if progress_floor is not None and use_backend_floor:
        try:
            stabilized = max(stabilized, int(progress_floor))
        except (TypeError, ValueError):
            pass
    last_displayed = job.get('display_progress_percent')
    if last_displayed is not None:
        try:
            stabilized = max(stabilized, int(last_displayed))
        except (TypeError, ValueError):
            pass
    if job.get('status') == 'completed':
        stabilized = 100
    job['display_progress_percent'] = stabilized
    return stabilized


def build_progress_label(
    job: dict[str, Any],
    *,
    stage: str,
    total_sites: int,
    completed_sites: int,
    elapsed_seconds: int,
    estimate_seconds: int,
    remaining_seconds: int | None,
) -> str:
    progress_message = str(job.get('progress_message') or '').strip()
    stage_labels = {
        'queued': 'Queued',
        'setup': 'Preparing',
        'crawl_site': 'Crawling sources',
        'run': 'Running',
        'promo': 'Supplemental search',
        'translate': 'Translating',
        'survey_filter': 'AI filtering',
        'ai_dedupe': 'AI dedupe',
        'write_outputs': 'Writing outputs',
        'capability_refresh': 'Refreshing source cache',
        'finalize': 'Finalizing',
        'done': 'Done',
    }
    if job.get('status') == 'completed':
        return 'Completed'
    if job.get('status') == 'failed':
        return 'Failed'
    label = progress_message or stage_labels.get(stage, stage or 'Running')
    if total_sites > 0:
        label = f'{label} ({completed_sites}/{total_sites})'
    if remaining_seconds is None:
        return f'{label} · elapsed {format_seconds(elapsed_seconds)}'
    return f'{label} · elapsed {format_seconds(elapsed_seconds)} · remaining {format_seconds(remaining_seconds)}'

def job_status_payload(job: dict[str, Any]) -> dict[str, Any]:
    started_at = float(job.get('started_at') or time.time())
    finished_at = job.get('finished_at')
    elapsed_seconds = int((finished_at or time.time()) - started_at)
    estimate_seconds, remaining_seconds, is_overdue = estimate_dynamic_timing(job, elapsed_seconds)
    total_sites = int(job.get('total_sites') or 0)
    completed_sites = int(job.get('completed_sites') or 0)
    stage = str(job.get('stage') or '')

    if job.get('status') == 'completed':
        progress_percent = stabilize_progress_percent(job, 100)
        progress_label = str(job.get('progress_message') or 'Completed')
        remaining_seconds = 0
    elif job.get('status') == 'failed':
        progress_percent = stabilize_progress_percent(job, 100)
        progress_label = str(job.get('progress_message') or 'Failed')
        remaining_seconds = 0
    else:
        progress_percent = stabilize_progress_percent(
            job,
            calculate_progress_percent(job, elapsed_seconds, estimate_seconds),
        )
        progress_label = build_progress_label(
            job,
            stage=stage,
            total_sites=total_sites,
            completed_sites=completed_sites,
            elapsed_seconds=elapsed_seconds,
            estimate_seconds=estimate_seconds,
            remaining_seconds=remaining_seconds,
        )

    return {
        'job_id': job['job_id'],
        'status': job['status'],
        'stage': stage,
        'active_tab': job.get('active_tab') or 'home',
        'country_code': normalize_country_request(job.get('country_code') or DEFAULT_COUNTRY_CODE),
        'argv': redact_cli_args(job.get('argv') or []),
        'output': job.get('output') or '',
        'exit_code': job.get('exit_code'),
        'progress_percent': progress_percent,
        'progress_label': progress_label,
        'elapsed_seconds': elapsed_seconds,
        'remaining_seconds': remaining_seconds,
        'estimate_seconds': estimate_seconds,
        'initial_estimate_seconds': int(job.get('initial_estimate_seconds') or estimate_seconds),
        'is_overdue': is_overdue,
        'total_sites': total_sites,
        'completed_sites': completed_sites,
        'active_sites': int(job.get('active_sites') or 0),
        'current_site': job.get('current_site') or '',
        'last_completed_site': job.get('last_completed_site') or '',
        'selected_platform_count': int(job.get('selected_platform_count') or 0),
        'news_platforms_text': job.get('news_platforms_text'),
        'source_state': job.get('source_state'),
    }

def render_index_page(
    *,
    result: dict[str, Any] | None,
    active_tab: str,
    country_code: str = DEFAULT_COUNTRY_CODE,
    news_form_state: dict[str, Any] | None = None,
    news_platforms_text: str | None = None,
    news_sides: list[str] | None = None,
    news_filter_state: dict[str, Any] | None = None,
    source_state: dict[str, Any] | None = None,
    article_browser_state: dict[str, Any] | None = None,
):
    all_country_options = country_options()
    custom_country_options = [
        (code, label)
        for code, label in all_country_options
        if code not in COUNTRY_CONFIGS
    ]
    normalized_source_state = build_source_manager_state(source_state)
    normalized_news_form_state = build_news_form_state(
        news_form_state or {
            'news_platforms_text': news_platforms_text,
            'news_sides': news_sides,
        },
        country_code=country_code,
    )
    normalized_news_filter_state = build_news_filter_state(news_filter_state, country_code=country_code)
    country_config = get_country_config(country_code)
    if article_browser_state is None:
        if active_tab == 'articles':
            article_browser_state = build_article_browser_state(country_code=country_code)
        else:
            article_browser_state = build_article_browser_stub_state(country_code=country_code)
    news_platform_ui = build_news_platform_ui_state(normalized_news_form_state['news_platforms_text'], country_code)
    source_manager_known_platforms = xlsx_source_test.list_available_platform_labels(country_code)
    return render_template(
        'index.html',
        app_title=country_config['app_title'],
        selected_country_code=country_code,
        selected_country_label=country_config['label'],
        country_options=all_country_options,
        custom_country_options=custom_country_options,
        sides=SIDES,
        news_form_state=normalized_news_form_state,
        news_platforms_text=normalized_news_form_state['news_platforms_text'],
        news_platform_ui=news_platform_ui,
        platform_display_labels=user_platform_display_labels(news_platform_ui['known_platforms']),
        selected_news_sides=normalized_news_form_state['news_sides'],
        news_filter_state=normalized_news_filter_state,
        default_promo_search_keywords_text=xlsx_source_test.default_promo_search_keywords_text(country_code),
        default_related_news_search_keywords_text=xlsx_source_test.default_related_news_search_keywords_text(country_code),
        default_report_search_keywords_text=xlsx_source_test.default_report_search_keywords_text(country_code),
        default_survey_ai_system_prompt=xlsx_source_test.default_survey_ai_system_prompt(country_code),
        recent_outputs=list_recent_output_dirs(country_code=country_code),
        result=result,
        active_tab=active_tab,
        source_records=load_source_manager_records(country_code),
        source_manager_known_platforms=source_manager_known_platforms,
        source_manager_platform_display_labels=user_platform_display_labels(source_manager_known_platforms),
        source_state=normalized_source_state,
        source_record_cards=[],
        article_browser_state=article_browser_state,
        platform_alias_state=build_platform_alias_state(country_code),
        default_country_ai_prompt=default_country_ai_prompt_text(),
        country_ai_prompt=read_country_ai_prompt_setting() or default_country_ai_prompt_text(),
    )


def remember_view_state(
    *,
    result: dict[str, Any] | None,
    active_tab: str,
    country_code: str = DEFAULT_COUNTRY_CODE,
    news_form_state: dict[str, Any] | None = None,
    news_platforms_text: str | None = None,
    news_sides: list[str] | None = None,
    news_filter_state: dict[str, Any] | None = None,
    source_state: dict[str, Any] | None = None,
) -> None:
    WEB_VIEW_STATE['result'] = result
    WEB_VIEW_STATE['active_tab'] = active_tab
    WEB_VIEW_STATE['country_code'] = country_code
    normalized_news_form_state = build_news_form_state(
        news_form_state or {
            'news_platforms_text': news_platforms_text,
            'news_sides': news_sides,
        },
        country_code=country_code,
    )
    WEB_VIEW_STATE['news_form_state'] = normalized_news_form_state
    WEB_VIEW_STATE['news_platforms_text'] = normalized_news_form_state['news_platforms_text']
    WEB_VIEW_STATE['news_sides'] = list(normalized_news_form_state['news_sides'])
    WEB_VIEW_STATE['news_filter_state'] = build_news_filter_state(news_filter_state, country_code=country_code)
    WEB_VIEW_STATE['source_state'] = build_source_manager_state(source_state)


def current_view_state() -> dict[str, Any]:
    persisted_country_code = read_last_country_code()
    return {
        'result': WEB_VIEW_STATE.get('result'),
        'active_tab': WEB_VIEW_STATE.get('active_tab') or 'home',
        'country_code': normalize_country_request(WEB_VIEW_STATE.get('country_code') or persisted_country_code),
        'news_form_state': WEB_VIEW_STATE.get('news_form_state'),
        'news_platforms_text': WEB_VIEW_STATE.get('news_platforms_text'),
        'news_sides': WEB_VIEW_STATE.get('news_sides') or DEFAULT_NEWS_SIDES.copy(),
        'news_filter_state': WEB_VIEW_STATE.get('news_filter_state'),
        'source_state': WEB_VIEW_STATE.get('source_state'),
    }


def safe_next_url(value: str | None, fallback: str = '/') -> str:
    normalized = str(value or '').strip()
    if not normalized or not normalized.startswith('/') or normalized.startswith('//'):
        return fallback
    if normalized.startswith('/login') or normalized.startswith('/setup-admin'):
        return fallback
    return normalized


@app.route('/setup-admin', methods=['GET', 'POST'])
def setup_admin():
    if admin_password_configured() and request.method == 'GET':
        return redirect(url_for('login'))
    error = ''
    next_url = safe_next_url(request.values.get('next'), '/')
    if request.method == 'POST':
        password = request.form.get('password') or ''
        confirm_password = request.form.get('confirm_password') or ''
        if len(password) < 8:
            error = '密码至少需要 8 位。'
        elif password != confirm_password:
            error = '两次输入的密码不一致。'
        else:
            save_admin_password(password)
            mark_logged_in()
            return redirect(next_url)
    return render_template('setup_admin.html', error=error, next_url=next_url)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if not admin_password_configured():
        return redirect(url_for('setup_admin', next=request.values.get('next') or '/'))
    error = ''
    next_url = safe_next_url(request.values.get('next'), '/')
    if request.method == 'POST':
        password = request.form.get('password') or ''
        if verify_admin_password(password):
            mark_logged_in()
            return redirect(next_url)
        error = '密码不正确。'
    return render_template('login.html', error=error, next_url=next_url)

@app.route('/logout', methods=['GET'])
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/', methods=['GET'])
def index():
    view_state = current_view_state()
    previous_country = normalize_country_request(view_state.get('country_code') or DEFAULT_COUNTRY_CODE)
    requested_country = normalize_country_request(request.args.get('country_code') or previous_country)
    ensure_country_support_files(get_country_config(requested_country))
    if read_last_country_code() != requested_country:
        persist_app_settings(country_code=requested_country)
    news_form_state = build_news_form_state(
        view_state.get('news_form_state') if requested_country == previous_country else None,
        country_code=requested_country,
    )
    news_filter_state = build_news_filter_state(
        view_state.get('news_filter_state') if requested_country == previous_country else None,
        country_code=requested_country,
    )
    article_browser_state = build_article_browser_state(
        request.args,
        country_code=requested_country,
        user_friendly_sources=True,
    )
    remember_view_state(
        result=view_state.get('result'),
        active_tab='home',
        country_code=requested_country,
        news_form_state=news_form_state,
        news_filter_state=news_filter_state,
        source_state=view_state.get('source_state'),
    )
    country_config = get_country_config(requested_country)
    news_platform_ui = build_news_platform_ui_state(news_form_state['news_platforms_text'], requested_country)
    return render_template(
        'user_app.html',
        app_title=country_config['app_title'],
        selected_country_code=requested_country,
        selected_country_label=country_config['label'],
        country_options=country_options(),
        news_form_state=news_form_state,
        news_platform_ui=news_platform_ui,
        user_platform_display_labels=user_platform_display_labels(news_platform_ui['known_platforms']),
        news_filter_state=news_filter_state,
        ai_configured=all(str(news_filter_state.get(key) or '').strip() for key in SURVEY_API_SETTING_FIELDS),
        article_browser_state=article_browser_state,
    )


@app.route('/developer', methods=['GET'])
def developer():
    view_state = current_view_state()
    previous_country = normalize_country_request(view_state.get('country_code') or DEFAULT_COUNTRY_CODE)
    requested_tab = (request.args.get('tab') or '').strip()
    if requested_tab in {'home', 'news', 'sources', 'articles', 'countries'}:
        view_state['active_tab'] = requested_tab
    requested_country = normalize_country_request(request.args.get('country_code') or previous_country)
    ensure_country_support_files(get_country_config(requested_country))
    view_state['country_code'] = requested_country
    WEB_VIEW_STATE['country_code'] = requested_country
    if read_last_country_code() != requested_country:
        persist_app_settings(country_code=requested_country)
    news_form_overrides = view_state.get('news_form_state') if requested_country == previous_country else None
    news_filter_overrides = view_state.get('news_filter_state') if requested_country == previous_country else None
    view_state['news_form_state'] = build_news_form_state(news_form_overrides, country_code=requested_country)
    view_state['news_filter_state'] = build_news_filter_state(news_filter_overrides, country_code=requested_country)
    if view_state['active_tab'] == 'articles':
        view_state['article_browser_state'] = build_article_browser_state(request.args, country_code=requested_country)
    else:
        view_state['article_browser_state'] = build_article_browser_stub_state(request.args, country_code=requested_country)
    remember_view_state(
        result=view_state.get('result'),
        active_tab=view_state['active_tab'],
        country_code=requested_country,
        news_form_state=view_state['news_form_state'],
        news_filter_state=view_state['news_filter_state'],
        source_state=view_state.get('source_state'),
    )
    return render_index_page(**view_state)


@app.route('/api/countries', methods=['POST'])
def api_create_country():
    country_code, config, errors = build_country_config_from_form(request.form)
    known_country_codes = {code for code, _ in country_options()}
    if country_code and country_code in known_country_codes:
        errors.append('该国家代码已存在，请换一个代码。')
    if errors:
        return jsonify({'ok': False, 'errors': errors}), 400

    try:
        persist_global_survey_api_settings(extract_survey_api_settings_from_form(request.form))
        ensure_country_support_files(config)
        save_custom_country_config(country_code, config)
        persist_app_settings(country_code=country_code)
    except Exception as exc:
        return jsonify({'ok': False, 'errors': [str(exc) or exc.__class__.__name__]}), 500

    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'country_label': config['label'],
            'redirect_url': url_for('developer', country_code=country_code, tab='news'),
            'files': {
                key: config[key]
                for key in [
                    'extra_sources_path',
                    'adapter_configs_path',
                    'site_credentials_path',
                    'source_capability_cache_path',
                ]
            },
        }
    )


@app.route('/api/countries/delete', methods=['POST'])
def api_delete_country():
    raw_country_code = (request.form.get('country_code') or '').strip()
    country_code = normalize_new_country_code(raw_country_code)
    if not country_code:
        return jsonify({'ok': False, 'error': '请选择要删除的国家。'}), 400

    known_country_codes = {code for code, _ in country_options()}
    if country_code not in known_country_codes:
        return jsonify({'ok': False, 'error': '该国家不存在。'}), 404
    if country_code in COUNTRY_CONFIGS:
        return jsonify({'ok': False, 'error': '内置国家不能删除。'}), 400

    try:
        removed_config = delete_custom_country_config(country_code)
        cleanup_custom_country_generated_files(country_code, removed_config)
        remove_country_settings(country_code)
        if read_last_country_code() == country_code:
            persist_app_settings(country_code=DEFAULT_COUNTRY_CODE)
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 500

    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'redirect_url': url_for('developer', country_code=DEFAULT_COUNTRY_CODE, tab='countries'),
        }
    )

@app.route('/api/countries/suggest', methods=['POST'])
def api_suggest_country():
    try:
        fields = suggest_country_form_fields_with_ai(request.form)
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message}), 400
    return jsonify({'ok': True, 'fields': fields})


@app.route('/api/countries/aliases', methods=['POST'])
def api_save_country_aliases():
    country_code = normalize_country_request(request.form.get('country_code'))
    alias_map = parse_platform_alias_text(request.form.get('platform_aliases_text') or '')
    if not alias_map:
        return jsonify({'ok': False, 'error': '请至少填写一组品牌搜索别名。'}), 400
    normalized_alias_map = {
        platform: unique_texts([platform, *aliases])
        for platform, aliases in alias_map.items()
        if platform.strip()
    }
    normalized_alias_map = canonicalize_platform_alias_map(normalized_alias_map, country_code)
    next_text = format_platform_alias_text(normalized_alias_map, xlsx_source_test.list_available_platform_labels(country_code))
    try:
        backup_platform_aliases_previous_value(country_code, next_text)
        save_country_config_patch(country_code, {'platform_search_term_overrides': normalized_alias_map})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 500
    return jsonify({'ok': True, 'country_code': country_code, 'platform_count': len(normalized_alias_map), 'text': next_text})


@app.route('/api/countries/aliases/default', methods=['POST'])
def api_restore_country_aliases():
    country_code = normalize_country_request(request.form.get('country_code'))
    alias_map = build_default_platform_alias_map(country_code)
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'platform_count': len(alias_map),
            'text': format_platform_alias_text(alias_map, xlsx_source_test.list_available_platform_labels(country_code)),
            'saved': False,
        }
    )


@app.route('/api/countries/aliases/previous', methods=['POST'])
def api_restore_previous_country_aliases():
    country_code = normalize_country_request(request.form.get('country_code'))
    previous_text = read_country_previous_config_value(country_code, PLATFORM_ALIASES_PREVIOUS_FIELD)
    if not previous_text:
        return jsonify({'ok': False, 'error': '暂无上一版可恢复。'}), 404
    alias_map = parse_platform_alias_text(previous_text)
    return jsonify({'ok': True, 'country_code': country_code, 'platform_count': len(alias_map), 'text': previous_text, 'saved': False})


@app.route('/api/countries/aliases/suggest', methods=['POST'])
def api_suggest_country_aliases():
    try:
        result = suggest_platform_aliases_with_ai(request.form)
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message}), 400
    return jsonify({'ok': True, **result})


@app.route('/api/news-filter/keywords', methods=['POST'])
def api_save_news_filter_keywords():
    country_code = normalize_country_request(request.form.get('country_code'))
    existing_state = read_country_settings(country_code).get('news_filter_state')
    news_filter_state = build_news_filter_state(existing_state if isinstance(existing_state, dict) else None, country_code=country_code)
    previous_related = news_filter_state.get('related_news_search_keywords') or ''
    previous_report = news_filter_state.get('report_search_keywords') or ''
    news_filter_state['related_news_search_keywords'] = request.form.get('related_news_search_keywords') or ''
    news_filter_state['report_search_keywords'] = request.form.get('report_search_keywords') or ''
    related_auto_split_count = xlsx_source_test.keyword_auto_split_count(news_filter_state['related_news_search_keywords'])
    report_auto_split_count = xlsx_source_test.keyword_auto_split_count(news_filter_state['report_search_keywords'])
    news_filter_state = build_news_filter_state(news_filter_state, country_code=country_code)
    if previous_related != news_filter_state['related_news_search_keywords']:
        set_country_previous_config_value(country_code, 'related_news_search_keywords_previous', previous_related)
    if previous_report != news_filter_state['report_search_keywords']:
        set_country_previous_config_value(country_code, 'report_search_keywords_previous', previous_report)
    persist_app_settings(country_code=country_code, news_filter_state=news_filter_state)
    related_block_count = len(
        xlsx_source_test.normalize_related_news_search_keyword_blocks(
            news_filter_state['related_news_search_keywords'],
            country_code,
            xlsx_source_test.DEFAULT_RECALL_MODE,
        )
    )
    default_related_block_count = len(
        xlsx_source_test.normalize_related_news_search_keyword_blocks('', country_code, xlsx_source_test.DEFAULT_RECALL_MODE)
    )
    keyword_warning = ''
    if related_block_count < max(1, default_related_block_count // 2):
        keyword_warning = '关键词块较少，可能影响新闻入池数量。'
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'related_news_search_keywords': news_filter_state['related_news_search_keywords'],
            'report_search_keywords': news_filter_state['report_search_keywords'],
            'related_news_keyword_auto_split_count': related_auto_split_count,
            'report_keyword_auto_split_count': report_auto_split_count,
            'related_news_search_keyword_block_count': related_block_count,
            'keyword_warning': keyword_warning,
        }
    )


@app.route('/api/news-filter/keywords/default', methods=['POST'])
def api_restore_news_filter_keywords():
    country_code = normalize_country_request(request.form.get('country_code'))
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'related_news_search_keywords': xlsx_source_test.default_related_news_search_keywords_text(country_code),
            'report_search_keywords': xlsx_source_test.default_report_search_keywords_text(country_code),
            'saved': False,
        }
    )


@app.route('/api/news-filter/keywords/previous', methods=['POST'])
def api_restore_previous_news_filter_keywords():
    country_code = normalize_country_request(request.form.get('country_code'))
    previous_related = read_country_previous_config_value(country_code, 'related_news_search_keywords_previous')
    previous_report = read_country_previous_config_value(country_code, 'report_search_keywords_previous')
    if not previous_related and not previous_report:
        return jsonify({'ok': False, 'error': '暂无上一版可恢复。'}), 404
    current_state = build_news_filter_state(None, country_code=country_code)
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'related_news_search_keywords': previous_related or current_state['related_news_search_keywords'],
            'report_search_keywords': previous_report or current_state['report_search_keywords'],
            'saved': False,
        }
    )


@app.route('/api/news-filter/keywords/suggest', methods=['POST'])
def api_suggest_news_filter_keywords():
    try:
        result = suggest_news_filter_keywords_with_ai(request.form)
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message}), 400
    return jsonify({'ok': True, **result})


@app.route('/api/news-filter/survey-prompt', methods=['POST'])
def api_save_news_filter_survey_prompt():
    country_code = normalize_country_request(request.form.get('country_code'))
    existing_state = read_country_settings(country_code).get('news_filter_state')
    news_filter_state = build_news_filter_state(existing_state if isinstance(existing_state, dict) else None, country_code=country_code)
    previous_prompt = news_filter_state.get('survey_system_prompt') or ''
    news_filter_state['survey_system_prompt'] = request.form.get('survey_system_prompt') or ''
    news_filter_state = build_news_filter_state(news_filter_state, country_code=country_code)
    if previous_prompt != news_filter_state['survey_system_prompt']:
        set_country_previous_config_value(country_code, 'survey_system_prompt_previous', previous_prompt)
    persist_app_settings(country_code=country_code, news_filter_state=news_filter_state)
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'survey_system_prompt': news_filter_state['survey_system_prompt'],
            'survey_system_prompt_source': news_filter_state['survey_system_prompt_source'],
            'survey_system_prompt_source_label': news_filter_state['survey_system_prompt_source_label'],
        }
    )


@app.route('/api/news-filter/survey-prompt/default', methods=['POST'])
def api_restore_news_filter_survey_prompt():
    country_code = normalize_country_request(request.form.get('country_code'))
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'survey_system_prompt': xlsx_source_test.default_survey_ai_system_prompt(country_code),
            'survey_system_prompt_source': 'system_default',
            'survey_system_prompt_source_label': xlsx_source_test.survey_system_prompt_source_label('', country_code),
            'saved': False,
        }
    )


@app.route('/api/news-filter/survey-prompt/previous', methods=['POST'])
def api_restore_previous_news_filter_survey_prompt():
    country_code = normalize_country_request(request.form.get('country_code'))
    previous_prompt = read_country_previous_config_value(country_code, 'survey_system_prompt_previous')
    if not previous_prompt:
        return jsonify({'ok': False, 'error': '暂无上一版可恢复。'}), 404
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'survey_system_prompt': previous_prompt,
            'survey_system_prompt_source': xlsx_source_test.survey_system_prompt_source(previous_prompt, country_code),
            'survey_system_prompt_source_label': xlsx_source_test.survey_system_prompt_source_label(previous_prompt, country_code),
            'saved': False,
        }
    )


@app.route('/api/news-filter/survey-prompt/suggest', methods=['POST'])
def api_suggest_news_filter_survey_prompt():
    try:
        result = suggest_news_filter_survey_prompt_with_ai(request.form)
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message}), 400
    return jsonify({'ok': True, **result})


@app.route('/api/countries/ai-prompt', methods=['POST'])
def api_save_country_ai_prompt():
    prompt = request.form.get('country_ai_prompt') or ''
    persist_country_ai_prompt_setting(prompt)
    return jsonify({'ok': True, 'country_ai_prompt': read_country_ai_prompt_setting() or default_country_ai_prompt_text()})


@app.route('/api/countries/ai-prompt/default', methods=['POST'])
def api_restore_country_ai_prompt():
    return jsonify({'ok': True, 'country_ai_prompt': default_country_ai_prompt_text(), 'saved': False})

@app.route('/api/news-form-state', methods=['POST'])
def api_save_news_form_state():
    country_code = normalize_country_request(request.form.get('country_code'))
    selected_known_platforms = [
        item.strip()
        for item in request.form.getlist('news_builtin_platforms')
        if item.strip()
    ]
    custom_platforms_text = (request.form.get('news_custom_platforms') or '').strip()
    normalized_known, normalized_custom = split_news_platforms_for_country(
        '\n'.join([*selected_known_platforms, *parse_platform_text(custom_platforms_text)]),
        country_code,
    )
    selected_platforms: list[str] = []
    for platform in normalized_known + normalized_custom:
        if platform not in selected_platforms:
            selected_platforms.append(platform)

    selected_sides = [
        item.strip()
        for item in request.form.getlist('news_sides')
        if item.strip() in {'media', 'buyer', 'seller'}
    ] or DEFAULT_NEWS_SIDES.copy()

    news_form_state = build_news_form_state(
        {
            'date_mode': request.form.get('date_mode') or 'days',
            'days': request.form.get('days') or '7',
            'start_date': request.form.get('start_date') or '',
            'end_date': request.form.get('end_date') or '',
            'translate_to': request.form.get('translate_to') or 'zh-CN',
            'output_dir': request.form.get('output_dir') or 'outputs',
            'recall_mode': (
                'balanced'
                if 'balanced' in [str(item).strip().lower() for item in request.form.getlist('recall_mode')]
                else ('strict' if request.form.getlist('recall_mode') else xlsx_source_test.DEFAULT_RECALL_MODE)
            ),
            'news_platforms_text': '\n'.join(selected_platforms),
            'news_sides': selected_sides,
        },
        country_code=country_code,
    )

    filter_overrides: dict[str, Any] = {
        'survey_filter_mode': request.form.get('survey_filter_mode') or xlsx_source_test.DEFAULT_SURVEY_FILTER_MODE,
        'related_news_search_enabled': bool(request.form.get('related_news_search_enabled')),
        'report_ranking_search_enabled': bool(request.form.get('report_ranking_search_enabled')),
    }
    for field_name in ('survey_api_url', 'survey_api_key', 'survey_api_model', 'promo_search_engine'):
        if field_name in request.form:
            filter_overrides[field_name] = request.form.get(field_name) or ''

    news_filter_state = build_news_filter_state(filter_overrides, country_code=country_code)
    persist_app_settings(country_code=country_code, news_form_state=news_form_state, news_filter_state=news_filter_state)
    remember_view_state(
        result=WEB_VIEW_STATE.get('result'),
        active_tab='news',
        country_code=country_code,
        news_form_state=news_form_state,
        news_platforms_text=news_form_state['news_platforms_text'],
        news_sides=news_form_state['news_sides'],
        news_filter_state=news_filter_state,
        source_state=WEB_VIEW_STATE.get('source_state'),
    )
    return jsonify(
        {
            'ok': True,
            'country_code': country_code,
            'news_form_state': news_form_state,
            'news_filter_state': {
                key: value
                for key, value in news_filter_state.items()
                if key not in {
                    'survey_api_key',
                    'survey_system_prompt',
                    'promo_search_keywords',
                    'related_news_search_keywords',
                    'report_search_keywords',
                }
            },
        }
    )


@app.route('/api/run/news-crawler', methods=['POST'])
def api_run_news_crawler():
    country_code = normalize_country_request(request.form.get('country_code'))
    try:
        argv, news_form_state, news_filter_state = build_news_crawler_argv(request.form)
        persist_app_settings(country_code=country_code, news_form_state=news_form_state, news_filter_state=news_filter_state)
        total_sites, completed_sites, active_sites, selected_platform_count = estimate_news_site_counts(argv)
        timing_profile = build_news_timing_profile(argv)
        job_id = create_job(
            active_tab='news',
            argv=argv,
            estimate_seconds=int(timing_profile.get('total_seconds') or estimate_news_seconds(argv)),
            country_code=country_code,
            task_kind='news',
            timing_profile=timing_profile,
        )
        update_job(
            job_id,
            stage='setup',
            total_sites=total_sites,
            completed_sites=completed_sites,
            active_sites=active_sites,
            current_site='准备启动新闻抓取',
            progress_message='抓取任务已启动',
            selected_platform_count=selected_platform_count,
        )
        run_job_async(
            job_id=job_id,
            func=xlsx_source_test.main,
            argv=argv,
            active_tab='news',
            country_code=country_code,
            news_form_state=news_form_state,
            news_platforms_text=news_form_state['news_platforms_text'],
            news_sides=news_form_state['news_sides'],
            news_filter_state=news_filter_state,
        )
        with JOB_STORE_LOCK:
            job = JOB_STORE[job_id]
        return jsonify(job_status_payload(job))
    except Exception as exc:
        detail = traceback.format_exc()
        print(detail)
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__, 'detail': detail, 'country_code': country_code}), 500


@app.route('/run/news-crawler', methods=['POST'])
def run_news_crawler():
    country_code = normalize_country_request(request.form.get('country_code'))
    argv, news_form_state, news_filter_state = build_news_crawler_argv(request.form)
    persist_app_settings(country_code=country_code, news_form_state=news_form_state, news_filter_state=news_filter_state)
    result = run_cli(xlsx_source_test.main, argv)
    remember_view_state(
        result=result,
        active_tab='news',
        country_code=country_code,
        news_form_state=news_form_state,
        news_platforms_text=news_form_state['news_platforms_text'],
        news_sides=news_form_state['news_sides'],
        news_filter_state=news_filter_state,
    )
    return redirect(url_for('developer'))


@app.route('/api/run/source-manager', methods=['POST'])
def api_run_source_manager():
    try:
        country_code = normalize_country_request(request.form.get('country_code'))
        argv, source_state = build_source_manager_argv(request.form)
        persist_app_settings(country_code=country_code, source_state=source_state)
        total_sites, completed_sites, active_sites, selected_platform_count = estimate_source_manager_site_counts(source_state)
        timing_profile = build_source_manager_timing_profile(source_state)
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc), 'detail': traceback.format_exc()}), 400
    job_id = create_job(
        active_tab='sources',
        argv=argv,
        estimate_seconds=int(timing_profile.get('total_seconds') or estimate_source_manager_seconds(source_state)),
        country_code=country_code,
        form_state=source_state,
        task_kind='sources',
        timing_profile=timing_profile,
    )
    update_job(
        job_id,
        stage='setup',
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=active_sites,
        current_site='准备启动来源管理',
        progress_message='来源管理任务已启动',
        selected_platform_count=selected_platform_count,
    )
    run_job_async(
        job_id=job_id,
        func=source_manager.main,
        argv=argv,
        active_tab='sources',
        country_code=country_code,
        source_state=source_state,
    )
    with JOB_STORE_LOCK:
        job = JOB_STORE[job_id]
    return jsonify(job_status_payload(job))


@app.route('/api/test/survey-api', methods=['POST'])
def api_test_survey_api():
    api_url = (request.form.get('survey_api_url') or '').strip()
    api_key = (request.form.get('survey_api_key') or '').strip()
    api_model = (request.form.get('survey_api_model') or '').strip()
    try:
        result = test_survey_api_connection(api_url=api_url, api_key=api_key, api_model=api_model)
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        return jsonify(
            {
                'ok': False,
                'error': str(exc) or exc.__class__.__name__,
                'detail': detail,
                'api_url': xlsx_source_test.normalize_chat_completions_url(api_url),
                'api_model': api_model,
            }
        ), 400
    return jsonify(
        {
            'ok': True,
            'message': 'API 测试成功。',
            'api_url': result['api_url'],
            'api_model': result['api_model'],
            'elapsed_ms': result['elapsed_ms'],
            'payload': result['payload'],
        }
    )

@app.route('/run/source-manager', methods=['POST'])
def run_source_manager():
    country_code = normalize_country_request(request.form.get('country_code'))
    try:
        argv, source_state = build_source_manager_argv(request.form)
    except Exception as exc:
        source_state = build_source_manager_state(
            {
                'source_action': request.form.get('source_action', 'list'),
                'sm_side': (request.form.get('sm_side') or '').strip(),
                'sm_platform_text': (request.form.get('sm_platform_text') or '').strip(),
                'sm_url_text': (request.form.get('sm_url_text') or '').strip(),
                'sm_platform_select': (request.form.get('sm_platform_select') or '').strip(),
                'sm_url_select': (request.form.get('sm_url_select') or '').strip(),
                'show_inactive': bool(request.form.get('show_inactive')),
                'skip_api': bool(request.form.get('skip_api')),
                'force_api': bool(request.form.get('force_api')),
                'api_url': (request.form.get('api_url') or '').strip(),
                'api_key': (request.form.get('api_key') or '').strip(),
                'api_model': (request.form.get('api_model') or '').strip(),
            }
        )
        remember_view_state(
            result={'exit_code': 1, 'output': str(exc)},
            active_tab='sources',
            country_code=country_code,
            source_state=source_state,
        )
        return redirect(url_for('developer'))
    persist_app_settings(country_code=country_code, source_state=source_state)
    result = run_cli(source_manager.main, argv)
    remember_view_state(
        result=result,
        active_tab='sources',
        country_code=country_code,
        source_state=source_state,
    )
    return redirect(url_for('developer'))


@app.route('/api/jobs/<job_id>', methods=['GET'])
def api_job_status(job_id: str):
    with JOB_STORE_LOCK:
        job = JOB_STORE.get(job_id)
    if not job:
        return jsonify({'error': 'job_not_found'}), 404
    return jsonify(job_status_payload(job))


@app.route('/api/article-sources', methods=['GET'])
def api_article_sources():
    country_code = normalize_country_request(request.args.get('country_code'))
    return jsonify(build_article_source_payload(country_code))


@app.route('/api/article-stars/export', methods=['GET'])
def api_export_article_stars():
    country_code = normalize_country_request(request.args.get('country_code'))
    raw_article_source = (request.args.get('article_source') or '').strip()
    selected_platform = (request.args.get('article_platform') or '').strip()
    keyword = (request.args.get('article_keyword') or '').strip()
    scoped_export_requested = bool(raw_article_source or selected_platform or keyword)
    csv_path = None if is_db_article_source(raw_article_source) else (resolve_article_csv_path(raw_article_source, country_code=country_code) if scoped_export_requested else None)
    rows = build_starred_export_rows(
        csv_path=csv_path,
        article_source=raw_article_source if is_db_article_source(raw_article_source) else '',
        selected_platform=selected_platform,
        keyword=keyword,
        country_code=country_code,
    )
    fieldnames = [
        'article_id',
        'starred_updated_at',
        'platform_label',
        'title',
        'title_original',
        'summary',
        'source_name',
        'source_url',
        'article_url',
        'published_at',
        'category',
        'survey_dimensions',
        'survey_question_ids',
        'survey_indicator_examples',
        'survey_ai_reason_raw',
        'survey_ai_reason_translated',
        'survey_filter_confidence',
        'volume_fill',
        'briefing_sentiment',
        'briefing_sentiment_reason',
        'industry_trend_flag',
        'industry_trend_category',
        'industry_trend_impact',
        'industry_trend_reason',
        'source_file',
    ]
    buffer = io.StringIO()
    buffer.write('\ufeff')
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(rows)
    filename = f"starred_articles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        buffer.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.route('/api/articles/briefing-table', methods=['POST'])
def api_generate_article_briefing_table():
    country_code = normalize_country_request(request.form.get('country_code'))
    raw_article_source = (request.form.get('article_source') or '').strip()
    selected_platform = (request.form.get('article_platform') or '').strip()
    keyword = (request.form.get('article_keyword') or '').strip()
    raw_star_filter = (request.form.get('article_starred_only') or '').strip().lower()
    star_filter = raw_star_filter if raw_star_filter in {'starred', 'unstarred'} else ''

    try:
        if is_db_article_source(raw_article_source):
            rows, _, _, _, _ = load_article_rows_from_db_source(
                raw_article_source,
                selected_platform=selected_platform,
                keyword=keyword,
                star_filter=star_filter,
                limit=None,
                hydrate_remote_summary=False,
            )
        else:
            csv_path = resolve_article_csv_path(raw_article_source, country_code=country_code)
            rows, _, _, _, _ = load_article_rows(
                csv_path,
                selected_platform=selected_platform,
                keyword=keyword,
                star_filter=star_filter,
                limit=None,
                hydrate_remote_summary=False,
            )
        rows = filter_rows_by_selected_article_ids(rows, request.form)
        matching_rows = len(rows)
        if not rows:
            return jsonify({'ok': False, 'error': '当前筛选条件下没有新闻可生成资讯表。'}), 400

        country_settings = read_country_settings(country_code)
        api_settings = extract_survey_api_settings(
            country_settings.get('news_filter_state') if isinstance(country_settings.get('news_filter_state'), dict) else None
        )
        if not all(api_settings.get(key) for key in SURVEY_API_SETTING_FIELDS):
            global_settings = read_global_survey_api_settings()
            api_settings = {**global_settings, **api_settings}
        api_settings['survey_api_url'] = xlsx_source_test.normalize_chat_completions_url(api_settings.get('survey_api_url') or '')
        missing_fields = [key for key in SURVEY_API_SETTING_FIELDS if not str(api_settings.get(key) or '').strip()]
        if missing_fields:
            return jsonify({'ok': False, 'error': '请先填写并保存 AI API URL、Key 和模型名，再生成资讯表。'}), 400

        ai_stats: dict[str, int] = {}
        output_path = build_article_briefing_table(rows, country_code=country_code, api_settings=api_settings, ai_stats=ai_stats)
    except ValueError as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 400
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message, 'detail': traceback.format_exc()}), 500

    data = output_path.read_bytes()
    return Response(
        data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{output_path.name}"',
            'X-Article-Matching-Rows': str(matching_rows),
            'X-Briefing-AI-Batch-Count': str(int(ai_stats.get('batch_count', 0) or 0)),
            'X-Briefing-AI-Failed-Batch-Count': str(int(ai_stats.get('failed_batch_count', 0) or 0)),
            'X-Briefing-AI-Failed-Row-Count': str(int(ai_stats.get('failed_row_count', 0) or 0)),
            'X-Briefing-AI-Completed-Row-Count': str(int(ai_stats.get('ai_completed_row_count', 0) or 0)),
        },
    )


@app.route('/api/articles/news-summary', methods=['POST'])
def api_generate_article_news_summary():
    country_code = normalize_country_request(request.form.get('country_code'))
    raw_article_source = (request.form.get('article_source') or '').strip()
    selected_platform = (request.form.get('article_platform') or '').strip()
    keyword = (request.form.get('article_keyword') or '').strip()
    raw_star_filter = (request.form.get('article_starred_only') or '').strip().lower()
    star_filter = raw_star_filter if raw_star_filter in {'starred', 'unstarred'} else ''

    try:
        if is_db_article_source(raw_article_source):
            rows, _, _, _, _ = load_article_rows_from_db_source(
                raw_article_source,
                selected_platform=selected_platform,
                keyword=keyword,
                star_filter=star_filter,
                limit=None,
                hydrate_remote_summary=False,
            )
        else:
            csv_path = resolve_article_csv_path(raw_article_source, country_code=country_code)
            rows, _, _, _, _ = load_article_rows(
                csv_path,
                selected_platform=selected_platform,
                keyword=keyword,
                star_filter=star_filter,
                limit=None,
                hydrate_remote_summary=False,
            )
        rows = filter_rows_by_selected_article_ids(rows, request.form)
        matching_rows = len(rows)
        if not rows:
            return jsonify({'ok': False, 'error': '当前筛选条件下没有新闻可生成新闻总结。'}), 400

        country_settings = read_country_settings(country_code)
        api_settings = extract_survey_api_settings(
            country_settings.get('news_filter_state') if isinstance(country_settings.get('news_filter_state'), dict) else None
        )
        if not all(api_settings.get(key) for key in SURVEY_API_SETTING_FIELDS):
            global_settings = read_global_survey_api_settings()
            api_settings = {**global_settings, **api_settings}
        api_settings['survey_api_url'] = xlsx_source_test.normalize_chat_completions_url(api_settings.get('survey_api_url') or '')
        missing_fields = [key for key in SURVEY_API_SETTING_FIELDS if not str(api_settings.get(key) or '').strip()]
        if missing_fields:
            return jsonify({'ok': False, 'error': '请先填写并保存 AI API URL、Key 和模型名，再生成新闻总结。'}), 400

        ai_stats: dict[str, int] = {}
        news_summary_utils.NEWS_SUMMARY_OUTPUT_DIR = NEWS_SUMMARY_OUTPUT_DIR
        result = news_summary_utils.generate_news_summary(rows, country_code=country_code, api_settings=api_settings, stats=ai_stats)
    except ValueError as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 400
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message, 'detail': traceback.format_exc()}), 500

    return jsonify({'ok': True, 'text': result.text, 'filename': result.output_path.name, 'matching_rows': matching_rows, 'stats': result.stats})

@app.route('/api/articles/manual/suggest', methods=['POST'])
def api_suggest_manual_article():
    country_code = normalize_country_request(request.form.get('country_code'))
    raw_article_source = (request.form.get('article_source') or '').strip()
    article_url = (request.form.get('article_url') or '').strip()
    platform_hint = (request.form.get('platform_hint') or request.form.get('platform_label') or '').strip()
    try:
        validate_manual_article_source(raw_article_source)
        if not article_url:
            return jsonify({'ok': False, 'error': '请先填写新闻 URL。'}), 400
        api_settings = saved_survey_api_settings_for_country(country_code)
        missing_fields = [key for key in SURVEY_API_SETTING_FIELDS if not str(api_settings.get(key) or '').strip()]
        if missing_fields:
            return jsonify({'ok': False, 'error': '请先填写并保存 AI API URL、Key 和模型名，再使用 AI 自动填写。'}), 400
        article_context = fetch_manual_article_url_context(article_url)
        payload = xlsx_source_test.call_survey_filter_api(
            manual_article_suggestion_messages(
                country_code=country_code,
                article_url=article_url,
                article_context=article_context,
                platform_hint=platform_hint,
            ),
            api_settings['survey_api_url'],
            api_settings['survey_api_key'],
            api_settings['survey_api_model'],
        )
        suggestion = normalize_manual_article_suggestion(payload, article_url=article_url, country_code=country_code)
    except ValueError as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400
    except Exception as exc:
        detail = extract_api_error_detail(exc)
        message = str(exc) or exc.__class__.__name__
        if detail and detail not in message:
            message = f'{message}: {detail}'
        return jsonify({'ok': False, 'error': message, 'detail': traceback.format_exc()}), 500
    return jsonify({'ok': True, **suggestion})

@app.route('/api/articles/manual', methods=['POST'])
def api_save_manual_article():
    try:
        run_id, payload = manual_article_payload_from_form(request.form)
        manual_article_id, updated_existing = db_store.save_manual_article(
            run_id,
            payload.get('country_code') or normalize_country_request(request.form.get('country_code')),
            payload,
            manual_article_id=(request.form.get('manual_article_id') or '').strip() or None,
        )
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 400
    return jsonify(
        {
            'ok': True,
            'manual_article_id': manual_article_id,
            'updated_existing': updated_existing,
        }
    )


@app.route('/api/articles/manual/<manual_article_id>/disable', methods=['POST'])
def api_disable_manual_article(manual_article_id: str):
    if not manual_article_id:
        return jsonify({'ok': False, 'error': 'missing_manual_article_id'}), 400
    try:
        changed = db_store.set_manual_article_enabled(manual_article_id, False)
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc) or exc.__class__.__name__}), 400
    if not changed:
        return jsonify({'ok': False, 'error': 'manual_article_not_found'}), 404
    return jsonify({'ok': True, 'manual_article_id': manual_article_id})


@app.route('/api/source-records', methods=['GET'])
def api_source_records():
    country_code = normalize_country_request(request.args.get('country_code'))
    return jsonify(
        build_source_record_payload(
            country_code=country_code,
            show_inactive=parse_bool_arg(request.args.get('show_inactive')),
            platform=(request.args.get('platform') or '').strip(),
            side=(request.args.get('side') or '').strip(),
            source_url=(request.args.get('source_url') or '').strip(),
        )
    )


@app.route('/api/article-stars', methods=['POST'])
def api_article_stars():
    payload = request.get_json(silent=True) or {}
    article_id = str(payload.get('article_id') or '').strip()
    if not article_id:
        return jsonify({'error': 'missing_article_id'}), 400

    starred = bool(payload.get('starred'))
    with ARTICLE_STAR_STORE_LOCK:
        read_article_star_store()
        if starred:
            star_payload = {
                'starred': True,
                'updated_at': int(time.time()),
                'country_code': str(payload.get('country_code') or ''),
                'title': str(payload.get('title') or ''),
                'article_url': str(payload.get('article_url') or ''),
            }
            db_store.set_article_star(article_id, star_payload, True)
        else:
            db_store.set_article_star(article_id, {}, False)
        store = db_store.load_star_store()

    return jsonify(
        {
            'ok': True,
            'article_id': article_id,
            'starred': starred,
            'starred_count': len(store),
        }
    )

if __name__ == '__main__':
    install_terminal_signal_guards()
    app.run(host='0.0.0.0', port=8000, debug=False, use_reloader=False)














