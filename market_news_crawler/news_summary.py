from __future__ import annotations

import json
import re
import html
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import xlsx_source_test
import runtime_paths


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = runtime_paths.outputs_dir()
NEWS_SUMMARY_OUTPUT_DIR = OUTPUT_DIR / 'generated_news_summaries'

PLATFORM_ORDER = ['EU行业', '行业', 'AMZ', 'Shein', 'Temu', 'TTS', 'eBay', 'IG']
COUNTRY_ZH_LABELS = {
    'japan': '日本',
    'italy': '意大利',
    'france': '法国',
    'germany': '德国',
    'spain': '西班牙',
}
SENTIMENT_ZH_LABELS = {
    'Positive': '正向',
    'Neutral': '中性',
    'Negative': '负向',
}
SUMMARY_TAG_OPTIONS = [
    '行业生态',
    '卖家生态',
    '平台生态',
    '平台竞争',
    '平台发展',
    '卖家规范',
    '卖家权益',
    '商品品类',
    '商品品质',
    '商品价格',
    '商品丰富度',
    '商品合规',
    '广告',
    '直播',
    '促销活动',
    '支付手段',
    '物流仓储',
    '运费竞争',
    '发货时效',
    '功能',
    '推荐',
    '搜索',
    '订单查询',
    '行业发展',
    '消费趋势',
    '品类趋势',
    '银发经济',
    'Z世代消费',
    '循环时尚',
    '知识产权',
    '政策监管',
    '平台政策',
    '平台合规',
    '报告',
    '品牌形象',
    '商家赋能',
    '电商合规',
    '平台监管',
    '支付方式',
    '消费者权益',
    '直播内容',
    '生成式AI',
    '市场份额',
    '电商生态',
]

DIMENSION_TAG_ALIASES = {
    'features': '功能',
    'feature': '功能',
    'price': '商品价格',
    'content': '直播内容',
    'quality': '商品品质',
    'logistics': '物流仓储',
    'customer': '消费者权益',
    'post-purchase': '消费者权益',
    'post purchase': '消费者权益',
    'brand image': '品牌形象',
    'nps': '品牌形象',
}

TAG_ALIASES = {
    **DIMENSION_TAG_ALIASES,
    'promotion': '促销活动',
    'promo': '促销活动',
    'coupon': '促销活动',
    'discount': '促销活动',
    'payment': '支付方式',
    'checkout': '支付方式',
    'compliance': '平台合规',
    'regulation': '政策监管',
    'regulatory': '政策监管',
    'policy': '政策监管',
    'safety': '商品合规',
    'consumer protection': '消费者权益',
    'platform development': '平台发展',
    'market share': '市场份额',
    'ecosystem': '电商生态',
    'report': '报告',
    'ai': '生成式AI',
    'live shopping': '直播内容',
    'live': '直播内容',
}

SURVEY_ARTIFACT_PATTERNS = [
    r'https?://\S+',
    r'\bB\d+_\d+\s*[:：][^。；;|]*[|]?',
    r'指标映射\s*[:：][^。；;]*',
    r'关联说明\s*[:：]?',
    r'\[[^\]]*平台[^\]]*\]',
]


@dataclass
class NewsSummaryResult:
    text: str
    output_path: Path
    stats: dict[str, int]
    excel_output_path: Path | None = None
    html: str = ''
    html_output_path: Path | None = None


def compact_summary_text(value: Any, max_chars: int = 1000) -> str:
    text = re.sub(r'\s+', ' ', str(value or '')).strip()
    return text[:max_chars].strip()


def country_zh_label(country_code: str) -> str:
    return COUNTRY_ZH_LABELS.get(str(country_code or '').strip().lower(), str(country_code or '').strip() or '目标国家')


def normalize_sentiment(value: Any) -> str:
    return xlsx_source_test.normalize_briefing_sentiment(value)


def sentiment_zh(value: Any) -> str:
    return SENTIMENT_ZH_LABELS.get(normalize_sentiment(value), '中性')


def parse_publish_dt(value: Any) -> datetime | None:
    text = str(value or '').strip()
    if not text:
        return None
    normalized = text.replace('Z', '+00:00')
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        pass
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    return None


def format_publish_date_zh(value: Any) -> str:
    parsed = parse_publish_dt(value)
    if parsed is None:
        return str(value or '').strip()
    return f'{parsed.year}年{parsed.month}月{parsed.day}日'


def primary_dimension(row: dict[str, Any]) -> str:
    dimensions = str(row.get('survey_dimensions') or '').strip()
    if not dimensions:
        return '品牌形象'
    first = re.split(r'\s*[|,;/]\s*', dimensions)[0].strip()
    return first or '品牌形象'


def summary_business_tag(value: Any, fallback: str = '') -> str:
    text = compact_summary_text(value, 80).strip('【】[]()（） ')
    if not text:
        return fallback
    lowered = text.lower()
    for key, mapped in TAG_ALIASES.items():
        if key in lowered:
            return mapped
    if re.fullmatch(r'[A-Za-z /&+-]+', text):
        return fallback
    return text


def primary_metric_tag(row: dict[str, Any], value: Any = '') -> str:
    return summary_business_tag(value or primary_dimension(row), '品牌形象')


def has_eu_scope(row: dict[str, Any]) -> bool:
    text = ' '.join(
        str(row.get(key) or '')
        for key in [
            'title',
            'title_original',
            'title_display_zh',
            'summary',
            'article_body_excerpt',
            'industry_trend_reason',
            'industry_trend_category',
        ]
    ).lower()
    return any(
        token in text
        for token in [
            'eu ',
            'europe',
            'european',
            'european commission',
            'digital services act',
            'dsa',
            '欧盟',
            '欧洲',
            '泛欧',
        ]
    )


def platform_summary_label(row: dict[str, Any]) -> str:
    if xlsx_source_test.normalize_industry_trend_flag(row.get('industry_trend_flag')):
        return 'EU行业' if has_eu_scope(row) else '行业'
    platform = str(row.get('platform_label') or '').strip()
    if not platform:
        return '行业'
    upper = platform.upper()
    if upper in {'TIKTOK SHOP', 'TTS'}:
        return 'TTS'
    if upper in {'INSTAGRAM', 'INSTAGRAM SHOPPING', 'IG'}:
        return 'IG'
    if upper == 'EBAY':
        return 'eBay'
    if upper == 'TEMU':
        return 'Temu'
    if upper == 'SHEIN':
        return 'Shein'
    if upper == 'AMAZON':
        return 'AMZ'
    return platform


def infer_tags(row: dict[str, Any]) -> list[str]:
    text = ' '.join(
        str(row.get(key) or '')
        for key in [
            'title',
            'title_original',
            'title_display_zh',
            'summary',
            'survey_dimensions',
            'survey_indicator_examples',
            'industry_trend_category',
            'industry_trend_reason',
        ]
    ).lower()
    checks = [
        ('政策监管', ['regulation', 'policy', 'commission', 'eu ', 'european', '监管', '政策', '法规', '调查']),
        ('平台合规', ['compliance', 'illegal', 'safety', 'privacy', 'data', '合规', '非法', '安全', '隐私', '数据']),
        ('商品价格', ['price', 'discount', 'deal', 'coupon', 'sale', '促销', '折扣', '优惠', '价格']),
        ('支付手段', ['payment', 'cash', 'checkout', '支付', '现金']),
        ('物流仓储', ['delivery', 'logistics', 'shipping', 'warehouse', '物流', '配送', '仓储']),
        ('广告', ['ad ', 'ads', 'advertising', '广告']),
        ('直播', ['live', 'creator', 'reels', 'tiktok', '直播', '达人', '创作者']),
        ('消费趋势', ['consumer', 'generation z', 'gen z', 'z世代', '消费', '人群']),
        ('报告', ['report', 'survey', 'study', 'data', '报告', '调研', '数据']),
        ('平台竞争', ['competition', 'rival', 'market share', '竞争', '份额']),
        ('品牌形象', ['trust', 'quality', 'brand', '信任', '品牌', '形象']),
        ('商品合规', ['pfas', 'counterfeit', 'illegal products', '假货', '正品', '商品合规']),
        ('平台生态', ['marketplace', 'seller', 'merchant', 'ecosystem', '卖家', '商家', '生态']),
        ('功能', ['feature', 'tool', 'function', '功能', '工具']),
    ]
    tags: list[str] = []
    for tag, needles in checks:
        if any(needle in text for needle in needles):
            tags.append(tag)
    if xlsx_source_test.normalize_industry_trend_flag(row.get('industry_trend_flag')):
        tags.insert(0, '行业发展')
    if not tags:
        tags.append(primary_dimension(row))
    output: list[str] = []
    for tag in tags:
        normalized = tag if tag in SUMMARY_TAG_OPTIONS else tag.strip()
        if normalized and normalized not in output:
            output.append(normalized)
        if len(output) >= 4:
            break
    return output


def limit_core_claim(value: Any, max_chars: int = 50) -> str:
    text = compact_summary_text(value, 200)
    text = text.strip('。？！?；; ')
    if len(text) <= max_chars:
        return text
    return text[:max_chars].rstrip('，、；;：: ') + '…'


def article_title(row: dict[str, Any]) -> str:
    return compact_summary_text(
        row.get('title_display_zh')
        or row.get('title_translated')
        or row.get('title')
        or row.get('title_original')
        or row.get('summary')
        or '未命名新闻',
        300,
    )


def extract_readable_article_text(html: str, max_chars: int = 5000) -> str:
    soup = BeautifulSoup(html or '', 'lxml')
    for node in soup(['script', 'style', 'noscript', 'svg', 'form', 'nav', 'footer', 'header']):
        node.decompose()
    candidates: list[str] = []
    for selector in ['article', 'main', '[role="main"]']:
        for node in soup.select(selector):
            text = compact_summary_text(node.get_text(' ', strip=True), max_chars)
            if len(text) >= 200:
                candidates.append(text)
    if not candidates:
        paragraphs = [
            compact_summary_text(node.get_text(' ', strip=True), 1000)
            for node in soup.find_all(['p', 'li'])
        ]
        candidates.append(' '.join(part for part in paragraphs if len(part) >= 30))
    return compact_summary_text(max(candidates, key=len) if candidates else '', max_chars)


def is_search_engine_article_url(value: str) -> bool:
    host = urlparse(value or '').netloc.lower()
    return any(host.endswith(domain) for domain in ['news.google.com', 'bing.com', 'www.bing.com'])


def extract_url_query_candidate(value: str) -> str:
    parsed = urlparse(value or '')
    query = parse_qs(parsed.query)
    for key in ['url', 'u', 'r', 'target', 'redirect']:
        for candidate in query.get(key, []):
            decoded = unquote(candidate or '').strip()
            if decoded.startswith(('http://', 'https://')) and not is_search_engine_article_url(decoded):
                return decoded
    return ''


def extract_external_url_from_html(page_url: str, html_text: str) -> str:
    soup = BeautifulSoup(html_text or '', 'lxml')
    parsed_page = urlparse(page_url or '')
    page_host = parsed_page.netloc.lower()
    for selector in [
        ('link', {'rel': re.compile(r'canonical', re.I)}),
        ('meta', {'property': re.compile(r'^(og:url|twitter:url)$', re.I)}),
    ]:
        node = soup.find(selector[0], attrs=selector[1])
        if not node:
            continue
        candidate = compact_summary_text(node.get('href') or node.get('content'), 500)
        if candidate.startswith(('http://', 'https://')) and not is_search_engine_article_url(candidate):
            return candidate
    for anchor in soup.find_all('a', href=True):
        candidate = compact_summary_text(anchor.get('href'), 500)
        if not candidate.startswith(('http://', 'https://')):
            continue
        host = urlparse(candidate).netloc.lower()
        if host and host != page_host and not is_search_engine_article_url(candidate) and 'google.' not in host:
            return candidate
    return ''


def resolve_article_url(article_url: str, timeout: int = 10) -> tuple[str, str]:
    normalized = compact_summary_text(article_url, 500)
    if not normalized:
        return '', 'missing_url'
    query_candidate = extract_url_query_candidate(normalized)
    if query_candidate:
        return query_candidate, 'query_param'
    if not is_search_engine_article_url(normalized):
        return normalized, 'direct'
    try:
        response = requests.get(
            normalized,
            timeout=timeout,
            allow_redirects=True,
            headers={'User-Agent': 'Mozilla/5.0 (compatible; MarketNewsCrawler/1.0)'},
        )
        final_url = compact_summary_text(response.url, 500)
        if final_url and final_url != normalized and not is_search_engine_article_url(final_url):
            return final_url, 'http_redirect'
        html_candidate = extract_external_url_from_html(final_url or normalized, response.text[:250000])
        if html_candidate:
            return html_candidate, 'html_external_link'
    except Exception:
        return normalized, 'resolve_failed'
    return normalized, 'unresolved_search_url'


def summary_link_url(row: dict[str, Any]) -> str:
    return compact_summary_text(row.get('resolved_article_url') or row.get('article_url'), 500)


def fetch_article_body_context(row: dict[str, Any], timeout: int = 12) -> dict[str, str]:
    original_article_url = compact_summary_text(row.get('article_url'), 500)
    article_url, resolve_status = resolve_article_url(original_article_url)
    context: dict[str, str] = {
        'original_article_url': original_article_url,
        'resolved_article_url': article_url,
        'resolved_url_status': resolve_status,
    }
    if not article_url:
        context['article_body_fetch_status'] = 'missing_url'
        return context
    try:
        session = requests.Session()
        session.headers.update({'User-Agent': 'Mozilla/5.0 (compatible; MarketNewsCrawler/1.0)'})
        published_hint = parse_publish_dt(row.get('published_at'))
        metadata = xlsx_source_test.extract_article_metadata(
            article_url,
            session,
            datetime(2000, 1, 1),
            datetime(2100, 1, 1),
            published_at_hint=published_hint,
            title_hint=article_title(row),
            source_discovery='news_summary',
        )
        if isinstance(metadata, dict):
            metadata_url = compact_summary_text(metadata.get('article_url'), 500)
            body_excerpt = compact_summary_text(metadata.get('body_excerpt') or metadata.get('summary'), 5000)
            context.update({
                'resolved_article_url': metadata_url or article_url,
                'article_body_fetch_status': 'ok' if body_excerpt else 'metadata_empty',
                'article_page_title': compact_summary_text(metadata.get('title'), 300),
                'article_meta_description': compact_summary_text(metadata.get('summary'), 800),
                'article_body_excerpt': body_excerpt,
                'article_body_fetch_method': 'extract_article_metadata',
            })
            return context
        context['article_metadata_status'] = 'empty'
    except Exception as exc:
        context['article_metadata_status'] = 'failed'
        context['article_metadata_error'] = str(exc)[:200]
    try:
        response = requests.get(
            article_url,
            timeout=timeout,
            headers={'User-Agent': 'Mozilla/5.0 (compatible; MarketNewsCrawler/1.0)'},
        )
        response.raise_for_status()
        html = response.text[:400000]
        soup = BeautifulSoup(html, 'lxml')
        title = compact_summary_text(soup.title.get_text(' ', strip=True) if soup.title else '', 300)
        description = ''
        description_node = (
            soup.find('meta', attrs={'name': re.compile(r'^description$', re.I)})
            or soup.find('meta', attrs={'property': re.compile(r'^(og:description|twitter:description)$', re.I)})
        )
        if description_node:
            description = compact_summary_text(description_node.get('content'), 800)
        body = extract_readable_article_text(html)
        context.update({
            'article_body_fetch_status': 'ok' if body else 'empty',
            'article_page_title': title,
            'article_meta_description': description,
            'article_body_excerpt': body,
        })
        return context
    except Exception as exc:
        context.update({
            'article_body_fetch_status': 'failed',
            'article_body_fetch_error': str(exc)[:200],
        })
        return context


def enrich_rows_with_article_body(
    rows: list[dict[str, Any]],
    *,
    stats: dict[str, int] | None = None,
    hydrate_article_body: bool = True,
    max_workers: int = 5,
) -> list[dict[str, Any]]:
    enriched_rows = [dict(row) for row in rows]
    if not hydrate_article_body:
        return enriched_rows

    def apply_context(index: int, context: dict[str, str]) -> None:
        enriched_rows[index].update({key: value for key, value in context.items() if value})
        status = context.get('article_body_fetch_status') or ''
        if stats is not None:
            stats['article_fetch_attempt_count'] += 1 if rows[index].get('article_url') else 0
            if context.get('resolved_article_url') and context.get('resolved_article_url') != context.get('original_article_url'):
                stats['resolved_url_count'] += 1
            elif context.get('resolved_url_status') in {'unresolved_search_url', 'resolve_failed'}:
                stats['unresolved_source_url_count'] += 1
            if context.get('article_body_fetch_method') == 'extract_article_metadata':
                stats['article_metadata_success_count'] += 1
            if status == 'ok':
                stats['article_fetch_success_count'] += 1
                stats['article_body_success_count'] += 1
            elif status in {'failed', 'empty', 'missing_url', 'metadata_empty'}:
                stats['article_fetch_failed_count'] += 1
                reason = context.get('resolved_url_status') or status or 'unknown'
                failed_by_reason = stats.setdefault('article_fetch_failed_by_reason', {})
                failed_by_reason[reason] = failed_by_reason.get(reason, 0) + 1

    with ThreadPoolExecutor(max_workers=max(1, int(max_workers or 1))) as executor:
        future_to_index = {
            executor.submit(fetch_article_body_context, row): index
            for index, row in enumerate(enriched_rows)
        }
        for future in as_completed(future_to_index):
            index = future_to_index[future]
            try:
                apply_context(index, future.result())
            except Exception as exc:
                apply_context(index, {'article_body_fetch_status': 'failed', 'article_body_fetch_error': str(exc)[:200]})
    return enriched_rows


def default_summary_item(row: dict[str, Any], country_code: str) -> dict[str, Any]:
    platform = platform_summary_label(row)
    country = country_zh_label(country_code)
    date_text = format_publish_date_zh(row.get('published_at')) or '近期'
    source_name = compact_summary_text(row.get('source_name'), 80) or '媒体'
    dimension = primary_metric_tag(row)
    tags = infer_tags(row)
    sentiment = sentiment_zh(row.get('briefing_sentiment'))
    trend_reason = compact_summary_text(row.get('industry_trend_reason'), 220)
    body_hint = clean_summary_generated_text(
        row.get('article_body_excerpt')
        or row.get('article_meta_description')
        or row.get('summary')
        or '',
        260,
    )
    title = article_title(row)
    core = limit_core_claim(title)
    implication = clean_summary_generated_text(trend_reason, 220) or f'后续可能影响{country}市场对平台{dimension}表现的判断'
    facts = body_hint or title
    detail = f'{date_text}{source_name}消息，{facts}。{implication}。'
    return {
        'article_id': str(row.get('article_id') or ''),
        'platform_label': platform,
        'core_claim': core,
        'detail': detail,
        'sentiment': sentiment,
        'primary_metric': dimension,
        'tags': tags,
        'evidence_points': [],
        'needs_review': True,
    }


def summary_payload_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ['items', 'summaries', 'articles', 'rows']:
        value = payload.get(key) if isinstance(payload, dict) else None
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def clean_summary_generated_text(value: Any, max_chars: int = 1200) -> str:
    text = compact_summary_text(value, max_chars)
    if not text:
        return ''
    for pattern in SURVEY_ARTIFACT_PATTERNS:
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    text = re.sub(r'【(?:正向|中性|负向|Positive|Neutral|Negative|Quality|Price|Features|Content|Logistics|Customer[^】]*)】', '', text, flags=re.IGNORECASE)
    text = text.replace('指标映射', '').replace('关联说明', '').replace('新闻URL', '')
    text = re.sub(r'\s*[|]\s*', '，', text)
    text = re.sub(r'\s+', ' ', text).strip(' ，,。；;|')
    return compact_summary_text(text, max_chars)


def has_summary_artifacts(value: Any) -> bool:
    text = str(value or '')
    if re.search(r'https?://', text, re.IGNORECASE):
        return True
    if re.search(r'\bB\d+_\d+\b', text):
        return True
    return any(token in text for token in ['指标映射', '关联说明', '问卷题项', 'News URL'])


def normalize_tags(value: Any, row: dict[str, Any]) -> list[str]:
    raw_items: list[Any]
    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value or '').strip()
        raw_items = re.split(r'[|,，、]\s*', text) if text else []
    tags: list[str] = []
    metric_tag = primary_metric_tag(row)
    if metric_tag:
        tags.append(metric_tag)
    for item in raw_items:
        tag = summary_business_tag(item)
        tag = re.sub(r'\bB\d+_\d+\b.*', '', tag).strip()
        if tag and tag not in tags:
            tags.append(tag)
        if len(tags) >= 3:
            break
    if len(tags) < 2:
        for inferred in infer_tags(row):
            mapped = summary_business_tag(inferred)
            if mapped and mapped not in tags:
                tags.append(mapped)
            if len(tags) >= 3:
                break
    return tags[:3]


def normalize_evidence_points(value: Any) -> list[str]:
    raw_items: list[Any]
    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value or '').strip()
        raw_items = re.split(r'[|;；\n]\s*', text) if text else []
    points: list[str] = []
    for item in raw_items:
        point = compact_summary_text(item, 180)
        if point and point not in points:
            points.append(point)
        if len(points) >= 5:
            break
    return points


def summary_item_needs_review(item: dict[str, Any], row: dict[str, Any]) -> bool:
    body_status = str(row.get('article_body_fetch_status') or '').strip()
    detail = compact_summary_text(item.get('detail'), 1200)
    if has_summary_artifacts(detail) or has_summary_artifacts(item.get('core_claim')) or has_summary_artifacts(item.get('primary_metric')):
        return True
    if body_status != 'ok':
        return True
    evidence_points = normalize_evidence_points(item.get('evidence_points'))
    if len(detail) < 90 or len(evidence_points) < 2:
        return True
    title = article_title(row)
    if title and detail and title in detail and len(detail) < len(title) + 80:
        return True
    return False


def normalize_summary_item(item: dict[str, Any], row: dict[str, Any], country_code: str) -> dict[str, Any]:
    fallback = default_summary_item(row, country_code)
    platform = str(item.get('platform_label') or item.get('platform') or fallback['platform_label']).strip()
    if platform.startswith('【') and platform.endswith('】'):
        platform = platform.strip('【】')
    platform = platform_summary_label({**row, 'platform_label': platform, 'industry_trend_flag': row.get('industry_trend_flag')})
    sentiment = str(item.get('sentiment') or item.get('sentiment_zh') or fallback['sentiment']).strip()
    if sentiment in {'Positive', '正面', '利好', '正向'}:
        sentiment = '正向'
    elif sentiment in {'Negative', '负面', '利空', '负向'}:
        sentiment = '负向'
    elif sentiment in {'Neutral', '中立', '中性'}:
        sentiment = '中性'
    elif sentiment not in {'正向', '中性', '负向'}:
        sentiment = fallback['sentiment']
    detail = clean_summary_generated_text(item.get('detail') or item.get('explanation') or fallback['detail'], 900)
    core_claim = clean_summary_generated_text(item.get('core_claim') or item.get('headline') or fallback['core_claim'], 120)
    primary_metric = primary_metric_tag(row, item.get('primary_metric') or item.get('metric') or fallback['primary_metric'])
    normalized = {
        'article_id': str(item.get('article_id') or row.get('article_id') or ''),
        'platform_label': platform or fallback['platform_label'],
        'core_claim': limit_core_claim(core_claim or fallback['core_claim']),
        'detail': detail or fallback['detail'],
        'sentiment': sentiment,
        'primary_metric': primary_metric,
        'tags': normalize_tags(item.get('tags'), row),
        'evidence_points': normalize_evidence_points(item.get('evidence_points')),
        'needs_review': False,
    }
    normalized['needs_review'] = summary_item_needs_review(normalized, row)
    return normalized


def build_news_summary_ai_messages(rows: list[dict[str, Any]], country_code: str) -> list[dict[str, str]]:
    country = country_zh_label(country_code)
    articles = []
    for index, row in enumerate(rows, start=1):
        articles.append(
            {
                'article_id': str(row.get('article_id') or index),
                'platform': row.get('platform_label') or '',
                'title_original': row.get('title_original') or row.get('title') or '',
                'title_zh': row.get('title_display_zh') or row.get('title_translated') or '',
                'summary': row.get('summary') or '',
                'original_article_url': row.get('original_article_url') or row.get('article_url') or '',
                'resolved_article_url': row.get('resolved_article_url') or row.get('article_url') or '',
                'resolved_url_status': row.get('resolved_url_status') or '',
                'article_page_title': row.get('article_page_title') or '',
                'article_meta_description': row.get('article_meta_description') or '',
                'body_fetch_status': row.get('article_body_fetch_status') or '',
                'article_body_excerpt': compact_summary_text(row.get('article_body_excerpt'), 4500),
                'published_at': row.get('published_at') or '',
                'source_name': row.get('source_name') or '',
                'nps_dimension': row.get('survey_dimensions') or '',
                'survey_indicator_examples': row.get('survey_indicator_examples') or '',
                'sentiment': row.get('briefing_sentiment') or '',
                'industry_trend_flag': bool(xlsx_source_test.normalize_industry_trend_flag(row.get('industry_trend_flag'))),
                'industry_trend_category': row.get('industry_trend_category') or '',
                'industry_trend_impact': row.get('industry_trend_impact') or '',
                'industry_trend_reason': row.get('industry_trend_reason') or '',
            }
        )
    schema = {
        'items': [
            {
                'article_id': 'same article_id from input',
                'platform_label': 'EU行业|行业|AMZ|Shein|Temu|TTS|eBay|IG',
                'core_claim': 'Chinese core judgement, <=50 Chinese characters',
                'detail': 'Chinese paragraph starting with YYYY年M月D日来源消息，with facts + implication, no URL and no tail tags',
                'sentiment': '正向|中性|负向',
                'primary_metric': 'Chinese business tag, not English NPS dimension',
                'tags': ['2-4 values from allowed_tags'],
                'evidence_points': ['at least two concrete facts from article_body_excerpt when available'],
            }
        ]
    }
    return [
        {
            'role': 'system',
            'content': (
                'You are a senior market intelligence analyst writing client-ready Chinese news summaries. '
                'Read the supplied article body excerpts before writing. Return JSON only. Do not omit any input article.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_zh': country,
                    'output_schema': schema,
                    'allowed_tags': SUMMARY_TAG_OPTIONS,
                    'rules': [
                        'Write in concise professional Chinese, strictly like the user-provided client briefing templates.',
                        'Base the summary on article_body_excerpt, article_page_title and article_meta_description when available; do not merely rewrite card titles, summary, survey linkage, or NPS wording.',
                        'Never output survey question ids such as B3_8, 指标映射, 关联说明, questionnaire item wording, raw URLs, 原文链接, or 媒体 in core_claim/detail.',
                        'When article_body_excerpt is available, include at least two concrete facts in evidence_points, such as dates, amounts, organizations, campaign names, percentages, fines, features, market data, product categories, or regulatory actions.',
                        'Each detail must start with a time/source phrase like 2026年5月28日Source消息，then summarize facts, data/action/entity, and cautious implications.',
                        'core_claim must be no more than 50 Chinese characters and should state the main judgement or implication.',
                        'detail must not be generic; avoid empty phrases such as 影响消费者感知 unless supported by specific facts.',
                        'Do not include article URL, 原文链接, 媒体, or final bracket tags in detail; the system will append 媒体 and tags later.',
                        'Prioritize macro policy, industry trends, consumer data, campaign results, compliance, platform functions, services, and brand-level impact.',
                        'If the article implies but does not explicitly state an impact, infer cautiously from the facts and NPS linkage.',
                        'Do not write Chinese users, 中国用户, or anything suggesting the market is China unless the input country is China.',
                        'Use platform_label EU行业 for Europe/EU-wide policy, regulation, or market trend news; use 行业 for other macro/industry/report/policy news; otherwise use AMZ, Shein, Temu, TTS, eBay, or IG.',
                        'Use only Chinese business tags from allowed_tags. Do not output English dimensions such as Quality, Price, Features, Content, Logistics, Customer.',
                    ],
                    'articles': articles,
                },
                ensure_ascii=False,
                indent=2,
            ),
        },
    ]


def summary_reference_text(row: dict[str, Any]) -> str:
    article_url = summary_link_url(row)
    if article_url:
        return '媒体'
    source_name = compact_summary_text(row.get('source_name'), 80)
    if source_name:
        return f'来源：{source_name}'
    return ''


def format_summary_line(item: dict[str, Any], row: dict[str, Any], country_code: str) -> str:
    return format_summary_copy_text(item, row, country_code)


def format_summary_copy_text(item: dict[str, Any], row: dict[str, Any], country_code: str) -> str:
    platform = str(item.get('platform_label') or platform_summary_label(row)).strip() or '行业'
    core = limit_core_claim(item.get('core_claim') or article_title(row))
    detail = clean_summary_generated_text(item.get('detail'), 1200)
    if not detail:
        detail = default_summary_item(row, country_code)['detail']
    detail = detail.rstrip('。')
    sentiment = str(item.get('sentiment') or sentiment_zh(row.get('briefing_sentiment'))).strip() or '中性'
    tags = normalize_tags(item.get('tags'), row)
    tag_text = ''.join(f'【{tag}】' for tag in [sentiment, *tags] if tag)
    article_url = summary_link_url(row)
    source_name = compact_summary_text(row.get('source_name'), 80)
    reference = '媒体' if article_url else (f'来源：{source_name}' if source_name else '')
    reference_text = reference if reference else ''
    return f'【{platform}】{core}。{detail}{reference_text}{tag_text}'.strip()


def format_summary_html_line(item: dict[str, Any], row: dict[str, Any], country_code: str) -> str:
    text = format_summary_copy_text(item, row, country_code)
    link_url = summary_link_url(row)
    if not link_url or '媒体' not in text:
        return html.escape(text)
    before, after = text.rsplit('媒体', 1)
    return (
        f'{html.escape(before)}'
        f'<a href="{html.escape(link_url, quote=True)}">媒体</a>'
        f'{html.escape(after)}'
    )


def build_news_summary_html(items: list[dict[str, Any]], rows: list[dict[str, Any]], country_code: str) -> str:
    paragraphs = [
        f'<p>{format_summary_html_line(item, row, country_code)}</p>'
        for item, row in zip(items, rows)
    ]
    return (
        '<!doctype html><html><head><meta charset="utf-8">'
        '<title>新闻总结</title>'
        '<style>body{font-family:"Microsoft YaHei",Arial,sans-serif;line-height:1.75;color:#111827;}'
        'p{margin:0 0 14px;}a{color:#2563eb;text-decoration:underline;}</style>'
        '</head><body>'
        + '\n'.join(paragraphs)
        + '</body></html>'
    )


def summary_excel_rows(items: list[dict[str, Any]], rows: list[dict[str, Any]], country_code: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for item, row in zip(items, rows):
        tags = normalize_tags(item.get('tags'), row)
        resolved_url = summary_link_url(row)
        original_url = compact_summary_text(row.get('original_article_url') or row.get('article_url'), 500)
        source_name = compact_summary_text(row.get('source_name'), 120)
        output.append(
            {
                '平台/行业': str(item.get('platform_label') or platform_summary_label(row)).strip() or '行业',
                '核心判断': limit_core_claim(item.get('core_claim') or article_title(row)),
                '详细说明': clean_summary_generated_text(item.get('detail') or default_summary_item(row, country_code)['detail'], 1200).rstrip('。') + '。',
                '媒体': '媒体' if resolved_url else (f'来源：{source_name}' if source_name else ''),
                '情绪': str(item.get('sentiment') or sentiment_zh(row.get('briefing_sentiment'))).strip() or '中性',
                '主要指标': primary_metric_tag(row, item.get('primary_metric') or primary_dimension(row)),
                '标签': '、'.join(tags),
                '发布时间': format_publish_date_zh(row.get('published_at')),
                '来源名称': source_name,
                '真实新闻URL': resolved_url,
                '原始抓取URL': original_url,
                '完整汇报文案': format_summary_copy_text(item, row, country_code),
            }
        )
    return output


def save_news_summary_excel(items: list[dict[str, Any]], rows: list[dict[str, Any]], country_code: str, output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '新闻总结'
    headers = ['完整汇报文案', '平台/行业', '核心判断', '详细说明', '媒体', '情绪', '主要指标', '标签', '发布时间', '来源名称', '真实新闻URL', '原始抓取URL']
    worksheet.append(headers)
    header_fill = PatternFill('solid', fgColor='E8F1EF')
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='12312B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_data in summary_excel_rows(items, rows, country_code):
        worksheet.append([row_data.get(header, '') for header in headers])
        row_index = worksheet.max_row
        url = str(row_data.get('真实新闻URL') or '').strip()
        copy_cell = worksheet.cell(row=row_index, column=headers.index('完整汇报文案') + 1)
        if url and copy_cell.value:
            copy_cell.hyperlink = url
        media_cell = worksheet.cell(row=row_index, column=headers.index('媒体') + 1)
        if url and media_cell.value:
            media_cell.hyperlink = url
            media_cell.style = 'Hyperlink'
        for cell in worksheet[row_index]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        copy_cell.font = Font(bold=False, color='111827')
        worksheet.cell(row=row_index, column=headers.index('核心判断') + 1).font = Font(bold=True)

    widths = {
        'A': 120,
        'B': 12,
        'C': 34,
        'D': 72,
        'E': 12,
        'F': 10,
        'G': 18,
        'H': 28,
        'I': 16,
        'J': 24,
        'K': 36,
        'L': 36,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = 'A2'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def summary_sort_key(row: dict[str, Any]) -> tuple[int, int, float]:
    platform = platform_summary_label(row)
    is_industry = 0 if platform in {'EU行业', '行业'} else 1
    try:
        platform_index = PLATFORM_ORDER.index(platform)
    except ValueError:
        platform_index = len(PLATFORM_ORDER)
    parsed = parse_publish_dt(row.get('published_at'))
    timestamp = parsed.timestamp() if parsed is not None else 0
    return (is_industry, platform_index, -timestamp)


def sorted_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=summary_sort_key)


def generate_news_summary(
    rows: list[dict[str, Any]],
    *,
    country_code: str,
    api_settings: dict[str, str],
    batch_size: int = 5,
    stats: dict[str, int] | None = None,
    hydrate_article_body: bool = True,
) -> NewsSummaryResult:
    if stats is not None:
        stats.setdefault('batch_count', 0)
        stats.setdefault('failed_batch_count', 0)
        stats.setdefault('failed_row_count', 0)
        stats.setdefault('ai_completed_row_count', 0)
        stats.setdefault('fallback_row_count', 0)
        stats.setdefault('article_fetch_attempt_count', 0)
        stats.setdefault('article_fetch_success_count', 0)
        stats.setdefault('article_fetch_failed_count', 0)
        stats.setdefault('resolved_url_count', 0)
        stats.setdefault('unresolved_source_url_count', 0)
        stats.setdefault('article_metadata_success_count', 0)
        stats.setdefault('article_body_success_count', 0)
        stats.setdefault('summary_quality_retry_count', 0)
        stats.setdefault('summary_needs_review_count', 0)
        stats.setdefault('article_fetch_failed_by_reason', {})

    ordered_rows = enrich_rows_with_article_body(
        sorted_summary_rows(rows),
        stats=stats,
        hydrate_article_body=hydrate_article_body,
    )
    items_by_id = {
        str(row.get('article_id') or index): default_summary_item(row, country_code)
        for index, row in enumerate(ordered_rows, start=1)
    }
    row_by_id = {
        str(row.get('article_id') or index): row
        for index, row in enumerate(ordered_rows, start=1)
    }

    for offset in range(0, len(ordered_rows), batch_size):
        batch = ordered_rows[offset:offset + batch_size]
        if stats is not None:
            stats['batch_count'] += 1
        try:
            payload = xlsx_source_test.call_survey_filter_api(
                build_news_summary_ai_messages(batch, country_code),
                api_settings['survey_api_url'],
                api_settings['survey_api_key'],
                api_settings['survey_api_model'],
            )
            seen_ids: set[str] = set()
            for item in summary_payload_items(payload):
                article_id = str(item.get('article_id') or '').strip()
                if article_id and article_id in row_by_id:
                    normalized_item = normalize_summary_item(item, row_by_id[article_id], country_code)
                    items_by_id[article_id] = normalized_item
                    seen_ids.add(article_id)
                    if stats is not None:
                        stats['ai_completed_row_count'] += 1
            missing = [
                str(row.get('article_id') or index)
                for index, row in enumerate(batch, start=offset + 1)
                if str(row.get('article_id') or index) not in seen_ids
            ]
            if stats is not None and missing:
                stats['fallback_row_count'] += len(missing)
        except Exception:
            if stats is not None:
                stats['failed_batch_count'] += 1
                stats['failed_row_count'] += len(batch)
                stats['fallback_row_count'] += len(batch)

    retry_rows = [
        row
        for index, row in enumerate(ordered_rows, start=1)
        if items_by_id.get(str(row.get('article_id') or index), {}).get('needs_review')
        and str(row.get('article_body_fetch_status') or '') == 'ok'
    ]
    if retry_rows:
        if stats is not None:
            stats['summary_quality_retry_count'] += len(retry_rows)
        for offset in range(0, len(retry_rows), max(1, min(batch_size, 3))):
            batch = retry_rows[offset:offset + max(1, min(batch_size, 3))]
            try:
                payload = xlsx_source_test.call_survey_filter_api(
                    build_news_summary_ai_messages(batch, country_code),
                    api_settings['survey_api_url'],
                    api_settings['survey_api_key'],
                    api_settings['survey_api_model'],
                )
                for item in summary_payload_items(payload):
                    article_id = str(item.get('article_id') or '').strip()
                    if article_id and article_id in row_by_id:
                        normalized_item = normalize_summary_item(item, row_by_id[article_id], country_code)
                        current_item = items_by_id.get(article_id) or {}
                        if not normalized_item.get('needs_review') or current_item.get('needs_review'):
                            items_by_id[article_id] = normalized_item
            except Exception:
                continue

    lines: list[str] = []
    ordered_items: list[dict[str, Any]] = []
    for index, row in enumerate(ordered_rows, start=1):
        article_id = str(row.get('article_id') or index)
        item = items_by_id.get(article_id) or default_summary_item(row, country_code)
        ordered_items.append(item)
        lines.append(format_summary_line(item, row, country_code))
    if stats is not None:
        stats['summary_needs_review_count'] = sum(1 for item in ordered_items if item.get('needs_review'))
    summary_text = '\n\n'.join(lines).strip()
    NEWS_SUMMARY_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"news_summary_{country_code}_{timestamp}.txt"
    excel_filename = f"news_summary_{country_code}_{timestamp}.xlsx"
    html_filename = f"news_summary_{country_code}_{timestamp}.html"
    output_path = NEWS_SUMMARY_OUTPUT_DIR / filename
    excel_output_path = NEWS_SUMMARY_OUTPUT_DIR / excel_filename
    html_output_path = NEWS_SUMMARY_OUTPUT_DIR / html_filename
    summary_html = build_news_summary_html(ordered_items, ordered_rows, country_code)
    output_path.write_text(summary_text, encoding='utf-8-sig')
    html_output_path.write_text(summary_html, encoding='utf-8')
    save_news_summary_excel(ordered_items, ordered_rows, country_code, excel_output_path)
    return NewsSummaryResult(summary_text, output_path, stats or {}, excel_output_path, summary_html, html_output_path)
