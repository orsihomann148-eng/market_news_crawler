import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

import news_summary  # noqa: E402


class NewsSummaryTest(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self._original_output_dir = news_summary.NEWS_SUMMARY_OUTPUT_DIR
        news_summary.NEWS_SUMMARY_OUTPUT_DIR = Path(self._tmpdir.name)
        self.addCleanup(lambda: setattr(news_summary, 'NEWS_SUMMARY_OUTPUT_DIR', self._original_output_dir))

    def test_ai_summary_is_sorted_and_saved_with_rich_media_reference(self) -> None:
        rows = [
            {
                'article_id': 'platform-1',
                'platform_label': 'Amazon',
                'title_display_zh': 'Amazon 推出新的现金支付点',
                'published_at': '2026-05-18T16:44:40+00:00',
                'source_name': 'Test Media',
                'article_url': 'https://example.com/amazon-cash',
                'survey_dimensions': 'Features',
                'briefing_sentiment': 'Positive',
            },
            {
                'article_id': 'industry-1',
                'platform_label': 'TEMU',
                'title_display_zh': '欧盟推进电商平台监管',
                'published_at': '2026-05-20',
                'source_name': 'Policy Daily',
                'article_url': 'https://example.com/eu-regulation',
                'survey_dimensions': 'Quality | Customer / post-purchase service',
                'briefing_sentiment': 'Neutral',
                'industry_trend_flag': 'true',
            },
        ]

        def fake_call_api(messages, api_url, api_key, api_model):
            user_payload = messages[1]['content']
            self.assertIn('"resolved_article_url": "https://example.com/amazon-cash"', user_payload)
            return {
                'items': [
                    {
                        'article_id': 'platform-1',
                        'platform_label': 'Amazon',
                        'core_claim': '意大利现金支付便利性提升',
                        'detail': '2026年5月18日，意大利/Test Media消息，Amazon 扩展现金支付点，可能提升支付便利性与下单转化。',
                        'sentiment': '正向',
                        'primary_metric': 'Features',
                        'tags': ['支付手段', '功能'],
                        'evidence_points': ['扩展现金支付点', '提升支付便利性'],
                    },
                    {
                        'article_id': 'industry-1',
                        'platform_label': '行业',
                        'core_claim': '欧盟监管强化平台合规压力',
                        'detail': '2026年5月20日，意大利/Policy Daily消息，欧盟推进电商平台监管，可能影响 Temu 等平台合规成本与消费者信任。',
                        'sentiment': '中性',
                        'primary_metric': 'Quality',
                        'tags': ['政策监管', '平台合规'],
                        'evidence_points': ['欧盟推进监管', '影响平台合规成本'],
                    },
                ]
            }

        original_call_api = news_summary.xlsx_source_test.call_survey_filter_api
        news_summary.xlsx_source_test.call_survey_filter_api = fake_call_api
        self.addCleanup(lambda: setattr(news_summary.xlsx_source_test, 'call_survey_filter_api', original_call_api))

        stats: dict[str, int] = {}
        result = news_summary.generate_news_summary(
            rows,
            country_code='italy',
            api_settings={
                'survey_api_url': 'https://example.test/v1/chat/completions',
                'survey_api_key': 'key',
                'survey_api_model': 'model',
            },
            stats=stats,
            hydrate_article_body=False,
        )

        lines = [line for line in result.text.splitlines() if line.strip()]
        self.assertTrue(lines[0].startswith('【EU行业】'))
        self.assertIn('2026年5月20日', lines[0])
        self.assertIn('媒体【中性】【商品品质】【政策监管】【平台合规】', lines[0])
        self.assertNotIn('https://example.com/eu-regulation', lines[0])
        self.assertTrue(lines[1].startswith('【AMZ】'))
        self.assertIn('媒体【正向】【功能】【支付手段】', lines[1])
        self.assertNotIn('https://example.com/amazon-cash', lines[1])
        self.assertEqual(result.output_path.read_text(encoding='utf-8-sig'), result.text)
        self.assertIsNotNone(result.excel_output_path)
        workbook = load_workbook(result.excel_output_path)
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, '完整汇报文案')
        self.assertIn('【EU行业】欧盟监管强化平台合规压力', worksheet['A2'].value)
        self.assertIn('媒体【中性】【商品品质】【政策监管】【平台合规】', worksheet['A2'].value)
        self.assertNotIn('https://example.com/eu-regulation', worksheet['A2'].value)
        self.assertEqual(worksheet['A2'].hyperlink.target, 'https://example.com/eu-regulation')
        self.assertEqual(worksheet['E2'].value, '媒体')
        self.assertEqual(worksheet['E2'].hyperlink.target, 'https://example.com/eu-regulation')
        self.assertEqual(worksheet['E3'].value, '媒体')
        self.assertEqual(worksheet['E3'].hyperlink.target, 'https://example.com/amazon-cash')
        self.assertEqual(worksheet['K1'].value, '真实新闻URL')
        self.assertEqual(worksheet['L1'].value, '原始抓取URL')
        self.assertEqual(worksheet['K2'].value, 'https://example.com/eu-regulation')
        self.assertEqual(worksheet['L2'].value, 'https://example.com/eu-regulation')
        self.assertIn('<a href="https://example.com/eu-regulation">媒体</a>', result.html)
        self.assertEqual(result.html_output_path.read_text(encoding='utf-8'), result.html)
        self.assertEqual(stats['failed_row_count'], 0)
        self.assertEqual(stats['ai_completed_row_count'], 2)

    def test_failed_ai_batch_uses_fallback_summary_with_source_reference(self) -> None:
        rows = [
            {
                'article_id': 'row-1',
                'platform_label': 'IG',
                'title_display_zh': 'Instagram 增加购物广告工具',
                'published_at': '2026-05-18',
                'source_name': 'Media',
                'survey_dimensions': 'Content',
                'briefing_sentiment': 'Neutral',
            }
        ]

        def failing_call_api(messages, api_url, api_key, api_model):
            raise RuntimeError('mock ai failed')

        original_call_api = news_summary.xlsx_source_test.call_survey_filter_api
        news_summary.xlsx_source_test.call_survey_filter_api = failing_call_api
        self.addCleanup(lambda: setattr(news_summary.xlsx_source_test, 'call_survey_filter_api', original_call_api))

        stats: dict[str, int] = {}
        result = news_summary.generate_news_summary(
            rows,
            country_code='italy',
            api_settings={
                'survey_api_url': 'https://example.test/v1/chat/completions',
                'survey_api_key': 'key',
                'survey_api_model': 'model',
            },
            stats=stats,
            hydrate_article_body=False,
        )

        self.assertIn('【IG】', result.text)
        self.assertIn('2026年5月18日', result.text)
        self.assertIn('来源：Media【中性】【直播内容】', result.text)
        self.assertNotIn('媒体【', result.text)
        self.assertEqual(stats['failed_batch_count'], 1)
        self.assertEqual(stats['failed_row_count'], 1)
        self.assertEqual(stats['fallback_row_count'], 1)
        workbook = load_workbook(result.excel_output_path)
        self.assertIn('来源：Media【中性】【直播内容】', workbook.active['A2'].value)
        self.assertIsNone(workbook.active['A2'].hyperlink)
        self.assertEqual(workbook.active['E2'].value, '来源：Media')
        self.assertIsNone(workbook.active['E2'].hyperlink)
        self.assertNotIn('<a ', result.html)

    def test_summary_without_url_or_source_omits_media_placeholder(self) -> None:
        row = {
            'article_id': 'row-2',
            'platform_label': 'SHEIN',
            'title_display_zh': 'SHEIN 产品合规受到关注',
            'published_at': '2026-05-19',
            'survey_dimensions': 'Quality',
            'briefing_sentiment': 'Negative',
        }
        item = news_summary.default_summary_item(row, 'italy')
        line = news_summary.format_summary_line(item, row, 'italy')

        self.assertIn('【Shein】', line)
        self.assertIn('【负向】【商品品质】', line)
        self.assertNotIn('媒体【', line)
        self.assertNotIn('来源：', line)

    def test_core_claim_is_limited_to_50_chars(self) -> None:
        long_text = '这是一个非常长的新闻核心判断句用于测试摘要标题不会超过五十个汉字并且会自动截断保留可读性'
        self.assertLessEqual(len(news_summary.limit_core_claim(long_text)), 51)

    def test_resolve_article_url_uses_query_param_candidate(self) -> None:
        resolved, status = news_summary.resolve_article_url('https://www.bing.com/news/apiclick.aspx?url=https%3A%2F%2Fexample.com%2Foriginal')

        self.assertEqual(resolved, 'https://example.com/original')
        self.assertEqual(status, 'query_param')

    def test_resolved_url_is_used_in_outputs_instead_of_original_google_url(self) -> None:
        row = {
            'article_id': 'row-resolved',
            'platform_label': 'Amazon',
            'title_display_zh': 'Amazon Prime Day 定档',
            'article_url': 'https://news.google.com/rss/articles/fake',
            'resolved_article_url': 'https://example.com/original-amazon',
            'published_at': '2026-06-02',
            'source_name': 'Media',
            'survey_dimensions': 'Price',
            'briefing_sentiment': 'Positive',
        }
        item = {
            'platform_label': 'Amazon',
            'core_claim': 'Prime Day定档提振促销吸引力',
            'detail': '2026年6月2日，Amazon 宣布 Prime Day 时间安排，平台促销声量提升。',
            'sentiment': '正向',
            'primary_metric': '促销吸引力',
            'tags': ['促销活动'],
        }

        self.assertNotIn('https://example.com/original-amazon', news_summary.format_summary_line(item, row, 'italy'))
        self.assertIn('媒体【正向】', news_summary.format_summary_line(item, row, 'italy'))
        self.assertNotIn('news.google.com', news_summary.format_summary_copy_text(item, row, 'italy'))
        excel_row = news_summary.summary_excel_rows([item], [row], 'italy')[0]
        self.assertEqual(excel_row['真实新闻URL'], 'https://example.com/original-amazon')
        self.assertEqual(excel_row['原始抓取URL'], 'https://news.google.com/rss/articles/fake')

    def test_card_artifacts_are_removed_from_summary_outputs(self) -> None:
        row = {
            'article_id': 'row-artifact',
            'platform_label': 'eBay',
            'title_display_zh': 'eBay 六月优惠券',
            'article_url': 'https://example.com/ebay-coupon',
            'published_at': '2026-06-01',
            'source_name': 'Media',
            'survey_dimensions': 'Price',
            'briefing_sentiment': 'Positive',
        }
        item = {
            'platform_label': 'eBay',
            'core_claim': 'eBay优惠券增强价格竞争力 https://bad.example',
            'detail': '2026年6月1日Media消息，关联说明：B3_8: [平台]的促销和折扣对我很有吸引力 | eBay推出优惠券，https://bad.example 影响消费者评价。',
            'sentiment': 'Positive',
            'primary_metric': 'Price',
            'tags': ['Price', '商品价格', '商品价格', '平台竞争'],
            'evidence_points': ['6月优惠券', '最高可省30欧元'],
        }

        normalized = news_summary.normalize_summary_item(item, row, 'italy')
        line = news_summary.format_summary_copy_text(normalized, row, 'italy')

        self.assertNotIn('https://bad.example', line)
        self.assertNotIn('B3_8', line)
        self.assertNotIn('关联说明', line)
        self.assertNotIn('【Price】', line)
        self.assertIn('媒体【正向】【商品价格】【平台竞争】', line)

    def test_ai_payload_includes_fetched_article_body_excerpt(self) -> None:
        rows = [
            {
                'article_id': 'row-body',
                'platform_label': 'eBay',
                'title_display_zh': 'eBay 推出直播购物活动',
                'article_url': 'https://example.com/ebay-live',
                'published_at': '2026-05-22',
                'survey_dimensions': 'Content',
                'briefing_sentiment': 'Neutral',
            }
        ]

        def fake_fetch(row):
            return {
                'original_article_url': 'https://news.google.com/rss/articles/ebay-live',
                'resolved_article_url': 'https://example.com/ebay-live',
                'resolved_url_status': 'http_redirect',
                'article_body_fetch_status': 'ok',
                'article_page_title': 'eBay Live shopping expands',
                'article_meta_description': 'eBay expands live shopping for collectors.',
                'article_body_excerpt': 'eBay Live is expanding its shopping events with sellers and creators. The move may improve product discovery and marketplace engagement.',
            }

        def fake_call_api(messages, api_url, api_key, api_model):
            user_payload = messages[1]['content']
            self.assertIn('"article_body_excerpt": "eBay Live is expanding', user_payload)
            self.assertIn('"resolved_article_url": "https://example.com/ebay-live"', user_payload)
            return {
                'items': [
                    {
                        'article_id': 'row-body',
                        'platform_label': 'eBay',
                        'core_claim': 'eBay直播购物强化内容发现',
                        'detail': '2026年5月22日，意大利消息，eBay 扩大直播购物活动，可提升商品发现效率与平台互动。',
                        'sentiment': '中性',
                        'primary_metric': 'Content',
                        'tags': ['直播', '平台生态'],
                        'evidence_points': ['sellers and creators', 'product discovery'],
                    }
                ]
            }

        original_fetch = news_summary.fetch_article_body_context
        original_call_api = news_summary.xlsx_source_test.call_survey_filter_api
        news_summary.fetch_article_body_context = fake_fetch
        news_summary.xlsx_source_test.call_survey_filter_api = fake_call_api
        self.addCleanup(lambda: setattr(news_summary, 'fetch_article_body_context', original_fetch))
        self.addCleanup(lambda: setattr(news_summary.xlsx_source_test, 'call_survey_filter_api', original_call_api))

        stats: dict[str, int] = {}
        result = news_summary.generate_news_summary(
            rows,
            country_code='italy',
            api_settings={
                'survey_api_url': 'https://example.test/v1/chat/completions',
                'survey_api_key': 'key',
                'survey_api_model': 'model',
            },
            stats=stats,
        )

        self.assertIn('eBay直播购物强化内容发现', result.text)
        self.assertEqual(stats['article_fetch_success_count'], 1)

    def test_extract_article_metadata_success_is_used_before_fallback_fetch(self) -> None:
        rows = [
            {
                'article_id': 'row-meta',
                'platform_label': 'TTS',
                'title_display_zh': 'TikTok Shop 直播活动',
                'article_url': 'https://example.com/tiktok-live',
                'published_at': '2026-05-26',
                'survey_dimensions': 'Content',
                'briefing_sentiment': 'Positive',
            }
        ]

        def fake_metadata(article_url, session, start, end, **kwargs):
            return {
                'article_url': 'https://media.example.com/tiktok-live',
                'title': 'QVC Italia boosts TikTok Shop live commerce',
                'published_at': datetime(2026, 5, 26).isoformat(),
                'summary': 'QVC Italia opened a dedicated studio and prepared a 10-hour TikTok Mega Live.',
                'body_excerpt': 'QVC Italia opened a dedicated TikTok Shop studio. The channel has 80% new customers, live sales account for 25%, affiliate creators account for 47%, and average order value is 28 euros.',
            }

        def fake_call_api(messages, api_url, api_key, api_model):
            user_payload = messages[1]['content']
            self.assertIn('"resolved_article_url": "https://media.example.com/tiktok-live"', user_payload)
            self.assertIn('affiliate creators account for 47%', user_payload)
            return {
                'items': [
                    {
                        'article_id': 'row-meta',
                        'platform_label': 'TTS',
                        'core_claim': 'QVC Italia强化TikTok Shop布局',
                        'detail': '5月26日，QVC Italia宣布加大TikTok Shop投入，开设专属演播室并计划10小时直播，TikTok渠道80%为新客户、直播销售占25%。',
                        'sentiment': '正向',
                        'primary_metric': '直播内容',
                        'tags': ['直播'],
                        'evidence_points': ['80% new customers', 'live sales account for 25%'],
                    }
                ]
            }

        original_metadata = news_summary.xlsx_source_test.extract_article_metadata
        original_call_api = news_summary.xlsx_source_test.call_survey_filter_api
        news_summary.xlsx_source_test.extract_article_metadata = fake_metadata
        news_summary.xlsx_source_test.call_survey_filter_api = fake_call_api
        self.addCleanup(lambda: setattr(news_summary.xlsx_source_test, 'extract_article_metadata', original_metadata))
        self.addCleanup(lambda: setattr(news_summary.xlsx_source_test, 'call_survey_filter_api', original_call_api))

        stats: dict[str, int] = {}
        result = news_summary.generate_news_summary(
            rows,
            country_code='italy',
            api_settings={
                'survey_api_url': 'https://example.test/v1/chat/completions',
                'survey_api_key': 'key',
                'survey_api_model': 'model',
            },
            stats=stats,
        )

        self.assertNotIn('https://media.example.com/tiktok-live', result.text)
        self.assertIn('<a href="https://media.example.com/tiktok-live">媒体</a>', result.html)
        self.assertEqual(stats['article_metadata_success_count'], 1)
        self.assertEqual(stats['article_body_success_count'], 1)


if __name__ == '__main__':
    unittest.main()
