import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

import web_app  # noqa: E402
import db_store  # noqa: E402
import xlsx_source_test  # noqa: E402


class BriefingTableTest(unittest.TestCase):
    def test_article_published_date_display_keeps_date_only(self) -> None:
        cases = [
            ("2026-05-18T16:44:40+00:00", "2026-05-18"),
            ("2026-05-18T16:44:40Z", "2026-05-18"),
            ("2026-05-18", "2026-05-18"),
            ("", ""),
            ("not a date", "not a date"),
        ]

        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(web_app.format_article_published_date(value), expected)

    def test_italy_ig_related_search_uses_commerce_qualified_terms(self) -> None:
        query = xlsx_source_test.build_related_news_search_query(
            "IG",
            None,
            "",
            {},
            '(offerta OR sconto OR promozione)',
            country_code="italy",
        )

        self.assertIn("Instagram", query)
        self.assertIn('"Instagram Shop"', query)
        self.assertIn('"Instagram ecommerce"', query)
        self.assertIn('"shopping su Instagram"', query)
        self.assertIn('"creator shop"', query)
        self.assertIn("(offerta OR sconto OR promozione)", query)
        self.assertIn("(Italy OR Italian OR Italia", query)

    def test_balanced_recall_adds_more_related_keyword_blocks(self) -> None:
        strict_blocks = xlsx_source_test.normalize_related_news_search_keyword_blocks(
            "",
            "italy",
            recall_mode="strict",
        )
        balanced_blocks = xlsx_source_test.normalize_related_news_search_keyword_blocks(
            "",
            "italy",
            recall_mode="balanced",
        )

        self.assertGreater(len(balanced_blocks), len(strict_blocks))
        balanced_text = "\n".join(block for _, block in balanced_blocks)
        self.assertIn("pagamenti", balanced_text)
        self.assertIn("seller", balanced_text)
        self.assertIn("social commerce", balanced_text)

    def test_legacy_default_survey_prompt_migrates_to_current_default(self) -> None:
        current_default = xlsx_source_test.default_survey_ai_system_prompt("italy")
        legacy_like_prompt = (
            current_default
            .replace("industry_trend_flag", "")
            .replace("industry_trend_category", "")
            .replace("industry_trend_impact", "")
            .replace("industry_trend_reason", "")
        )

        state = web_app.build_news_filter_state(
            {"survey_system_prompt": legacy_like_prompt},
            country_code="italy",
        )

        self.assertEqual(state["survey_system_prompt"], current_default)
        self.assertEqual(state["survey_system_prompt_source"], "system_default")
        self.assertIn("系统默认提示词", state["survey_system_prompt_source_label"])
        self.assertIn("industry_trend_flag", state["survey_system_prompt"])

    def test_custom_survey_prompt_is_not_migrated(self) -> None:
        custom_prompt = "请只保留和意大利平台 NPS 指标明确相关的新闻，并用 JSON 返回。"

        state = web_app.build_news_filter_state(
            {"survey_system_prompt": custom_prompt},
            country_code="italy",
        )

        self.assertEqual(state["survey_system_prompt"], custom_prompt)
        self.assertEqual(state["survey_system_prompt_source"], "custom")
        self.assertIn("自定义提示词", state["survey_system_prompt_source_label"])

    def test_ai_filter_persists_confidence(self) -> None:
        rows = [
            {
                "title": "Amazon expands cash payment methods in Italy",
                "summary": "The platform adds cash payment points for shoppers.",
                "published_at": "2026-05-18",
                "platform_label": "Amazon",
                "country_code": "italy",
            }
        ]
        indicators = [
            xlsx_source_test.SurveyIndicator(
                dimension="Features",
                question_id="B5_18",
                prompt_en="The platform supports my preferred payment methods.",
                prompt_zh="[平台]支持我偏好的支付方式",
            )
        ]

        def fake_call_survey_filter_api(messages, api_url, api_key, api_model):
            return {
                "decisions": [
                    {
                        "article_id": "0",
                        "relevant": True,
                        "matched_dimensions": ["Features"],
                        "matched_question_ids": ["B5_18"],
                        "reason": "现金支付方式扩展会影响消费者对支付便利性的评价。",
                        "confidence": "medium",
                    }
                ]
            }

        original_call = xlsx_source_test.call_survey_filter_api
        try:
            xlsx_source_test.call_survey_filter_api = fake_call_survey_filter_api
            filtered, _, _, stats = xlsx_source_test.apply_survey_indicator_filter(
                rows,
                indicators,
                mode="ai",
                api_url="https://example.com/v1/chat/completions",
                api_key="key",
                api_model="model",
                country_code="italy",
                recall_mode="balanced",
            )
        finally:
            xlsx_source_test.call_survey_filter_api = original_call

        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0]["survey_filter_confidence"], "medium")
        self.assertEqual(stats["recall_mode"], "balanced")

    def test_low_volume_fill_marks_medium_confidence_rows(self) -> None:
        indicators = [
            xlsx_source_test.SurveyIndicator(
                dimension="Features",
                question_id="B5_18",
                prompt_en="The platform supports my preferred payment methods.",
                prompt_zh="[平台]支持我偏好的支付方式",
            )
        ]
        source_rows = [
            {
                "title": "Amazon Italy expands cash payment methods",
                "summary": "Amazon adds payment options for Italian shoppers.",
                "published_at": "2026-05-18",
                "platform_label": "Amazon",
                "matched_brands": ["Amazon"],
                "country_code": "italy",
            }
        ]

        filled, stats = xlsx_source_test.apply_low_volume_fill(
            source_rows,
            [],
            indicators,
            country_code="italy",
            recall_mode="balanced",
            target_per_brand_week=7,
        )

        self.assertEqual(len(filled), 1)
        self.assertEqual(stats["added_count"], 1)
        self.assertEqual(filled[0]["volume_fill"], "true")
        self.assertEqual(filled[0]["survey_filter_confidence"], "medium")

    def test_italy_ig_results_require_commerce_context(self) -> None:
        noisy_article = {
            "title": "Instagram lancia una nuova app fotografica in Italia",
            "summary": "La nuova funzione social arriva agli utenti italiani.",
        }
        commerce_article = {
            "title": "Instagram Shop cresce in Italia con nuovi strumenti per creator shop",
            "summary": "La piattaforma punta su social commerce e acquisti su Instagram.",
        }

        self.assertFalse(xlsx_source_test.article_matches_instagram_commerce_context(noisy_article, "italy"))
        self.assertTrue(xlsx_source_test.article_matches_instagram_commerce_context(commerce_article, "italy"))

    def test_italy_platform_aliases_are_not_misread_as_custom_brands(self) -> None:
        state = web_app.build_news_platform_ui_state(
            "Instagram\nTikTok Shop\nTemu\nAmazon\nAliExpress",
            "italy",
        )

        self.assertEqual(state["selected_known_platforms"], ["IG", "TTS", "TEMU", "Amazon"])
        self.assertEqual(state["custom_platforms_text"], "AliExpress")

    def test_news_crawler_argv_canonicalizes_platform_aliases_for_country(self) -> None:
        form = {
            "country_code": "italy",
            "date_mode": "days",
            "days": "7",
            "translate_to": "zh-CN",
            "output_dir": "outputs",
            "news_builtin_platforms": ["Amazon", "Instagram"],
            "news_custom_platforms": "TikTok Shop\nTemu",
            "news_sides": ["media"],
            "survey_filter_mode": "keyword",
        }

        class FormData(dict):
            def getlist(self, key):
                value = self.get(key)
                if isinstance(value, list):
                    return value
                return [] if value is None else [value]

        argv, news_form_state, _ = web_app.build_news_crawler_argv(FormData(form))

        self.assertIn("IG", news_form_state["news_platforms_text"].splitlines())
        self.assertIn("TTS", news_form_state["news_platforms_text"].splitlines())
        self.assertIn("TEMU", news_form_state["news_platforms_text"].splitlines())
        self.assertNotIn("Instagram", news_form_state["news_platforms_text"].splitlines())
        self.assertNotIn("TikTok Shop", news_form_state["news_platforms_text"].splitlines())
        self.assertIn("--platform", argv)

    def test_saved_text_configs_keep_previous_version(self) -> None:
        original_settings_path = web_app.APP_SETTINGS_PATH
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                web_app.APP_SETTINGS_PATH = Path(temp_dir) / "settings.json"
                web_app.persist_app_settings(
                    country_code="italy",
                    news_filter_state={
                        "related_news_search_keywords": "(old related keywords)",
                        "report_search_keywords": "(old report keywords)",
                        "survey_system_prompt": "Old custom survey prompt",
                    },
                )
                web_app.persist_app_settings(
                    country_code="italy",
                    news_filter_state={
                        "related_news_search_keywords": "(new related keywords)",
                        "report_search_keywords": "(new report keywords)",
                        "survey_system_prompt": "New custom survey prompt",
                    },
                )

                country_settings = web_app.read_country_settings("italy")
                self.assertEqual(country_settings["related_news_search_keywords_previous"], "(old related keywords)")
                self.assertEqual(country_settings["report_search_keywords_previous"], "(old report keywords)")
                self.assertEqual(country_settings["survey_system_prompt_previous"], "Old custom survey prompt")

                web_app.save_admin_password("password123")
                client = web_app.app.test_client()
                with client.session_transaction() as session:
                    session[web_app.AUTH_SESSION_KEY] = True
                keywords_response = client.post("/api/news-filter/keywords/previous", data={"country_code": "italy"})
                prompt_response = client.post("/api/news-filter/survey-prompt/previous", data={"country_code": "italy"})
                self.assertEqual(keywords_response.status_code, 200)
                self.assertEqual(prompt_response.status_code, 200)
                self.assertEqual(keywords_response.get_json()["related_news_search_keywords"], "(old related keywords)")
                self.assertEqual(prompt_response.get_json()["survey_system_prompt"], "Old custom survey prompt")
        finally:
            web_app.APP_SETTINGS_PATH = original_settings_path

    def test_previous_alias_endpoint_returns_saved_snapshot_or_clear_error(self) -> None:
        original_settings_path = web_app.APP_SETTINGS_PATH
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                web_app.APP_SETTINGS_PATH = Path(temp_dir) / "settings.json"
                web_app.save_admin_password("password123")
                client = web_app.app.test_client()
                with client.session_transaction() as session:
                    session[web_app.AUTH_SESSION_KEY] = True

                missing_response = client.post("/api/countries/aliases/previous", data={"country_code": "italy"})
                self.assertEqual(missing_response.status_code, 404)
                self.assertIn("暂无上一版", missing_response.get_json()["error"])

                web_app.set_country_previous_config_value(
                    "italy",
                    web_app.PLATFORM_ALIASES_PREVIOUS_FIELD,
                    "Amazon | Amazon.it\nTEMU | Temu Italy",
                )
                response = client.post("/api/countries/aliases/previous", data={"country_code": "italy"})
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                self.assertEqual(payload["text"], "Amazon | Amazon.it\nTEMU | Temu Italy")
                self.assertFalse(payload["saved"])
        finally:
            web_app.APP_SETTINGS_PATH = original_settings_path

    def test_article_browser_can_read_sqlite_source(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "test.db"
            original_db_path = db_store.DEFAULT_DB_PATH
            try:
                db_store.DEFAULT_DB_PATH = db_path
                db_store.save_run(
                    Path(temp_dir) / "italy_sqlite_run",
                    {"country_code": "italy", "country_label": "意大利"},
                    [{"title": "Before DB article", "platform_label": "IG"}],
                    [
                        {
                            "title": "After DB article",
                            "title_translated": "数据库新闻",
                            "platform_label": "Amazon",
                            "published_at": "2026-05-18T16:44:40+00:00",
                            "article_url": "https://example.com/db",
                            "survey_dimensions": "Features",
                        }
                    ],
                )

                state = web_app.build_article_browser_state(
                    {"article_source": "db:italy_sqlite_run:after", "article_limit": "all"},
                    country_code="italy",
                )

                self.assertTrue(state["file_exists"])
                self.assertEqual(state["selected_source"], "db:italy_sqlite_run:after")
                self.assertEqual(state["matching_rows"], 1)
                self.assertEqual(state["rows"][0]["title_original"], "After DB article")
                self.assertEqual(state["rows"][0]["published_at_display"], "2026-05-18")
            finally:
                db_store.DEFAULT_DB_PATH = original_db_path

    def test_article_browser_merges_manual_sqlite_articles(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "test.db"
            original_db_path = db_store.DEFAULT_DB_PATH
            try:
                db_store.DEFAULT_DB_PATH = db_path
                run_id = db_store.save_run(
                    Path(temp_dir) / "italy_manual_run",
                    {"country_code": "italy", "country_label": "Italy"},
                    [],
                    [{"title": "Crawler article", "platform_label": "Amazon", "published_at": "2026-05-18"}],
                )
                manual_id, _ = db_store.save_manual_article(
                    run_id,
                    "italy",
                    {
                        "title": "Manual article",
                        "platform_label": "SHEIN",
                        "article_url": "https://example.com/manual",
                        "published_at": "2026-05-19",
                        "manual_added": True,
                    },
                )

                state = web_app.build_article_browser_state(
                    {"article_source": "db:italy_manual_run:after", "article_limit": "all"},
                    country_code="italy",
                )

                self.assertEqual(state["matching_rows"], 2)
                manual_rows = [row for row in state["rows"] if row.get("manual_added")]
                self.assertEqual(len(manual_rows), 1)
                self.assertEqual(manual_rows[0]["manual_article_id"], manual_id)
                self.assertEqual(manual_rows[0]["platform_label"], "SHEIN")
            finally:
                db_store.DEFAULT_DB_PATH = original_db_path

    def sample_rows(self) -> list[dict[str, str]]:
        return [
            {
                "article_id": "article-up",
                "platform_label": "Amazon",
                "title_original": "Amazon expands cash payment points in Italy",
                "title": "Amazon expands cash payment points in Italy",
                "title_display_zh": "Amazon expands cash payment points in Italy",
                "summary": "Amazon expands payment options for Italian shoppers.",
                "source_name": "Example News",
                "article_url": "https://example.com/amazon-cash",
                "source_url": "https://example.com",
                "published_at": "2026-03-23",
                "category": "news",
                "survey_dimensions": "Features",
                "survey_question_ids": "B5_18",
                "survey_indicator_examples": "B5_18: payment options",
                "briefing_sentiment": "Positive",
                "briefing_sentiment_reason": "Payment convenience may improve platform perception.",
            },
            {
                "article_id": "article-down",
                "platform_label": "SHEIN",
                "title_original": "EU investigates illegal products on SHEIN",
                "title": "EU investigates illegal products on SHEIN",
                "title_display_zh": "EU investigates illegal products on SHEIN",
                "summary": "EU regulators investigate product compliance.",
                "source_name": "Example Wire",
                "article_url": "https://example.com/shein-eu",
                "source_url": "https://example.com",
                "published_at": "2026-04-02",
                "category": "news",
                "survey_dimensions": "Quality | Customer / post-purchase service",
                "survey_question_ids": "B1_3 | B7_5",
                "survey_indicator_examples": "B1_3: seller legitimacy | B7_5: consumer protection",
                "briefing_sentiment": "Negative",
                "briefing_sentiment_reason": "Regulatory scrutiny may hurt trust perception.",
            },
        ]

    def test_briefing_table_uses_template_and_maps_nps_direction(self) -> None:
        rows = self.sample_rows()

        def fake_call_survey_filter_api(messages, api_url, api_key, api_model):
            payload = json.loads(messages[1]["content"])
            items = []
            for article in payload["articles"]:
                article_id = article["article_id"]
                direction = "Up" if article_id == "article-up" else "Down"
                metric = "Trust" if article_id == "article-down" else "Content"
                items.append(
                    {
                        "article_id": article_id,
                        "Key User Segment Affected": "Test segment",
                        "Core Summary": f"Summary for {article_id}",
                        "Sentiment (Positive / Neutral / Negative)": "Negative" if direction == "Up" else "Positive",
                        "Affected Side (Buyer / Seller / Both)": "Buyer",
                        "Impact Direction on NPS (Up / Down / Neutral)": direction,
                        "Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)": metric,
                        "Impact Level (High / Medium / Low)": "Medium",
                        "Short-term or Lagged Effect": "Immediate",
                        "Quant Linkage Hypothesis": f"Hypothesis for {article_id}",
                        "Analyst Note": f"Analyst note for {article_id}",
                        "Include in Client Readout (Y/N)": "Y",
                        "Follow-up Needed (Y/N)": "Y",
                    }
                )
            return {"items": items}

        original_call = web_app.xlsx_source_test.call_survey_filter_api
        original_output_dir = web_app.BRIEFING_TABLE_OUTPUT_DIR
        try:
            web_app.xlsx_source_test.call_survey_filter_api = fake_call_survey_filter_api
            with tempfile.TemporaryDirectory() as temp_dir:
                web_app.BRIEFING_TABLE_OUTPUT_DIR = Path(temp_dir)
                output_path = web_app.build_article_briefing_table(
                    rows,
                    country_code="italy",
                    api_settings={
                        "survey_api_url": "https://example.com/v1/chat/completions",
                        "survey_api_key": "test-key",
                        "survey_api_model": "test-model",
                    },
                )

                self.assertTrue(output_path.exists())
                workbook = load_workbook(output_path)
                worksheet = workbook["Sheet1"]
                try:
                    headers = [worksheet.cell(1, col).value for col in range(1, 23)]
                    self.assertEqual(headers, web_app.BRIEFING_TABLE_HEADERS)
                    self.assertEqual(worksheet.max_row, 3)

                    titles = [worksheet.cell(row, 7).value for row in range(2, 4)]
                    self.assertEqual(titles, [rows[0]["title_display_zh"], rows[1]["title_display_zh"]])
                    self.assertNotIn("TikTok Shop 日本站加强直播带货工具与内容推荐", titles)

                    self.assertEqual(worksheet.cell(2, 4).value, "↑")
                    self.assertEqual(worksheet.cell(3, 4).value, "↓")
                    self.assertEqual(worksheet.cell(2, 13).value, "Positive")
                    self.assertEqual(worksheet.cell(3, 13).value, "Negative")
                    self.assertEqual(worksheet.cell(2, 15).value, "Up")
                    self.assertEqual(worksheet.cell(3, 15).value, "Down")
                finally:
                    workbook.close()
        finally:
            web_app.xlsx_source_test.call_survey_filter_api = original_call
            web_app.BRIEFING_TABLE_OUTPUT_DIR = original_output_dir

    def test_briefing_ai_failure_stats_are_reported(self) -> None:
        rows = self.sample_rows()

        def failing_call_survey_filter_api(messages, api_url, api_key, api_model):
            raise RuntimeError("mock AI outage")

        original_call = web_app.xlsx_source_test.call_survey_filter_api
        try:
            web_app.xlsx_source_test.call_survey_filter_api = failing_call_survey_filter_api
            stats: dict[str, int] = {}
            fields = web_app.generate_briefing_ai_fields(
                rows,
                country_code="italy",
                api_settings={
                    "survey_api_url": "https://example.com/v1/chat/completions",
                    "survey_api_key": "test-key",
                    "survey_api_model": "test-model",
                },
                batch_size=10,
                stats=stats,
            )

            self.assertEqual(stats["batch_count"], 1)
            self.assertEqual(stats["failed_batch_count"], 1)
            self.assertEqual(stats["failed_row_count"], 2)
            self.assertEqual(stats["ai_completed_row_count"], 0)
            self.assertIn("AI生成失败，需人工复核", fields["article-up"]["Analyst Note"])
            self.assertEqual(fields["article-up"]["Impact Direction on NPS (Up / Down / Neutral)"], "Neutral")
        finally:
            web_app.xlsx_source_test.call_survey_filter_api = original_call

    def test_briefing_ai_payload_and_defaults_use_industry_trend_fields(self) -> None:
        row = {
            "article_id": "trend-1",
            "platform_label": "TTS",
            "title_display_zh": "意大利近五分之一消费者使用 TikTok Shop",
            "summary": "报告称 TikTok Shop 在意大利社交电商采用率提升。",
            "survey_dimensions": "Content",
            "survey_indicator_examples": "社交电商使用率提升，可能影响平台内容体验。",
            "briefing_sentiment": "Positive",
            "industry_trend_flag": "true",
            "industry_trend_category": "market_adoption",
            "industry_trend_impact": "Positive",
            "industry_trend_reason": "新闻涉及 TikTok Shop 在意大利的市场采用率提升。",
        }

        messages = web_app.build_briefing_ai_messages([row], "italy")
        payload = json.loads(messages[1]["content"])
        article = payload["articles"][0]
        self.assertTrue(article["industry_trend_flag"])
        self.assertEqual(article["industry_trend_category"], "market_adoption")
        self.assertTrue(any("industry_trend_flag" in rule for rule in payload["rules"]))

        defaults = web_app.default_briefing_ai_fields(row, country_code="italy")
        self.assertEqual(defaults["Impact Direction on NPS (Up / Down / Neutral)"], "Up")
        self.assertIn("行业趋势标记", defaults["Analyst Note"])

    def test_briefing_sentiment_fields_are_generated_for_after_rows(self) -> None:
        rows = [
            {
                "platform": "Amazon",
                "title": "Amazon expands cash payment points in Italy",
                "summary": "Payment convenience improves for shoppers.",
                "survey_dimensions": "Features",
            },
            {
                "platform": "SHEIN",
                "title": "EU investigates illegal products on SHEIN",
                "summary": "Regulators investigate product safety and compliance.",
                "survey_dimensions": "Quality",
            },
        ]

        def fake_call_filter_api(messages, api_url, api_key, api_model):
            payload = json.loads(messages[1]["content"])
            self.assertEqual(len(payload["articles"]), 2)
            return {
                "items": [
                    {"article_id": "0", "sentiment": "Positive", "reason": "支付便利性提升。"},
                    {"article_id": "1", "sentiment": "Negative", "reason": "非法商品调查损害信任。"},
                ]
            }

        updated_rows, stats = xlsx_source_test.add_briefing_sentiment_fields(
            rows,
            api_url="https://example.com/v1/chat/completions",
            api_key="test-key",
            api_model="test-model",
            country_code="italy",
            batch_size=10,
            call_filter_api=fake_call_filter_api,
        )

        self.assertIs(updated_rows, rows)
        self.assertEqual(stats["request_count"], 1)
        self.assertEqual(stats["evaluated_count"], 2)
        self.assertEqual(stats["api_error_count"], 0)
        self.assertEqual(rows[0]["briefing_sentiment"], "Positive")
        self.assertEqual(rows[1]["briefing_sentiment"], "Negative")
        self.assertEqual(rows[0]["briefing_sentiment_reason"], "支付便利性提升。")

    def test_briefing_sentiment_defaults_to_neutral_without_api_config(self) -> None:
        rows = [{"title": "Ambiguous article", "briefing_sentiment": "bad-value"}]

        updated_rows, stats = xlsx_source_test.add_briefing_sentiment_fields(
            rows,
            api_url="",
            api_key="",
            api_model="",
            country_code="italy",
        )

        self.assertIs(updated_rows, rows)
        self.assertEqual(stats["request_count"], 0)
        self.assertEqual(stats["defaulted_count"], 1)
        self.assertEqual(rows[0]["briefing_sentiment"], "Neutral")

    def test_invalid_key_user_segment_is_rewritten_for_target_country(self) -> None:
        row = self.sample_rows()[0]
        item = {
            "Key User Segment Affected": "Chinese user segment",
            "Impact Direction on NPS (Up / Down / Neutral)": "Up",
        }

        normalized = web_app.normalize_briefing_ai_item(item, row, "italy")

        self.assertEqual(normalized["Key User Segment Affected"], "偏好现金支付的意大利买家")

    def test_china_user_segment_is_never_kept_for_foreign_news(self) -> None:
        row = self.sample_rows()[1]
        item = {
            "Key User Segment Affected": "中国用户",
            "Impact Direction on NPS (Up / Down / Neutral)": "Down",
        }

        normalized = web_app.normalize_briefing_ai_item(item, row, "italy")

        self.assertEqual(normalized["Key User Segment Affected"], "意大利品质与合规敏感买家")

    def test_valid_target_country_segment_is_kept(self) -> None:
        row = self.sample_rows()[0]
        item = {
            "Key User Segment Affected": "意大利 Prime 会员",
            "Impact Direction on NPS (Up / Down / Neutral)": "Up",
        }

        normalized = web_app.normalize_briefing_ai_item(item, row, "italy")

        self.assertEqual(normalized["Key User Segment Affected"], "意大利 Prime 会员")

    def test_segment_fallback_rules_cover_report_friendly_cases(self) -> None:
        cases = [
            (
                {
                    "platform_label": "Amazon",
                    "title_original": "Se hai Amazon Prime puoi scegliere subito 4 regali gratis",
                    "survey_dimensions": "Features",
                    "survey_indicator_examples": "Prime会员可获得免费礼物",
                },
                "意大利 Prime 会员",
            ),
            (
                {
                    "platform_label": "Amazon",
                    "title_original": "Amazon.it expands cash payment points across Italy",
                    "survey_dimensions": "Features",
                    "survey_indicator_examples": "现金支付点扩展",
                },
                "偏好现金支付的意大利买家",
            ),
            (
                {
                    "platform_label": "SHEIN",
                    "title_original": "EU investigates Shein illegal products and PFAS issues",
                    "survey_dimensions": "Quality | Customer / post-purchase service",
                    "survey_indicator_examples": "非法商品和PFAS合规调查",
                },
                "意大利品质与合规敏感买家",
            ),
            (
                {
                    "platform_label": "eBay",
                    "title_original": "eBay changes seller policy and merchant commission rules",
                    "survey_dimensions": "Customer / post-purchase service",
                    "survey_indicator_examples": "卖家政策变化",
                },
                "意大利平台卖家",
            ),
        ]

        for row, expected in cases:
            with self.subTest(expected=expected):
                self.assertEqual(web_app.infer_briefing_segment(row, "italy"), expected)


if __name__ == "__main__":
    unittest.main()
