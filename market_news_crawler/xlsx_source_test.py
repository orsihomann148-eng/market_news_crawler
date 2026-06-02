#!/usr/bin/env python3
from __future__ import annotations

import argparse
import ast
import csv
import difflib
import json
import os
import re
import time
from copy import deepcopy
import urllib.parse
import xml.etree.ElementTree as ET
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import dedupe as dedupe_utils
import db_store
import output_writer
import survey_filter as survey_filter_utils
from country_config import (
    DEFAULT_COUNTRY_CODE,
    available_country_codes,
    country_dict_setting,
    country_list_setting,
    country_setting,
    default_country_file_paths,
    get_country_config,
    normalize_country_code,
)

format_brand_labels = dedupe_utils.format_brand_labels
normalize_article_title_for_dedupe = dedupe_utils.normalize_article_title_for_dedupe
strip_title_source_suffix = dedupe_utils.strip_title_source_suffix
from news_crawler import (
    DEFAULT_SITE_CREDENTIALS_PATH,
    NPS_DIMENSION_RULES,
    SUMMARY_MIN_PARAGRAPH_CHARS,
    TOKYO_TZ,
    TextTranslator,
    add_translation_fields,
    apply_site_credentials,
    abs_url,
    build_session,
    clean_text,
    contains_chinese_chars,
    detect_source_status,
    extract_meta_description,
    fetch,
    guess_country,
    html_to_text,
    infer_nps_dimension,
    load_site_credentials,
    merge_summary_candidates,
    normalize_translation_target,
    parse_date_range,
    parse_dt,
    trim_summary_text,
)


URL_PATTERN = re.compile(r"https?://[^\s\n\u3000]+")
FEED_TYPES = {
    "application/rss+xml",
    "application/atom+xml",
    "application/xml",
    "text/xml",
}
COMMON_FEED_PATHS = ["/feed", "/rss", "/feed.xml", "/rss.xml", "/atom.xml"]
DEFAULT_COUNTRY_FILE_PATHS = default_country_file_paths(DEFAULT_COUNTRY_CODE)
DEFAULT_EXTRA_SOURCES_PATH = DEFAULT_COUNTRY_FILE_PATHS["extra_sources_path"]
DEFAULT_ADAPTER_CONFIGS_PATH = DEFAULT_COUNTRY_FILE_PATHS["adapter_configs_path"]
DEFAULT_SURVEY_FILTER_MODE = "keyword"
DEFAULT_SURVEY_API_URL_ENV = "NEWS_SURVEY_API_URL"
DEFAULT_SURVEY_API_KEY_ENV = "NEWS_SURVEY_API_KEY"
DEFAULT_SURVEY_API_MODEL_ENV = "NEWS_SURVEY_API_MODEL"
DEFAULT_SURVEY_AI_BATCH_SIZE = 8
DEFAULT_RECALL_MODE = "balanced"
RECALL_MODE_CHOICES = {"strict", "balanced"}
RECALL_TARGET_PER_BRAND_WEEK = 7
SOURCE_SHORT_WINDOW_DAYS = 8
SOURCE_SHORT_WINDOW_MAX_TERMS_PER_PLATFORM = 2
SOURCE_SHALLOW_MAX_LINKS = 6
DEDICATED_GOOGLE_PLATFORM_LABELS = {"TEMU", "TTS", "IG"}
BROAD_ENTRY_PLATFORM_LABELS = {"TEMU", "TTS", "IG"}
BROAD_ENTRY_MARKET_SCOPE = "europe_or_global_relevant"
NEWS_PROGRESS_SETUP = 3
NEWS_PROGRESS_CRAWL_START = 5
NEWS_PROGRESS_CRAWL_END = 26
NEWS_PROGRESS_POST_CRAWL = 28
NEWS_PROGRESS_PROMO_END = 34
NEWS_PROGRESS_TRANSLATION_START = 38
NEWS_PROGRESS_TRANSLATION_AI_CAP = 72
NEWS_PROGRESS_TRANSLATION_KEYWORD_CAP = 80
NEWS_PROGRESS_FILTER_AI_START = 76
NEWS_PROGRESS_FILTER_AI_CAP = 90
NEWS_PROGRESS_FILTER_KEYWORD_START = 84
NEWS_PROGRESS_FILTER_KEYWORD_CAP = 92
NEWS_PROGRESS_DEDUPE_PREP = 92
NEWS_PROGRESS_DEDUPE_CAP = 96
NEWS_PROGRESS_OUTPUT_PREP = 97
NEWS_PROGRESS_JSON_WRITE = 98
NEWS_PROGRESS_CSV_WRITE = 99
FOREIGN_MARKET_TERMS = ["indonesia", "indonesian", "印度尼西亚", "印尼", "greece", "greek", "希腊", "india", "indian", "印度", "thailand", "thai", "泰国", "vietnam", "vietnamese", "越南", "malaysia", "malaysian", "马来西亚", "singapore", "singaporean", "新加坡", "philippines", "philippine", "菲律宾", "usa", "u.s.", "united states", "american", "美国", "uk", "united kingdom", "britain", "british", "英国", "europe", "european", "欧洲", "germany", "german", "德国", "france", "french", "法国", "italy", "italian", "意大利", "spain", "spanish", "西班牙", "korea", "south korea", "korean", "韩国", "china", "chinese", "中国", "hong kong", "香港", "taiwan", "台湾", "australia", "australian", "澳大利亚", "canada", "canadian", "加拿大", "brazil", "brazilian", "巴西", "mexico", "mexican", "墨西哥", "turkey", "turkish", "土耳其", "uae", "dubai", "阿联酋", "迪拜", "saudi", "saudi arabia", "沙特"]
LOW_IMPACT_LAUNCH_TERMS = ["launch", "launches", "launched", "release", "released", "introduce", "introduced", "introduces", "debut", "debuts", "unveil", "unveils", "new product", "new item", "推出", "发布", "上市", "上新", "発売", "新発売", "新商品", "新产品", "新作"]
LOW_IMPACT_PRODUCT_TERMS = ["product", "item", "sku", "collection", "collaboration", "capsule", "phone", "smartphone", "bag", "shoe", "shoes", "cosmetic", "cosmetics", "beauty", "lipstick", "snack", "toy", "model", "series", "merch", "merchandise", "单品", "商品", "款式", "系列", "联名", "新品"]
SINGLE_PRODUCT_PROMO_TERMS = [
    "deal", "deals", "sale", "on sale", "discount", "coupon", "special price", "flash sale", "price drop", "markdown",
    "cheap", "cheaper", "lowest price", "good deal", "bargain", "promo", "promotion", "offer", "limited offer",
    "特价", "特価", "优惠", "優惠", "折扣", "值", "便宜", "好价", "好價", "促销", "促銷", "优惠价", "優惠價",
    "値下げ", "安い", "お得", "セール", "タイムセール", "クーポン", "期間限定", "特別価格",
]
SINGLE_PRODUCT_CATEGORY_TERMS = [
    "earphone", "earphones", "earbud", "earbuds", "headphone", "headphones", "speaker", "camera", "action camera",
    "smartwatch", "watch", "tablet", "laptop", "notebook", "pc", "monitor", "tv", "television", "projector",
    "router", "keyboard", "mouse", "microphone", "console", "game", "vacuum", "fan", "air fryer", "fridge",
    "coffee maker", "printer", "bag", "shoe", "shoes", "sneaker", "shirt", "jacket", "coat", "dress",
    "cosmetic", "cosmetics", "skincare", "lipstick", "serum", "cream", "toy", "figure", "lego", "snack",
    "drink", "bottle", "protein", "supplement", "diaper", "stroller", "furniture",
    "耳机", "耳機", "耳塞", "耳麦", "耳麥", "相机", "相機", "运动相机", "運動相機", "手表", "手錶", "平板", "笔记本", "筆記本",
    "电脑", "電腦", "显示器", "顯示器", "电视", "電視", "投影", "键盘", "鍵盤", "鼠标", "滑鼠", "麦克风", "麥克風",
    "游戏机", "遊戲機", "吸尘器", "吸塵器", "风扇", "風扇", "冰箱", "咖啡机", "咖啡機", "打印机", "印表機",
    "包", "鞋", "球鞋", "服饰", "服飾", "外套", "裙", "化妆品", "化妝品", "护肤", "護膚", "口红", "口紅",
    "玩具", "积木", "積木", "零食", "饮料", "飲料", "奶粉", "纸尿裤", "紙尿褲", "婴儿车", "嬰兒車", "家具",
    "イヤホン", "ヘッドホン", "スピーカー", "カメラ", "アクションカメラ", "スマートウォッチ", "腕時計", "タブレット",
    "ノートpc", "ノートパソコン", "モニター", "テレビ", "プロジェクター", "キーボード", "マウス", "マイク",
    "ゲーム機", "掃除機", "扇風機", "冷蔵庫", "コーヒーメーカー", "プリンター", "バッグ", "シューズ", "スニーカー",
    "ジャケット", "コート", "ワンピース", "化粧品", "スキンケア", "口紅", "玩具", "おもちゃ", "菓子", "飲料", "家具",
]
BROAD_PROMO_SCOPE_TERMS = [
    "sitewide", "all users", "all products", "across the site", "platform-wide", "shopping event", "membership day",
    "prime day", "black friday", "cyber monday", "double 11", "double 12", "11.11", "12.12", "coupon campaign",
    "point campaign", "points up", "cashback event", "festival campaign",
    "全站", "全场", "全場", "全品类", "全品類", "平台活动", "平台活動", "大促", "满减", "滿減", "购物节", "購物節",
    "黑五", "雙11", "双11", "雙12", "双12", "优惠券活动", "優惠券活動", "积分活动", "積分活動", "返现活动", "返現活動",
    "全商品", "全ショップ", "全カテゴリ", "サイト全体", "全会員", "全會員", "大型セール", "ポイントアップ", "ポイント還元",
    "プライムデー", "ブラックフライデー", "サイバーマンデー", "メガ割", "スーパーsale", "お買い物マラソン",
]
MATERIAL_IMPACT_TERMS = ["policy", "rule", "terms", "fee", "fees", "shipping", "delivery", "logistics", "return", "refund", "customer service", "support", "coupon", "discount", "promotion", "price", "pricing", "payment", "checkout", "search", "recommendation", "algorithm", "app", "feature", "service", "sitewide", "all users", "livestream", "live shopping", "authentic", "counterfeit", "quality", "seller policy", "platform rule", "政策", "规则", "条款", "费用", "运费", "物流", "配送", "退货", "退款", "客服", "售后", "优惠券", "折扣", "促销", "价格", "支付", "搜索", "推荐", "算法", "功能", "服务", "直播", "正品", "假货", "质量"]
MEDIA_PLATFORM_CONTEXT_TERMS = MATERIAL_IMPACT_TERMS + [
    "ec",
    "ecommerce",
    "e-commerce",
    "online shop",
    "online store",
    "marketplace",
    "shopping",
    "shop",
    "mall",
    "retail",
    "merchant",
    "seller",
    "buyer",
    "consumer",
    "official store",
    "app store",
    "cart",
    "order",
    "coupon",
    "points",
    "bank",
    "wallet",
    "security",
    "fraud",
    "通販",
    "电商",
    "電商",
    "购物",
    "購物",
    "网购",
    "網購",
    "商家",
    "卖家",
    "賣家",
    "买家",
    "買家",
    "用户",
    "用戶",
    "消费者",
    "消費者",
    "商城",
    "店铺",
    "店鋪",
    "网店",
    "網店",
    "官方店",
    "购物车",
    "購物車",
    "订单",
    "訂單",
    "积分",
    "積分",
    "支付安全",
    "不正利用",
    "セキュリティ",
    "不正",
    "不正利用",
    "決済",
    "通販サイト",
    "オンラインショップ",
    "オンラインストア",
    "ショッピング",
    "モール",
    "出店",
    "購入",
    "注文",
    "楽天市場",
    "amazon.co.jp",
    "tiktok shop",
    "live commerce",
    "ライブコマース",
]
NOISY_TOPIC_TERMS = [
    "baseball",
    "soccer",
    "football",
    "basketball",
    "volleyball",
    "tennis",
    "golf",
    "boxing",
    "pitcher",
    "batter",
    "home run",
    "match",
    "tournament",
    "coach",
    "goal",
    "player",
    "league",
    "赛马",
    "比赛",
    "球员",
    "球队",
    "联赛",
    "投手",
    "打者",
    "本垒打",
    "野球",
    "野球部",
    "サッカー",
    "バスケ",
    "バレー",
    "テニス",
    "ゴルフ",
    "ボクシング",
    "試合",
    "大会",
    "選手",
    "監督",
    "打線",
    "先発",
    "登板",
    "本塁打",
    "プロ野球",
]
NETKEIZAI_SEARCH_PAGE_SIZE = 10
NETKEIZAI_MAX_SEARCH_PAGES = 6
ARTICLE_HINTS = [
    "/news",
    "/article",
    "/articles",
    "/press",
    "/business",
    "/story",
    "/stories",
    "/entry/",
    "/archives/",
    "/2026/",
    "/2025/",
]
EXCLUDE_HINTS = [
    "login",
    "signin",
    "signup",
    "account",
    "privacy",
    "terms",
    "contact",
    "help",
    "support",
    "faq",
    "cart",
    "wishlist",
    "javascript:",
    "#",
]

CANONICAL_PLATFORM_ALIASES = {
    "TikTok/TikTok Shop": [
        "tiktok shop",
        "tiktok shop",
        "tiktokshop",
        "tik tok shop",
        "ティックトックショップ",
        "tiktok",
        "ティックトック",
    ],
    "Amazon": [
        "amazon japan",
        "amazon japan",
        "amazon.co.jp",
        "アマゾンジャパン",
        "amazon",
        "アマゾン",
    ],
    "Rakuten Ichiba": [
        "rakuten ichiba",
        "楽天市場",
        "らくてん市場",
        "rakuten",
        "楽天",
    ],
    "Qoo10": [
        "qoo10",
        "큐텐",
        "キューテン",
    ],
    "TEMU": [
        "temu",
        "ティームー",
    ],
    "Shein": [
        "shein group",
        "she in",
        "SHEIN",
        "shein",
        "シーイン",
    ],
}

DISPLAY_TO_SOURCE_PLATFORM = {
    "TikTok Shop": "TikTok/TikTok Shop",
    "Amazon": "Amazon",
    "Rakuten Ichiba": "Rakuten Ichiba",
    "Qoo10": "Qoo10",
    "TEMU": "TEMU",
    "SHEIN": "Shein",
}

SOURCE_TO_DISPLAY_PLATFORM = {value: key for key, value in DISPLAY_TO_SOURCE_PLATFORM.items()}

PLATFORM_SEARCH_TERMS = {
    "TikTok/TikTok Shop": ["TikTok Shop", "TikTok"],
    "Amazon": ["Amazon Japan", "Amazon", "アマゾン"],
    "Rakuten Ichiba": ["楽天市場", "Rakuten", "Rakuten Ichiba"],
    "Qoo10": ["Qoo10"],
    "TEMU": ["TEMU"],
    "Shein": ["SHEIN", "Shein"],
}
PLATFORM_PROMO_EVENT_TERMS = {
    "TikTok/TikTok Shop": ["ライブコマース", "LIVE shopping", "coupon", "campaign", "春セール", "holiday sale"],
    "Amazon": ["amazon.co.jp", "タイムセール", "プライムデー", "ブラックフライデー", "新生活SALE", "初売り", "サイバーマンデー"],
    "Rakuten Ichiba": ["楽天スーパーSALE", "お買い物マラソン", "スーパーDEAL", "ポイントアップ", "39ショップ", "楽天イーグルス感謝祭", "超ポイントバック祭"],
    "Qoo10": ["メガ割", "メガポ", "タイムセール", "クーポン", "MOVE", "ビューティーフェス"],
    "TEMU": ["coupon", "クーポン", "sale", "キャンペーン", "black friday", "holiday deal"],
    "Shein": ["SHEIN sale", "coupon", "クーポン", "セール", "campaign", "holiday", "black friday"],
}
PROMO_SEARCH_ENGINE_CHOICES = {"google", "bing", "both"}
PROMO_SIGNAL_TERMS = [
    "sale",
    "discount",
    "coupon",
    "campaign",
    "promotion",
    "promo",
    "deal",
    "deals",
    "special offer",
    "limited offer",
    "limited time",
    "limited-time",
    "price cut",
    "markdown",
    "clearance",
    "reward",
    "rewards",
    "points",
    "point up",
    "point boost",
    "point campaign",
    "point back",
    "cashback",
    "offer",
    "offers",
    "festival",
    "shopping event",
    "flash sale",
    "bundle",
    "voucher",
    "セール",
    "フェア",
    "期間限定",
    "特典",
    "割引",
    "値引",
    "値下げ",
    "クーポン",
    "キャンペーン",
    "特価",
    "ポイント",
    "ポイントアップ",
    "ポイント還元",
    "還元",
    "优惠",
    "折扣",
    "促销",
    "活动",
    "优惠券",
    "满减",
    "double 11",
    "double 12",
    "11.11",
    "12.12",
    "black friday",
    "cyber monday",
    "prime day",
    "boxing day",
    "holiday sale",
    "seasonal sale",
    "spring sale",
    "summer sale",
    "winter sale",
    "new year sale",
    "christmas",
    "valentine",
    "white day",
    "mothers day",
    "father's day",
    "back to school",
    "bonus season",
    "gift",
    "gifting",
    "shopping festival",
    "shopping season",
    "festival campaign",
    "年末年始",
    "初売り",
    "新春",
    "福袋",
    "ゴールデンウィーク",
    "gw",
    "母の日",
    "父の日",
    "敬老の日",
    "クリスマス",
    "年末",
    "年始",
    "お盆",
    "バレンタイン",
    "ホワイトデー",
    "新生活",
    "新生活応援",
    "入学",
    "卒業",
    "夏休み",
    "冬休み",
    "双11",
    "双十二",
    "双12",
    "618",
    "节日",
    "節日",
    "节庆",
    "節慶",
    "礼物",
    "禮物",
    "送礼",
    "送禮",
    "消费季",
    "消費期",
]
SUPPLEMENTAL_SEARCH_TYPE_CHOICES = {"related_news", "report_ranking"}

REPORT_SIGNAL_TERMS = [
    "ranking",
    "rankings",
    "rank",
    "benchmark",
    "report",
    "survey",
    "study",
    "index",
    "data",
    "stats",
    "market share",
    "leaderboard",
    "comparison",
    "comparative",
    "classment",
    "classement",
    "rapport",
    "barometre",
    "baromètre",
    "etude",
    "étude",
    "donnees",
    "données",
    "comparatif",
    "palmares",
    "palmarès",
    "排名",
    "榜单",
    "报告",
    "数据",
    "调查",
    "指数",
    "统计",
    "市占率",
    "ランキング",
    "レポート",
    "調査",
    "データ",
    "指標",
]

@dataclass
class SourceEntry:
    platform: str
    side: str
    source_url: str


@dataclass
class SourceResult:
    platform: str
    side: str
    source_url: str
    fetch_url: str
    source_site: str
    final_url: str | None
    status: str
    http_status: int | None
    recent_article_count: int
    note: str
    elapsed_seconds: float = 0.0
    candidate_count: int = 0
    matched_brand_count: int = 0
    recent_yield: str = ""
    zero_yield_reason: str = ""
    source_recommendation: str = ""
    search_endpoint_status: str = ""
    selector_match_count: int = 0
    parsed_date_count: int = 0


@dataclass
class SurveyIndicator:
    dimension: str
    question_id: str
    prompt_en: str
    prompt_zh: str


@dataclass
class ExplicitMediaRun:
    status: str
    note: str
    articles: list[dict[str, Any]]
    final_url: str | None = None
    search_endpoint_status: str = ""
    selector_match_count: int = 0
    parsed_date_count: int = 0


@dataclass
class SurveyAIFilterDecision:
    relevant: bool
    matched_dimensions: list[str]
    matched_question_ids: list[str]
    reason: str
    confidence: str = "high"
    method: str = "ai"
    industry_trend_flag: bool = False
    industry_trend_category: str = ""
    industry_trend_impact: str = "Neutral"
    industry_trend_reason: str = ""


def interpolate_progress(start: int, end: int, completed: int, total: int) -> int:
    if total <= 0:
        return int(start)
    ratio = max(0.0, min(1.0, completed / max(total, 1)))
    return int(start) + int(ratio * max(0, int(end) - int(start)))


country_consumer_research_phrase = survey_filter_utils.country_consumer_research_phrase
country_market_terms = survey_filter_utils.country_market_terms
country_market_search_block = survey_filter_utils.country_market_search_block
country_prompt_platform_examples = survey_filter_utils.country_prompt_platform_examples
default_survey_ai_system_prompt = survey_filter_utils.default_survey_ai_system_prompt
legacy_default_survey_ai_system_prompts = survey_filter_utils.legacy_default_survey_ai_system_prompts
survey_system_prompt_has_current_default_markers = survey_filter_utils.survey_system_prompt_has_current_default_markers
looks_like_legacy_default_survey_system_prompt = survey_filter_utils.looks_like_legacy_default_survey_system_prompt
survey_system_prompt_source = survey_filter_utils.survey_system_prompt_source
survey_system_prompt_source_label = survey_filter_utils.survey_system_prompt_source_label
normalize_survey_system_prompt = survey_filter_utils.normalize_survey_system_prompt
promo_search_query_blocks = survey_filter_utils.promo_search_query_blocks
default_promo_search_keywords_text = survey_filter_utils.default_promo_search_keywords_text
related_news_query_blocks = survey_filter_utils.related_news_query_blocks
recall_enhanced_related_news_query_blocks = survey_filter_utils.recall_enhanced_related_news_query_blocks
build_related_news_query_blocks = survey_filter_utils.build_related_news_query_blocks
report_query_blocks = survey_filter_utils.report_query_blocks
default_related_news_search_keywords_text = survey_filter_utils.default_related_news_search_keywords_text
default_report_search_keywords_text = survey_filter_utils.default_report_search_keywords_text
_split_keyword_lines = survey_filter_utils._split_keyword_lines
split_keyword_line_into_blocks = survey_filter_utils.split_keyword_line_into_blocks
split_keyword_text_into_blocks = survey_filter_utils.split_keyword_text_into_blocks
keyword_auto_split_count = survey_filter_utils.keyword_auto_split_count
normalize_keyword_blocks_for_storage = survey_filter_utils.normalize_keyword_blocks_for_storage
_matches_legacy_country_keyword_text = survey_filter_utils._matches_legacy_country_keyword_text
_is_default_keyword_prefix_subset = survey_filter_utils._is_default_keyword_prefix_subset
normalize_promo_search_keywords_text = survey_filter_utils.normalize_promo_search_keywords_text
normalize_related_news_search_keywords_text = survey_filter_utils.normalize_related_news_search_keywords_text
normalize_report_search_keywords_text = survey_filter_utils.normalize_report_search_keywords_text


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="读取国家子目录中的 source_survey.xlsx 里的 3.1 来源，并测试近一周新闻抓取情况。")
    parser.add_argument("--country", choices=available_country_codes(), default=DEFAULT_COUNTRY_CODE, help="国家代码，控制网站库、搜索地区和默认时区。")
    parser.add_argument("--xlsx", default="", help="问卷/来源表格路径；留空则使用国家子目录中的默认文件")
    parser.add_argument("--days", type=int, default=7, help="向前回溯天数，默认 7")
    parser.add_argument("--start-date", help="起始日期，格式 YYYY-MM-DD。")
    parser.add_argument("--end-date", help="结束日期，格式 YYYY-MM-DD。默认是当前时间。")
    parser.add_argument("--timezone", default="", help="时区；留空则使用国家默认时区")
    parser.add_argument("--output-dir", default="outputs", help="输出目录")
    parser.add_argument("--extra-sources", default="", help="额外来源配置 JSON；留空则使用国家子目录中的默认文件")
    parser.add_argument("--adapter-configs", default="", help="站点适配配置 JSON；留空则使用国家子目录中的默认文件")
    parser.add_argument("--site-credentials", default="", help="站点账号/凭据配置 JSON；留空则使用国家子目录中的默认文件")
    parser.add_argument("--max-links-per-source", type=int, default=8, help="单个站点最多跟进多少个候选文章链接")
    parser.add_argument("--workers", type=int, default=8, help="并发抓取线程数")
    parser.add_argument("--translate-to", default="zh-CN", help="翻译输出语言，默认 zh-CN，可选 zh-CN / en")
    parser.add_argument(
        "--platform",
        dest="platforms",
        action="append",
        help="限制抓取的品牌，可重复传入；支持页面展示名、来源表里的品牌名或常见别名。",
    )
    parser.add_argument(
        "--sides",
        nargs="+",
        choices=["media", "buyer", "seller"],
        default=["media", "buyer", "seller"],
        help="优先抓取哪些来源侧别，默认媒体侧/买家侧/卖家侧全选。",
    )
    parser.add_argument(
        "--survey-filter-mode",
        choices=["keyword", "ai"],
        default=DEFAULT_SURVEY_FILTER_MODE,
        help="指标筛选方式：keyword=关键词匹配，ai=调用大模型理解新闻后筛选。",
    )
    parser.add_argument("--survey-api-url", default="", help="AI 筛选接口地址，建议填写到 /chat/completions。")
    parser.add_argument("--survey-api-key", default="", help="AI 筛选接口密钥。")
    parser.add_argument("--survey-api-model", default="", help="AI 筛选模型名。")
    parser.add_argument("--survey-system-prompt", default="", help="AI 理解筛选时的系统提示词；留空则使用内置默认提示词。")
    parser.add_argument("--survey-ai-workers", type=int, default=4, help="AI 筛选并发数，默认 4。")
    parser.add_argument("--survey-ai-batch-size", type=int, default=DEFAULT_SURVEY_AI_BATCH_SIZE, help="AI 批量筛选时每次请求包含多少条新闻，默认 8。")
    parser.add_argument(
        "--recall-mode",
        choices=sorted(RECALL_MODE_CHOICES),
        default=DEFAULT_RECALL_MODE,
        help="新闻召回策略：balanced=扩大检索覆盖并保留中等置信度相关新闻，strict=沿用更严格筛选。",
    )
    parser.add_argument("--promo-search", action="store_true", help="抓完来源网站后，再到搜索引擎按品牌补充搜索全网新闻，后续再由筛选链路判断相关性。")
    parser.add_argument("--promo-search-keywords", default="", help="全网补充搜索时使用的关键词块；留空则使用内置默认关键词。")
    parser.add_argument("--search-related-news", action="store_true", help="补充检索：按品牌 x 指标搜索相关新闻，并关注往期对比与变化。")
    parser.add_argument("--related-news-search-keywords", default="", help="相关新闻检索时使用的关键词块；留空则使用内置默认关键词。")
    parser.add_argument("--search-report-ranking", action="store_true", help="补充检索：按指标搜索品牌间同期排名、数据、报告与研究。")
    parser.add_argument("--report-search-keywords", default="", help="数据/报告检索时使用的关键词块；留空则使用内置默认关键词。")
    parser.add_argument(
        "--promo-search-engine",
        choices=sorted(PROMO_SEARCH_ENGINE_CHOICES),
        default="both",
        help="进行全网补充搜索时使用的搜索引擎：google / bing / both。",
    )
    args = parser.parse_args(argv)
    args.country = normalize_country_code(getattr(args, "country", DEFAULT_COUNTRY_CODE))
    country_config = get_country_config(args.country)
    args.xlsx = clean_text(getattr(args, "xlsx", "")) or str(country_config["xlsx_path"])
    args.timezone = clean_text(getattr(args, "timezone", "")) or str(country_config["timezone"])
    args.extra_sources = clean_text(getattr(args, "extra_sources", "")) or str(country_config["extra_sources_path"])
    args.adapter_configs = clean_text(getattr(args, "adapter_configs", "")) or str(country_config["adapter_configs_path"])
    args.site_credentials = clean_text(getattr(args, "site_credentials", "")) or str(country_config["site_credentials_path"])
    if bool(getattr(args, "promo_search", False)) and not bool(getattr(args, "search_related_news", False)) and not bool(getattr(args, "search_report_ranking", False)):
        args.search_related_news = True
        args.search_report_ranking = True
    if not clean_text(getattr(args, "related_news_search_keywords", "")) and clean_text(getattr(args, "promo_search_keywords", "")):
        args.related_news_search_keywords = clean_text(getattr(args, "promo_search_keywords", ""))
    return args


def normalize_url(url: str) -> str:
    normalized = clean_text(url)
    normalized = normalized.split("（", 1)[0]
    normalized = normalized.split(")", 1)[0] if " " in normalized else normalized
    normalized = normalized.rstrip("）),，。；;")
    return normalized


def canonicalize_source_url(url: str) -> str:
    lowered = url.lower().rstrip("/")
    replacements = {
        "https://www.aboutamazon.co.jp": "https://www.aboutamazon.jp",
        "https://aboutamazon.co.jp": "https://www.aboutamazon.jp",
    }
    return replacements.get(lowered, url)


def read_json_file(path: str) -> Any:
    file_path = Path(path)
    if not file_path.exists():
        return None
    return json.loads(file_path.read_text(encoding="utf-8"))


def load_sources_from_xlsx(path: str) -> list[SourceEntry]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries: list[SourceEntry] = []
    seen: set[tuple[str, str, str]] = set()
    for row in range(4, 10):
        platform = clean_text(ws.cell(row, 2).value)
        for column, side in [(3, "seller"), (4, "buyer"), (5, "media")]:
            cell_value = str(ws.cell(row, column).value or "")
            for raw_url in URL_PATTERN.findall(cell_value):
                source_url = normalize_url(raw_url)
                source_url = canonicalize_source_url(source_url)
                key = (platform, side, source_url)
                if not source_url or key in seen:
                    continue
                seen.add(key)
                entries.append(SourceEntry(platform=platform, side=side, source_url=source_url))
    return entries


def load_extra_sources(path: str) -> list[SourceEntry]:
    payload = read_json_file(path)
    if not payload:
        return []
    if isinstance(payload, dict):
        payload = payload.get("sources", [])
    if not isinstance(payload, list):
        return []
    entries: list[SourceEntry] = []
    seen: set[tuple[str, str, str]] = set()
    for row in payload:
        if not isinstance(row, dict):
            continue
        if row.get("active") is False:
            continue
        platform = clean_text(row.get("platform"))
        side = clean_text(row.get("side"))
        source_url = canonicalize_source_url(normalize_url(str(row.get("source_url") or "")))
        key = (platform, side, source_url)
        if side == "media" and not platform:
            platform = "General Media"
            key = (platform, side, source_url)
        if not platform or side not in {"media", "buyer", "seller"} or not source_url or key in seen:
            continue
        seen.add(key)
        entries.append(SourceEntry(platform=platform, side=side, source_url=source_url))
    return entries


def merge_source_entries(*groups: list[SourceEntry]) -> list[SourceEntry]:
    merged: list[SourceEntry] = []
    seen: set[tuple[str, str, str]] = set()
    for group in groups:
        for entry in group:
            key = (entry.platform, entry.side, canonicalize_source_url(entry.source_url))
            if key in seen:
                continue
            seen.add(key)
            merged.append(entry)
    return merged


def load_adapter_configs(path: str) -> dict[str, dict[str, Any]]:
    payload = read_json_file(path)
    if not isinstance(payload, dict):
        return {}
    normalized: dict[str, dict[str, Any]] = {}
    for domain, config in payload.items():
        if isinstance(config, dict):
            normalized[clean_text(domain).lower()] = config
    return normalized


def filter_sources(entries: list[SourceEntry], allowed_sides: list[str]) -> list[SourceEntry]:
    allowed = set(allowed_sides)
    return [entry for entry in entries if entry.side in allowed]


SURVEY_DIMENSION_KEYWORDS = {
    'Product quality': ['quality', 'authentic', 'counterfeit', 'description', 'consistent', 'safety', 'quality control', '品质', '正品', '正規品', '偽物', '质量', '描述一致', '安全'],
    'Seller quality': ['seller', 'merchant', 'trustworthy', 'legitimate', 'customer service', 'customer support', 'responsive', 'helpful', '卖家', '商家', '客服', '信赖', '信任', '服务'],
    'Product variety': ['variety', 'selection', 'assortment', 'brand', 'brands', 'trend', 'unique', 'interesting', 'recommendation', 'recommendations', 'preferences', 'discover', 'catalog', '品类', '品牌', '推荐', '偏好', '兴趣', '潮流', '丰富', '多样', '独特', '发现'],
    'Price': ['price', 'pricing', 'discount', 'coupon', 'promotion', 'sale', 'voucher', 'shipping cost', 'shipping voucher', 'competitive', 'prices are stable', '价格', '折扣', '优惠', '优惠券', '运费', '运费券', '促销', '售价', '价格稳定'],
    'Content': ['video', 'live', 'livestream', 'short video', 'short videos', 'shoppable', 'host', 'interactive', 'respond promptly', 'product introduction', 'purchase links', '内容', '直播', '短视频', '视频', '主播', '互动', '讲解', '内容丰富', '发现产品', '发现品牌'],
    'Logistics': ['shipping', 'delivery', 'delivered', 'tracking', 'parcel', 'courier', 'ship on time', 'estimated delivery', 'fulfillment', '物流', '配送', '发货', '送达', '追踪', '包裹', '快递', '按时发货', '运输'],
    'Post-purchase service': ['refund', 'return', 'cancel', 'cancellation', 'after-sales', 'post-purchase', 'return shipping fee', 'refund request', 'pick-up service', '退款', '退货', '取消订单', '售后', '退货运费', '上门取件'],
    'Product feature': ['feature', 'function', 'app', 'search', 'payment', 'checkout', 'membership', 'review', 'rating', 'ui', 'ux', 'security', 'recommendation engine', 'algorithm', '功能', '应用', '搜索', '支付', '结账', '会员', '评价', '评分', '安全', '推荐机制'],
}


def normalize_question_id(value: str) -> str:
    return clean_text(value).replace('（', '(').replace('）', ')')


def looks_like_question_id(value: str) -> bool:
    return bool(re.match(r'^[A-Za-z]\d+(?:[_-]\d+)?(?:\([^)]+\))?$', normalize_question_id(value)))


def normalize_survey_dimension(value: str) -> str:
    normalized = clean_text(value)
    normalized = re.sub(r'^[A-Za-z]\d+\s+', '', normalized).strip()
    return normalized


def load_survey_indicators_from_xlsx(path: str) -> list[SurveyIndicator]:
    workbook = load_workbook(path, data_only=True)
    for worksheet in workbook.worksheets:
        section_row = None
        for row_index in range(1, worksheet.max_row + 1):
            first_cell = clean_text(worksheet.cell(row_index, 1).value)
            if '资讯需映射到的指标' in first_cell:
                section_row = row_index
                break
        if section_row is None:
            continue

        indicators: list[SurveyIndicator] = []
        current_dimension = ''
        blank_streak = 0
        for row_index in range(section_row + 2, worksheet.max_row + 1):
            dimension = normalize_survey_dimension(worksheet.cell(row_index, 1).value)
            second_cell = clean_text(worksheet.cell(row_index, 2).value)
            third_cell = clean_text(worksheet.cell(row_index, 3).value)
            fourth_cell = clean_text(worksheet.cell(row_index, 4).value)
            fifth_cell = clean_text(worksheet.cell(row_index, 5).value)

            if looks_like_question_id(third_cell):
                question_id = normalize_question_id(third_cell)
                prompt_en = fourth_cell
                prompt_zh = fifth_cell
            else:
                question_id = normalize_question_id(second_cell)
                prompt_en = third_cell
                prompt_zh = fourth_cell
            if not any([dimension, question_id, prompt_en, prompt_zh]):
                blank_streak += 1
                if indicators and blank_streak >= 2:
                    break
                continue
            blank_streak = 0
            if dimension:
                current_dimension = dimension
            is_dimension_header = looks_like_question_id(question_id) and '_' not in question_id and not prompt_en and not prompt_zh
            if not current_dimension or not question_id or is_dimension_header or not (prompt_en or prompt_zh):
                continue
            indicators.append(SurveyIndicator(
                dimension=current_dimension,
                question_id=question_id,
                prompt_en=prompt_en,
                prompt_zh=prompt_zh,
            ))
        if indicators:
            return indicators
    return []


def group_survey_indicators_by_dimension(indicators: list[SurveyIndicator]) -> dict[str, list[SurveyIndicator]]:
    grouped: dict[str, list[SurveyIndicator]] = {}
    for indicator in indicators:
        grouped.setdefault(indicator.dimension, []).append(indicator)
    return grouped


def build_survey_indicator_question_lookup(survey_indicators: list[SurveyIndicator]) -> dict[str, SurveyIndicator]:
    lookup: dict[str, SurveyIndicator] = {}
    for indicator in survey_indicators:
        lookup[normalize_question_id(indicator.question_id).lower()] = indicator
    return lookup


def build_dimension_keyword_lookup(survey_indicators: list[SurveyIndicator]) -> dict[str, list[str]]:
    merged: dict[str, list[str]] = {dimension: list(keywords) for dimension, keywords in SURVEY_DIMENSION_KEYWORDS.items()}
    for dimension, keywords in NPS_DIMENSION_RULES:
        merged.setdefault(dimension, [])
        for keyword in keywords:
            normalized = clean_text(keyword)
            if normalized and normalized not in merged[dimension]:
                merged[dimension].append(normalized)
    for indicator in survey_indicators:
        merged.setdefault(indicator.dimension, [])
        for prompt_text in [indicator.prompt_en, indicator.prompt_zh]:
            normalized_prompt = clean_text(prompt_text)
            if not normalized_prompt:
                continue
            for keyword in re.findall(r"[A-Za-z][A-Za-z\-']{2,}|[一-鿿]{2,}|[぀-ヿ]{2,}", normalized_prompt):
                normalized = clean_text(keyword)
                if normalized and normalized.lower() not in {'platform', 'products', 'product', 'sellers', 'seller'} and normalized not in merged[indicator.dimension]:
                    merged[indicator.dimension].append(normalized)
    return merged


def survey_keyword_matches(text: str, keyword: str) -> bool:
    normalized_text = clean_text(text)
    normalized_keyword = clean_text(keyword)
    if not normalized_text or not normalized_keyword:
        return False
    if normalized_keyword.isascii():
        pattern = re.compile(rf'(?<![A-Za-z0-9]){re.escape(normalized_keyword)}(?![A-Za-z0-9])', re.IGNORECASE)
        return bool(pattern.search(normalized_text))
    return normalized_keyword in normalized_text


def detect_survey_dimensions(
    article: dict[str, Any],
    indicator_dimensions: set[str],
    dimension_keyword_lookup: dict[str, list[str]],
) -> list[str]:
    haystack = ' '.join(
        clean_text(str(article.get(field) or ''))
        for field in ['title', 'summary', 'category', 'title_translated', 'summary_translated', 'body_excerpt']
    )
    matched: list[str] = []

    inferred_dimensions: list[str] = []
    for title_key, summary_key in [('title', 'summary'), ('title_translated', 'summary_translated')]:
        inferred = infer_nps_dimension(str(article.get(title_key) or ''), str(article.get(summary_key) or '') or None)
        if inferred and inferred in indicator_dimensions and inferred not in inferred_dimensions:
            inferred_dimensions.append(inferred)

    for dimension in inferred_dimensions:
        matched.append(dimension)

    search_focus_dimension = clean_text(article.get('search_focus_dimension'))
    if search_focus_dimension in indicator_dimensions and search_focus_dimension not in matched:
        matched.append(search_focus_dimension)

    for dimension, keywords in dimension_keyword_lookup.items():
        if dimension not in indicator_dimensions or dimension in matched:
            continue
        hit_count = sum(1 for keyword in keywords if survey_keyword_matches(haystack, keyword))
        if hit_count >= 1:
            matched.append(dimension)
    return matched


def normalize_string_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value).strip()
        parsed_items = None
        if text.startswith('[') and text.endswith(']'):
            try:
                candidate = ast.literal_eval(text)
            except (SyntaxError, ValueError):
                candidate = None
            if isinstance(candidate, list):
                parsed_items = candidate
        raw_items = parsed_items if parsed_items is not None else re.split(r"[|,;\n]+", text)
    ordered: list[str] = []
    for item in raw_items:
        normalized = clean_text(str(item))
        if normalized and normalized not in ordered:
            ordered.append(normalized)
    return ordered


def normalize_chat_completions_url(url: str) -> str:
    normalized = clean_text(url).rstrip('/')
    if not normalized:
        return ''
    if '/chat/completions' in normalized:
        return normalized
    return f'{normalized}/chat/completions'


def resolve_survey_api_credentials(args: argparse.Namespace) -> tuple[str, str, str]:
    api_url = normalize_chat_completions_url(clean_text(args.survey_api_url) or clean_text(os.environ.get(DEFAULT_SURVEY_API_URL_ENV)))
    api_key = clean_text(args.survey_api_key) or clean_text(os.environ.get(DEFAULT_SURVEY_API_KEY_ENV))
    api_model = clean_text(args.survey_api_model) or clean_text(os.environ.get(DEFAULT_SURVEY_API_MODEL_ENV))
    return api_url, api_key, api_model


def normalize_ai_dimension_matches(values: Any, grouped_indicators: dict[str, list[SurveyIndicator]]) -> list[str]:
    lookup = {clean_text(dimension).lower(): dimension for dimension in grouped_indicators}
    matched: list[str] = []
    for raw in normalize_string_list(values):
        key = clean_text(raw).lower()
        dimension = lookup.get(key)
        if not dimension:
            for lookup_key, candidate in lookup.items():
                if key and (key in lookup_key or lookup_key in key):
                    dimension = candidate
                    break
        if dimension and dimension not in matched:
            matched.append(dimension)
    return matched


def normalize_ai_question_ids(values: Any, question_lookup: dict[str, SurveyIndicator]) -> list[str]:
    matched: list[str] = []
    for raw in normalize_string_list(values):
        key = normalize_question_id(raw).lower()
        indicator = question_lookup.get(key)
        if indicator and indicator.question_id not in matched:
            matched.append(indicator.question_id)
    return matched


def infer_question_ids_from_article(
    row: dict[str, Any],
    matched_dimensions: list[str],
    grouped_indicators: dict[str, list[SurveyIndicator]],
) -> list[str]:
    haystack = article_relevance_haystack(row).lower()
    if not haystack:
        return []

    question_hint_terms = {
        'payment_method': [
            'payment method', 'payment methods', 'pay in cash', 'cash payment', 'cash payments',
            'cash', 'contanti', 'pagamento', 'pagamenti', 'pagare', 'metodo di pagamento',
            'metodi di pagamento', 'contrassegno', 'checkout', '支付', '付款',
        ],
        'payment_security': [
            'secure payment', 'payment security', 'secure checkout', 'safe payment',
            'sicuro', 'sicura', 'sicurezza', '安全', '安心',
        ],
        'loyalty': [
            'loyalty', 'reward', 'rewards', 'cashback', 'membership',
            'fedelt', 'punti fedelt', 'programma fedelt', 'premio', 'premi', 'ricompensa', '会员', '积分',
        ],
        'recommendation': [
            'recommendation', 'recommendations', 'recommended', 'personalized',
            'preferences', 'interests', 'raccomand', 'consigli', '推荐', '偏好',
        ],
        'search': ['search', 'seller search', 'product search', 'ricerca', 'cerca', '检索', '搜索'],
        'order': ['order', 'orders', 'check my order', 'ordine', 'ordini', '订单'],
        'review': ['review', 'reviews', 'rating', 'ratings', 'recensioni', 'valutazioni', '评价'],
        'authenticity': [
            'copyright', 'copyrights', 'intellectual property', 'ip infringement', 'infringement',
            'counterfeit', 'fake', 'authentic', 'original', 'trademark', 'piracy',
            'contraffazione', 'contraffatti', 'autentico', 'autentici', 'violazione',
            'proprietà intellettuale', '版权', '侵权', '正品', '假货',
        ],
        'seller_trust': [
            'reliable', 'trustworthy', 'legitimate', 'legit', 'scam', 'fraud', 'safe',
            'affidabile', 'sicuro', 'sicura', 'truffa', 'frodi', 'legittimo',
            '可靠', '安全可靠吗', '信任', '合法',
        ],
        'description_match': [
            'matches the description', 'as described', 'product description', 'not as described',
            'descrizione', 'come descritto', 'corrisponde', '描述一致', '实物与描述',
        ],
    }
    question_id_hints = {
        'payment_method': ['B5_18'],
        'payment_security': ['B5_7', 'B5_17'],
        'loyalty': ['B5_16'],
        'recommendation': ['B5_1', 'B5_9', 'B5_10'],
        'search': ['B5_3'],
        'order': ['B5_4'],
        'review': ['B5_15'],
        'authenticity': ['B1_2'],
        'seller_trust': ['B1_3', 'B1_4'],
        'description_match': ['B1_7'],
    }

    candidate_ids: list[str] = []
    for hint_name, terms in question_hint_terms.items():
        if any(term in haystack for term in terms):
            for question_id in question_id_hints.get(hint_name, []):
                if question_id not in candidate_ids:
                    candidate_ids.append(question_id)

    payment_location_terms = [
        'cash payment', 'cash payments', 'pay in cash', 'physical points', 'payment points',
        'pickup points', 'collection points', 'punti in tutta italia', 'punti vendita',
        'punti fisici', 'pagamenti in contanti', 'pagare in contanti',
    ]
    loyalty_specific_terms = [
        'loyalty', 'reward', 'rewards', 'cashback', 'membership', 'fedelt',
        'punti fedelt', 'programma fedelt', 'premio', 'premi', 'ricompensa', '积分', '会员',
    ]
    if any(term in haystack for term in payment_location_terms) and not any(term in haystack for term in loyalty_specific_terms):
        candidate_ids = [question_id for question_id in candidate_ids if question_id != 'B5_16']

    allowed_question_ids = {
        indicator.question_id
        for dimension in matched_dimensions
        for indicator in grouped_indicators.get(dimension, [])
    }
    return [question_id for question_id in candidate_ids if question_id in allowed_question_ids]


def build_survey_indicator_examples(
    matched_dimensions: list[str],
    matched_question_ids: list[str],
    grouped_indicators: dict[str, list[SurveyIndicator]],
    question_lookup: dict[str, SurveyIndicator],
    *,
    fallback_reason: str = '',
) -> str:
    dimension_templates = {
        'Price': '新闻涉及价格、折扣、优惠券、积分或促销变化，可能直接影响用户对平台价格竞争力的感受。',
        'Logistics': '新闻涉及配送、物流或履约体验变化，可能直接影响用户对平台收货效率和服务稳定性的感受。',
        'Post-purchase service': '新闻涉及退换货、退款、客服或售后流程变化，可能直接影响用户对平台售后体验的感受。',
        'Content': '新闻涉及平台内容展示、直播、短视频或导购体验变化，可能直接影响用户对平台内容吸引力和种草效率的感受。',
        'Seller quality': '新闻涉及商家治理、卖家质量或平台信任机制，可能直接影响用户对平台商家可靠性的感受。',
        'Product quality': '新闻涉及正品、质量、安全或商品可靠性，可能直接影响用户对平台商品质量的感受。',
        'Product variety': '新闻涉及品牌、品类或商品丰富度变化，可能直接影响用户对平台可选商品范围的感受。',
        'Product feature': '新闻涉及搜索、支付、下单、推荐或平台功能流程变化，可能直接影响用户对平台使用便利性的感受。',
    }

    indicator_texts: list[str] = []
    for question_id in matched_question_ids:
        indicator = question_lookup.get(normalize_question_id(question_id).lower())
        if not indicator:
            continue
        prompt_text = indicator.prompt_zh or indicator.prompt_en
        if prompt_text:
            indicator_text = f'{indicator.question_id}: {prompt_text}'
            if indicator_text not in indicator_texts:
                indicator_texts.append(indicator_text)

    reason_text = clean_text(fallback_reason)
    example_texts = indicator_texts[:3]
    if reason_text and reason_text not in example_texts:
        example_texts.append(reason_text)

    for dimension in matched_dimensions:
        if len(example_texts) >= 4:
            break
        template = dimension_templates.get(dimension)
        if template and template not in example_texts:
            example_texts.append(template)
    return ' | '.join(example_texts[:4])


def build_survey_reason_fallback(
    row: dict[str, Any],
    matched_dimensions: list[str],
    matched_question_ids: list[str],
) -> str:
    title = clean_text(row.get('title_translated') or row.get('title') or '')
    dimensions = set(matched_dimensions)
    question_ids = set(matched_question_ids)
    if {'Quality', 'Customer / post-purchase service'} & dimensions or {'B1_2', 'B1_3', 'B1_4', 'B7_5'} & question_ids:
        return '新闻涉及平台商品合规、卖家合法性或消费者权益保障，可能影响消费者对平台可信度和售后保障的评价。'
    if 'Price' in dimensions or any(question_id.startswith('B3_') for question_id in question_ids):
        return '新闻涉及平台价格、折扣或促销机制，可能影响消费者对价格竞争力和优惠吸引力的评价。'
    if 'Features' in dimensions or any(question_id.startswith('B5_') for question_id in question_ids):
        return '新闻涉及平台功能、支付方式、会员权益或使用流程，可能影响消费者对平台便利性和功能吸引力的评价。'
    if 'Variety' in dimensions or any(question_id.startswith('B2_') for question_id in question_ids):
        return '新闻涉及平台商品品类或供给范围变化，可能影响消费者对商品丰富度和选择空间的评价。'
    if 'Logistics' in dimensions or any(question_id.startswith('B6_') for question_id in question_ids):
        return '新闻涉及配送、履约或收货体验，可能影响消费者对平台物流效率和稳定性的评价。'
    if 'Content' in dimensions:
        return '新闻涉及平台内容展示、导购或互动体验，可能影响消费者对平台内容吸引力的评价。'
    if title:
        return '新闻内容与命中的问卷指标相关，可能影响消费者对平台体验和整体评价的判断。'
    return ''


def enrich_row_with_survey_match(
    row: dict[str, Any],
    *,
    matched_dimensions: list[str],
    matched_question_ids: list[str],
    grouped_indicators: dict[str, list[SurveyIndicator]],
    question_lookup: dict[str, SurveyIndicator],
    explanation: str = '',
    method: str = 'keyword',
    confidence: str = 'high',
    volume_fill: bool = False,
    industry_trend_flag: bool = False,
    industry_trend_category: str = '',
    industry_trend_impact: str = '',
    industry_trend_reason: str = '',
) -> dict[str, Any]:
    canonical_dimensions: list[str] = []
    for dimension in matched_dimensions:
        cleaned = clean_text(dimension)
        if cleaned and cleaned not in canonical_dimensions:
            canonical_dimensions.append(cleaned)

    canonical_question_ids: list[str] = []
    for question_id in matched_question_ids:
        normalized_key = normalize_question_id(question_id).lower()
        indicator = question_lookup.get(normalized_key)
        value = indicator.question_id if indicator else normalize_question_id(question_id)
        if value and value not in canonical_question_ids:
            canonical_question_ids.append(value)

    if not canonical_dimensions and canonical_question_ids:
        for question_id in canonical_question_ids:
            indicator = question_lookup.get(normalize_question_id(question_id).lower())
            if indicator and indicator.dimension not in canonical_dimensions:
                canonical_dimensions.append(indicator.dimension)

    if not canonical_question_ids:
        for question_id in infer_question_ids_from_article(row, canonical_dimensions, grouped_indicators):
            if question_id not in canonical_question_ids:
                canonical_question_ids.append(question_id)

    enriched = dict(row)
    enriched['survey_relevant'] = 'yes'
    enriched['survey_dimensions'] = ' | '.join(canonical_dimensions)
    enriched['survey_question_ids'] = ' | '.join(canonical_question_ids)
    enriched['survey_indicator_count'] = len(canonical_question_ids) or len(canonical_dimensions)
    explanation_text = clean_text(explanation)
    if method.startswith('ai') and not explanation_text:
        explanation_text = build_survey_reason_fallback(enriched, canonical_dimensions, canonical_question_ids)
    if method.startswith('ai') and explanation_text:
        enriched['survey_ai_reason_raw'] = explanation_text
        enriched['survey_ai_reason_translated'] = explanation_text
    enriched['survey_indicator_examples'] = build_survey_indicator_examples(
        canonical_dimensions,
        canonical_question_ids,
        grouped_indicators,
        question_lookup,
        fallback_reason=explanation_text,
    )
    enriched['survey_filter_method'] = method
    enriched['survey_filter_confidence'] = normalize_survey_filter_confidence(confidence)
    enriched['volume_fill'] = 'true' if volume_fill else clean_text(enriched.get('volume_fill') or '')
    enriched = apply_industry_trend_fields(
        enriched,
        flag=industry_trend_flag,
        category=industry_trend_category,
        impact=industry_trend_impact,
        reason=industry_trend_reason,
    )
    return enriched


def keyword_filter_single_article(
    row: dict[str, Any],
    grouped_indicators: dict[str, list[SurveyIndicator]],
    indicator_dimensions: set[str],
    dimension_keyword_lookup: dict[str, list[str]],
    question_lookup: dict[str, SurveyIndicator],
    *,
    method: str = 'keyword',
    explanation: str = '',
) -> tuple[dict[str, Any] | None, list[str]]:
    matched_dimensions = detect_survey_dimensions(row, indicator_dimensions, dimension_keyword_lookup)
    if not matched_dimensions:
        return None, []

    matched_question_ids: list[str] = []
    for question_id in infer_question_ids_from_article(row, matched_dimensions, grouped_indicators):
        if question_id not in matched_question_ids:
            matched_question_ids.append(question_id)
    if not matched_question_ids:
        return None, []

    enriched = enrich_row_with_survey_match(
        row,
        matched_dimensions=matched_dimensions,
        matched_question_ids=matched_question_ids,
        grouped_indicators=grouped_indicators,
        question_lookup=question_lookup,
        explanation=explanation,
        method=method,
    )
    return enriched, matched_dimensions


def article_relevance_haystack(row: dict[str, Any]) -> str:
    return ' '.join(
        clean_text(str(row.get(field) or ''))
        for field in ['title', 'summary', 'title_translated', 'summary_translated', 'category', 'body_excerpt']
    )


def contains_any_relevance_keyword(text: str, keywords: list[str]) -> bool:
    return any(survey_keyword_matches(text, keyword) for keyword in keywords)


def is_foreign_market_only_article(row: dict[str, Any]) -> bool:
    haystack = article_relevance_haystack(row)
    country_code = normalize_country_code(row.get('country_code') or DEFAULT_COUNTRY_CODE)
    return contains_any_relevance_keyword(haystack, FOREIGN_MARKET_TERMS) and not contains_any_relevance_keyword(haystack, country_market_terms(country_code))


def is_minor_product_launch_article(row: dict[str, Any]) -> bool:
    haystack = article_relevance_haystack(row)
    if not contains_any_relevance_keyword(haystack, LOW_IMPACT_LAUNCH_TERMS):
        return False
    if not contains_any_relevance_keyword(haystack, LOW_IMPACT_PRODUCT_TERMS):
        return False
    if contains_any_relevance_keyword(haystack, MATERIAL_IMPACT_TERMS):
        return False
    if contains_any_relevance_keyword(haystack, PROMO_SIGNAL_TERMS):
        return False
    return True


def title_looks_like_single_product_coverage(row: dict[str, Any]) -> bool:
    title = clean_text(row.get('title_translated') or row.get('title') or '')
    if not title:
        return False

    if re.search(r'[“"「『][^”"」』]{2,80}[”"」』]', title):
        return True
    if re.search(r'\b[A-Za-z]{1,8}[- ]?[A-Za-z0-9]{1,16}\d[A-Za-z0-9\-]*\b', title):
        return True
    if re.search(r'\b[A-Z]{2,}[A-Za-z0-9\-]{1,}\b', title):
        return True
    return False


def is_single_product_promo_article(row: dict[str, Any]) -> bool:
    haystack = article_relevance_haystack(row)
    if not contains_any_relevance_keyword(haystack, SINGLE_PRODUCT_PROMO_TERMS):
        return False
    if contains_any_relevance_keyword(haystack, BROAD_PROMO_SCOPE_TERMS):
        return False
    if contains_any_relevance_keyword(haystack, MATERIAL_IMPACT_TERMS) and not contains_any_relevance_keyword(haystack, LOW_IMPACT_PRODUCT_TERMS + SINGLE_PRODUCT_CATEGORY_TERMS):
        return False

    product_signal_count = 0
    if contains_any_relevance_keyword(haystack, LOW_IMPACT_PRODUCT_TERMS):
        product_signal_count += 1
    if contains_any_relevance_keyword(haystack, SINGLE_PRODUCT_CATEGORY_TERMS):
        product_signal_count += 1
    if title_looks_like_single_product_coverage(row):
        product_signal_count += 1

    return product_signal_count >= 2


def survey_hard_filter_reason(row: dict[str, Any]) -> str:
    country_config = get_country_config(row.get('country_code') or DEFAULT_COUNTRY_CODE)
    consumer_label = country_config["consumer_label"]
    market_label = country_config["market_label"]
    if is_foreign_market_only_article(row):
        return f'新闻主要针对非{market_label}，难以直接影响{consumer_label}对平台的总体感受'
    if is_minor_product_launch_article(row):
        return f'新闻主要是单个商品或小范围发布，影响面较小，难以改变{consumer_label}对平台的总体感受'
    if is_single_product_promo_article(row):
        return f'新闻主要是单个商品的特价、优惠或好价信息，影响面较小，难以改变{consumer_label}对平台的总体感受'
    return ''


def has_direct_user_impact_signal(row: dict[str, Any]) -> bool:
    haystack = article_relevance_haystack(row)
    return contains_any_relevance_keyword(haystack, MATERIAL_IMPACT_TERMS + PROMO_SIGNAL_TERMS)


def has_platform_level_balanced_signal(row: dict[str, Any]) -> bool:
    haystack = article_relevance_haystack(row).lower()
    platform_level_terms = [
        "prime day", "black friday", "cyber monday", "spring deal", "deal days", "shopping festival",
        "membership", "loyalty", "points", "cashback", "payment", "checkout", "cash payment",
        "delivery", "shipping", "return", "refund", "privacy", "data", "security", "safety",
        "investigation", "illegal", "counterfeit", "compliance", "pfas", "seller", "merchant",
        "commission", "policy", "marketplace", "algorithm", "recommendation", "app", "feature",
        "trend", "report", "survey", "study", "market share", "penetration", "adoption",
        "consumer behavior", "consumers use", "growth", "social commerce", "ecommerce growth",
        "brand performance", "market performance", "customer experience", "consumer experience",
        "consumer trust", "complaints", "satisfaction", "regulation", "regulator", "antitrust",
        "consumer protection", "creator economy", "retail media", "platform ecosystem",
        "pagamento", "pagamenti", "contanti", "consegna", "spedizione", "reso", "rimborso",
        "privacy", "dati", "sicurezza", "indagine", "illegale", "venditore", "venditori",
        "commissioni", "regole", "mercato", "social commerce", "shopping su instagram",
        "rapporto", "ricerca", "studio", "dati", "tendenza", "crescita", "consumatori",
        "quota di mercato", "esperienza cliente", "esperienza di acquisto", "fiducia",
        "soddisfazione", "reclami", "inchiesta", "tutela dei consumatori",
    ]
    return any(term in haystack for term in platform_level_terms)


def should_apply_hard_filter(row: dict[str, Any], hard_reason: str, recall_mode: str) -> bool:
    if not hard_reason:
        return False
    if clean_text(recall_mode).lower() != "balanced":
        return True
    if is_foreign_market_only_article(row):
        return True
    return not has_platform_level_balanced_signal(row)


def categorize_ai_exclusion_reason(row: dict[str, Any], decision: SurveyAIFilterDecision | None = None) -> str:
    reason = clean_text(decision.reason if decision else "")
    haystack = f"{article_relevance_haystack(row)} {reason}".lower()
    confidence = normalize_survey_filter_confidence(decision.confidence, default="low") if decision else "low"
    if any(term in haystack for term in ["foreign market", "other country", "fuori mercato", "non target", "stati uniti", "usa", "uk", "china", "france", "germany", "spain"]):
        return "foreign_market_or_non_target_market"
    if any(term in haystack for term in ["entertainment", "music", "movie", "video", "concert", "celebrity", "reality", "serie tv", "film", "musica", "concerto"]):
        return "entertainment_or_content_noise"
    if any(term in haystack for term in [
        "single product", "sku", "shopping guide", "buyers guide", "best deals", "gift ideas",
        "classifica", "migliori", "guida all'acquisto", "codice sconto", "coupon",
        "one-sku", "one sku", "single-product", "product listicle", "listicle", "offerta", "sconto",
    ]):
        return "single_product_or_coupon"
    if any(term in haystack for term in ["no platform", "target platform", "not about", "brand not", "platform not", "non riguarda"]):
        return "target_platform_not_material"
    if any(term in haystack for term in [
        "creator economy", "creator marketplace", "creator commerce", "brand partnership",
        "shoppable posts", "reels shopping", "instagram business", "meta commerce",
        "ai shopping", "shopping agent", "brand discovery", "collaborazioni", "creator",
    ]):
        return "creator_commerce_weak_link"
    if any(term in haystack for term in [
        "acquisition", "takeover", "capital market", "investor", "shareholder", "board",
        "gamestop", "offerta", "acquisizione", "cda", "wall street", "bitcoin exposure",
    ]):
        return "capital_market_event_weak_link"
    if any(term in haystack for term in [
        "platform ecosystem", "marketplace strategy", "seller ecosystem", "merchant ecosystem",
        "ebay live", "live shopping", "live commerce", "marketplace performance",
        "platform-level", "shopping festival", "major promotion", "platform strategy",
    ]):
        return "platform_ecosystem_weak_link"
    if confidence == "low":
        return "low_confidence_or_weak_link"
    if any(term in haystack for term in ["b2b", "corporate", "partnership", "investor", "organization", "financing"]):
        return "b2b_or_corporate_only"
    return "weak_nps_link"


def chunk_items(items: list[Any], chunk_size: int) -> list[list[Any]]:
    size = max(1, chunk_size)
    return [items[index:index + size] for index in range(0, len(items), size)]


def build_survey_ai_batch_messages(
    batch_items: list[tuple[int, dict[str, Any]]],
    survey_indicators: list[SurveyIndicator],
    system_prompt: str | None = None,
) -> list[dict[str, str]]:
    indicator_payload = [
        {
            'dimension': indicator.dimension,
            'question_id': indicator.question_id,
            'prompt_zh': indicator.prompt_zh,
            'prompt_en': indicator.prompt_en,
        }
        for indicator in survey_indicators
    ]
    article_payload = []
    for index, article in batch_items:
        article_payload.append(
            {
                'article_id': str(index),
                'platform': clean_text(article.get('platform') or article.get('platform_label') or ''),
                'source_site': clean_text(article.get('source_site') or ''),
                'published_at': clean_text(article.get('published_at') or ''),
                'title': clean_text(article.get('title') or ''),
                'title_translated': clean_text(article.get('title_translated') or ''),
                'summary': clean_text(article.get('summary') or ''),
                'summary_translated': clean_text(article.get('summary_translated') or ''),
                'category': clean_text(article.get('category') or ''),
                'body_excerpt': clean_text(article.get('body_excerpt') or ''),
                'tags': [clean_text(tag) for tag in (article.get('tags') or []) if clean_text(tag)],
                'article_url': clean_text(article.get('article_url') or ''),
                'broad_entry': clean_text(article.get('broad_entry') or ''),
                'broad_entry_reason': clean_text(article.get('broad_entry_reason') or ''),
                'market_scope': clean_text(article.get('market_scope') or ''),
                'search_query_strategy': clean_text(article.get('search_query_strategy') or ''),
            }
        )
    normalized_system_prompt = clean_text(system_prompt) or default_survey_ai_system_prompt()
    return [
        {
            'role': 'system',
            'content': normalized_system_prompt,
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'task': '批量判断新闻是否与问卷指标相关，并给出每条新闻命中的维度、题号与简短关联说明。',
                    'strict_rules': [
                        'matched_question_ids must only contain exact question_id values from survey_indicators, such as B5_18. Do not return dimension names, translated dimension labels, broad section IDs such as B5, or invented IDs.',
                        'If an article is about payment methods, cash payment, payment locations, checkout, or payment convenience, prefer payment method / payment process question IDs. Do not map it to product recommendation questions.',
                        'If an article is about points, loyalty, rewards, membership benefits, or cashback, prefer loyalty / membership program question IDs.',
                        'Only use product recommendation question IDs when the article explicitly discusses recommendation accuracy, recommendation diversity, repetitive recommendations, personalization, preferences, or interests.',
                        'The reason must connect the article fact to the specific question_id. If the article only broadly fits a dimension but cannot fit a concrete question, set relevant=false.',
                        'The reason field must be written in Simplified Chinese. Do not write English reasons.',
                        'Return confidence as high, medium, or low. Use medium when the article has credible platform/NPS relevance but the impact is indirect or requires interpretation.',
                        'Industry trend, market report, penetration/adoption, consumer usage, regulatory/compliance trend, social commerce trend, platform ecosystem, seller ecosystem, logistics/payment ecosystem, or brand-level performance news can be relevant when it concerns the target platform/brand and target market.',
                        'In balanced recall mode, do not reject target-brand industry trend, market share, brand growth/decline, platform ecosystem, regulation/compliance, consumer trust, or long-term perception news only because the impact is indirect or lagged. Mark it relevant=true with confidence=medium when the connection is plausible and explain the linkage.',
                        'If broad_entry=true or market_scope=europe_or_global_relevant, treat the article as a European/global platform-impact candidate. Do not reject it only because the title lacks the target country, but keep it only when the article can plausibly affect target-country consumers through regulation, safety, privacy, product quality, social commerce, or long-term brand perception.',
                        'Multi-platform regulatory, compliance, product-safety, customs, low-value parcel, privacy, or illegal-product news can affect every named platform. If Temu, SHEIN, AliExpress, Instagram, Meta, TikTok Shop, or another target platform is named, judge each named platform separately instead of assigning the event only to the headline-leading brand.',
                        'For TikTok Shop/TTS, social commerce, live shopping, creator economy, seller ecosystem, platform policy, payment/logistics, consumer protection, regulation, and compliance news can affect consumer platform perception. In balanced mode, keep these as relevant=true with confidence=medium when the link is plausible, even if the article discusses platform ecosystem or merchants rather than a single consumer feature.',
                        'For Instagram/IG, Meta commerce, Instagram business tools, social commerce, creator marketplace, shoppable posts, ads-commerce, and brand partnership news can be relevant when they affect shopping, brand discovery, merchants, or consumer purchase journeys. Reject pure entertainment, celebrity posts, ordinary account activity, and non-commerce social media features.',
                        'For eBay, eBay Live, live shopping, seller ecosystem, platform-level promotions, marketplace strategy, platform policy, acquisition or takeover news can be relevant when they affect platform ecosystem, brand image, consumer trust, or long-term platform stability. In balanced mode, keep these as relevant=true with confidence=medium when the linkage is plausible.',
                        'Keep platform-level promotions, major shopping festivals, live shopping events, creator commerce partnerships, AI shopping agents, and brand discovery tools. Still reject ordinary coupons, single-product discounts, one-SKU deals, editorial shopping guides, and product listicles.',
                        'For acquisition, capital-market, or platform-strategy events involving the target platform, keep one representative event when it affects platform strategy, stability, brand image, or consumer trust; duplicate media coverage will be handled by final AI dedupe.',
                        'If the article discusses target-brand consumer experience, complaints, trust, safety, privacy, logistics, payment, seller policy, social commerce, creator commerce, marketplace performance, or brand-level market performance in the target country, keep it for NPS review unless it is clearly a single-product guide or entertainment noise.',
                        'For industry trend or brand-level impact news, return industry_trend_flag=true, one industry_trend_category from market_adoption, brand_performance, regulatory_risk, platform_ecosystem, consumer_behavior, competitive_landscape, other, one industry_trend_impact from Positive, Negative, Mixed, Neutral, and a Simplified Chinese industry_trend_reason.',
                        'Do not mark pure macro ecommerce news with no target brand/platform and no target-market impact as relevant.',
                    ],
                    'articles': article_payload,
                    'survey_indicators': indicator_payload,
                    'output_schema': {
                        'decisions': [
                            {
                                'article_id': '0',
                                'relevant': True,
                                'matched_dimensions': ['Product quality'],
                                'matched_question_ids': ['C1_1(key)'],
                                'reason': '新闻提到平台加强正品审核与质量保障，会影响用户对商品质量和正品性的评价。',
                                'confidence': 'high',
                                'industry_trend_flag': True,
                                'industry_trend_category': 'regulatory_risk',
                                'industry_trend_impact': 'Negative',
                                'industry_trend_reason': '新闻涉及平台层面的监管或合规趋势，可能影响消费者对平台可信度的整体认知。',
                            }
                        ]
                    },
                },
                ensure_ascii=False,
            ),
        },
    ]


def normalize_ai_response_content(content: Any) -> str:
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, dict):
        return json.dumps(content, ensure_ascii=False)
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
                continue
            if not isinstance(item, dict):
                continue
            for key in ['text', 'content', 'output_text']:
                value = item.get(key)
                if isinstance(value, str) and value.strip():
                    parts.append(value)
                    break
        return '\n'.join(part.strip() for part in parts if part and part.strip())
    return clean_text(str(content))


def strip_markdown_code_fence(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith('```') and stripped.endswith('```'):
        lines = stripped.splitlines()
        if len(lines) >= 3:
            return '\n'.join(lines[1:-1]).strip()
    return stripped


def decode_first_json_value(text: str) -> Any:
    decoder = json.JSONDecoder()
    for index, char in enumerate(text):
        if char not in '{[':
            continue
        try:
            value, _ = decoder.raw_decode(text[index:])
            return value
        except json.JSONDecodeError:
            continue
    raise json.JSONDecodeError('No JSON object or array found', text, 0)


def coerce_survey_filter_payload(raw_value: Any) -> dict[str, Any]:
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, list):
        return {'decisions': raw_value}
    if isinstance(raw_value, str):
        stripped = strip_markdown_code_fence(raw_value)
        attempts = [stripped]
        fenced_match = re.search(r"```(?:json)?\s*(.*?)```", stripped, re.IGNORECASE | re.DOTALL)
        if fenced_match:
            attempts.insert(0, fenced_match.group(1).strip())
        for candidate in attempts:
            if not candidate:
                continue
            try:
                parsed = json.loads(candidate)
            except json.JSONDecodeError:
                try:
                    parsed = decode_first_json_value(candidate)
                except json.JSONDecodeError:
                    continue
            if isinstance(parsed, str) and parsed != candidate:
                try:
                    parsed = json.loads(parsed)
                except json.JSONDecodeError:
                    try:
                        parsed = decode_first_json_value(parsed)
                    except json.JSONDecodeError:
                        pass
            if isinstance(parsed, dict):
                return parsed
            if isinstance(parsed, list):
                return {'decisions': parsed}
    raise RuntimeError('survey filter API did not return a JSON object')


def call_survey_filter_api(messages: list[dict[str, str]], api_url: str, api_key: str, api_model: str) -> dict[str, Any]:
    if not api_url or not api_key or not api_model:
        raise RuntimeError('survey filter API credentials are incomplete')
    response = post_json_request_with_proxy_fallback(
        api_url,
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        json_payload={
            'model': api_model,
            'response_format': {'type': 'json_object'},
            'messages': messages,
        },
        timeout=180,
    )
    response.raise_for_status()
    body = response.json()

    raw_content: Any = None
    if isinstance(body, dict):
        choices = body.get('choices')
        if isinstance(choices, list) and choices:
            first_choice = choices[0] if isinstance(choices[0], dict) else {}
            message = first_choice.get('message') if isinstance(first_choice, dict) else {}
            if isinstance(message, dict) and 'content' in message:
                raw_content = message.get('content')
            elif isinstance(first_choice, dict) and 'text' in first_choice:
                raw_content = first_choice.get('text')

    normalized_content = normalize_ai_response_content(raw_content)
    if not normalized_content:
        raise RuntimeError('survey filter API returned empty content')
    return coerce_survey_filter_payload(normalized_content)


BRIEFING_SENTIMENT_VALUES = {"Positive", "Neutral", "Negative"}
INDUSTRY_TREND_CATEGORIES = {
    "market_adoption",
    "brand_performance",
    "regulatory_risk",
    "platform_ecosystem",
    "consumer_behavior",
    "competitive_landscape",
    "other",
}
INDUSTRY_TREND_IMPACTS = {"Positive", "Negative", "Mixed", "Neutral"}


def normalize_briefing_sentiment(value: Any) -> str:
    normalized = clean_text(value)
    lowered = normalized.lower()
    for sentiment in BRIEFING_SENTIMENT_VALUES:
        if lowered == sentiment.lower():
            return sentiment
    if lowered in {"positive", "pos", "up", "good", "favorable"}:
        return "Positive"
    if lowered in {"negative", "neg", "down", "bad", "unfavorable"}:
        return "Negative"
    return "Neutral"


def normalize_industry_trend_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = clean_text(value).lower()
    return normalized in {"1", "true", "yes", "y", "是", "行业趋势", "trend", "industry_trend"}


def normalize_industry_trend_category(value: Any) -> str:
    normalized = clean_text(value).lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "adoption": "market_adoption",
        "market": "market_adoption",
        "penetration": "market_adoption",
        "performance": "brand_performance",
        "growth": "brand_performance",
        "regulation": "regulatory_risk",
        "regulatory": "regulatory_risk",
        "compliance": "regulatory_risk",
        "ecosystem": "platform_ecosystem",
        "seller_ecosystem": "platform_ecosystem",
        "consumer": "consumer_behavior",
        "behavior": "consumer_behavior",
        "competition": "competitive_landscape",
        "competitive": "competitive_landscape",
    }
    normalized = aliases.get(normalized, normalized)
    return normalized if normalized in INDUSTRY_TREND_CATEGORIES else "other"


def normalize_industry_trend_impact(value: Any) -> str:
    normalized = clean_text(value).lower()
    for impact in INDUSTRY_TREND_IMPACTS:
        if normalized == impact.lower():
            return impact
    if normalized in {"positive", "pos", "up", "good", "favorable"}:
        return "Positive"
    if normalized in {"negative", "neg", "down", "bad", "risk", "unfavorable"}:
        return "Negative"
    if normalized in {"mixed", "both", "双向", "mixed/neutral"}:
        return "Mixed"
    return "Neutral"


def infer_industry_trend_from_article(row: dict[str, Any]) -> dict[str, Any]:
    haystack = article_relevance_haystack(row).lower()
    search_task_type = clean_text(row.get("search_task_type")).lower()
    if not haystack:
        return {
            "industry_trend_flag": False,
            "industry_trend_category": "",
            "industry_trend_impact": "Neutral",
            "industry_trend_reason": "",
        }

    regulatory_terms = [
        "regulation", "regulatory", "investigation", "probe", "illegal", "compliance", "privacy", "data",
        "counterfeit", "unsafe", "pfas", "eu", "commission", "authority", "antitrust",
        "indagine", "inchiesta", "illegale", "privacy", "dati", "sicurezza", "contraffazione",
        "sostanze", "autorità", "commissione europea", "accusa", "accuse",
    ]
    market_terms = [
        "market share", "penetration", "adoption", "1 in", "one in", "consumers use", "consumer use",
        "e-commerce", "ecommerce", "social commerce", "growth", "grows", "market", "trend", "report",
        "survey", "study", "data", "stats", "index",
        "mercato", "crescita", "consumatori", "rapporto", "ricerca", "studio", "dati", "tendenza",
        "trend", "classifica", "quota", "adozione", "penetrazione",
    ]
    performance_terms = [
        "sales", "revenue", "gmv", "growth", "decline", "loss", "profit", "expands", "expansion",
        "vendite", "ricavi", "fatturato", "crescita", "calo", "espansione",
    ]
    ecosystem_terms = [
        "seller", "merchant", "marketplace", "logistics", "delivery", "shipping", "payment", "checkout",
        "creator", "creators", "shop", "business tools", "platform tools", "ecosystem",
        "venditori", "merchant", "marketplace", "logistica", "spedizioni", "pagamenti", "creator",
        "strumenti", "ecosistema",
    ]
    consumer_terms = [
        "consumer behavior", "shoppers", "consumer", "customers", "usage", "users", "gen z",
        "comportamento", "consumatori", "clienti", "utenti", "acquisti", "shopping",
    ]
    competitive_terms = [
        "competitor", "competition", "versus", "vs", "ranking", "rankings", "comparison", "benchmark",
        "concorrente", "concorrenza", "classifica", "confronto", "comparazione", "benchmark",
    ]
    market_priority_terms = [
        "penetration", "adoption", "1 in", "one in", "consumers use", "consumer use",
        "survey", "study", "report", "data", "stats",
        "adozione", "penetrazione", "consumatori", "rapporto", "ricerca", "studio", "dati",
    ]

    category = ""
    if contains_any_relevance_keyword(haystack, regulatory_terms):
        category = "regulatory_risk"
    elif contains_any_relevance_keyword(haystack, competitive_terms):
        category = "competitive_landscape"
    elif contains_any_relevance_keyword(haystack, market_priority_terms):
        category = "market_adoption"
    elif contains_any_relevance_keyword(haystack, ecosystem_terms):
        category = "platform_ecosystem"
    elif contains_any_relevance_keyword(haystack, market_terms):
        category = "market_adoption"
    elif contains_any_relevance_keyword(haystack, performance_terms):
        category = "brand_performance"
    elif contains_any_relevance_keyword(haystack, consumer_terms):
        category = "consumer_behavior"

    is_report = search_task_type == "report_ranking" or article_looks_like_report(row)
    has_trend_signal = bool(category) or is_report
    if not has_trend_signal:
        return {
            "industry_trend_flag": False,
            "industry_trend_category": "",
            "industry_trend_impact": "Neutral",
            "industry_trend_reason": "",
        }

    if not category:
        category = "other"
    impact = "Neutral"
    if category in {"market_adoption", "brand_performance", "platform_ecosystem"} and not contains_any_relevance_keyword(haystack, ["decline", "loss", "calo", "rischio", "risk"]):
        impact = "Positive"
    if category == "regulatory_risk":
        impact = "Negative"
    if category == "competitive_landscape":
        impact = "Mixed"

    category_labels = {
        "market_adoption": "市场渗透率、消费者采用或市场增长",
        "brand_performance": "品牌整体增长、业绩或市场表现",
        "regulatory_risk": "监管、合规、隐私或商品安全风险",
        "platform_ecosystem": "平台生态、卖家、物流、支付或工具能力变化",
        "consumer_behavior": "消费者行为、使用习惯或购物偏好变化",
        "competitive_landscape": "竞争格局、排名或品牌对比变化",
        "other": "行业报告或整体趋势",
    }
    reason = f"新闻涉及{category_labels.get(category, '行业趋势')}，可能影响消费者对该平台或品牌的整体认知。"
    return {
        "industry_trend_flag": True,
        "industry_trend_category": category,
        "industry_trend_impact": impact,
        "industry_trend_reason": reason,
    }


def apply_industry_trend_fields(
    row: dict[str, Any],
    *,
    flag: Any = None,
    category: Any = "",
    impact: Any = "",
    reason: Any = "",
) -> dict[str, Any]:
    enriched = dict(row)
    inferred = infer_industry_trend_from_article(enriched)
    explicit_flag = normalize_industry_trend_flag(flag)
    existing_flag = normalize_industry_trend_flag(enriched.get("industry_trend_flag"))
    should_flag = explicit_flag or existing_flag or bool(inferred.get("industry_trend_flag"))
    if not should_flag:
        enriched["industry_trend_flag"] = ""
        enriched["industry_trend_category"] = clean_text(enriched.get("industry_trend_category") or "")
        enriched["industry_trend_impact"] = normalize_industry_trend_impact(enriched.get("industry_trend_impact"))
        enriched["industry_trend_reason"] = clean_text(enriched.get("industry_trend_reason") or "")
        return enriched

    trend_category = normalize_industry_trend_category(category or enriched.get("industry_trend_category") or inferred.get("industry_trend_category"))
    trend_impact = normalize_industry_trend_impact(impact or enriched.get("industry_trend_impact") or inferred.get("industry_trend_impact"))
    trend_reason = clean_text(reason or enriched.get("industry_trend_reason") or inferred.get("industry_trend_reason"))
    if not trend_reason:
        trend_reason = "新闻涉及品牌或平台整体层面的行业趋势，可能影响消费者对平台的长期认知。"
    enriched["industry_trend_flag"] = "true"
    enriched["industry_trend_category"] = trend_category
    enriched["industry_trend_impact"] = trend_impact
    enriched["industry_trend_reason"] = trend_reason
    return enriched


def normalize_survey_filter_confidence(value: Any, default: str = "high") -> str:
    normalized = clean_text(value).lower()
    if normalized in {"high", "medium", "low"}:
        return normalized
    if normalized in {"strong", "certain", "confident"}:
        return "high"
    if normalized in {"mid", "moderate", "possible"}:
        return "medium"
    if normalized in {"weak", "uncertain"}:
        return "low"
    return default if default in {"high", "medium", "low"} else "high"


def build_briefing_sentiment_messages(batch_items: list[tuple[int, dict[str, Any]]], country_code: str) -> list[dict[str, str]]:
    article_payload = []
    for index, article in batch_items:
        article_payload.append(
            {
                "article_id": str(index),
                "platform": clean_text(article.get("platform") or article.get("platform_label") or ""),
                "published_at": clean_text(article.get("published_at") or ""),
                "title": clean_text(article.get("title") or ""),
                "title_translated": clean_text(article.get("title_translated") or ""),
                "summary": clean_text(article.get("summary") or ""),
                "summary_translated": clean_text(article.get("summary_translated") or ""),
                "nps_dimension": clean_text(article.get("survey_dimensions") or ""),
                "survey_question_ids": clean_text(article.get("survey_question_ids") or ""),
                "survey_indicator_examples": clean_text(article.get("survey_indicator_examples") or ""),
                "article_url": clean_text(article.get("article_url") or ""),
                "source_site": clean_text(article.get("source_site") or article.get("source_name") or ""),
            }
        )
    return [
        {
            "role": "system",
            "content": (
                "You are an NPS market intelligence analyst. Return JSON only. "
                "Judge each article's likely directional sentiment toward platform/NPS perception. "
                "Use exactly one of Positive, Neutral, Negative."
            ),
        },
        {
            "role": "user",
            "content": json.dumps(
                {
                    "country_code": country_code,
                    "rules": [
                        "Positive: the news is likely to improve platform perception or NPS, such as better benefits, discounts, payment convenience, features, assortment, logistics, trust, or service.",
                        "Negative: the news is likely to hurt platform perception or NPS, such as regulatory investigation, illegal or unsafe products, counterfeit risk, privacy/data risk, service failure, seller trust risk, or consumer protection issues.",
                        "Neutral: the news is informational, mixed, ambiguous, or has no clear effect on NPS perception.",
                        "Return every input article exactly once.",
                    ],
                    "articles": article_payload,
                    "output_schema": {
                        "items": [
                            {
                                "article_id": "0",
                                "sentiment": "Positive|Neutral|Negative",
                                "reason": "short Chinese reason",
                            }
                        ]
                    },
                },
                ensure_ascii=False,
            ),
        },
    ]


def extract_briefing_sentiment_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ["items", "decisions", "rows", "articles"]:
        value = payload.get(key) if isinstance(payload, dict) else None
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def add_briefing_sentiment_fields(
    articles: list[dict[str, Any]],
    *,
    api_url: str,
    api_key: str,
    api_model: str,
    country_code: str,
    batch_size: int = DEFAULT_SURVEY_AI_BATCH_SIZE,
    call_filter_api=call_survey_filter_api,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    stats = {
        "enabled": bool(articles),
        "configured": bool(api_url and api_key and api_model),
        "request_count": 0,
        "evaluated_count": 0,
        "api_error_count": 0,
        "defaulted_count": 0,
    }
    if not articles:
        return articles, stats

    for article in articles:
        article["briefing_sentiment"] = normalize_briefing_sentiment(article.get("briefing_sentiment"))
        article.setdefault("briefing_sentiment_reason", "")

    if not (api_url and api_key and api_model):
        stats["defaulted_count"] = len(articles)
        return articles, stats

    indexed_articles = list(enumerate(articles))
    for batch_items in chunk_items(indexed_articles, max(1, batch_size)):
        stats["request_count"] += 1
        try:
            payload = call_filter_api(
                build_briefing_sentiment_messages(batch_items, country_code),
                api_url,
                api_key,
                api_model,
            )
            items = extract_briefing_sentiment_items(payload)
            item_by_id = {
                clean_text(item.get("article_id")): item
                for item in items
                if clean_text(item.get("article_id"))
            }
            for index, article in batch_items:
                item = item_by_id.get(str(index))
                if not item:
                    stats["defaulted_count"] += 1
                    continue
                article["briefing_sentiment"] = normalize_briefing_sentiment(item.get("sentiment"))
                article["briefing_sentiment_reason"] = clean_text(item.get("reason"))
                stats["evaluated_count"] += 1
        except Exception:
            stats["api_error_count"] += 1
            stats["defaulted_count"] += len(batch_items)
            for _, article in batch_items:
                article["briefing_sentiment"] = "Neutral"
                article["briefing_sentiment_reason"] = "AI sentiment generation failed; defaulted to Neutral."

    return articles, stats


def title_translation_failed(row: dict[str, Any], target_language: str = "zh-CN") -> bool:
    title = clean_text(row.get("title"))
    translated = clean_text(row.get("title_translated"))
    normalized_target = normalize_translation_target(target_language)
    if not title:
        return False
    if normalized_target == "zh-CN":
        if contains_chinese_chars(title):
            return False
        return not translated or translated == title or not contains_chinese_chars(translated)
    return not translated or translated == title


def count_failed_title_translations(rows: list[dict[str, Any]], target_language: str = "zh-CN") -> int:
    return sum(1 for row in rows if title_translation_failed(row, target_language))


def build_title_translation_messages(items: list[tuple[int, str]]) -> list[dict[str, str]]:
    payload = {
        "task": "Translate these ecommerce news titles into Simplified Chinese. Preserve brand names, source names, numbers, and product/platform names. Return only JSON.",
        "titles": [
            {
                "article_id": str(index),
                "title": title,
            }
            for index, title in items
        ],
        "output_schema": {
            "translations": [
                {
                    "article_id": "0",
                    "title_zh": "中文标题",
                }
            ]
        },
    }
    return [
        {
            "role": "system",
            "content": (
                "You are a translation assistant for a Chinese market research UI. "
                "Translate news titles into natural Simplified Chinese. "
                "Do not summarize, classify, or add commentary. "
                "Keep brand names such as Amazon, eBay, SHEIN, TEMU, TikTok Shop, IG and source names such as Webnews unchanged when appropriate. "
                "Return only a JSON object with a translations array."
            ),
        },
        {
            "role": "user",
            "content": json.dumps(payload, ensure_ascii=False),
        },
    ]


def normalize_title_translation_payload(payload: dict[str, Any]) -> dict[int, str]:
    raw_translations = payload.get("translations") or payload.get("title_translations") or payload.get("titles")
    normalized: dict[int, str] = {}
    if isinstance(raw_translations, dict):
        for raw_id, value in raw_translations.items():
            try:
                article_id = int(str(raw_id).strip())
            except ValueError:
                continue
            if isinstance(value, dict):
                value = value.get("title_zh") or value.get("title_translated") or value.get("translation") or value.get("title")
            translated = clean_text(str(value) if value is not None else "")
            if translated:
                normalized[article_id] = translated
        return normalized
    if not isinstance(raw_translations, list):
        return normalized
    for item in raw_translations:
        if not isinstance(item, dict):
            continue
        raw_id = item.get("article_id") or item.get("id") or item.get("index")
        try:
            article_id = int(str(raw_id).strip())
        except (TypeError, ValueError):
            continue
        raw_translated = item.get("title_zh") or item.get("title_translated") or item.get("translation") or item.get("title")
        translated = clean_text(str(raw_translated) if raw_translated is not None else "")
        if translated:
            normalized[article_id] = translated
    return normalized


def apply_ai_title_translation_fallback(
    rows: list[dict[str, Any]],
    *,
    api_url: str,
    api_key: str,
    api_model: str,
    target_language: str = "zh-CN",
    batch_size: int = 20,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    failed_items = [
        (index, clean_text(row.get("title")))
        for index, row in enumerate(rows)
        if title_translation_failed(row, target_language)
    ]
    stats = {
        "failed_count": len(failed_items),
        "ai_fallback_count": 0,
        "ai_fallback_error_count": 0,
    }
    if not failed_items or normalize_translation_target(target_language) != "zh-CN":
        return rows, stats
    if not api_url or not api_key or not api_model:
        return rows, stats

    enriched_rows = [dict(row) for row in rows]
    for start in range(0, len(failed_items), max(1, batch_size)):
        batch_items = failed_items[start:start + max(1, batch_size)]
        try:
            payload = call_survey_filter_api(build_title_translation_messages(batch_items), api_url, api_key, api_model)
        except Exception:
            stats["ai_fallback_error_count"] += 1
            continue
        translations = normalize_title_translation_payload(payload)
        for index, original_title in batch_items:
            translated = translations.get(index)
            if not translated:
                continue
            if translated == original_title or not contains_chinese_chars(translated):
                continue
            enriched_rows[index]["title_translated"] = translated
            stats["ai_fallback_count"] += 1
    return enriched_rows, stats


def request_error_chain_text(exc: BaseException) -> str:
    parts: list[str] = []
    current: BaseException | None = exc
    seen_ids: set[int] = set()
    while current is not None and id(current) not in seen_ids:
        seen_ids.add(id(current))
        message = str(current or '').strip()
        if message:
            parts.append(f'{type(current).__name__}: {message}')
        next_exc = getattr(current, '__cause__', None) or getattr(current, '__context__', None)
        current = next_exc if isinstance(next_exc, BaseException) else None
    return ' | '.join(parts).lower()


def should_retry_without_proxy(exc: requests.exceptions.RequestException) -> bool:
    error_text = request_error_chain_text(exc)
    if isinstance(exc, requests.exceptions.ProxyError):
        return True
    if isinstance(exc, requests.exceptions.SSLError) and (
        'eof occurred in violation of protocol' in error_text
        or 'wrong version number' in error_text
        or 'tlsv1 alert' in error_text
    ):
        return True
    if isinstance(exc, requests.exceptions.ConnectionError) and 'proxy' in error_text:
        return True
    return 'unable to connect to proxy' in error_text or 'proxyerror' in error_text


def post_json_request_with_proxy_fallback(
    url: str,
    *,
    headers: dict[str, str],
    json_payload: dict[str, Any],
    timeout: int | float,
) -> requests.Response:
    try:
        return requests.post(
            url,
            headers=headers,
            json=json_payload,
            timeout=timeout,
        )
    except requests.exceptions.RequestException as exc:
        if not should_retry_without_proxy(exc):
            raise

        session = requests.Session()
        session.trust_env = False
        try:
            return session.post(
                url,
                headers=headers,
                json=json_payload,
                timeout=timeout,
            )
        except requests.exceptions.RequestException as retry_exc:
            raise retry_exc from exc


def classify_articles_with_ai(
    batch_items: list[tuple[int, dict[str, Any]]],
    survey_indicators: list[SurveyIndicator],
    grouped_indicators: dict[str, list[SurveyIndicator]],
    question_lookup: dict[str, SurveyIndicator],
    api_url: str,
    api_key: str,
    api_model: str,
    system_prompt: str | None = None,
) -> dict[int, SurveyAIFilterDecision]:
    payload = call_survey_filter_api(
        build_survey_ai_batch_messages(batch_items, survey_indicators, system_prompt),
        api_url,
        api_key,
        api_model,
    )
    raw_decisions = payload.get('decisions')
    if not isinstance(raw_decisions, list):
        raise RuntimeError('survey filter API did not return a valid decisions array')

    decisions_by_index: dict[int, SurveyAIFilterDecision] = {}
    valid_indices = {index for index, _ in batch_items}

    for item in raw_decisions:
        if not isinstance(item, dict):
            continue
        try:
            index = int(str(item.get('article_id') or '').strip())
        except ValueError:
            continue
        if index not in valid_indices:
            continue

        matched_question_ids = normalize_ai_question_ids(item.get('matched_question_ids'), question_lookup)
        matched_dimensions = normalize_ai_dimension_matches(item.get('matched_dimensions'), grouped_indicators)
        if not matched_dimensions and matched_question_ids:
            for question_id in matched_question_ids:
                indicator = question_lookup.get(normalize_question_id(question_id).lower())
                if indicator and indicator.dimension not in matched_dimensions:
                    matched_dimensions.append(indicator.dimension)
        if matched_dimensions:
            article_row = next((row for candidate_index, row in batch_items if candidate_index == index), {})
            inferred_question_ids = infer_question_ids_from_article(article_row, matched_dimensions, grouped_indicators)
            if inferred_question_ids:
                matched_question_ids = inferred_question_ids

        if not matched_question_ids:
            matched_dimensions = []
        relevant = bool(matched_question_ids)
        reason = clean_text(item.get('reason'))
        confidence = normalize_survey_filter_confidence(item.get('confidence'), default="high" if relevant else "low")
        decisions_by_index[index] = SurveyAIFilterDecision(
            relevant=relevant,
            matched_dimensions=matched_dimensions,
            matched_question_ids=matched_question_ids,
            reason=reason,
            confidence=confidence,
            method='ai',
            industry_trend_flag=normalize_industry_trend_flag(
                item.get('industry_trend_flag')
                or item.get('is_industry_trend')
                or item.get('trend_flag')
            ),
            industry_trend_category=normalize_industry_trend_category(
                item.get('industry_trend_category')
                or item.get('trend_category')
            ),
            industry_trend_impact=normalize_industry_trend_impact(
                item.get('industry_trend_impact')
                or item.get('trend_impact')
            ),
            industry_trend_reason=clean_text(
                item.get('industry_trend_reason')
                or item.get('trend_reason')
            ),
        )

    return decisions_by_index


def ai_keyword_bridge_single_article(
    row: dict[str, Any],
    grouped_indicators: dict[str, list[SurveyIndicator]],
    indicator_dimensions: set[str],
    dimension_keyword_lookup: dict[str, list[str]],
    question_lookup: dict[str, SurveyIndicator],
    *,
    explanation: str,
    method: str,
) -> tuple[dict[str, Any] | None, list[str]]:
    return keyword_filter_single_article(
        row,
        grouped_indicators,
        indicator_dimensions,
        dimension_keyword_lookup,
        question_lookup,
        method=method,
        explanation=explanation,
    )


def apply_survey_indicator_keyword_filter(
    rows: list[dict[str, Any]],
    survey_indicators: list[SurveyIndicator],
    *,
    progress_callback=None,
    total_sites: int = 0,
    completed_sites: int = 0,
) -> tuple[list[dict[str, Any]], Counter, int, dict[str, Any]]:
    if not survey_indicators:
        return rows, Counter(), len(rows), {'mode': 'none', 'api_error_count': 0, 'ai_evaluated_count': 0, 'keyword_fallback_count': 0}

    grouped_indicators = group_survey_indicators_by_dimension(survey_indicators)
    question_lookup = build_survey_indicator_question_lookup(survey_indicators)
    indicator_dimensions = set(grouped_indicators)
    dimension_keyword_lookup = build_dimension_keyword_lookup(survey_indicators)
    filtered_rows: list[dict[str, Any]] = []
    dimension_counts: Counter = Counter()
    original_count = len(rows)
    progress_step = max(1, original_count // 20) if original_count else 1

    for index, row in enumerate(rows, start=1):
        enriched, matched_dimensions = keyword_filter_single_article(
            row,
            grouped_indicators,
            indicator_dimensions,
            dimension_keyword_lookup,
            question_lookup,
        )
        if not enriched:
            if progress_callback is not None and (index <= 3 or index == original_count or index % progress_step == 0):
                emit_progress(
                    progress_callback,
                    stage='survey_filter',
                    total_sites=total_sites,
                    completed_sites=completed_sites,
                    active_sites=0,
                    current_site=f'关键词筛选新闻（已检查 {index}/{original_count} 条）',
                    last_completed_site=clean_text(row.get('title_translated') or row.get('title') or row.get('article_url') or '')[:120],
                    message=f'正在按关键词判断新闻与指标的相关性，已检查 {index}/{original_count} 条',
                    progress_percent=min(
                        NEWS_PROGRESS_FILTER_KEYWORD_CAP,
                        interpolate_progress(
                            NEWS_PROGRESS_FILTER_KEYWORD_START,
                            NEWS_PROGRESS_FILTER_KEYWORD_CAP,
                            index,
                            original_count,
                        ),
                    ),
                )
            continue
        for dimension in matched_dimensions:
            dimension_counts[dimension] += 1
        filtered_rows.append(enriched)
        if progress_callback is not None and (index <= 3 or index == original_count or index % progress_step == 0):
            emit_progress(
                progress_callback,
                    stage='survey_filter',
                total_sites=total_sites,
                completed_sites=completed_sites,
                active_sites=0,
                current_site=f'关键词筛选新闻（已检查 {index}/{original_count} 条）',
                last_completed_site=clean_text(row.get('title_translated') or row.get('title') or row.get('article_url') or '')[:120],
                message=f'正在按关键词判断新闻与指标的相关性，已检查 {index}/{original_count} 条',
                    progress_percent=min(
                        NEWS_PROGRESS_FILTER_KEYWORD_CAP,
                        interpolate_progress(
                            NEWS_PROGRESS_FILTER_KEYWORD_START,
                            NEWS_PROGRESS_FILTER_KEYWORD_CAP,
                            index,
                            original_count,
                        ),
                    ),
            )

    stats = {
        'mode': 'keyword',
        'api_error_count': 0,
        'ai_evaluated_count': 0,
        'keyword_fallback_count': 0,
    }
    return filtered_rows, dimension_counts, original_count, stats


def apply_survey_indicator_ai_filter(
    rows: list[dict[str, Any]],
    survey_indicators: list[SurveyIndicator],
    *,
    api_url: str,
    api_key: str,
    api_model: str,
    system_prompt: str = '',
    ai_workers: int = 4,
    ai_batch_size: int = DEFAULT_SURVEY_AI_BATCH_SIZE,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
    progress_callback=None,
    total_sites: int = 0,
    completed_sites: int = 0,
) -> tuple[list[dict[str, Any]], Counter, int, dict[str, Any]]:
    if not survey_indicators:
        return rows, Counter(), len(rows), {'mode': 'none', 'api_error_count': 0, 'ai_evaluated_count': 0, 'keyword_fallback_count': 0, 'ai_request_count': 0, 'ai_batch_size': ai_batch_size}
    if not api_url or not api_key or not api_model:
        raise RuntimeError('已选择 AI 指标筛选，但缺少 API URL / API Key / 模型名。')

    grouped_indicators = group_survey_indicators_by_dimension(survey_indicators)
    question_lookup = build_survey_indicator_question_lookup(survey_indicators)
    indicator_dimensions = set(grouped_indicators)
    dimension_keyword_lookup = build_dimension_keyword_lookup(survey_indicators)

    original_count = len(rows)
    filtered_rows_by_index: dict[int, dict[str, Any]] = {}
    dimension_counts: Counter = Counter()
    api_error_batch_count = 0
    keyword_fallback_count = 0
    ai_bridge_count = 0
    processed_count = 0
    ai_excluded_reason_counts: Counter = Counter()
    broad_entry_ai_excluded_reason_counts: dict[str, Counter] = {}

    if not rows:
        return [], Counter(), 0, {
            'mode': 'ai',
            'api_error_count': 0,
            'ai_evaluated_count': 0,
            'keyword_fallback_count': 0,
            'ai_bridge_count': 0,
            'ai_request_count': 0,
            'ai_batch_size': ai_batch_size,
        }

    batch_size = max(1, ai_batch_size)
    hard_filtered_count = 0
    indexed_rows: list[tuple[int, dict[str, Any]]] = []
    for index, row in enumerate(rows):
        if clean_text(row.get('search_task_type')) == 'report_ranking':
            focus_dimension = clean_text(row.get('search_focus_dimension'))
            if focus_dimension and focus_dimension in grouped_indicators:
                matched_question_ids = infer_question_ids_from_article(row, [focus_dimension], grouped_indicators)
                if not matched_question_ids:
                    indexed_rows.append((index, row))
                    continue
                enriched = enrich_row_with_survey_match(
                    row,
                    matched_dimensions=[focus_dimension],
                    matched_question_ids=matched_question_ids,
                    grouped_indicators=grouped_indicators,
                    question_lookup=question_lookup,
                    explanation='数据/报告检索任务已按指标定向构造，直接保留用于品牌间同期比较。',
                    method='report_search_direct',
                )
                filtered_rows_by_index[index] = enriched
                dimension_counts[focus_dimension] += 1
                continue
        hard_reason = survey_hard_filter_reason(row)
        if should_apply_hard_filter(row, hard_reason, recall_mode):
            hard_filtered_count += 1
            if clean_text(row.get("broad_entry")).lower() == "true":
                for platform in broad_entry_platform_labels(row):
                    bucket = broad_entry_ai_excluded_reason_counts.setdefault(platform, Counter())
                    bucket[f"hard_filter:{hard_reason or 'unknown'}"] += 1
            continue
        indexed_rows.append((index, row))

    if not indexed_rows:
        return [], Counter(), original_count, {
            'mode': 'ai',
            'api_error_count': 0,
            'ai_evaluated_count': 0,
            'keyword_fallback_count': 0,
            'ai_bridge_count': 0,
            'ai_request_count': 0,
            'ai_batch_size': batch_size,
            'hard_filtered_count': hard_filtered_count,
            'recall_mode': recall_mode,
        }

    candidate_count = len(indexed_rows)
    batches = chunk_items(indexed_rows, batch_size)
    total_batches = len(batches)
    worker_count = max(1, min(ai_workers, total_batches, 6))
    consumer_label = get_country_config(country_code)["consumer_label"]

    def worker(batch_index: int, batch_items: list[tuple[int, dict[str, Any]]]) -> tuple[int, list[tuple[int, dict[str, Any]]], dict[int, SurveyAIFilterDecision] | None, str | None]:
        try:
            decisions = classify_articles_with_ai(
                batch_items,
                survey_indicators,
                grouped_indicators,
                question_lookup,
                api_url,
                api_key,
                api_model,
                system_prompt,
            )
            return batch_index, batch_items, decisions, None
        except Exception as exc:
            return batch_index, batch_items, None, clean_text(str(exc)) or exc.__class__.__name__

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {executor.submit(worker, batch_index, batch_items): batch_index for batch_index, batch_items in enumerate(batches, start=1)}
        completed_batches = 0
        for future in as_completed(futures):
            batch_index, batch_items, decisions, error_message = future.result()
            completed_batches += 1
            processed_count += len(batch_items)

            if decisions is not None:
                for index, row in batch_items:
                    decision = decisions.get(index)
                    if decision and decision.relevant and (decision.matched_dimensions or decision.matched_question_ids):
                        enriched = enrich_row_with_survey_match(
                            row,
                            matched_dimensions=decision.matched_dimensions,
                            matched_question_ids=decision.matched_question_ids,
                            grouped_indicators=grouped_indicators,
                            question_lookup=question_lookup,
                            explanation=decision.reason,
                            method='ai_batch',
                            confidence=decision.confidence,
                            industry_trend_flag=decision.industry_trend_flag,
                            industry_trend_category=decision.industry_trend_category,
                            industry_trend_impact=decision.industry_trend_impact,
                            industry_trend_reason=decision.industry_trend_reason,
                        )
                        filtered_rows_by_index[index] = enriched
                        for dimension in normalize_string_list(enriched.get('survey_dimensions')):
                            dimension_counts[dimension] += 1
                        continue

                    if decision and decision.relevant:
                        bridge_explanation = decision.reason or 'AI 判定新闻与指标相关，已结合关键词补全维度'
                        enriched, matched_dimensions = ai_keyword_bridge_single_article(
                            row,
                            grouped_indicators,
                            indicator_dimensions,
                            dimension_keyword_lookup,
                            question_lookup,
                            explanation=bridge_explanation,
                            method='ai_keyword_bridge',
                        )
                        if enriched and (has_direct_user_impact_signal(row) or decision.industry_trend_flag):
                            enriched = apply_industry_trend_fields(
                                enriched,
                                flag=decision.industry_trend_flag,
                                category=decision.industry_trend_category,
                                impact=decision.industry_trend_impact,
                                reason=decision.industry_trend_reason,
                            )
                            filtered_rows_by_index[index] = enriched
                            ai_bridge_count += 1
                            for dimension in matched_dimensions:
                                dimension_counts[dimension] += 1
                            continue

                    if decision is None:
                        fallback_explanation = 'AI 批量结果未返回该新闻判断，已回退为关键词判定'
                        enriched, matched_dimensions = keyword_filter_single_article(
                            row,
                            grouped_indicators,
                            indicator_dimensions,
                            dimension_keyword_lookup,
                            question_lookup,
                            method='keyword_fallback',
                            explanation=fallback_explanation,
                        )
                        if enriched:
                            filtered_rows_by_index[index] = enriched
                            keyword_fallback_count += 1
                            for dimension in matched_dimensions:
                                dimension_counts[dimension] += 1
                    else:
                        excluded_reason = categorize_ai_exclusion_reason(row, decision)
                        ai_excluded_reason_counts[excluded_reason] += 1
                        if clean_text(row.get("broad_entry")).lower() == "true":
                            for platform in broad_entry_platform_labels(row):
                                bucket = broad_entry_ai_excluded_reason_counts.setdefault(platform, Counter())
                                bucket[excluded_reason] += 1
            else:
                api_error_batch_count += 1
                for index, row in batch_items:
                    fallback_explanation = f'AI 批量接口异常，已回退为关键词判定：{error_message}'
                    enriched, matched_dimensions = keyword_filter_single_article(
                        row,
                        grouped_indicators,
                        indicator_dimensions,
                        dimension_keyword_lookup,
                        question_lookup,
                        method='keyword_fallback',
                        explanation=fallback_explanation,
                    )
                    if enriched:
                        filtered_rows_by_index[index] = enriched
                        keyword_fallback_count += 1
                        for dimension in matched_dimensions:
                            dimension_counts[dimension] += 1

            remaining_batches = max(0, total_batches - completed_batches)
            progress_percent = interpolate_progress(
                NEWS_PROGRESS_FILTER_AI_START,
                NEWS_PROGRESS_FILTER_AI_CAP,
                processed_count,
                original_count,
            )
            last_titles = [clean_text(row.get('title_translated') or row.get('title') or row.get('article_url') or '') for _, row in batch_items]
            last_title = ' / '.join(title for title in last_titles[:2] if title)
            emit_progress(
                progress_callback,
                stage='survey_filter',
                total_sites=total_sites,
                completed_sites=completed_sites,
                active_sites=min(worker_count, remaining_batches),
                current_site=f'AI 批量筛选新闻（候选 {processed_count}/{len(indexed_rows)} 条，预先排除 {hard_filtered_count} 条，{completed_batches}/{total_batches} 批）',
                last_completed_site=last_title[:120],
                message=f'正在使用 AI 批量判定{consumer_label}总体感受相关性：候选 {processed_count}/{len(indexed_rows)} 条，预先排除 {hard_filtered_count} 条，批次 {completed_batches}/{total_batches}',
                progress_percent=min(NEWS_PROGRESS_FILTER_AI_CAP, progress_percent),
            )

    filtered_rows = [filtered_rows_by_index[index] for index in sorted(filtered_rows_by_index)]
    stats = {
        'mode': 'ai',
        'api_error_count': api_error_batch_count,
        'ai_evaluated_count': candidate_count,
        'keyword_fallback_count': keyword_fallback_count,
        'ai_bridge_count': ai_bridge_count,
        'ai_request_count': total_batches,
        'ai_batch_size': batch_size,
        'ai_error_batch_count': api_error_batch_count,
        'hard_filtered_count': hard_filtered_count,
        'recall_mode': recall_mode,
        'ai_excluded_reason_counts': dict(ai_excluded_reason_counts),
        'broad_entry_ai_excluded_reason_counts': {
            platform: dict(counter)
            for platform, counter in broad_entry_ai_excluded_reason_counts.items()
        },
    }
    return filtered_rows, dimension_counts, original_count, stats


def apply_survey_indicator_filter(
    rows: list[dict[str, Any]],
    survey_indicators: list[SurveyIndicator],
    *,
    mode: str = DEFAULT_SURVEY_FILTER_MODE,
    api_url: str = '',
    api_key: str = '',
    api_model: str = '',
    system_prompt: str = '',
    ai_workers: int = 4,
    ai_batch_size: int = DEFAULT_SURVEY_AI_BATCH_SIZE,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
    progress_callback=None,
    total_sites: int = 0,
    completed_sites: int = 0,
) -> tuple[list[dict[str, Any]], Counter, int, dict[str, Any]]:
    normalized_mode = clean_text(mode).lower() or DEFAULT_SURVEY_FILTER_MODE
    if normalized_mode == 'ai':
        return apply_survey_indicator_ai_filter(
            rows,
            survey_indicators,
            api_url=api_url,
            api_key=api_key,
            api_model=api_model,
            system_prompt=system_prompt,
            ai_workers=ai_workers,
            ai_batch_size=ai_batch_size,
            country_code=country_code,
            recall_mode=recall_mode,
            progress_callback=progress_callback,
            total_sites=total_sites,
            completed_sites=completed_sites,
        )
    return apply_survey_indicator_keyword_filter(
        rows,
        survey_indicators,
        progress_callback=progress_callback,
        total_sites=total_sites,
        completed_sites=completed_sites,
    )


def article_identity_key(row: dict[str, Any]) -> str:
    url = clean_text(row.get("verification_final_url") or row.get("article_url") or "")
    if url:
        return f"url:{url.lower()}"
    title = normalize_article_title_for_dedupe(clean_text(row.get("title") or row.get("title_translated") or ""))
    published_at = clean_text(row.get("published_at") or "")[:10]
    platform = clean_text(row.get("platform_label") or row.get("platform") or "")
    return f"text:{platform.lower()}|{published_at}|{title}"


def article_week_start(row: dict[str, Any]) -> str:
    parsed = parse_dt(row.get("published_at"))
    if not parsed:
        return ""
    week_start = parsed.date() - timedelta(days=parsed.weekday())
    return week_start.isoformat()


def article_brand_labels(row: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    labels: list[str] = []
    for brand in normalize_string_list(row.get("matched_brands")):
        if brand and brand not in labels:
            labels.append(brand)
    platform_label = clean_text(row.get("platform_label") or row.get("platform") or "")
    if platform_label:
        for item in [part.strip() for part in re.split(r"[,|/]+", platform_label) if part.strip()]:
            if item and item not in labels:
                labels.append(item)
    canonical: list[str] = []
    lookup = requested_display_platform_lookup(country_code)
    for label in labels:
        mapped = lookup.get(label.lower(), label)
        if mapped and mapped not in canonical:
            canonical.append(mapped)
    return canonical


def brand_week_counts(rows: list[dict[str, Any]], country_code: str = DEFAULT_COUNTRY_CODE) -> dict[tuple[str, str], int]:
    counts: Counter = Counter()
    for row in rows:
        week = article_week_start(row)
        if not week:
            continue
        for brand in article_brand_labels(row, country_code):
            counts[(brand, week)] += 1
    return dict(counts)


def week_starts_between(start: datetime, end: datetime) -> list[str]:
    current = start.date() - timedelta(days=start.date().weekday())
    final = end.date() - timedelta(days=end.date().weekday())
    weeks: list[str] = []
    while current <= final:
        weeks.append(current.isoformat())
        current = current + timedelta(days=7)
    return weeks


def build_brand_stage_funnel_summary(
    selected_platforms: list[str],
    *,
    start: datetime,
    end: datetime,
    country_code: str,
    raw_rows: list[dict[str, Any]],
    initial_deduped_rows: list[dict[str, Any]],
    survey_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    raw_counts = brand_week_counts(raw_rows, country_code)
    initial_counts = brand_week_counts(initial_deduped_rows, country_code)
    survey_counts = brand_week_counts(survey_rows, country_code)
    final_counts = brand_week_counts(final_rows, country_code)
    rows: list[dict[str, Any]] = []
    for brand in selected_platforms:
        for week in week_starts_between(start, end):
            final_count = final_counts.get((brand, week), 0)
            rows.append(
                {
                    "brand": brand,
                    "week_start": week,
                    "raw_count": raw_counts.get((brand, week), 0),
                    "initial_dedupe_count": initial_counts.get((brand, week), 0),
                    "survey_filter_count": survey_counts.get((brand, week), 0),
                    "final_count": final_count,
                }
            )
    return rows


def diagnose_brand_stage_funnel(row: dict[str, Any]) -> str:
    raw_count = int(row.get("raw_count", 0) or 0)
    initial_count = int(row.get("initial_dedupe_count", 0) or 0)
    survey_count = int(row.get("survey_filter_count", 0) or 0)
    final_count = int(row.get("final_count", 0) or 0)

    if raw_count <= 2:
        return "候选较少，问题主要发生在搜索入池阶段"
    if initial_count and survey_count <= max(1, int(initial_count * 0.2)):
        return "入池候选有量，但多数在 AI 指标筛选阶段被排除"
    if survey_count and final_count <= max(1, int(survey_count * 0.5)):
        return "筛选后仍有候选，但最终去重/合并移除较多"
    if final_count == 0:
        return "有少量候选，但最终未保留可汇报新闻"
    return "各阶段有保留，可继续关注最终新闻质量"


def build_brand_stage_total_summary(
    selected_platforms: list[str],
    weekly_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {
        brand: {
            "brand": brand,
            "raw_count": 0,
            "initial_dedupe_count": 0,
            "survey_filter_count": 0,
            "final_count": 0,
        }
        for brand in selected_platforms
    }
    for row in weekly_rows:
        brand = clean_text(row.get("brand") or "")
        if not brand:
            continue
        if brand not in totals:
            totals[brand] = {
                "brand": brand,
                "raw_count": 0,
                "initial_dedupe_count": 0,
                "survey_filter_count": 0,
                "final_count": 0,
            }
        for key in ("raw_count", "initial_dedupe_count", "survey_filter_count", "final_count"):
            totals[brand][key] += int(row.get(key, 0) or 0)

    summary_rows: list[dict[str, Any]] = []
    for brand, row in totals.items():
        summary = dict(row)
        summary["diagnosis"] = diagnose_brand_stage_funnel(summary)
        summary_rows.append(summary)
    return summary_rows


def broad_entry_platform_labels(row: dict[str, Any]) -> list[str]:
    labels = []
    for value in normalize_string_list(row.get("matched_brands")):
        if value in BROAD_ENTRY_PLATFORM_LABELS and value not in labels:
            labels.append(value)
    for key in ("platform_label", "platform"):
        value = clean_text(row.get(key) or "")
        if value in BROAD_ENTRY_PLATFORM_LABELS and value not in labels:
            labels.append(value)
    return labels


def build_broad_entry_platform_summary(
    promo_search_stats: dict[str, Any],
    source_rows: list[dict[str, Any]] | None = None,
    survey_rows: list[dict[str, Any]] | None = None,
    final_rows: list[dict[str, Any]] | None = None,
    survey_filter_stats: dict[str, Any] | None = None,
) -> dict[str, dict[str, Any]]:
    query_summary = promo_search_stats.get("platform_query_summary") or {}
    summary: dict[str, dict[str, Any]] = {}
    if not isinstance(query_summary, dict):
        return summary
    for row in query_summary.values():
        if not isinstance(row, dict):
            continue
        platform = clean_text(row.get("platform_label") or "")
        if platform not in BROAD_ENTRY_PLATFORM_LABELS:
            continue
        bucket = summary.setdefault(
            platform,
            {
                "target_market_feed_count": 0,
                "target_market_kept_count": 0,
                "broad_feed_count": 0,
                "broad_kept_count": 0,
                "broad_short_feed_count": 0,
                "broad_short_kept_count": 0,
                "broad_entry_count": 0,
                "multi_platform_event_kept_count": 0,
                "broad_entry_before_filter_count": 0,
                "broad_entry_ai_kept_count": 0,
                "broad_entry_final_count": 0,
                "broad_entry_ai_excluded_count": 0,
                "skipped_reason_count": {},
                "allowed_reason_count": {},
                "ai_excluded_reason_count": {},
            },
        )
        strategy = clean_text(row.get("query_strategy") or "")
        if strategy in {"platform_broad_eu_global", "platform_broad_short"}:
            bucket["broad_feed_count"] += int(row.get("feed_count") or 0)
            bucket["broad_kept_count"] += int(row.get("kept_count") or 0)
            bucket["broad_entry_count"] += int(row.get("broad_entry_count") or 0)
            if strategy == "platform_broad_short":
                bucket["broad_short_feed_count"] += int(row.get("feed_count") or 0)
                bucket["broad_short_kept_count"] += int(row.get("kept_count") or 0)
        else:
            bucket["target_market_feed_count"] += int(row.get("feed_count") or 0)
            bucket["target_market_kept_count"] += int(row.get("kept_count") or 0)
        skipped = row.get("skipped_reason_count") or {}
        if isinstance(skipped, dict):
            merged = Counter(bucket.get("skipped_reason_count") or {})
            merged.update({str(key): int(value or 0) for key, value in skipped.items()})
            bucket["skipped_reason_count"] = dict(merged)
        allowed = row.get("allowed_reason_count") or {}
        if isinstance(allowed, dict):
            merged_allowed = Counter(bucket.get("allowed_reason_count") or {})
            merged_allowed.update({str(key): int(value or 0) for key, value in allowed.items()})
            bucket["allowed_reason_count"] = dict(merged_allowed)
    source_rows = source_rows or []
    survey_rows = survey_rows or []
    final_rows = final_rows or []
    survey_keys = {article_identity_key(row) for row in survey_rows}
    final_keys = {article_identity_key(row) for row in final_rows}
    for row in source_rows:
        if clean_text(row.get("broad_entry")).lower() != "true":
            continue
        for platform in broad_entry_platform_labels(row):
            bucket = summary.setdefault(
                platform,
                {
                    "target_market_feed_count": 0,
                    "target_market_kept_count": 0,
                    "broad_feed_count": 0,
                    "broad_kept_count": 0,
                    "broad_short_feed_count": 0,
                    "broad_short_kept_count": 0,
                    "broad_entry_count": 0,
                    "multi_platform_event_kept_count": 0,
                    "broad_entry_before_filter_count": 0,
                    "broad_entry_ai_kept_count": 0,
                    "broad_entry_final_count": 0,
                    "broad_entry_ai_excluded_count": 0,
                    "skipped_reason_count": {},
                    "allowed_reason_count": {},
                    "ai_excluded_reason_count": {},
                },
            )
            bucket["broad_entry_before_filter_count"] += 1
            identity = article_identity_key(row)
            if identity in survey_keys:
                bucket["broad_entry_ai_kept_count"] += 1
            else:
                bucket["broad_entry_ai_excluded_count"] += 1
            if identity in final_keys:
                bucket["broad_entry_final_count"] += 1
            if clean_text(row.get("broad_entry_reason")) == "multi_platform_regulatory_or_market_event":
                bucket["multi_platform_event_kept_count"] += 1
    broad_ai_excluded = (survey_filter_stats or {}).get("broad_entry_ai_excluded_reason_counts") or {}
    if isinstance(broad_ai_excluded, dict):
        for platform, reasons in broad_ai_excluded.items():
            if platform not in summary or not isinstance(reasons, dict):
                continue
            merged_reasons = Counter(summary[platform].get("ai_excluded_reason_count") or {})
            merged_reasons.update({str(reason): int(count or 0) for reason, count in reasons.items()})
            summary[platform]["ai_excluded_reason_count"] = dict(merged_reasons)
    return summary


def build_low_volume_platform_diagnostics(
    brand_stage_total_summary: list[dict[str, Any]],
    promo_search_stats: dict[str, Any],
    broad_entry_platform_summary: dict[str, dict[str, Any]],
    platforms: list[str] | None = None,
) -> dict[str, dict[str, Any]]:
    target_platforms = platforms or ["TTS", "IG"]
    engine_summary = promo_search_stats.get("platform_engine_summary") or {}
    query_summary = promo_search_stats.get("platform_query_summary") or {}
    brand_lookup = {
        clean_text(row.get("brand")): row
        for row in brand_stage_total_summary
        if isinstance(row, dict) and clean_text(row.get("brand"))
    }
    diagnostics: dict[str, dict[str, Any]] = {}
    for platform in target_platforms:
        brand_row = brand_lookup.get(platform) or {}
        google_bucket = engine_summary.get(f"{platform}|google") if isinstance(engine_summary, dict) else {}
        if not isinstance(google_bucket, dict):
            google_bucket = {}
        skipped_counter: Counter = Counter()
        feed_count = int(google_bucket.get("feed_count") or 0)
        kept_count = int(google_bucket.get("kept_count") or 0)
        task_count = int(google_bucket.get("task_count") or 0)
        if isinstance(query_summary, dict):
            for row in query_summary.values():
                if not isinstance(row, dict) or clean_text(row.get("platform_label")) != platform:
                    continue
                skipped = row.get("skipped_reason_count") or {}
                if isinstance(skipped, dict):
                    skipped_counter.update({str(reason): int(count or 0) for reason, count in skipped.items()})
        broad_row = broad_entry_platform_summary.get(platform) or {}
        diagnostics[platform] = {
            "google_task_count": task_count,
            "search_feed_count": feed_count,
            "entry_kept_count": kept_count,
            "raw_count": int(brand_row.get("raw_count") or 0),
            "initial_dedupe_count": int(brand_row.get("initial_dedupe_count") or 0),
            "survey_filter_count": int(brand_row.get("survey_filter_count") or 0),
            "final_count": int(brand_row.get("final_count") or 0),
            "top_skipped_reason_count": dict(skipped_counter.most_common(6)),
            "broad_entry_before_filter_count": int(broad_row.get("broad_entry_before_filter_count") or broad_row.get("broad_entry_count") or 0),
            "broad_entry_ai_kept_count": int(broad_row.get("broad_entry_ai_kept_count") or 0),
            "broad_entry_final_count": int(broad_row.get("broad_entry_final_count") or 0),
        }
    return diagnostics


def apply_low_volume_fill(
    source_rows: list[dict[str, Any]],
    kept_rows: list[dict[str, Any]],
    survey_indicators: list[SurveyIndicator],
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
    target_per_brand_week: int = RECALL_TARGET_PER_BRAND_WEEK,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if clean_text(recall_mode).lower() != "balanced" or not source_rows or not survey_indicators:
        return kept_rows, {"enabled": False, "added_count": 0, "target_per_brand_week": target_per_brand_week}

    grouped_indicators = group_survey_indicators_by_dimension(survey_indicators)
    question_lookup = build_survey_indicator_question_lookup(survey_indicators)
    indicator_dimensions = set(grouped_indicators)
    dimension_keyword_lookup = build_dimension_keyword_lookup(survey_indicators)
    existing_keys = {article_identity_key(row) for row in kept_rows}
    counts = Counter(brand_week_counts(kept_rows, country_code))
    added_rows: list[dict[str, Any]] = []
    candidates = [
        row for row in source_rows
        if article_identity_key(row) not in existing_keys and has_platform_level_balanced_signal(row)
    ]
    candidates.sort(key=lambda row: clean_text(row.get("published_at") or ""), reverse=True)

    for row in candidates:
        week = article_week_start(row)
        if not week:
            continue
        brands = article_brand_labels(row, country_code)
        low_brands = [brand for brand in brands if counts[(brand, week)] < target_per_brand_week]
        if not low_brands:
            continue
        enriched, _ = keyword_filter_single_article(
            row,
            grouped_indicators,
            indicator_dimensions,
            dimension_keyword_lookup,
            question_lookup,
            method="low_volume_fill",
            explanation="低量补充复评：该新闻具有平台体验或NPS相关信号，用于补足品牌周度新闻量。",
        )
        if not enriched:
            continue
        enriched["survey_filter_confidence"] = "medium"
        enriched["volume_fill"] = "true"
        added_rows.append(enriched)
        existing_keys.add(article_identity_key(row))
        for brand in low_brands:
            counts[(brand, week)] += 1

    merged_rows = kept_rows + added_rows
    merged_rows.sort(key=lambda row: clean_text(row.get("published_at") or ""), reverse=True)
    return merged_rows, {
        "enabled": True,
        "added_count": len(added_rows),
        "target_per_brand_week": target_per_brand_week,
        "added_brand_week_counts": {f"{brand}|{week}": count for (brand, week), count in brand_week_counts(added_rows, country_code).items()},
    }


def official_source_target(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, str]:
    country_sources = country_dict_setting(country_code, "official_sources")
    fallback_sources = country_dict_setting(DEFAULT_COUNTRY_CODE, "official_sources")
    target = country_sources.get(platform) or fallback_sources.get(platform) or {}
    return {
        "label": clean_text(target.get("label")),
        "url": clean_text(target.get("url")),
    }


def default_official_source_url(platform: str) -> str:
    target = country_dict_setting(DEFAULT_COUNTRY_CODE, "official_sources").get(platform, {})
    return clean_text(target.get("url")) if isinstance(target, dict) else ""


def country_for_official_source_url(source_url: str, default: str = DEFAULT_COUNTRY_CODE) -> str:
    parsed_source = urllib.parse.urlparse(source_url)
    source_host = parsed_source.netloc.lower()
    source_path = parsed_source.path.rstrip("/")
    best_country = normalize_country_code(default)
    best_score = 0

    for country_code in available_country_codes():
        for target in country_dict_setting(country_code, "official_sources").values():
            if not isinstance(target, dict):
                continue
            target_url = clean_text(target.get("url"))
            if not target_url:
                continue
            parsed_target = urllib.parse.urlparse(target_url)
            if parsed_target.netloc.lower() != source_host:
                continue
            target_path = parsed_target.path.rstrip("/")
            score = 1
            if target_path and (source_path == target_path or source_path.startswith(f"{target_path}/")):
                score += len(target_path)
            if parsed_target.query and parsed_target.query.lower() == parsed_source.query.lower():
                score += len(parsed_target.query)
            if score > best_score:
                best_country = country_code
                best_score = score

    return best_country


def source_matches_official_platform(source_url: str, platform: str) -> bool:
    source_host = urllib.parse.urlparse(source_url).netloc.lower()
    if not source_host:
        return False
    for country_code in available_country_codes():
        target = country_dict_setting(country_code, "official_sources").get(platform)
        if not isinstance(target, dict):
            continue
        target_url = clean_text(target.get("url"))
        target_host = urllib.parse.urlparse(target_url).netloc.lower()
        if source_host == target_host:
            return True
    return False


def list_available_platform_labels(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    overrides = [str(item) for item in country_list_setting(country_code, "available_platform_labels") if str(item).strip()]
    if overrides:
        return overrides
    return [display_platform_label(platform, country_code) for platform in DISPLAY_TO_SOURCE_PLATFORM.values()]


def display_platform_label(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    overrides = country_dict_setting(country_code, "platform_display_overrides")
    if platform in overrides:
        return str(overrides[platform])
    return SOURCE_TO_DISPLAY_PLATFORM.get(platform, platform)


def platform_search_terms(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    overrides = country_dict_setting(country_code, "platform_search_term_overrides")
    if platform in overrides:
        return [clean_text(item) for item in overrides[platform] if clean_text(item)]
    normalized_platform = clean_text(platform).lower()
    for override_key, override_terms in overrides.items():
        if clean_text(override_key).lower() == normalized_platform and isinstance(override_terms, list):
            return [clean_text(item) for item in override_terms if clean_text(item)]
    source_platform = DISPLAY_TO_SOURCE_PLATFORM.get(platform)
    if source_platform is None:
        source_platform = next(
            (
                source
                for label, source in DISPLAY_TO_SOURCE_PLATFORM.items()
                if clean_text(label).lower() == normalized_platform
            ),
            None,
        )
    if source_platform is not None:
        if source_platform in overrides:
            return [clean_text(item) for item in overrides[source_platform] if clean_text(item)]
        for override_key, override_terms in overrides.items():
            if clean_text(override_key).lower() == clean_text(source_platform).lower() and isinstance(override_terms, list):
                return [clean_text(item) for item in override_terms if clean_text(item)]
        if source_platform in PLATFORM_SEARCH_TERMS:
            return [clean_text(item) for item in PLATFORM_SEARCH_TERMS.get(source_platform, []) if clean_text(item)]
    return [clean_text(item) for item in PLATFORM_SEARCH_TERMS.get(platform, [platform]) if clean_text(item)]


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", str(value or "")))


def normalize_platform_alias_token(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = re.sub(r"\([^)]*\)", " ", normalized)
    normalized = re.sub(r"[^A-Za-z0-9]+", " ", normalized).strip().lower()
    return re.sub(r"\s+", " ", normalized)


def platform_alias_concept(platform: str) -> str:
    normalized = normalize_platform_alias_token(platform)
    collapsed = normalized.replace(" ", "")
    if not normalized:
        return ""
    if (
        "instagram" in normalized
        or collapsed in {"ig", "ins", "insdtc", "insshop", "instagramshop", "instagramshopping"}
        or ("ins" in normalized and "dtc" in normalized)
    ):
        return "instagram_shopping"
    if "tiktok" in normalized or collapsed in {"tts", "ttsop", "tiktokshop", "ttshop"}:
        return "tiktok_shop"
    if "amazon" in normalized:
        return "amazon"
    if "zalando" in normalized:
        return "zalando"
    if "shein" in normalized:
        return "shein"
    if "temu" in normalized:
        return "temu"
    if "rakuten" in normalized:
        return "rakuten"
    return ""


def platform_alias_seed_terms(platform: str) -> list[str]:
    platform_name = str(platform or "").strip()
    concept = platform_alias_concept(platform_name)
    concept_terms = {
        "instagram_shopping": [
            platform_name,
            "Instagram Shopping",
            "Instagram Shop",
            "Instagram",
            "Instagram DTC",
            "IG",
            "INS",
            "INS-DTC",
        ],
        "tiktok_shop": [
            platform_name,
            "TikTok Shop",
            "TikTok",
            "TTS",
            "TT Shop",
        ],
        "amazon": [platform_name, "Amazon"],
        "zalando": [platform_name, "Zalando"],
        "shein": [platform_name, "SHEIN", "Shein"],
        "temu": [platform_name, "Temu", "TEMU"],
        "rakuten": [platform_name, "Rakuten", "Rakuten Ichiba"],
    }
    return unique_clean_texts(concept_terms.get(concept, [platform_name]))


def custom_platform_country_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    config = get_country_config(country_code)
    terms: list[str] = []
    skip_tokens = [
        "market",
        "markets",
        "consumer",
        "consumers",
        "consommateur",
        "consommateurs",
        "mercado",
        "mercati",
        "mercato",
        "markt",
    ]
    for term in country_market_terms(country_code):
        normalized = clean_text(term)
        lowered = normalized.lower()
        if (
            not normalized
            or contains_cjk(normalized)
            or any(token in lowered for token in skip_tokens)
        ):
            continue
        if len(normalized.split()) > 2:
            continue
        if normalized.islower():
            normalized = normalized.title()
        if normalized not in terms:
            terms.append(normalized)
    google_gl = clean_text(str(config.get("google_news_gl") or "")).upper()
    if google_gl and google_gl not in terms:
        terms.append(google_gl)
    title_code = clean_text(country_code).replace("_", " ").title()
    if len(title_code) > 2 and title_code not in terms:
        terms.append(title_code)
    return terms


def generated_platform_search_terms(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    platform_name = clean_text(platform)
    if not platform_name:
        return []
    config = get_country_config(country_code)
    google_gl = clean_text(str(config.get("google_news_gl") or "")).upper()
    country_terms = custom_platform_country_terms(country_code)
    primary_country = country_terms[0] if country_terms else clean_text(country_code).title()
    seed_terms = platform_alias_seed_terms(platform_name)
    candidates: list[str] = []

    def add_candidate(value: str) -> None:
        normalized = clean_text(value)
        if normalized and not contains_cjk(normalized) and normalized not in candidates:
            candidates.append(normalized)

    for seed in seed_terms:
        add_candidate(seed)
        if primary_country:
            add_candidate(f"{seed} {primary_country}")
        if google_gl:
            add_candidate(f"{seed} {google_gl}")

    concept = platform_alias_concept(platform_name)
    if concept in {"amazon", "zalando", "shein", "temu"} and google_gl and len(google_gl) == 2:
        canonical = {
            "amazon": "Amazon",
            "zalando": "Zalando",
            "shein": "shein",
            "temu": "temu",
        }[concept]
        add_candidate(f"{canonical}.{google_gl.lower()}")
    elif concept == "instagram_shopping":
        for term in country_terms[:3]:
            add_candidate(f"Instagram Shopping {term}")
            add_candidate(f"Instagram Shop {term}")
            add_candidate(f"Instagram {term}")
    elif concept == "tiktok_shop":
        for term in country_terms[:3]:
            add_candidate(f"TikTok Shop {term}")
            add_candidate(f"TikTok {term}")

    return candidates


def expanded_platform_search_terms(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    configured_terms = platform_search_terms(platform, country_code)
    normalized_platform = clean_text(platform)
    if configured_terms and any(clean_text(term).lower() != normalized_platform.lower() for term in configured_terms):
        return unique_clean_texts(configured_terms)
    return unique_clean_texts([normalized_platform, *configured_terms, *generated_platform_search_terms(platform, country_code)])


def platform_alias_effective_summary(
    platforms: list[str],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    full_name_signals = {
        "TTS": ["tiktok shop", "tiktok"],
        "TikTok Shop": ["tiktok shop", "tiktok"],
        "IG": ["instagram", "instagram shop", "instagram shopping"],
        "Instagram": ["instagram", "instagram shop", "instagram shopping"],
        "Instagram Shopping": ["instagram", "instagram shop", "instagram shopping"],
    }
    abbreviation_signals = {
        "TTS": {"tts", "tt shop"},
        "TikTok Shop": {"tts", "tt shop"},
        "IG": {"ig", "ins", "ins-dtc"},
        "Instagram": {"ig", "ins", "ins-dtc"},
        "Instagram Shopping": {"ig", "ins", "ins-dtc"},
    }
    for platform in platforms:
        normalized_platform = clean_text(platform)
        terms = expanded_platform_search_terms(normalized_platform, country_code)
        lowered_terms = [clean_text(term).lower() for term in terms]
        concept = platform_alias_concept(normalized_platform)
        canonical_key = "TTS" if concept == "tiktok_shop" else ("IG" if concept == "instagram_shopping" else normalized_platform)
        full_signals = full_name_signals.get(canonical_key, [normalized_platform.lower()])
        abbrev_signals = abbreviation_signals.get(canonical_key, set())
        has_full_name = any(any(signal in term for signal in full_signals) for term in lowered_terms)
        abbreviation_only = bool(terms) and not has_full_name and all(term in abbrev_signals for term in lowered_terms)
        warnings: list[str] = []
        if canonical_key in {"TTS", "IG"} and not has_full_name:
            warnings.append("品牌别名未展开到全称，可能影响搜索入池。")
        if abbreviation_only:
            warnings.append("品牌别名似乎只包含缩写，建议补充平台全称。")
        summary[normalized_platform] = {
            "alias_count": len(terms),
            "has_full_name": has_full_name,
            "abbreviation_only": abbreviation_only,
            "representative_terms": terms[:12],
            "warnings": warnings,
        }
    return summary


def requested_platform_lookup(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for source_platform, aliases in CANONICAL_PLATFORM_ALIASES.items():
        mapping[clean_text(source_platform).lower()] = source_platform
        mapping[clean_text(display_platform_label(source_platform, country_code)).lower()] = source_platform
        normalized_aliases = aliases
        alias_exclude_tokens = [clean_text(str(token)).lower() for token in country_list_setting(country_code, "platform_alias_exclude_tokens")]
        if alias_exclude_tokens:
            normalized_aliases = [
                alias
                for alias in aliases
                if not any(token in clean_text(alias).lower() for token in alias_exclude_tokens)
            ]
        for alias in normalized_aliases:
            normalized = clean_text(alias).lower()
            if normalized:
                mapping[normalized] = source_platform
    for source_platform in DISPLAY_TO_SOURCE_PLATFORM.values():
        display_label = display_platform_label(source_platform, country_code)
        for alias in [
            source_platform,
            display_label,
            *platform_search_terms(source_platform, country_code),
            *platform_search_terms(display_label, country_code),
        ]:
            normalized = clean_text(alias).lower()
            if normalized:
                mapping[normalized] = source_platform
    return mapping


def requested_display_platform_lookup(country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for display_label in list_available_platform_labels(country_code):
        for alias in expanded_platform_search_terms(display_label, country_code):
            normalized = clean_text(alias).lower()
            if normalized:
                mapping[normalized] = display_label
    return mapping


def resolve_requested_platforms(
    requested_platforms: list[str] | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> tuple[list[str], list[str], list[str]]:
    requested = [clean_text(item) for item in (requested_platforms or []) if clean_text(item)]
    if not requested:
        source_platforms = list(DISPLAY_TO_SOURCE_PLATFORM.values())
        display_platforms = [display_platform_label(item, country_code) for item in source_platforms]
        return display_platforms, source_platforms, []

    lookup = requested_platform_lookup(country_code)
    display_lookup = requested_display_platform_lookup(country_code)
    selected_source_platforms: list[str] = []
    selected_display_platforms: list[str] = []
    custom_platforms: list[str] = []
    for item in requested:
        source_platform = lookup.get(item.lower())
        if source_platform:
            if source_platform not in selected_source_platforms:
                selected_source_platforms.append(source_platform)
            display_label = display_platform_label(source_platform, country_code)
            if display_label not in selected_display_platforms:
                selected_display_platforms.append(display_label)
            continue
        display_label = display_lookup.get(item.lower())
        if display_label:
            source_platform = DISPLAY_TO_SOURCE_PLATFORM.get(display_label)
            if source_platform:
                if source_platform not in selected_source_platforms:
                    selected_source_platforms.append(source_platform)
                normalized_display = display_platform_label(source_platform, country_code)
                if normalized_display not in selected_display_platforms:
                    selected_display_platforms.append(normalized_display)
            elif display_label not in custom_platforms:
                custom_platforms.append(display_label)
            continue
        if item not in custom_platforms:
            custom_platforms.append(item)
    return selected_display_platforms, selected_source_platforms, custom_platforms


def build_media_search_terms(requested_platforms: list[str] | None, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    _, selected_source_platforms, custom_platforms = resolve_requested_platforms(requested_platforms, country_code)
    terms: list[str] = []
    for source_platform in selected_source_platforms:
        for term in platform_search_terms(source_platform, country_code):
            normalized = clean_text(term)
            if normalized and normalized not in terms:
                terms.append(normalized)
    for platform in custom_platforms:
        for term in expanded_platform_search_terms(platform, country_code):
            normalized = clean_text(term)
            if normalized and normalized not in terms:
                terms.append(normalized)
    return terms


def build_source_search_terms(
    requested_platforms: list[str] | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    *,
    max_terms_per_platform: int = SOURCE_SHORT_WINDOW_MAX_TERMS_PER_PLATFORM,
) -> list[str]:
    """Build a compact term set for configured source-site search.

    Source adapters run a query per term. For short windows, using every alias
    makes slow/low-yield sources dominate runtime without improving recency
    yield. Keep brand coverage, but cap aliases per selected platform.
    """
    selected_display_platforms, selected_source_platforms, custom_platforms = resolve_requested_platforms(
        requested_platforms,
        country_code,
    )
    terms: list[str] = []

    for display_label, source_platform in zip(selected_display_platforms, selected_source_platforms):
        candidates = [display_label, source_platform]
        candidates.extend(platform_search_terms(source_platform, country_code))
        for term in unique_clean_texts(candidates)[:max(1, max_terms_per_platform)]:
            if term not in terms:
                terms.append(term)

    for platform in custom_platforms:
        if clean_text(platform) == "TTS":
            candidates = ["TikTok Shop", "TikTok"]
        elif clean_text(platform) == "IG":
            candidates = ["Instagram Shopping", "Instagram"]
        else:
            candidates = [platform]
            candidates.extend(expanded_platform_search_terms(platform, country_code))
        for term in unique_clean_texts(candidates)[:max(1, max_terms_per_platform)]:
            if term not in terms:
                terms.append(term)

    return terms or build_media_search_terms(requested_platforms, country_code)


def media_match_targets(requested_platforms: list[str] | None, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    _, selected_source_platforms, custom_platforms = resolve_requested_platforms(requested_platforms, country_code)
    return selected_source_platforms + custom_platforms


def unique_clean_texts(values: list[str]) -> list[str]:
    ordered: list[str] = []
    for value in values:
        normalized = clean_text(value)
        if normalized and normalized not in ordered:
            ordered.append(normalized)
    return ordered


def normalize_promo_search_keyword_blocks(
    raw_text: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[tuple[str, str]]:
    source_text = normalize_promo_search_keywords_text(raw_text, country_code)
    blocks: list[tuple[str, str]] = []
    for line_index, raw_line in enumerate(source_text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        label = f'custom_{line_index}'
        query_block = line
        for separator in ['=', ':']:
            if separator in line:
                possible_label, possible_block = line.split(separator, 1)
                if possible_block.strip():
                    label = clean_text(possible_label) or label
                    query_block = possible_block.strip()
                    break
        query_block = query_block.strip()
        if not query_block:
            continue
        block_key = clean_text(query_block)
        if not block_key:
            continue
        if not any(existing_block == query_block for _, existing_block in blocks):
            blocks.append((label, query_block))
    return blocks


def normalize_related_news_search_keyword_blocks(
    raw_text: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
) -> list[tuple[str, str]]:
    source_text = normalize_related_news_search_keywords_text(raw_text, country_code, recall_mode)
    blocks: list[tuple[str, str]] = []
    for line_index, raw_line in enumerate(source_text.splitlines(), start=1):
        for split_index, line in enumerate(split_keyword_line_into_blocks(raw_line), start=1):
            line = line.strip()
            if not line:
                continue
            label = f"related_news_{line_index}"
            if split_index > 1:
                label = f"{label}_{split_index}"
            query_block = line
            for separator in ['=', ':']:
                if separator in line:
                    possible_label, possible_block = line.split(separator, 1)
                    if possible_block.strip():
                        label = clean_text(possible_label) or label
                        query_block = possible_block.strip()
                        break
            query_block = query_block.strip()
            if not query_block:
                continue
            if not any(existing_block == query_block for _, existing_block in blocks):
                blocks.append((label, query_block))
    if clean_text(recall_mode).lower() == "balanced":
        for label, query_block in recall_enhanced_related_news_query_blocks(country_code):
            if not any(clean_text(existing_block) == clean_text(query_block) for _, existing_block in blocks):
                blocks.append((label, query_block))
    return blocks


def normalize_report_search_keyword_blocks(
    raw_text: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[tuple[str, str]]:
    source_text = normalize_report_search_keywords_text(raw_text, country_code)
    blocks: list[tuple[str, str]] = []
    for line_index, raw_line in enumerate(source_text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        label = f"report_{line_index}"
        query_block = line
        for separator in ['=', ':']:
            if separator in line:
                possible_label, possible_block = line.split(separator, 1)
                if possible_block.strip():
                    label = clean_text(possible_label) or label
                    query_block = possible_block.strip()
                    break
        query_block = query_block.strip()
        if not query_block:
            continue
        if not any(existing_block == query_block for _, existing_block in blocks):
            blocks.append((label, query_block))
    industry_blocks = [
        (
            "industry_trend",
            '(trend OR trends OR "market trend" OR "industry trend" OR adoption OR penetration OR "consumer behavior" OR "ecommerce growth" OR "social commerce" OR "marketplace report")',
        ),
    ]
    if clean_text(country_code).lower() == "italy":
        industry_blocks.append(
            (
                "industry_trend_local",
                '(tendenza OR tendenze OR mercato OR adozione OR penetrazione OR "comportamento dei consumatori" OR "crescita ecommerce" OR "social commerce" OR "rapporto marketplace")',
            )
        )
    for label, query_block in industry_blocks:
        if not any(existing_block == query_block for _, existing_block in blocks):
            blocks.append((label, query_block))
    return blocks


def select_promo_search_engines(engine_mode: str) -> list[str]:
    normalized = clean_text(engine_mode).lower()
    if normalized == "google":
        return ["google"]
    if normalized == "bing":
        return ["bing"]
    return ["google", "bing"]


def build_promo_brand_terms(
    display_label: str,
    source_platform: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[str]:
    terms: list[str] = []

    def add_term(value: str) -> None:
        normalized = clean_text(value)
        if normalized and normalized not in terms:
            terms.append(normalized)

    add_term(display_label)
    for alias in expanded_platform_search_terms(display_label, country_code):
        add_term(alias)
    if source_platform:
        normalized_source_platform = clean_text(source_platform).lower()
        alias_exclude_tokens = [clean_text(str(token)).lower() for token in country_list_setting(country_code, "platform_alias_exclude_tokens")]
        should_include_source_platform = '/' not in clean_text(source_platform)
        if should_include_source_platform and (
            not alias_exclude_tokens
            or not any(token in normalized_source_platform for token in alias_exclude_tokens)
        ):
            add_term(source_platform)
        for alias in platform_search_terms(source_platform, country_code):
            add_term(alias)
        for alias in platform_match_aliases(source_platform, country_code):
            add_term(alias)

    preferred: list[str] = []
    for term in terms:
        lowered = term.lower()
        if any(token in lowered for token in [".co.jp", "shop", "市場", "amazon", "rakuten", "qoo10", "temu", "shein", "tiktok"]):
            if term not in preferred:
                preferred.append(term)
    for term in terms:
        if term not in preferred:
            preferred.append(term)
        if len(preferred) >= 5:
            break
    return preferred[:5]


def is_instagram_platform_label(label: str | None, source_platform: str | None = None) -> bool:
    values = [clean_text(label), clean_text(source_platform)]
    return any(platform_alias_concept(value) == "instagram_shopping" for value in values if value)


def is_tiktok_shop_platform_label(label: str | None, source_platform: str | None = None) -> bool:
    values = [clean_text(label), clean_text(source_platform)]
    return any(platform_alias_concept(value) == "tiktok_shop" for value in values if value)


def instagram_commerce_brand_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    country_terms = custom_platform_country_terms(country_code)
    primary_country = country_terms[0] if country_terms else clean_text(country_code).title()
    candidates = [
        "Instagram",
        "Instagram Shop",
        "Instagram Shopping",
        "Instagram ecommerce",
        "Instagram social commerce",
        "Instagram creator shop",
    ]
    if primary_country:
        candidates.extend(
            [
                f"Instagram {primary_country}",
                f"Instagram Shop {primary_country}",
                f"Instagram Shopping {primary_country}",
                f"Instagram ecommerce {primary_country}",
            ]
        )
    if clean_text(country_code).lower() == "italy":
        candidates.extend(
            [
                "Instagram Italia",
                "Instagram Shop Italia",
                "Instagram Shopping Italia",
            ]
        )
    return unique_clean_texts(candidates)


def instagram_commerce_qualifier_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    candidates = [
        "Instagram Shop",
        "Instagram Shopping",
        "Instagram ecommerce",
        "Instagram social commerce",
        "Instagram creator shop",
        "Instagram business",
        "Instagram checkout",
        "Meta shopping tools",
        "Meta commerce",
        "Meta commerce tools",
        "Meta business tools",
        "Meta ads commerce",
        "creator economy",
        "reels shopping",
        "shoppable posts",
        "brand partnership",
        "creator marketplace",
        "business tools",
        "social commerce",
        "creator shop",
        "ecommerce",
        "shopping",
        "shop",
        "checkout",
    ]
    if clean_text(country_code).lower() == "italy":
        candidates.extend(
            [
                "shopping su Instagram",
                "acquisti su Instagram",
                "comprare su Instagram",
                "negozio Instagram",
                "social commerce Italia",
                "ecommerce Instagram Italia",
                "creator economy Italia",
                "strumenti business Instagram",
                "shopping nei Reels",
                "post acquistabili",
            ]
        )
    return unique_clean_texts(candidates)


def instagram_strong_commerce_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    terms = [
        "Instagram Shop",
        "Instagram Shopping",
        "Instagram ecommerce",
        "Instagram social commerce",
        "Instagram creator shop",
        "Instagram checkout",
        "Meta shopping tools",
        "Meta commerce",
        "Meta commerce tools",
        "Meta business tools",
        "Meta ads commerce",
        "social commerce",
        "creator shop",
        "reels shopping",
        "shoppable posts",
        "checkout",
    ]
    if clean_text(country_code).lower() == "italy":
        terms.extend([
            "shopping su Instagram",
            "acquisti su Instagram",
            "comprare su Instagram",
            "negozio Instagram",
            "social commerce Italia",
            "shopping nei Reels",
            "post acquistabili",
        ])
    return unique_clean_texts(terms)


def instagram_weak_commerce_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    terms = [
        "Instagram business",
        "Instagram business tools",
        "Meta commerce",
        "Meta commerce tools",
        "Meta business tools",
        "Meta ads commerce",
        "creator economy",
        "brand partnership",
        "creator marketplace",
        "shoppable posts",
        "reels shopping",
        "business tools",
        "small business",
        "merchant",
        "advertising",
        "ads",
        "commerce",
    ]
    if clean_text(country_code).lower() == "italy":
        terms.extend([
            "creator economy Italia",
            "strumenti business Instagram",
            "piccole imprese",
            "commercianti",
            "pubblicità",
        ])
    return unique_clean_texts(terms)


def temu_eu_global_impact_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    terms = [
        "EU",
        "European Union",
        "European Commission",
        "Europe",
        "European",
        "DSA",
        "Digital Services Act",
        "consumer protection",
        "consumer rights",
        "customs",
        "low value parcel",
        "low-value parcels",
        "de minimis",
        "online marketplace",
        "platform regulation",
        "product safety",
        "privacy",
        "data protection",
        "GDPR",
        "investigation",
        "compliance",
        "counterfeit",
        "illegal products",
        "market share",
        "growth",
    ]
    if clean_text(country_code).lower() == "italy":
        terms.extend(
            [
                "Unione Europea",
                "Commissione europea",
                "Europa",
                "protezione consumatori",
                "tutela dei consumatori",
                "dogane",
                "pacchi di basso valore",
                "marketplace online",
                "sicurezza prodotti",
                "privacy",
                "dati personali",
                "indagine",
                "inchiesta",
                "conformità",
                "contraffazione",
                "prodotti illegali",
            ]
        )
    return unique_clean_texts(terms)


def article_text_for_matching(article: dict[str, Any]) -> str:
    return clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )


def article_matches_temu_broad_context(article: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> bool:
    combined_text = article_text_for_matching(article)
    if not combined_text:
        return False
    if not any(alias_matches(combined_text, alias) for alias in expanded_platform_search_terms("TEMU", country_code)):
        return False
    return any(alias_matches(combined_text, term) for term in temu_eu_global_impact_terms(country_code))


def article_matches_instagram_broad_context(article: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> bool:
    combined_text = article_text_for_matching(article)
    if not combined_text:
        return False
    platform_signal = any(alias_matches(combined_text, term) for term in ["Instagram", "Meta"])
    commerce_signal = any(
        alias_matches(combined_text, term)
        for term in unique_clean_texts(instagram_strong_commerce_terms(country_code) + instagram_weak_commerce_terms(country_code))
    )
    return platform_signal and commerce_signal


def tiktok_shop_broad_context_terms(country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    terms = [
        "TikTok Shop",
        "social commerce",
        "live shopping",
        "live commerce",
        "creator economy",
        "creator marketplace",
        "seller",
        "sellers",
        "merchant",
        "merchants",
        "marketplace",
        "payment",
        "delivery",
        "shipping",
        "return",
        "refund",
        "consumer protection",
        "regulation",
        "compliance",
        "privacy",
        "data",
        "product safety",
    ]
    if clean_text(country_code).lower() == "italy":
        terms.extend([
            "acquisti su TikTok",
            "negozio TikTok",
            "venditore",
            "venditori",
            "commercianti",
            "pagamento",
            "consegna",
            "spedizione",
            "reso",
            "rimborso",
            "tutela consumatori",
            "conformità",
            "sicurezza",
        ])
    return unique_clean_texts(terms)


def article_matches_tiktok_shop_broad_context(article: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> bool:
    combined_text = article_text_for_matching(article)
    if not combined_text:
        return False
    platform_signal = any(alias_matches(combined_text, alias) for alias in expanded_platform_search_terms("TTS", country_code))
    if not platform_signal:
        return False
    return any(alias_matches(combined_text, term) for term in tiktok_shop_broad_context_terms(country_code))


def article_matches_multi_platform_regulatory_event(article: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> bool:
    combined_text = article_text_for_matching(article)
    if not combined_text:
        return False
    has_temu = any(alias_matches(combined_text, alias) for alias in expanded_platform_search_terms("TEMU", country_code))
    if not has_temu:
        return False
    multi_platform_terms = [
        "Shein",
        "SHEIN",
        "AliExpress",
        "Chinese marketplace",
        "Chinese marketplaces",
        "Chinese platforms",
        "Chinese ecommerce",
        "piattaforme cinesi",
        "marketplace cinesi",
        "ecommerce cinese",
        "Shein e Temu",
        "AliExpress, Shein, Temu",
    ]
    has_multi_platform_signal = any(alias_matches(combined_text, term) for term in multi_platform_terms)
    has_regulatory_signal = any(alias_matches(combined_text, term) for term in temu_eu_global_impact_terms(country_code))
    return has_multi_platform_signal and has_regulatory_signal


def broad_entry_reason_for_article(
    article: dict[str, Any],
    task: dict[str, Any],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    platform_label = clean_text(task.get("platform_label"))
    if platform_label == "TEMU" and article_matches_multi_platform_regulatory_event(article, country_code):
        return "multi_platform_regulatory_or_market_event"
    if platform_label == "TEMU" and article_matches_temu_broad_context(article, country_code):
        return "TEMU EU/Europe/global platform impact signal"
    if is_tiktok_shop_platform_label(platform_label, task.get("source_platform")) and article_matches_tiktok_shop_broad_context(article, country_code):
        return "TTS/TikTok Shop platform ecosystem or social commerce signal"
    if is_instagram_platform_label(platform_label, task.get("source_platform")) and article_matches_instagram_broad_context(article, country_code):
        return "IG/Meta indirect commerce or creator economy signal"
    return ""


def build_or_block(terms: list[str]) -> str:
    normalized_terms = unique_clean_texts(terms)
    quoted_terms = [f'"{term}"' if " " in term else term for term in normalized_terms]
    if not quoted_terms:
        return ""
    if len(quoted_terms) == 1:
        return quoted_terms[0]
    return f"({' OR '.join(quoted_terms)})"


def build_platform_promo_event_block(source_platform: str | None) -> str | None:
    if not source_platform:
        return None
    terms = [clean_text(term) for term in PLATFORM_PROMO_EVENT_TERMS.get(source_platform, []) if clean_text(term)]
    if not terms:
        return None
    return build_or_block(terms)


def build_promo_search_query(
    display_label: str,
    source_platform: str | None = None,
    keyword_block: str | None = None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    include_market_block: bool = True,
    primary_term_only: bool = False,
) -> str:
    brand_terms = build_promo_brand_terms(display_label, source_platform, country_code)
    if primary_term_only and brand_terms:
        brand_terms = brand_terms[:1]
    quoted_terms = [f'"{term}"' if " " in term else term for term in brand_terms]
    brand_block = " OR ".join(quoted_terms)
    if len(quoted_terms) > 1:
        brand_block = f"({brand_block})"
    default_blocks = promo_search_query_blocks(country_code)
    promo_block = keyword_block or default_blocks[0][1]
    query_parts = [brand_block, promo_block]
    if include_market_block:
        query_parts.append(country_market_search_block(country_code))
    return " ".join(part for part in query_parts if part)


def build_brand_only_search_query(
    display_label: str,
    source_platform: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    if is_instagram_platform_label(display_label, source_platform):
        return build_or_block(instagram_commerce_brand_terms(country_code))
    brand_terms = build_promo_brand_terms(display_label, source_platform, country_code)
    return build_or_block(brand_terms) or clean_text(display_label)


def list_survey_dimensions(survey_indicators: list[SurveyIndicator]) -> list[str]:
    ordered: list[str] = []
    for indicator in survey_indicators:
        dimension = clean_text(indicator.dimension)
        if dimension and dimension not in ordered:
            ordered.append(dimension)
    return ordered


def dimension_search_terms(
    dimension: str,
    dimension_keyword_lookup: dict[str, list[str]],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[str]:
    ordered: list[str] = []
    country_overrides = country_dict_setting(country_code, "dimension_search_term_overrides")
    dimension_overrides = country_overrides.get(dimension, [])
    for term in dimension_overrides if isinstance(dimension_overrides, list) else []:
        normalized = clean_text(term)
        if normalized and normalized not in ordered:
            ordered.append(normalized)
    normalized_dimension = clean_text(dimension)
    if normalized_dimension and normalized_dimension not in ordered:
        ordered.append(normalized_dimension)
    for term in dimension_keyword_lookup.get(dimension, []):
        normalized = clean_text(term)
        if normalized and normalized not in ordered:
            ordered.append(normalized)
    return ordered[:6]


def build_dimension_search_block(
    dimension: str,
    dimension_keyword_lookup: dict[str, list[str]],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    terms = dimension_search_terms(dimension, dimension_keyword_lookup, country_code)
    if not terms:
        return ""
    quoted_terms = [f'"{term}"' if " " in term else term for term in terms]
    if len(quoted_terms) == 1:
        return quoted_terms[0]
    return f"({' OR '.join(quoted_terms)})"


def build_selected_brand_scope_block(
    requested_platforms: list[str] | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    selected_display_platforms, selected_source_platforms, custom_platforms = resolve_requested_platforms(requested_platforms, country_code)
    terms: list[str] = []
    for platform in selected_display_platforms + custom_platforms:
        for term in expanded_platform_search_terms(platform, country_code):
            normalized = clean_text(term)
            if normalized and normalized not in terms:
                terms.append(normalized)
    for source_platform in selected_source_platforms:
        for term in build_promo_brand_terms(display_platform_label(source_platform, country_code), source_platform, country_code):
            normalized = clean_text(term)
            if normalized and normalized not in terms:
                terms.append(normalized)
    quoted_terms = [f'"{term}"' if " " in term else term for term in terms[:8]]
    if not quoted_terms:
        return ""
    if len(quoted_terms) == 1:
        return quoted_terms[0]
    return f"({' OR '.join(quoted_terms)})"


def build_related_news_search_query(
    display_label: str,
    source_platform: str | None,
    dimension: str,
    dimension_keyword_lookup: dict[str, list[str]],
    keyword_block: str | None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    brand_block = build_brand_only_search_query(display_label, source_platform, country_code)
    query_parts = [brand_block]
    if is_instagram_platform_label(display_label, source_platform):
        query_parts.append(build_or_block(instagram_commerce_qualifier_terms(country_code)))
    query_parts.extend([keyword_block, country_market_search_block(country_code)])
    return " ".join(part for part in query_parts if part)


def dedicated_google_platform_blocks(
    display_label: str,
    source_platform: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[tuple[str, str]]:
    normalized_label = clean_text(display_label)
    if normalized_label == "TEMU":
        return [
            ("temu_news", '(news OR ecommerce OR marketplace OR app OR shopping OR retail OR consumatori)'),
            ("temu_italy_market", '("Temu Italia" OR "Temu Italy" OR "Temu in Italia" OR "Temu mercato italiano" OR "Temu consumatori italiani")'),
            ("temu_service", '(payment OR delivery OR shipping OR return OR refund OR customer service OR pagamento OR pagamenti OR consegna OR spedizione OR reso OR rimborso OR assistenza)'),
            ("temu_safety", '(safety OR privacy OR data OR investigation OR compliance OR counterfeit OR illegal OR consumer protection OR sicurezza OR privacy OR dati OR indagine OR inchiesta OR conformità OR contraffatto OR illegale)'),
            ("temu_seller", '(seller OR merchant OR marketplace OR commission OR policy OR venditore OR venditori OR commercianti OR commissioni OR regole)'),
            ("temu_reviews", '(reviews OR complaints OR trust OR reliability OR recensioni OR reclami OR affidabile OR fiducia OR "è sicuro")'),
            ("temu_trend", '(trend OR report OR market OR adoption OR penetration OR growth OR "market share" OR "consumer behavior" OR tendenza OR rapporto OR mercato OR crescita OR "quota di mercato")'),
        ]
    if normalized_label in {"TTS", "TikTok Shop"} or platform_alias_concept(source_platform or normalized_label) == "tiktok_shop":
        return [
            ("tts_news", '("TikTok Shop" OR "TikTok Shop Italia" OR "TikTok Shop Italy") (news OR ecommerce OR marketplace OR shopping OR retail)'),
            ("tts_social_commerce", '("TikTok Shop" OR TikTok) ("social commerce" OR "live shopping" OR livestream OR creator OR creators OR "creator shop" OR "creator economy" OR "live commerce")'),
            ("tts_service", '("TikTok Shop" OR TikTok) (payment OR delivery OR shipping OR return OR refund OR pagamento OR pagamenti OR consegna OR spedizione OR reso OR rimborso)'),
            ("tts_seller", '("TikTok Shop" OR TikTok) (seller OR merchant OR commission OR policy OR marketplace OR venditore OR venditori OR commercianti OR commissioni OR regole)'),
            ("tts_italy", '("TikTok Shop Italia" OR "TikTok Shop Italy" OR "TikTok shopping Italia" OR "acquisti su TikTok" OR "negozio TikTok")'),
            ("tts_safety", '("TikTok Shop" OR TikTok) (safety OR privacy OR data OR compliance OR consumer protection OR sicurezza OR privacy OR dati OR conformità OR tutela consumatori)'),
            ("tts_trend", '("TikTok Shop" OR TikTok) (trend OR report OR adoption OR penetration OR growth OR "consumer behavior" OR "market share" OR tendenza OR rapporto OR crescita OR consumatori OR "quota di mercato")'),
        ]
    if normalized_label == "IG" or platform_alias_concept(source_platform or normalized_label) == "instagram_shopping":
        return [
            ("ig_business", '("Instagram business" OR "Instagram for Business" OR "Instagram per aziende" OR "Meta business tools" OR "Meta Business Suite" OR "strumenti business Instagram")'),
            ("ig_shopping", '("Instagram Shopping" OR "Instagram Shop" OR "Instagram checkout" OR "shopping su Instagram" OR "acquisti su Instagram" OR "negozi Instagram")'),
            ("ig_social_commerce", '("Instagram social commerce" OR "social commerce Italia" OR "creator economy" OR "creator marketplace" OR "creator economy Italia")'),
            ("ig_reels_shopping", '("reels shopping" OR "shopping nei Reels" OR "shoppable posts" OR "post acquistabili" OR "shopping nei post")'),
            ("ig_meta_tools", '("Meta shopping tools" OR "Meta commerce" OR "Meta ads" OR "Meta per aziende" OR "brand partnership" OR "partnership con brand")'),
            ("ig_creator", '("Instagram creator shop" OR "creator shop" OR "brand partnership" OR "creator economy Italia" OR "strumenti per creator")'),
            ("ig_trend", '("Instagram" OR Meta) ("social commerce" OR trend OR report OR adoption OR "consumer behavior" OR tendenza OR rapporto OR consumatori OR "comportamento dei consumatori")'),
        ]
    return []


def broad_eu_global_platform_blocks(
    display_label: str,
    source_platform: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[tuple[str, str]]:
    normalized_label = clean_text(display_label)
    if normalized_label == "TEMU":
        return [
            ("temu_broad_eu_regulation", '(EU OR Europe OR "European Commission" OR "European Union" OR DSA OR "Digital Services Act" OR "Commissione europea" OR "Unione Europea") (regulation OR compliance OR investigation OR "consumer protection" OR "product safety" OR privacy OR data OR customs OR "low value parcel" OR "online marketplace")'),
            ("temu_broad_platform_risk", '(Europe OR European OR EU OR "Unione Europea" OR Europa) (Temu OR TEMU) (counterfeit OR "illegal products" OR safety OR privacy OR "consumer rights" OR dogane OR contraffazione OR sicurezza OR privacy)'),
            ("temu_broad_market_trend", '(Europe OR European OR EU OR Europa) (Temu OR TEMU) (growth OR market OR "market share" OR adoption OR consumers OR ecommerce OR marketplace OR crescita OR mercato OR consumatori)'),
        ]
    if normalized_label == "IG" or platform_alias_concept(source_platform or normalized_label) == "instagram_shopping":
        return [
            ("ig_broad_meta_commerce", '(Meta OR Instagram) ("shopping tools" OR commerce OR ecommerce OR "social commerce" OR checkout OR "business tools" OR "strumenti business" OR "Meta commerce")'),
            ("ig_broad_creator_commerce", '(Meta OR Instagram) ("creator economy" OR "creator marketplace" OR "brand partnership" OR "shoppable posts" OR "reels shopping" OR "shopping nei Reels" OR "post acquistabili")'),
            ("ig_broad_ads_conversion", '(Meta OR Instagram) (ads OR advertising OR merchants OR retailers OR brands OR conversion OR "business suite" OR aziende OR commercianti OR pubblicità) (commerce OR shopping OR ecommerce OR vendita OR vendite)'),
        ]
    if normalized_label in {"TTS", "TikTok Shop"} or platform_alias_concept(source_platform or normalized_label) == "tiktok_shop":
        return [
            ("tts_broad_eu_commerce", '("TikTok Shop" OR TikTok) (EU OR Europe OR European OR Europa) ("social commerce" OR "live shopping" OR marketplace OR ecommerce OR regulation OR compliance OR "consumer protection")'),
            ("tts_broad_creator_commerce", '("TikTok Shop" OR TikTok) ("creator economy" OR creators OR "live commerce" OR "brand partnership" OR sellers OR merchants OR venditori OR commercianti)'),
        ]
    return []


def broad_short_platform_blocks(
    display_label: str,
    source_platform: str | None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[tuple[str, str]]:
    normalized_label = clean_text(display_label)
    if normalized_label == "TEMU":
        return [
            ("temu_short_european_commission", '"European Commission"'),
            ("temu_short_product_safety", '"product safety"'),
            ("temu_short_consumer_protection", '"consumer protection"'),
            ("temu_short_customs", 'customs'),
            ("temu_short_privacy_data", '(privacy OR data OR GDPR)'),
            ("temu_short_digital_services_act", '("Digital Services Act" OR DSA)'),
            ("temu_short_illegal_products", '("illegal products" OR counterfeit OR contraffazione)'),
            ("temu_short_marketplace_regulation", '("marketplace regulation" OR compliance OR regulation)'),
        ]
    if normalized_label == "IG" or platform_alias_concept(source_platform or normalized_label) == "instagram_shopping":
        return [
            ("ig_short_instagram_shopping", '"Instagram shopping"'),
            ("ig_short_instagram_shopping_italy", '"Instagram shopping Italy"'),
            ("ig_short_instagram_checkout", '"Instagram checkout"'),
            ("ig_short_creator_marketplace", '"Instagram creator marketplace"'),
            ("ig_short_meta_commerce", '"Meta commerce"'),
            ("ig_short_meta_shopping_ads", '"Meta shopping ads"'),
            ("ig_short_instagram_ads_commerce", '"Instagram ads commerce"'),
            ("ig_short_social_commerce_instagram", '"social commerce Instagram"'),
            ("ig_short_shoppable_posts", '"Instagram shoppable posts"'),
            ("ig_short_business_tools", '"Instagram business tools"'),
        ]
    if normalized_label in {"TTS", "TikTok Shop"} or platform_alias_concept(source_platform or normalized_label) == "tiktok_shop":
        return [
            ("tts_short_tiktok_shop_italy", '"TikTok Shop Italy"'),
            ("tts_short_tiktok_shop_europe", '"TikTok Shop Europe"'),
            ("tts_short_tiktok_shop_sellers", '"TikTok Shop sellers"'),
            ("tts_short_tiktok_shop_live_shopping", '"TikTok Shop live shopping"'),
            ("tts_short_tiktok_shop_regulation", '"TikTok Shop regulation"'),
            ("tts_short_tiktok_shop_consumer_protection", '"TikTok Shop consumer protection"'),
        ]
    return []


def build_dedicated_google_search_query(
    display_label: str,
    source_platform: str | None,
    keyword_block: str,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
    include_market_block: bool = True,
) -> str:
    if clean_text(display_label) == "IG" or platform_alias_concept(source_platform or display_label) == "instagram_shopping":
        query_parts = [keyword_block]
        if include_market_block:
            query_parts.append(country_market_search_block(country_code))
        return " ".join(part for part in query_parts if part)
    brand_block = build_brand_only_search_query(display_label, source_platform, country_code)
    query_parts = [brand_block, keyword_block]
    if include_market_block:
        query_parts.append(country_market_search_block(country_code))
    return " ".join(part for part in query_parts if part)


def build_report_search_query(
    requested_platforms: list[str] | None,
    dimension: str,
    dimension_keyword_lookup: dict[str, list[str]],
    keyword_block: str | None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    brand_scope_block = build_selected_brand_scope_block(requested_platforms, country_code)
    query_parts = [brand_scope_block, keyword_block, country_market_search_block(country_code)]
    return " ".join(part for part in query_parts if part)


def build_promo_search_feed_url(engine: str, query: str, country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    country_config = get_country_config(country_code)
    encoded_query = urllib.parse.quote(query, safe="")
    if engine == "google":
        return (
            f"https://news.google.com/rss/search?q={encoded_query}"
            f"&hl={country_config['google_news_hl']}"
            f"&gl={country_config['google_news_gl']}"
            f"&ceid={country_config['google_news_ceid']}"
        )
    return f"https://www.bing.com/news/search?q={encoded_query}&format=rss&mkt={country_config['bing_news_market']}"


def build_promo_search_page_url(engine: str, query: str, country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    country_config = get_country_config(country_code)
    encoded_query = urllib.parse.quote(query, safe="")
    if engine == "google":
        return (
            f"https://news.google.com/search?q={encoded_query}"
            f"&hl={country_config['google_news_hl']}"
            f"&gl={country_config['google_news_gl']}"
            f"&ceid={country_config['google_news_ceid']}"
        )
    return f"https://www.bing.com/news/search?q={encoded_query}&mkt={country_config['bing_news_market']}"


def article_matches_requested_brand(
    article: dict[str, Any],
    display_label: str,
    source_platform: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> bool:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )
    if not combined_text:
        return False
    if source_platform:
        matched = find_matched_platforms(combined_text, [source_platform], country_code)
        return display_label in matched
    return any(alias_matches(combined_text, alias) for alias in expanded_platform_search_terms(display_label, country_code))


def article_matches_instagram_commerce_context(
    article: dict[str, Any],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> bool:
    return instagram_commerce_context_strength(article, country_code) in {"strong", "weak"}


def instagram_commerce_context_strength(
    article: dict[str, Any],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> str:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )
    if not combined_text:
        return ""
    if any(alias_matches(combined_text, term) for term in instagram_strong_commerce_terms(country_code)):
        return "strong"
    if any(alias_matches(combined_text, term) for term in instagram_weak_commerce_terms(country_code)):
        return "weak"
    return ""


def instagram_non_commerce_skip_reason(article: dict[str, Any], country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    ).lower()
    if not combined_text:
        return "instagram_commerce_context_missing"
    entertainment_terms = [
        "celebrity", "singer", "actor", "actress", "concert", "music", "movie", "film",
        "reality", "tv show", "serie tv", "musica", "concerto", "attore", "attrice",
        "cantante", "vip", "influencer gossip", "gossip",
    ]
    account_terms = [
        "post", "posts", "photo", "photos", "story", "stories", "account", "profile",
        "followers", "like", "likes", "comment", "comments", "reel", "reels",
        "foto", "storie", "profilo", "follower", "seguaci", "commenti",
    ]
    if any(term in combined_text for term in entertainment_terms):
        return "instagram_entertainment_noise"
    if any(term in combined_text for term in account_terms) and not any(
        alias_matches(combined_text, term)
        for term in unique_clean_texts(instagram_strong_commerce_terms(country_code) + instagram_weak_commerce_terms(country_code))
    ):
        return "instagram_account_or_influencer_noise"
    return "instagram_commerce_context_missing"


def matched_platform_labels_for_task(
    article: dict[str, Any],
    task: dict[str, Any],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[str]:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )
    if not combined_text:
        return []

    matched_labels: list[str] = []
    source_platforms = [clean_text(item) for item in task.get("match_source_platforms") or [] if clean_text(item)]
    if source_platforms:
        for source_platform in source_platforms:
            display_label = display_platform_label(source_platform, country_code)
            matched = find_matched_platforms(combined_text, [source_platform], country_code)
            if display_label in matched and display_label not in matched_labels:
                matched_labels.append(display_label)

    for label in [clean_text(item) for item in task.get("match_platform_labels") or [] if clean_text(item)]:
        aliases = expanded_platform_search_terms(label, country_code)
        if any(alias_matches(combined_text, alias) for alias in aliases) and label not in matched_labels:
            matched_labels.append(label)

    platform_label = clean_text(task.get("platform_label"))
    if (
        is_instagram_platform_label(platform_label, task.get("source_platform"))
        and article_matches_instagram_broad_context(article, country_code)
        and platform_label
        and platform_label not in matched_labels
    ):
        matched_labels.append(platform_label)

    return matched_labels


def article_looks_like_promo(article: dict[str, Any]) -> bool:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    ).lower()
    if not combined_text:
        return False
    return any(alias_matches(combined_text, term.lower()) for term in PROMO_SIGNAL_TERMS)


def article_looks_like_report(article: dict[str, Any]) -> bool:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    ).lower()
    if not combined_text:
        return False
    return any(alias_matches(combined_text, term.lower()) for term in REPORT_SIGNAL_TERMS)


def build_promo_search_tasks(
    requested_platforms: list[str] | None,
    engine_mode: str,
    keyword_blocks_text: str | None = None,
    *,
    survey_indicators: list[SurveyIndicator] | None = None,
    related_news_enabled: bool = True,
    related_news_keywords_text: str | None = None,
    report_ranking_enabled: bool = True,
    report_keywords_text: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
) -> list[dict[str, Any]]:
    selected_display_platforms, selected_source_platforms, custom_platforms = resolve_requested_platforms(requested_platforms, country_code)
    grouped_indicators = group_survey_indicators_by_dimension(survey_indicators or [])
    dimension_keyword_lookup = build_dimension_keyword_lookup(survey_indicators or [])
    dimensions = list(grouped_indicators) or list_survey_dimensions(survey_indicators or [])
    tasks: list[dict[str, Any]] = []

    if related_news_enabled:
        keyword_blocks = normalize_related_news_search_keyword_blocks(
            related_news_keywords_text if related_news_keywords_text is not None else keyword_blocks_text,
            country_code,
            recall_mode,
        )
        for source_platform in selected_source_platforms:
            display_label = display_platform_label(source_platform, country_code)
            dedicated_blocks = dedicated_google_platform_blocks(display_label, source_platform, country_code)
            use_dedicated_google = bool(dedicated_blocks) and clean_text(engine_mode).lower() != "bing"
            if use_dedicated_google:
                for query_label, keyword_block in dedicated_blocks:
                    tasks.append(
                        {
                            "task_type": "related_news",
                            "engine": "google",
                            "platform_label": display_label,
                            "source_platform": source_platform,
                            "query_label": query_label,
                            "query": build_dedicated_google_search_query(
                                display_label,
                                source_platform,
                                keyword_block,
                                country_code=country_code,
                            ),
                            "focus_dimension": "",
                            "match_platform_labels": [display_label],
                            "match_source_platforms": [source_platform],
                            "query_strategy": "platform_dedicated_google",
                            "market_scope": "target_market",
                            "broad_entry_allowed": clean_text(recall_mode).lower() == "balanced",
                        }
                    )
                if clean_text(recall_mode).lower() == "balanced":
                    for query_label, keyword_block in broad_eu_global_platform_blocks(display_label, source_platform, country_code):
                        tasks.append(
                            {
                                "task_type": "related_news",
                                "engine": "google",
                                "platform_label": display_label,
                                "source_platform": source_platform,
                                "query_label": query_label,
                                "query": build_dedicated_google_search_query(
                                    display_label,
                                    source_platform,
                                    keyword_block,
                                    country_code=country_code,
                                    include_market_block=False,
                                ),
                                "focus_dimension": "",
                                "match_platform_labels": [display_label],
                                "match_source_platforms": [source_platform],
                                "query_strategy": "platform_broad_eu_global",
                                "market_scope": BROAD_ENTRY_MARKET_SCOPE,
                            }
                        )
                    for query_label, keyword_block in broad_short_platform_blocks(display_label, source_platform, country_code):
                        tasks.append(
                            {
                                "task_type": "related_news",
                                "engine": "google",
                                "platform_label": display_label,
                                "source_platform": source_platform,
                                "query_label": query_label,
                                "query": build_dedicated_google_search_query(
                                    display_label,
                                    source_platform,
                                    keyword_block,
                                    country_code=country_code,
                                    include_market_block=False,
                                ),
                                "focus_dimension": "",
                                "match_platform_labels": [display_label],
                                "match_source_platforms": [source_platform],
                                "query_strategy": "platform_broad_short",
                                "market_scope": BROAD_ENTRY_MARKET_SCOPE,
                            }
                        )
                continue
            for engine in select_promo_search_engines(engine_mode):
                for query_index, (query_label, keyword_block) in enumerate(keyword_blocks, start=1):
                    tasks.append(
                        {
                            "task_type": "related_news",
                            "engine": engine,
                            "platform_label": display_label,
                            "source_platform": source_platform,
                            "query_label": f"{query_label or 'related_news'}_{query_index}",
                            "query": build_related_news_search_query(
                                display_label,
                                source_platform,
                                "",
                                dimension_keyword_lookup,
                                keyword_block,
                                country_code=country_code,
                            ),
                            "focus_dimension": "",
                            "match_platform_labels": [display_label],
                            "match_source_platforms": [source_platform],
                            "query_strategy": "generic_keyword_block",
                        }
                    )
        for platform in custom_platforms:
            dedicated_blocks = dedicated_google_platform_blocks(platform, None, country_code)
            use_dedicated_google = bool(dedicated_blocks) and clean_text(engine_mode).lower() != "bing"
            if use_dedicated_google:
                for query_label, keyword_block in dedicated_blocks:
                    tasks.append(
                        {
                            "task_type": "related_news",
                            "engine": "google",
                            "platform_label": platform,
                            "source_platform": "",
                            "query_label": query_label,
                            "query": build_dedicated_google_search_query(
                                platform,
                                None,
                                keyword_block,
                                country_code=country_code,
                            ),
                            "focus_dimension": "",
                            "match_platform_labels": [platform],
                            "match_source_platforms": [],
                            "query_strategy": "platform_dedicated_google",
                            "market_scope": "target_market",
                            "broad_entry_allowed": clean_text(recall_mode).lower() == "balanced",
                        }
                    )
                if clean_text(recall_mode).lower() == "balanced":
                    for query_label, keyword_block in broad_eu_global_platform_blocks(platform, None, country_code):
                        tasks.append(
                            {
                                "task_type": "related_news",
                                "engine": "google",
                                "platform_label": platform,
                                "source_platform": "",
                                "query_label": query_label,
                                "query": build_dedicated_google_search_query(
                                    platform,
                                    None,
                                    keyword_block,
                                    country_code=country_code,
                                    include_market_block=False,
                                ),
                                "focus_dimension": "",
                                "match_platform_labels": [platform],
                                "match_source_platforms": [],
                                "query_strategy": "platform_broad_eu_global",
                                "market_scope": BROAD_ENTRY_MARKET_SCOPE,
                            }
                        )
                    for query_label, keyword_block in broad_short_platform_blocks(platform, None, country_code):
                        tasks.append(
                            {
                                "task_type": "related_news",
                                "engine": "google",
                                "platform_label": platform,
                                "source_platform": "",
                                "query_label": query_label,
                                "query": build_dedicated_google_search_query(
                                    platform,
                                    None,
                                    keyword_block,
                                    country_code=country_code,
                                    include_market_block=False,
                                ),
                                "focus_dimension": "",
                                "match_platform_labels": [platform],
                                "match_source_platforms": [],
                                "query_strategy": "platform_broad_short",
                                "market_scope": BROAD_ENTRY_MARKET_SCOPE,
                            }
                        )
                continue
            for engine in select_promo_search_engines(engine_mode):
                for query_index, (query_label, keyword_block) in enumerate(keyword_blocks, start=1):
                    tasks.append(
                        {
                            "task_type": "related_news",
                            "engine": engine,
                            "platform_label": platform,
                            "source_platform": "",
                            "query_label": f"{query_label or 'related_news'}_{query_index}",
                            "query": build_related_news_search_query(
                                platform,
                                None,
                                "",
                                dimension_keyword_lookup,
                                keyword_block,
                                country_code=country_code,
                            ),
                            "focus_dimension": "",
                            "match_platform_labels": [platform],
                            "match_source_platforms": [],
                            "query_strategy": "generic_keyword_block",
                        }
                    )

    if report_ranking_enabled:
        report_keyword_blocks = normalize_report_search_keyword_blocks(report_keywords_text, country_code)
        brand_scope_labels = selected_display_platforms + custom_platforms
        for engine in select_promo_search_engines(engine_mode):
            for query_index, (query_label, keyword_block) in enumerate(report_keyword_blocks, start=1):
                tasks.append(
                    {
                        "task_type": "report_ranking",
                        "engine": engine,
                        "platform_label": "Cross-brand ranking",
                        "source_platform": "",
                        "query_label": f"{query_label or 'report_ranking'}_{query_index}",
                        "query": build_report_search_query(
                            requested_platforms,
                            "",
                            dimension_keyword_lookup,
                            keyword_block,
                            country_code=country_code,
                        ),
                        "focus_dimension": "",
                        "match_platform_labels": brand_scope_labels,
                        "match_source_platforms": selected_source_platforms,
                        "query_strategy": "report_ranking",
                    }
                )

    deduped_tasks: list[dict[str, Any]] = []
    seen_task_keys: set[tuple[str, str, str]] = set()
    for task in tasks:
        task_key = (
            str(task.get("task_type") or ""),
            str(task.get("engine") or ""),
            str(task.get("query") or ""),
        )
        if task_key in seen_task_keys:
            continue
        seen_task_keys.add(task_key)
        deduped_tasks.append(task)
    return deduped_tasks


def attach_search_engine_context(
    *,
    task_type: str,
    focus_dimension: str,
    query_label: str,
    query: str,
    platform_label: str,
    source_platform: str,
    engine: str,
    feed_url: str,
    search_page_url: str,
    articles: list[dict[str, Any]],
    query_strategy: str = "",
    matched_platform_labels: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    engine_site = urllib.parse.urlparse(search_page_url).netloc
    for article in articles:
        row = dict(article)
        matched_labels = [clean_text(item) for item in (matched_platform_labels or article.get("matched_brands") or []) if clean_text(item)]
        display_platform_label = dedupe_utils.format_brand_labels(matched_labels) if matched_labels else platform_label
        if task_type == "report_ranking":
            row["platform"] = display_platform_label if matched_labels else "Cross-brand report"
        else:
            row["platform"] = display_platform_label
        row["platform_label"] = display_platform_label
        row["source_platform"] = source_platform or platform_label
        row["side"] = "search_engine"
        row["source_url"] = search_page_url
        row["source_final_url"] = search_page_url
        row["source_feed_url"] = feed_url
        row["source_site"] = engine_site
        row["source_country_guess"] = guess_country(article["article_url"])
        row["matched_brands"] = matched_labels or [platform_label]
        row["search_task_type"] = task_type
        row["search_focus_dimension"] = focus_dimension
        row["search_query_label"] = query_label
        row["search_query_strategy"] = query_strategy
        row["search_query"] = query
        row["country_code"] = country_code
        enriched.append(row)
    return enriched


def collect_search_engine_promo_articles(
    requested_platforms: list[str] | None,
    *,
    engine_mode: str,
    keyword_blocks_text: str | None = None,
    survey_indicators: list[SurveyIndicator] | None = None,
    related_news_enabled: bool = True,
    related_news_keywords_text: str | None = None,
    report_ranking_enabled: bool = True,
    report_keywords_text: str | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
    start: datetime,
    end: datetime,
    progress_callback=None,
    total_sites: int = 0,
    completed_sites: int = 0,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    tasks = build_promo_search_tasks(
        requested_platforms,
        engine_mode,
        keyword_blocks_text,
        survey_indicators=survey_indicators,
        related_news_enabled=related_news_enabled,
        related_news_keywords_text=related_news_keywords_text,
        report_ranking_enabled=report_ranking_enabled,
        report_keywords_text=report_keywords_text,
        country_code=country_code,
        recall_mode=recall_mode,
    )
    if not tasks:
        return [], {"task_count": 0, "article_count": 0, "related_news_task_count": 0, "report_task_count": 0, "related_news_article_count": 0, "report_article_count": 0, "task_diagnostics": []}

    articles: list[dict[str, Any]] = []
    task_diagnostics: list[dict[str, Any]] = []
    completed_tasks = 0
    worker_count = max(1, min(6, len(tasks)))

    def worker(task: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
        session = build_session()
        feed_url = build_promo_search_feed_url(task["engine"], task["query"], country_code)
        search_page_url = build_promo_search_page_url(task["engine"], task["query"], country_code)
        feed_rows = parse_feed_articles(feed_url, session, start, end)
        kept_rows: list[dict[str, Any]] = []
        seen_urls: set[str] = set()
        skipped_reasons: Counter = Counter()
        allowed_reasons: Counter = Counter()
        for feed_row in feed_rows:
            article_url = clean_text(feed_row.get("article_url"))
            if not article_url or article_url in seen_urls:
                skipped_reasons["duplicate_or_empty_url"] += 1
                continue
            seen_urls.add(article_url)
            published_at = parse_dt(feed_row.get("published_at"))
            feed_title = clean_text(feed_row.get("title") or "")
            feed_summary = html_to_text(feed_row.get("summary") or "") or None
            source_discovery = f"{task['engine']}_promo_search"
            is_engine_redirect = matches_domain(article_url, "news.google.com", "bing.com", "www.bing.com")

            meta = None
            if not is_engine_redirect:
                meta = hydrate_search_result_article(
                    article_url,
                    feed_title,
                    published_at,
                    session,
                    start,
                    end,
                    summary=feed_summary,
                    category=feed_row.get("category"),
                    tags=feed_row.get("tags") or [],
                    source_discovery=source_discovery,
                    allow_search_only=True,
                )

            if not meta and published_at is not None and dt_in_range(published_at, start, end):
                meta = build_search_only_article(
                    article_url,
                    feed_title,
                    published_at,
                    summary=feed_summary,
                    category=feed_row.get("category"),
                    tags=feed_row.get("tags") or [],
                    source_discovery=f"{source_discovery}_search_only",
                )
            if not meta:
                skipped_reasons["no_metadata_or_out_of_range"] += 1
                continue
            matched_labels = matched_platform_labels_for_task(meta, task, country_code)
            broad_entry_reason = ""
            market_scope = clean_text(task.get("market_scope") or "")
            is_broad_entry_task = clean_text(task.get("query_strategy")) in {"platform_broad_eu_global", "platform_broad_short"}
            if is_broad_entry_task:
                broad_entry_reason = broad_entry_reason_for_article(meta, task, country_code)
            elif task.get("broad_entry_allowed") and is_instagram_platform_label(task.get("platform_label"), task.get("source_platform")):
                broad_entry_reason = broad_entry_reason_for_article(meta, task, country_code)
            if task.get("task_type") == "report_ranking":
                if not matched_labels or not article_looks_like_report(meta):
                    skipped_reasons["report_not_matched"] += 1
                    continue
            else:
                if task["platform_label"] not in matched_labels:
                    if broad_entry_reason:
                        matched_labels = [clean_text(task["platform_label"])]
                        allowed_reasons["broad_entry_platform_context_allowed"] += 1
                    else:
                        skipped_reasons["platform_not_matched"] += 1
                        continue
                if is_instagram_platform_label(task.get("platform_label"), task.get("source_platform")):
                    instagram_context = instagram_commerce_context_strength(meta, country_code)
                    if not instagram_context and broad_entry_reason:
                        instagram_context = "weak"
                    if not instagram_context:
                        skipped_reasons[instagram_non_commerce_skip_reason(meta, country_code)] += 1
                        continue
                    if instagram_context == "weak":
                        allowed_reasons["instagram_weak_commerce_allowed"] += 1
                if is_broad_entry_task and not broad_entry_reason:
                    skipped_reasons["broad_entry_context_not_matched"] += 1
                    continue
            if broad_entry_reason:
                meta["broad_entry"] = "true"
                meta["broad_entry_reason"] = broad_entry_reason
                meta["market_scope"] = market_scope or BROAD_ENTRY_MARKET_SCOPE
            meta["matched_brands"] = matched_labels
            meta["country_code"] = country_code
            kept_rows.append(meta)
        enriched_rows = attach_search_engine_context(
            task_type=str(task.get("task_type") or ""),
            focus_dimension=str(task.get("focus_dimension") or ""),
            query_label=str(task.get("query_label") or ""),
            query=str(task.get("query") or ""),
            platform_label=task["platform_label"],
            source_platform=task["source_platform"],
            engine=task["engine"],
            feed_url=feed_url,
            search_page_url=search_page_url,
            query_strategy=str(task.get("query_strategy") or ""),
            articles=kept_rows,
            country_code=country_code,
        )
        diagnostic = {
            "task_type": str(task.get("task_type") or ""),
            "engine": str(task.get("engine") or ""),
            "platform_label": str(task.get("platform_label") or ""),
            "source_platform": str(task.get("source_platform") or ""),
            "query_label": str(task.get("query_label") or ""),
            "query_strategy": str(task.get("query_strategy") or ""),
            "market_scope": str(task.get("market_scope") or ""),
            "query": str(task.get("query") or ""),
            "feed_count": len(feed_rows),
            "kept_count": len(enriched_rows),
            "broad_entry_count": sum(1 for row in enriched_rows if clean_text(row.get("broad_entry")).lower() == "true"),
            "skipped_reason_count": dict(skipped_reasons),
            "allowed_reason_count": dict(allowed_reasons),
        }
        return task, enriched_rows, diagnostic

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = [executor.submit(worker, task) for task in tasks]
        for future in as_completed(futures):
            task, task_articles, diagnostic = future.result()
            articles.extend(task_articles)
            task_diagnostics.append(diagnostic)
            completed_tasks += 1
            remaining_tasks = max(0, len(tasks) - completed_tasks)
            active_searches = min(worker_count, remaining_tasks)
            progress_percent = interpolate_progress(
                NEWS_PROGRESS_POST_CRAWL,
                NEWS_PROGRESS_PROMO_END,
                completed_tasks,
                len(tasks),
            )
            emit_progress(
                progress_callback,
                stage="promo_search",
                total_sites=total_sites,
                completed_sites=completed_sites,
                active_sites=active_searches,
                current_site=f"补充检索（已完成 {completed_tasks}/{len(tasks)} 组，正在搜索 {active_searches} 组）",
                last_completed_site=f"{task['engine']} · {task.get('task_type', 'search')} · {task.get('focus_dimension', '')}",
                message=f"正在通过 {task['engine']} 执行相关新闻/数据报告补充检索；已完成 {completed_tasks}/{len(tasks)} 组检索",
                progress_percent=min(NEWS_PROGRESS_PROMO_END, progress_percent),
            )

    related_news_tasks = sum(1 for task in tasks if task.get("task_type") == "related_news")
    report_tasks = sum(1 for task in tasks if task.get("task_type") == "report_ranking")
    related_news_articles = sum(1 for row in articles if clean_text(row.get("search_task_type")) == "related_news")
    report_articles = sum(1 for row in articles if clean_text(row.get("search_task_type")) == "report_ranking")
    platform_engine_summary: dict[str, dict[str, int]] = {}
    platform_query_summary: dict[str, dict[str, Any]] = {}
    for diagnostic in task_diagnostics:
        platform_key = f"{diagnostic.get('platform_label') or ''}|{diagnostic.get('engine') or ''}"
        bucket = platform_engine_summary.setdefault(
            platform_key,
            {"task_count": 0, "feed_count": 0, "kept_count": 0},
        )
        bucket["task_count"] += 1
        bucket["feed_count"] += int(diagnostic.get("feed_count") or 0)
        bucket["kept_count"] += int(diagnostic.get("kept_count") or 0)
        query_key = (
            f"{diagnostic.get('platform_label') or ''}|"
            f"{diagnostic.get('query_strategy') or ''}|"
            f"{diagnostic.get('query_label') or ''}"
        )
        query_bucket = platform_query_summary.setdefault(
            query_key,
            {
                "platform_label": diagnostic.get("platform_label") or "",
                "query_strategy": diagnostic.get("query_strategy") or "",
                "query_label": diagnostic.get("query_label") or "",
                "market_scope": diagnostic.get("market_scope") or "",
                "task_count": 0,
                "feed_count": 0,
                "kept_count": 0,
                "broad_entry_count": 0,
                "skipped_count": 0,
                "skipped_reason_count": {},
                "allowed_reason_count": {},
            },
        )
        query_bucket["task_count"] += 1
        query_bucket["feed_count"] += int(diagnostic.get("feed_count") or 0)
        query_bucket["kept_count"] += int(diagnostic.get("kept_count") or 0)
        query_bucket["broad_entry_count"] += int(diagnostic.get("broad_entry_count") or 0)
        skipped_reason_count = diagnostic.get("skipped_reason_count") or {}
        if isinstance(skipped_reason_count, dict):
            query_bucket["skipped_count"] += sum(int(value or 0) for value in skipped_reason_count.values())
            merged_reasons = Counter(query_bucket.get("skipped_reason_count") or {})
            merged_reasons.update({str(key): int(value or 0) for key, value in skipped_reason_count.items()})
            query_bucket["skipped_reason_count"] = dict(merged_reasons)
        allowed_reason_count = diagnostic.get("allowed_reason_count") or {}
        if isinstance(allowed_reason_count, dict):
            merged_allowed_reasons = Counter(query_bucket.get("allowed_reason_count") or {})
            merged_allowed_reasons.update({str(key): int(value or 0) for key, value in allowed_reason_count.items()})
            query_bucket["allowed_reason_count"] = dict(merged_allowed_reasons)
    return articles, {
        "task_count": len(tasks),
        "article_count": len(articles),
        "related_news_task_count": related_news_tasks,
        "report_task_count": report_tasks,
        "related_news_article_count": related_news_articles,
        "report_article_count": report_articles,
        "platform_engine_summary": platform_engine_summary,
        "platform_query_summary": platform_query_summary,
        "task_diagnostics": sorted(
            task_diagnostics,
            key=lambda item: (str(item.get("task_type") or ""), str(item.get("platform_label") or ""), str(item.get("engine") or ""), str(item.get("query_label") or "")),
        ),
    }


def dedupe_media_sources(entries: list[SourceEntry]) -> list[SourceEntry]:
    deduped: list[SourceEntry] = []
    seen_urls: set[str] = set()
    for entry in entries:
        canonical_url = canonicalize_source_url(entry.source_url)
        if canonical_url in seen_urls:
            continue
        seen_urls.add(canonical_url)
        deduped.append(SourceEntry(platform=entry.platform, side=entry.side, source_url=canonical_url))
    return deduped


def select_sources_for_run(
    entries: list[SourceEntry],
    requested_platforms: list[str] | None,
    allowed_sides: list[str],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[SourceEntry]:
    allowed = set(allowed_sides)
    requested_present = any(clean_text(item) for item in (requested_platforms or []))
    _, selected_source_platforms, _ = resolve_requested_platforms(requested_platforms, country_code)
    selected_platform_set = set(selected_source_platforms)

    selected_entries: list[SourceEntry] = []
    if "media" in allowed:
        media_entries = [entry for entry in entries if entry.side == "media"]
        selected_entries.extend(dedupe_media_sources(media_entries))

    for entry in entries:
        if entry.side not in allowed or entry.side == "media":
            continue
        if requested_present and entry.platform not in selected_platform_set:
            continue
        selected_entries.append(entry)

    return selected_entries


def source_site(url: str) -> str:
    return urllib.parse.urlparse(url).netloc


def alias_matches(text: str, alias: str) -> bool:
    normalized_text = clean_text(text)
    if not normalized_text:
        return False
    if alias.isascii():
        pattern = re.compile(rf"(?<![A-Za-z0-9]){re.escape(alias)}(?![A-Za-z0-9])", re.IGNORECASE)
        return bool(pattern.search(normalized_text))
    return alias in normalized_text


def platform_match_aliases(platform: str, country_code: str = DEFAULT_COUNTRY_CODE) -> list[str]:
    aliases = CANONICAL_PLATFORM_ALIASES.get(platform, [])
    ordered: list[str] = []
    alias_exclude_tokens = [clean_text(str(token)).lower() for token in country_list_setting(country_code, "platform_alias_exclude_tokens")]
    country_aliases = aliases
    if alias_exclude_tokens:
        country_aliases = [
            alias
            for alias in aliases
            if not any(token in clean_text(alias).lower() for token in alias_exclude_tokens)
        ]
    seed_items = [display_platform_label(platform, country_code)]
    normalized_platform = clean_text(platform).lower()
    should_include_platform = '/' not in clean_text(platform)
    if should_include_platform and (
        not alias_exclude_tokens
        or not any(token in normalized_platform for token in alias_exclude_tokens)
    ):
        seed_items.insert(0, platform)
    for item in seed_items + platform_search_terms(platform, country_code) + country_aliases:
        normalized = clean_text(item)
        if normalized and normalized not in ordered:
            ordered.append(normalized)
    return ordered


def find_matched_platforms(
    text: str,
    candidates: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[str]:
    matched: list[str] = []
    candidate_platforms = candidates or list(CANONICAL_PLATFORM_ALIASES.keys())
    for platform in candidate_platforms:
        aliases = platform_match_aliases(platform, country_code)
        if any(alias_matches(text, alias) for alias in aliases):
            display_label = display_platform_label(platform, country_code)
            if display_label not in matched:
                matched.append(display_label)
    return matched


def resolve_platform_for_article(entry: SourceEntry, matched_platforms: list[str], country_code: str = DEFAULT_COUNTRY_CODE) -> str:
    if entry.side != "media":
        return display_platform_label(entry.platform, country_code)
    unique = []
    for platform in matched_platforms:
        if platform not in unique:
            unique.append(platform)
    if not unique:
        return "Unassigned Media Match"
    return ", ".join(unique)


def article_has_media_platform_context(article: dict[str, Any]) -> bool:
    combined_text = clean_text(
        f"{article.get('title') or ''} {article.get('summary') or ''} "
        f"{article.get('category') or ''} {' '.join(article.get('tags') or [])} "
        f"{article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )
    if not combined_text:
        return False
    if contains_any_relevance_keyword(combined_text, NOISY_TOPIC_TERMS) and not contains_any_relevance_keyword(combined_text, MEDIA_PLATFORM_CONTEXT_TERMS):
        return False
    return contains_any_relevance_keyword(combined_text, MEDIA_PLATFORM_CONTEXT_TERMS)


def should_keep_article(
    entry: SourceEntry,
    article: dict[str, Any],
    media_targets: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> tuple[bool, list[str]]:
    title_text = clean_text(article.get("title"))
    summary_text = clean_text(article.get("summary"))
    combined_text = clean_text(
        f"{title_text} {summary_text} {article.get('category') or ''} "
        f"{' '.join(article.get('tags') or [])} {article.get('body_excerpt') or ''} {article.get('article_url') or ''}"
    )
    matched = find_matched_platforms(combined_text, media_targets if entry.side == "media" else None, country_code)
    if entry.side == "media":
        if not matched:
            return False, []
        if not article_has_media_platform_context(article):
            return False, []
        return True, matched

    official_brands = [display_platform_label(entry.platform, country_code)]
    if matched:
        return True, matched
    return True, official_brands


def dt_in_range(dt: datetime | None, start: datetime, end: datetime) -> bool:
    if dt is None:
        return False
    return start <= dt.astimezone(start.tzinfo) <= end


def source_domain(url: str) -> str:
    return urllib.parse.urlparse(url).netloc.lower()


def matches_domain(url: str, *domains: str) -> bool:
    domain = source_domain(url)
    return any(domain == item or domain.endswith(f".{item}") for item in domains)


def discover_feeds(soup: BeautifulSoup, base_url: str) -> list[str]:
    feeds: list[str] = []
    seen: set[str] = set()

    for link in soup.find_all("link", href=True):
        link_type = clean_text(link.get("type", "")).lower()
        rels = {rel.lower() for rel in link.get("rel", [])}
        if "alternate" in rels and link_type in FEED_TYPES:
            href = abs_url(base_url, link["href"])
            if href not in seen:
                seen.add(href)
                feeds.append(href)

    for anchor in soup.find_all("a", href=True):
        href = abs_url(base_url, anchor["href"])
        lowered = href.lower()
        if any(token in lowered for token in ["/feed", "/rss", "atom"]) and href not in seen:
            seen.add(href)
            feeds.append(href)

    parsed = urllib.parse.urlparse(base_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    for suffix in COMMON_FEED_PATHS:
        href = origin + suffix
        if href not in seen:
            seen.add(href)
            feeds.append(href)

    return feeds[:6]


def parse_feed_articles(feed_url: str, session, start: datetime, end: datetime) -> list[dict[str, Any]]:
    try:
        response = fetch(session, feed_url)
    except Exception:
        return []

    try:
        root = ET.fromstring(response.text)
    except ET.ParseError:
        return []

    articles: list[dict[str, Any]] = []
    namespaces = {
        "atom": "http://www.w3.org/2005/Atom",
        "content": "http://purl.org/rss/1.0/modules/content/",
    }

    if root.tag.endswith("rss"):
        channel = root.find("channel")
        if channel is None:
            return []
        for item in channel.findall("item"):
            title = clean_text(item.findtext("title"))
            link = clean_text(item.findtext("link"))
            date_text = item.findtext("pubDate") or item.findtext("date")
            published_at = parse_dt(date_text)
            if not link or not dt_in_range(published_at, start, end):
                continue
            summary = clean_text(item.findtext("description") or "")
            categories = [clean_text(node.text or "") for node in item.findall("category")]
            articles.append(
                {
                    "article_url": link,
                    "title": title,
                    "published_at": published_at.isoformat(),
                    "summary": summary or None,
                    "category": " / ".join(item for item in categories if item) or None,
                    "tags": [item for item in categories if item],
                    "source_discovery": "feed",
                }
            )
    else:
        for entry in root.findall("atom:entry", namespaces):
            title = clean_text(entry.findtext("atom:title", default="", namespaces=namespaces))
            link_node = entry.find("atom:link", namespaces)
            href = link_node.get("href") if link_node is not None else ""
            date_text = (
                entry.findtext("atom:published", default="", namespaces=namespaces)
                or entry.findtext("atom:updated", default="", namespaces=namespaces)
            )
            published_at = parse_dt(date_text)
            if not href or not dt_in_range(published_at, start, end):
                continue
            summary = clean_text(entry.findtext("atom:summary", default="", namespaces=namespaces))
            categories = [
                clean_text(node.get("term") or "")
                for node in entry.findall("atom:category", namespaces)
            ]
            articles.append(
                {
                    "article_url": href,
                    "title": title,
                    "published_at": published_at.isoformat(),
                    "summary": summary or None,
                    "category": " / ".join(item for item in categories if item) or None,
                    "tags": [item for item in categories if item],
                    "source_discovery": "feed",
                }
            )

    deduped: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for article in articles:
        if article["article_url"] in seen_urls:
            continue
        seen_urls.add(article["article_url"])
        deduped.append(article)
    return deduped


def score_candidate_link(base_url: str, href: str, text: str) -> int:
    score = 0
    parsed_base = urllib.parse.urlparse(base_url)
    parsed_href = urllib.parse.urlparse(href)
    if parsed_href.netloc == parsed_base.netloc:
        score += 3
    lowered_href = href.lower()
    lowered_text = text.lower()
    if any(token in lowered_href for token in ARTICLE_HINTS):
        score += 3
    if re.search(r"/20\d{2}/\d{2}/\d{2}", lowered_href) or re.search(r"\d{4}/\d{2}/\d{2}", lowered_href):
        score += 3
    if any(token in lowered_text for token in ["news", "article", "press", "release", "announcement", "business"]):
        score += 2
    if 16 <= len(text) <= 140:
        score += 1
    if any(token in lowered_href for token in EXCLUDE_HINTS):
        score -= 10
    return score


def discover_candidate_links(soup: BeautifulSoup, base_url: str, max_links: int) -> list[str]:
    candidates: list[tuple[int, str]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = abs_url(base_url, anchor["href"])
        text = clean_text(anchor.get_text(" ", strip=True))
        if not href.startswith("http") or href in seen:
            continue
        seen.add(href)
        score = score_candidate_link(base_url, href, text)
        if score > 0:
            candidates.append((score, href))
    candidates.sort(key=lambda item: item[0], reverse=True)
    ordered: list[str] = []
    for _, href in candidates:
        if href not in ordered:
            ordered.append(href)
        if len(ordered) >= max_links:
            break
    return ordered


def extract_json_ld_dates(soup: BeautifulSoup) -> list[str]:
    results: list[str] = []
    for script in soup.find_all("script", {"type": "application/ld+json"}):
        text = script.get_text(strip=True)
        if not text:
            continue
        try:
            payload = json.loads(text)
        except Exception:
            continue
        stack = [payload]
        while stack:
            obj = stack.pop()
            if isinstance(obj, dict):
                for key, value in obj.items():
                    if key in {"datePublished", "dateCreated", "uploadDate", "pubDate"} and isinstance(value, str):
                        results.append(value)
                    elif isinstance(value, (dict, list)):
                        stack.append(value)
            elif isinstance(obj, list):
                stack.extend(obj)
    return results


def extract_article_tags(soup: BeautifulSoup) -> list[str]:
    tags: list[str] = []
    for selector, attr in [
        ('meta[name="keywords"]', "content"),
        ('meta[property="article:tag"]', "content"),
    ]:
        for node in soup.select(selector):
            raw_value = clean_text(node.get(attr, ""))
            if not raw_value:
                continue
            for part in re.split(r"[,/|、，]\s*", raw_value):
                normalized = clean_text(part)
                if normalized and normalized not in tags:
                    tags.append(normalized)

    for node in soup.select('a[rel="tag"], .tag a, .tags a, [class*="tag"] a'):
        value = clean_text(node.get_text(" ", strip=True))
        if value and value not in tags:
            tags.append(value)
        if len(tags) >= 12:
            break
    return tags[:12]


def extract_article_category(soup: BeautifulSoup) -> str | None:
    for selector, attr in [
        ('meta[property="article:section"]', "content"),
        ('meta[name="article:section"]', "content"),
        ('meta[property="og:section"]', "content"),
    ]:
        node = soup.select_one(selector)
        if node and node.get(attr):
            value = clean_text(node.get(attr))
            if value:
                return value
    breadcrumb = soup.select_one('nav[aria-label*="breadcrumb"] li:last-child, .breadcrumb li:last-child, .breadcrumbs li:last-child')
    if breadcrumb:
        value = clean_text(breadcrumb.get_text(" ", strip=True))
        if value:
            return value
    return None


def extract_article_body_excerpt(soup: BeautifulSoup, max_chars: int = 500) -> str | None:
    containers = [
        soup.select_one("article"),
        soup.select_one("main"),
        soup.select_one('[class*="article"]'),
        soup.select_one('[class*="content"]'),
    ]
    text_parts: list[str] = []
    for container in containers:
        if container is None:
            continue
        for node in container.select("p, li"):
            text = trim_summary_text(node.get_text(" ", strip=True), max_chars=max_chars)
            if text and len(text) >= SUMMARY_MIN_PARAGRAPH_CHARS:
                text_parts.append(text)
            if len(" ".join(text_parts)) >= max_chars:
                break
        if text_parts:
            break
    return merge_summary_candidates(*text_parts, max_chars=max_chars)


def extract_article_metadata(
    article_url: str,
    session,
    start: datetime,
    end: datetime,
    *,
    published_at_hint: datetime | None = None,
    title_hint: str | None = None,
    source_discovery: str = "page_follow",
) -> dict[str, Any] | None:
    try:
        response = fetch(session, article_url)
    except Exception:
        return None

    soup = BeautifulSoup(response.text, "lxml")
    status, _ = detect_source_status(article_url, response.status_code, response.url, response.text[:50000])
    is_embedded_auth_public_article = (
        status == "login_required"
        and urllib.parse.urlparse(response.url).path.lower().startswith("/articles/detail/")
        and bool(soup.select_one('meta[property="og:title"]') or soup.select_one("h1"))
    )
    if status != "public" and not is_embedded_auth_public_article:
        return None

    published_candidates: list[str] = []

    for selector, attr in [
        ('meta[property="article:published_time"]', "content"),
        ('meta[name="article:published_time"]', "content"),
        ('meta[name="publish-date"]', "content"),
        ('meta[name="pubdate"]', "content"),
        ('meta[property="og:updated_time"]', "content"),
        ("time[datetime]", "datetime"),
    ]:
        node = soup.select_one(selector)
        if node and node.get(attr):
            published_candidates.append(node[attr])

    published_candidates.extend(extract_json_ld_dates(soup))

    published_at = None
    for candidate in published_candidates:
        published_at = parse_dt(candidate)
        if published_at is not None:
            break
    if published_at is None:
        published_at = published_at_hint

    if not dt_in_range(published_at, start, end):
        return None

    title = ""
    for selector, attr in [
        ('meta[property="og:title"]', "content"),
        ('meta[name="twitter:title"]', "content"),
    ]:
        node = soup.select_one(selector)
        if node and node.get(attr):
            title = clean_text(node[attr])
            break
    if not title:
        title_node = soup.select_one("h1") or soup.select_one("title")
        title = clean_text(title_node.get_text(" ", strip=True)) if title_node else ""
    if not title:
        title = clean_text(title_hint)
    if not title:
        return None

    body_excerpt = extract_article_body_excerpt(soup)
    return {
        "article_url": response.url,
        "title": title,
        "published_at": published_at.isoformat(),
        "summary": merge_summary_candidates(extract_meta_description(soup), body_excerpt),
        "category": extract_article_category(soup),
        "tags": extract_article_tags(soup),
        "body_excerpt": body_excerpt,
        "source_discovery": source_discovery,
    }


def build_search_only_article(
    article_url: str,
    title: str,
    published_at: datetime,
    *,
    summary: str | None = None,
    category: str | None = None,
    tags: list[str] | None = None,
    source_discovery: str = "site_search",
) -> dict[str, Any]:
    return {
        "article_url": article_url,
        "title": clean_text(title),
        "published_at": published_at.isoformat(),
        "summary": clean_text(summary) or None,
        "category": clean_text(category) or None,
        "tags": [clean_text(item) for item in (tags or []) if clean_text(item)],
        "source_discovery": source_discovery,
    }


def hydrate_search_result_article(
    article_url: str,
    title: str,
    published_at: datetime | None,
    session,
    start: datetime,
    end: datetime,
    *,
    summary: str | None = None,
    category: str | None = None,
    tags: list[str] | None = None,
    source_discovery: str = "site_search",
    allow_search_only: bool = False,
) -> dict[str, Any] | None:
    meta = extract_article_metadata(
        article_url,
        session,
        start,
        end,
        published_at_hint=published_at,
        title_hint=title,
        source_discovery=source_discovery,
    )
    if meta:
        meta["summary"] = merge_summary_candidates(
            meta.get("summary"),
            summary,
            meta.get("body_excerpt"),
        )
        category_values = []
        for value in [meta.get("category"), category]:
            normalized = clean_text(value)
            if normalized and normalized not in category_values:
                category_values.append(normalized)
        meta["category"] = " / ".join(category_values) if category_values else None
        merged_tags: list[str] = []
        for value in list(meta.get("tags") or []) + list(tags or []):
            normalized = clean_text(value)
            if normalized and normalized not in merged_tags:
                merged_tags.append(normalized)
        meta["tags"] = merged_tags
        return meta
    if allow_search_only and published_at is not None and dt_in_range(published_at, start, end):
        return build_search_only_article(
            article_url,
            title,
            published_at,
            summary=summary,
            category=category,
            tags=tags,
            source_discovery=f"{source_discovery}_search_only",
        )
    return None


def configured_search_terms(
    config: dict[str, Any],
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[str]:
    if search_terms:
        return search_terms
    raw_terms = config.get("search_terms")
    if isinstance(raw_terms, list):
        terms = [clean_text(str(item)) for item in raw_terms if clean_text(str(item))]
        if terms:
            return terms
    locale = clean_text(config.get("query_locale")).lower()
    if locale == "global":
        return build_media_search_terms(None, country_code)
    return build_media_search_terms(None, country_code)


def adapter_config_for_url(source_url: str, adapter_configs: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    domain = source_domain(source_url)
    if domain in adapter_configs:
        return adapter_configs[domain]
    for candidate, config in adapter_configs.items():
        if domain.endswith(f".{candidate}"):
            return config
    return None


def build_config_search_url(config: dict[str, Any], term: str, page_index: int) -> str:
    template = clean_text(config.get("search_url_template"))
    if not template:
        raise ValueError("missing search_url_template")
    page_mode = clean_text(config.get("page_mode")).lower() or "single"
    offset_step = int(config.get("offset_step", 10) or 10)
    return template.format(
        query=urllib.parse.quote(term, safe=""),
        raw_query=term,
        page=page_index + 1,
        offset=page_index * offset_step,
    )


def extract_configured_value(node, selector: str | None, attr: str | None = None) -> str:
    selector = clean_text(selector)
    if not selector:
        return ""
    target = node.select_one(selector)
    if target is None:
        return ""
    if attr:
        return clean_text(target.get(attr))
    return clean_text(target.get_text(" ", strip=True))


def parse_configured_date_text(value: str) -> datetime | None:
    text = clean_text(value)
    if not text:
        return None
    parsed = parse_dt(text)
    if parsed is not None:
        return parsed
    for fmt in ["%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"]:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=ZoneInfo("UTC"))
        except ValueError:
            continue
    return None


def collect_configured_adapter_articles(
    entry: SourceEntry,
    session,
    start: datetime,
    end: datetime,
    config: dict[str, Any],
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    diagnostics: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    item_selector = clean_text(config.get("item_selector"))
    link_selector = clean_text(config.get("link_selector"))
    title_selector = clean_text(config.get("title_selector"))
    date_selector = clean_text(config.get("date_selector"))
    summary_selector = clean_text(config.get("summary_selector"))
    date_attr = clean_text(config.get("date_attr")) or None
    link_attr = clean_text(config.get("link_attr")) or "href"
    title_attr = clean_text(config.get("title_attr")) or None
    summary_attr = clean_text(config.get("summary_attr")) or None
    source_discovery = clean_text(config.get("source_discovery")) or "configured_search"
    allow_search_only = bool(config.get("allow_search_only", False))
    max_pages = int(config.get("max_pages", 1) or 1)
    max_terms = int(config.get("max_terms", 0) or 0)
    max_items_per_term = int(config.get("max_items_per_term", 0) or 0)
    require_search_result_date = bool(config.get("require_search_result_date", False))
    link_pattern = clean_text(config.get("link_pattern"))
    if diagnostics is not None:
        diagnostics.setdefault("search_endpoint_status", "")
        diagnostics.setdefault("selector_match_count", 0)
        diagnostics.setdefault("parsed_date_count", 0)

    for feed_url in [clean_text(item) for item in config.get("feed_urls", []) if clean_text(item)]:
        for row in parse_feed_articles(feed_url, session, start, end):
            if not matches_any_search_term(f"{row.get('title') or ''} {row.get('summary') or ''}", search_terms):
                continue
            article_url = clean_text(row.get("article_url"))
            if not article_url or article_url in seen_urls:
                continue
            seen_urls.add(article_url)
            row["source_discovery"] = f"{source_discovery}_feed"
            articles.append(row)

    configured_terms = configured_search_terms(config, search_terms, country_code)
    if max_terms > 0:
        configured_terms = configured_terms[:max_terms]
    for term in configured_terms:
        for page_index in range(max_pages):
            search_url = build_config_search_url(config, term, page_index)
            response = fetch(session, search_url)
            if diagnostics is not None and not diagnostics.get("search_endpoint_status"):
                diagnostics["search_endpoint_status"] = f"http_{response.status_code}"
            soup = BeautifulSoup(response.text, "lxml")
            items = soup.select(item_selector) if item_selector else []
            if diagnostics is not None:
                diagnostics["selector_match_count"] = int(diagnostics.get("selector_match_count") or 0) + len(items)
            if not items:
                break
            page_match_count = 0
            for item_index, item in enumerate(items):
                if max_items_per_term > 0 and item_index >= max_items_per_term:
                    break
                anchor = item.select_one(link_selector) if link_selector else None
                if anchor is None or not anchor.get(link_attr):
                    continue
                href = abs_url(response.url, anchor.get(link_attr))
                if link_pattern and link_pattern not in href:
                    continue
                if href in seen_urls:
                    continue
                title = extract_configured_value(item, title_selector, title_attr) if title_selector else clean_text(anchor.get_text(" ", strip=True))
                date_text = extract_configured_value(item, date_selector, date_attr) if date_selector else ""
                summary = extract_configured_value(item, summary_selector, summary_attr) if summary_selector else ""
                published_at = parse_configured_date_text(date_text) if date_text else None
                if diagnostics is not None and published_at is not None:
                    diagnostics["parsed_date_count"] = int(diagnostics.get("parsed_date_count") or 0) + 1
                if require_search_result_date and published_at is None:
                    continue
                meta = hydrate_search_result_article(
                    href,
                    title,
                    published_at,
                    session,
                    start,
                    end,
                    summary=summary or None,
                    source_discovery=source_discovery,
                    allow_search_only=allow_search_only,
                )
                if not meta:
                    continue
                seen_urls.add(href)
                articles.append(meta)
                page_match_count += 1
            if page_match_count == 0:
                break
    return articles


def collect_eczine_articles(session, start: datetime, end: datetime, search_terms: list[str] | None = None) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None):
        search_url = f"https://eczine.jp/search/{urllib.parse.quote(term, safe='')}"
        response = fetch(session, search_url)
        soup = BeautifulSoup(response.text, "lxml")
        for item in soup.select("li.c-articleindex_listitem"):
            anchor = item.select_one("p.c-articleindex_item_heading a[href]")
            time_node = item.select_one("time[datetime]")
            if not anchor or not time_node:
                continue
            article_url = abs_url(response.url, anchor["href"])
            if article_url in seen_urls:
                continue
            published_at = parse_dt(time_node.get("datetime"))
            if not dt_in_range(published_at, start, end):
                continue
            seen_urls.add(article_url)
            meta = hydrate_search_result_article(
                article_url,
                anchor.get_text(" ", strip=True),
                published_at,
                session,
                start,
                end,
                source_discovery="eczine_search",
            )
            if meta:
                articles.append(meta)
    return articles


def collect_markezine_articles(session, start: datetime, end: datetime, search_terms: list[str] | None = None) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None):
        search_url = f"https://markezine.jp/search/{urllib.parse.quote(term, safe='')}"
        response = fetch(session, search_url)
        soup = BeautifulSoup(response.text, "lxml")
        for item in soup.select("li.c-articleindex_listitem"):
            anchor = item.select_one("p.c-articleindex_item_heading a[href]")
            time_node = item.select_one("time[datetime]")
            if not anchor or not time_node:
                continue
            article_url = abs_url(response.url, anchor["href"])
            if article_url in seen_urls:
                continue
            published_at = parse_dt(time_node.get("datetime"))
            if not dt_in_range(published_at, start, end):
                continue
            seen_urls.add(article_url)
            meta = hydrate_search_result_article(
                article_url,
                anchor.get_text(" ", strip=True),
                published_at,
                session,
                start,
                end,
                source_discovery="markezine_search",
            )
            if meta:
                articles.append(meta)
    return articles


def collect_netshop_articles(session, start: datetime, end: datetime, search_terms: list[str] | None = None) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None):
        search_url = f"https://netshop.impress.co.jp/search?keys={urllib.parse.quote(term, safe='')}"
        response = fetch(session, search_url)
        soup = BeautifulSoup(response.text, "lxml")
        for anchor in soup.select("a.article-link[href]"):
            title_node = anchor.select_one(".field--name-title")
            date_node = anchor.select_one(".datetimecreated")
            title = clean_text(title_node.get_text(" ", strip=True)) if title_node else ""
            published_text = date_node.get("title") if date_node else ""
            article_url = abs_url(response.url, anchor["href"])
            if not title or article_url in seen_urls:
                continue
            published_at = parse_dt(published_text)
            if not dt_in_range(published_at, start, end):
                continue
            seen_urls.add(article_url)
            meta = hydrate_search_result_article(
                article_url,
                title,
                published_at,
                session,
                start,
                end,
                source_discovery="netshop_search",
            )
            if meta:
                articles.append(meta)
    return articles


def collect_tiktok_newsroom_articles(
    session,
    start: datetime,
    end: datetime,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    target = official_source_target("TikTok/TikTok Shop", country_code)
    source = target["url"] or default_official_source_url("TikTok/TikTok Shop")
    response = fetch(session, source)
    fallback_lang = str(country_setting(country_code, "tiktok_newsroom_lang", "ja-JP"))
    lang = urllib.parse.parse_qs(urllib.parse.urlparse(source).query).get("lang", [fallback_lang])[0]
    soup = BeautifulSoup(response.text, "lxml")
    articles: list[dict[str, Any]] = []
    for script in soup.find_all("script", {"type": "application/json"}):
        payload = script.get_text(strip=True)
        if not payload.startswith("%7B%22url%22"):
            continue
        try:
            data = json.loads(urllib.parse.unquote(payload))
        except Exception:
            continue
        main_article = data.get("state", {}).get("loaderData", {}).get("routes/_app._index", {}).get("mainArticle", {})
        article_id = clean_text(main_article.get("id"))
        title = clean_text(main_article.get("title"))
        published_at = parse_dt(main_article.get("publishedDate"), default_tz="UTC")
        if not article_id or not title or not dt_in_range(published_at, start, end):
            continue
        articles.append(
            {
                "article_url": f"https://newsroom.tiktok.com/{article_id}?lang={lang}",
                "title": title,
                "published_at": published_at.isoformat(),
                "summary": clean_text(main_article.get("content", ""))[:600] or None,
                "source_discovery": "tiktok_newsroom_state",
            }
        )
        break
    return articles


def collect_aboutamazon_articles(
    session,
    start: datetime,
    end: datetime,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    target = official_source_target("Amazon", country_code)
    source = target["url"] or default_official_source_url("Amazon")
    response = fetch(session, source)
    soup = BeautifulSoup(response.text, "lxml")
    articles: list[dict[str, Any]] = []
    seen: set[str] = set()
    for card in soup.select(".promo-card-v2"):
        title_node = card.select_one(".promo-card-v2__title a")
        date_node = card.select_one(".card-meta__published")
        if not title_node or not date_node or not title_node.get("href"):
            continue
        href = abs_url(source, title_node["href"])
        if href in seen:
            continue
        seen.add(href)
        published_at = parse_dt(date_node.get_text(" ", strip=True))
        if not dt_in_range(published_at, start, end):
            continue
        articles.append(
            {
                "article_url": href,
                "title": clean_text(title_node.get_text(" ", strip=True)),
                "published_at": published_at.isoformat(),
                "summary": None,
                "source_discovery": "aboutamazon_news",
            }
        )
    return articles


def collect_netkeizai_search_results(session, search_term: str, start: datetime, end: datetime) -> list[dict[str, Any]]:
    encoded_term = urllib.parse.quote(search_term, safe="")
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    for page_index in range(NETKEIZAI_MAX_SEARCH_PAGES):
        offset = page_index * NETKEIZAI_SEARCH_PAGE_SIZE
        search_url = f"https://netkeizai.com/search/index/{encoded_term}"
        if offset:
            search_url = f"{search_url}/{offset}"

        try:
            response = fetch(session, search_url)
        except Exception:
            break

        soup = BeautifulSoup(response.text, "lxml")
        page_results: list[dict[str, Any]] = []
        oldest_in_page: datetime | None = None

        for anchor in soup.find_all("a", href=True):
            href = abs_url(response.url, anchor["href"])
            if "/articles/detail/" not in href or href in seen_urls:
                continue

            title_node = anchor.select_one(".tit")
            date_node = anchor.select_one(".data")
            title = clean_text(title_node.get_text(" ", strip=True)) if title_node else ""
            published_at = parse_dt(date_node.get_text(" ", strip=True) if date_node else "")
            if not title or published_at is None:
                continue

            if oldest_in_page is None or published_at < oldest_in_page:
                oldest_in_page = published_at
            if published_at > end or published_at < start:
                continue

            seen_urls.add(href)
            page_results.append(
                {
                    "article_url": href,
                    "title": title,
                    "published_at_hint": published_at,
                }
            )

        if not page_results and oldest_in_page is not None and oldest_in_page < start:
            break

        for item in page_results:
            meta = extract_article_metadata(
                item["article_url"],
                session,
                start,
                end,
                published_at_hint=item["published_at_hint"],
                title_hint=item["title"],
                source_discovery="site_search",
            )
            if meta:
                articles.append(meta)

        if oldest_in_page is not None and oldest_in_page < start:
            break

    return articles


def collect_nikkei_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None, country_code):
        search_url = f"https://www.nikkei.com/search?keyword={urllib.parse.quote(term, safe='')}"
        response = fetch(session, search_url)
        soup = BeautifulSoup(response.text, "lxml")
        for item in soup.select("div.nui-card__content"):
            anchor = item.select_one(".nui-card__title a[href*='/article/']")
            time_node = item.select_one(".nui-card__meta time[datetime]")
            excerpt = item.select_one(".nui-card__excerpt")
            if not anchor or not time_node:
                continue
            article_url = abs_url(response.url, anchor["href"])
            if article_url in seen_urls:
                continue
            published_at = parse_dt(time_node.get("datetime"))
            if not dt_in_range(published_at, start, end):
                continue
            seen_urls.add(article_url)
            meta = hydrate_search_result_article(
                article_url,
                anchor.get_text(" ", strip=True),
                published_at,
                session,
                start,
                end,
                summary=excerpt.get_text(" ", strip=True) if excerpt else None,
                source_discovery="nikkei_search",
                allow_search_only=True,
            )
            if meta:
                articles.append(meta)
    return articles


def collect_rakuten_press_articles(
    session,
    start: datetime,
    end: datetime,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    target = official_source_target("Rakuten Ichiba", country_code)
    source = target["url"] or default_official_source_url("Rakuten Ichiba")
    response = fetch(session, source)
    soup = BeautifulSoup(response.text, "lxml")
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    pattern = re.compile(str(country_setting(country_code, "rakuten_press_path_date_pattern", r"/news/press/(\d{4})/(\d{2})(\d{2})_\d+\.html$")))
    date_group_mode = str(country_setting(country_code, "rakuten_press_date_group_mode", "ymd_compact"))
    timezone_name = str(get_country_config(country_code)["timezone"])
    for anchor in soup.find_all("a", href=True):
        href = abs_url(response.url, anchor["href"])
        match = pattern.search(urllib.parse.urlparse(href).path)
        if not match or href in seen_urls:
            continue
        seen_urls.add(href)
        title = clean_text(anchor.get_text(" ", strip=True))
        if not title:
            continue
        if date_group_mode == "ymd":
            year, month, day = match.groups()
        else:
            year, month, day_text = match.groups()
            day = day_text[-2:]
        published_at = datetime(int(year), int(month), int(day), 0, 0, tzinfo=ZoneInfo(timezone_name))
        if not dt_in_range(published_at, start, end):
            continue
        meta = hydrate_search_result_article(
            href,
            title,
            published_at,
            session,
            start,
            end,
            source_discovery="rakuten_press_fallback",
            allow_search_only=True,
        )
        if meta:
            articles.append(meta)
            continue
        articles.append(
            {
                "article_url": href,
                "title": title,
                "published_at": published_at.isoformat(),
                "summary": None,
                "source_discovery": "rakuten_press_fallback",
            }
        )
    return articles


def collect_qoo10_feed_articles(session, start: datetime, end: datetime) -> list[dict[str, Any]]:
    rows = parse_feed_articles("https://article-university.qoo10.jp/feed", session, start, end)
    for row in rows:
        row["source_discovery"] = "qoo10_university_feed"
        if row.get("summary"):
            row["summary"] = html_to_text(row["summary"])
    return rows


def extract_temu_raw_data(html: str) -> dict[str, Any] | None:
    soup = BeautifulSoup(html, "lxml")
    for script in soup.find_all("script"):
        content = script.get_text()
        if "window.rawData=" not in content or ";document.dispatchEvent" not in content:
            continue
        start_index = content.index("window.rawData=") + len("window.rawData=")
        end_index = content.index(";document.dispatchEvent")
        try:
            return json.loads(content[start_index:end_index])
        except Exception:
            return None
    return None


def collect_temu_announcement_articles(session, start: datetime, end: datetime) -> list[dict[str, Any]]:
    response = fetch(session, "https://www.temu.com/about_temu_home.html")
    raw_data = extract_temu_raw_data(response.text)
    if not raw_data:
        return []

    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for entry in raw_data.get("store", {}).get("articleLists", []):
        article_url = abs_url(response.url, entry.get("customLink") or entry.get("seoLink") or entry.get("link", ""))
        if not article_url or article_url in seen_urls:
            continue
        seen_urls.add(article_url)
        try:
            detail_response = fetch(session, article_url)
        except Exception:
            continue
        detail_raw = extract_temu_raw_data(detail_response.text)
        detail = detail_raw.get("store", {}).get("detail", {}) if detail_raw else {}
        published_at = parse_dt(detail.get("showTime"))
        if not dt_in_range(published_at, start, end):
            continue
        articles.append(
            {
                "article_url": article_url,
                "title": clean_text(entry.get("title")),
                "published_at": published_at.isoformat(),
                "summary": clean_text(entry.get("briefText")) or None,
                "source_discovery": "temu_announcements_fallback",
            }
        )
    return articles


def collect_shein_newsroom_articles(session, start: datetime, end: datetime) -> list[dict[str, Any]]:
    response = fetch(session, "https://www.sheingroup.com/newsroom")
    soup = BeautifulSoup(response.text, "lxml")
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for article in soup.find_all("article"):
        title_node = article.select_one("a.h4")
        meta_node = article.find(string=re.compile(r"—"))
        if not title_node or not title_node.get("href") or not meta_node:
            continue
        article_url = abs_url(response.url, title_node["href"])
        if article_url in seen_urls:
            continue
        seen_urls.add(article_url)
        meta_text = clean_text(str(meta_node))
        if "—" not in meta_text:
            continue
        _, date_text = [clean_text(part) for part in meta_text.split("—", 1)]
        published_at = parse_dt(date_text)
        if not dt_in_range(published_at, start, end):
            continue
        articles.append(
            {
                "article_url": article_url,
                "title": clean_text(title_node.get_text(" ", strip=True)),
                "published_at": published_at.isoformat(),
                "summary": None,
                "source_discovery": "shein_newsroom_fallback",
            }
        )
    return articles


def collect_yahoo_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None, country_code):
        search_url = f"https://news.yahoo.co.jp/search?p={urllib.parse.quote(term, safe='')}"
        response = fetch(session, search_url)
        soup = BeautifulSoup(response.text, "lxml")
        for item in soup.select("ol.newsFeed_list > li"):
            anchor = item.select_one("a[href*='https://news.yahoo.co.jp/articles/']")
            time_node = item.find("time")
            if not anchor or not time_node:
                continue
            article_url = anchor["href"]
            if article_url in seen_urls:
                continue
            text_blocks = [clean_text(div.get_text(" ", strip=True)) for div in anchor.find_all("div")]
            text_blocks = [block for block in text_blocks if block]
            if not text_blocks:
                continue
            title = text_blocks[0]
            summary = text_blocks[1] if len(text_blocks) > 1 else None
            published_at = parse_dt(time_node.get_text(" ", strip=True))
            if not dt_in_range(published_at, start, end):
                continue
            seen_urls.add(article_url)
            meta = hydrate_search_result_article(
                article_url,
                title,
                published_at,
                session,
                start,
                end,
                summary=summary,
                source_discovery="yahoo_search",
                allow_search_only=True,
            )
            if meta:
                articles.append(meta)
    return articles


def collect_netkeizai_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for term in search_terms or build_media_search_terms(None, country_code):
        for article in collect_netkeizai_search_results(session, term, start, end):
            article_url = article["article_url"]
            if article_url in seen_urls:
                continue
            seen_urls.add(article_url)
            articles.append(article)
    return articles


def matches_any_search_term(text: str, search_terms: list[str] | None = None) -> bool:
    normalized = clean_text(text)
    if not normalized or not search_terms:
        return True
    return any(alias_matches(normalized, term) for term in search_terms if clean_text(term))


def collect_generic_listing_articles(
    listing_urls: list[str],
    session,
    start: datetime,
    end: datetime,
    *,
    source_discovery: str,
    search_terms: list[str] | None = None,
    max_links_per_page: int = 18,
) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    for listing_url in listing_urls:
        try:
            response = fetch(session, listing_url)
        except Exception:
            continue

        soup = BeautifulSoup(response.text, "lxml")
        listing_domain = source_domain(response.url)
        candidates: list[tuple[int, str, str]] = []
        seen_candidates: set[str] = set()

        for anchor in soup.find_all("a", href=True):
            href = abs_url(response.url, anchor["href"])
            if not href.startswith("http") or href in seen_candidates:
                continue
            if source_domain(href) != listing_domain:
                continue
            title_hint = clean_text(anchor.get_text(" ", strip=True))
            score = score_candidate_link(response.url, href, title_hint)
            if score <= 0:
                continue
            if not matches_any_search_term(f"{title_hint} {href}", search_terms):
                continue
            seen_candidates.add(href)
            candidates.append((score, href, title_hint))

        candidates.sort(key=lambda item: item[0], reverse=True)
        for _, href, title_hint in candidates[:max_links_per_page]:
            if href in seen_urls:
                continue
            seen_urls.add(href)
            meta = extract_article_metadata(
                href,
                session,
                start,
                end,
                title_hint=title_hint,
                source_discovery=source_discovery,
            )
            if meta:
                articles.append(meta)

    return articles


def collect_domain_limited_search_articles(
    session,
    start: datetime,
    end: datetime,
    *,
    search_terms: list[str] | None,
    domains: list[str],
    source_discovery: str,
    country_code: str = DEFAULT_COUNTRY_CODE,
    engines: list[str] | None = None,
    max_terms: int = 4,
) -> list[dict[str, Any]]:
    if not search_terms:
        return []

    normalized_terms = unique_clean_texts(search_terms)[:max_terms]
    normalized_domains = unique_clean_texts(domains)
    if not normalized_terms or not normalized_domains:
        return []

    selected_engines = engines or ["google", "bing"]
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    for term in normalized_terms:
        for domain in normalized_domains:
            query = f'{term} site:{domain}'
            for engine in selected_engines:
                feed_url = build_promo_search_feed_url(engine, query, country_code)
                for feed_row in parse_feed_articles(feed_url, session, start, end):
                    article_url = clean_text(feed_row.get("article_url"))
                    if not article_url or article_url in seen_urls:
                        continue

                    is_engine_redirect = matches_domain(article_url, "news.google.com", "bing.com", "www.bing.com")
                    if not is_engine_redirect and source_domain(article_url) not in normalized_domains:
                        continue

                    feed_title = clean_text(feed_row.get("title") or "")
                    feed_summary = html_to_text(feed_row.get("summary") or "") or None
                    if not matches_any_search_term(f"{feed_title} {feed_summary or ''}", normalized_terms):
                        continue

                    published_at = parse_dt(feed_row.get("published_at"))
                    meta = None
                    if not is_engine_redirect:
                        meta = hydrate_search_result_article(
                            article_url,
                            feed_title,
                            published_at,
                            session,
                            start,
                            end,
                            summary=feed_summary,
                            source_discovery=source_discovery,
                            allow_search_only=True,
                        )

                    if not meta and published_at is not None and dt_in_range(published_at, start, end):
                        meta = build_search_only_article(
                            article_url,
                            feed_title,
                            published_at,
                            summary=feed_summary,
                            source_discovery=f"{source_discovery}_search_only",
                        )

                    if not meta:
                        continue

                    seen_urls.add(article_url)
                    articles.append(meta)

    return articles


def merge_article_lists_by_url(*groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for group in groups:
        for article in group:
            article_url = clean_text(article.get("article_url"))
            if article_url and article_url in seen_urls:
                continue
            if article_url:
                seen_urls.add(article_url)
            merged.append(article)
    return merged


def collect_prtimes_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    listing_urls = [
        "https://prtimes.jp/",
        "https://prtimes.jp/technology/",
        "https://prtimes.jp/mobile/",
        "https://prtimes.jp/app/",
        "https://prtimes.jp/fashion/",
        "https://prtimes.jp/lifestyle/",
    ]
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    for listing_url in listing_urls:
        try:
            response = fetch(session, listing_url)
        except Exception:
            continue
        soup = BeautifulSoup(response.text, "lxml")
        for anchor in soup.find_all("a", href=True):
            href = abs_url(response.url, anchor["href"])
            title = clean_text(anchor.get_text(" ", strip=True))
            if "/main/html/rd/p/" not in href or not title or href in seen_urls:
                continue
            if not matches_any_search_term(f"{title} {href}", search_terms):
                continue
            seen_urls.add(href)
            meta = hydrate_search_result_article(
                href,
                title,
                None,
                session,
                start,
                end,
                source_discovery="prtimes_listing",
            )
            if meta:
                articles.append(meta)

    return articles


def collect_nbc_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    listing_urls = [
        "https://www.nbcnews.com/business",
        "https://www.nbcnews.com/tech-media",
        "https://www.nbcnews.com/business/consumer",
        "https://www.nbcnews.com/",
    ]
    listing_articles = collect_generic_listing_articles(
        listing_urls,
        session,
        start,
        end,
        source_discovery="nbc_listing",
        search_terms=search_terms,
        max_links_per_page=20,
    )
    if not search_terms or len(listing_articles) >= 3:
        return listing_articles

    fallback_articles = collect_domain_limited_search_articles(
        session,
        start,
        end,
        search_terms=search_terms,
        domains=["www.nbcnews.com", "nbcnews.com"],
        source_discovery="nbc_domain_search",
        country_code=country_code,
    )
    return merge_article_lists_by_url(listing_articles, fallback_articles)


def collect_cnn_articles(
    session,
    start: datetime,
    end: datetime,
    search_terms: list[str] | None = None,
    *,
    edition_base: str = "https://www.cnn.com",
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    listing_urls = [
        f"{edition_base}/business",
        f"{edition_base}/business/tech",
        f"{edition_base}/style",
        edition_base,
    ]
    listing_articles = collect_generic_listing_articles(
        listing_urls,
        session,
        start,
        end,
        source_discovery="cnn_listing",
        search_terms=search_terms,
        max_links_per_page=24,
    )
    if not search_terms or len(listing_articles) >= 3:
        return listing_articles

    edition_domain = urllib.parse.urlparse(edition_base).netloc
    fallback_articles = collect_domain_limited_search_articles(
        session,
        start,
        end,
        search_terms=search_terms,
        domains=[edition_domain],
        source_discovery="cnn_domain_search",
        country_code=country_code,
    )
    return merge_article_lists_by_url(listing_articles, fallback_articles)


def collect_fashionsnap_articles(session, start: datetime, end: datetime, search_terms: list[str] | None = None) -> list[dict[str, Any]]:
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()

    for row in parse_feed_articles("https://www.fashionsnap.com/rss.xml", session, start, end):
        row["source_discovery"] = "fashionsnap_feed"
        if row.get("summary"):
            row["summary"] = html_to_text(row["summary"])
        if not matches_any_search_term(f"{row.get('title') or ''} {row.get('summary') or ''}", search_terms):
            continue
        article_url = clean_text(row.get("article_url"))
        if not article_url or article_url in seen_urls:
            continue
        seen_urls.add(article_url)
        articles.append(row)

    if articles:
        return articles

    return collect_generic_listing_articles(
        ["https://www.fashionsnap.com/article/news/", "https://www.fashionsnap.com/article/inside/"],
        session,
        start,
        end,
        source_discovery="fashionsnap_listing",
        search_terms=search_terms,
        max_links_per_page=20,
    )


def collect_ilsole24ore_articles(session, start: datetime, end: datetime, search_terms: list[str] | None = None) -> list[dict[str, Any]]:
    feed_urls = [
        "https://www.ilsole24ore.com/rss/economia.xml",
        "https://www.ilsole24ore.com/rss/economia--consumi.xml",
        "https://www.ilsole24ore.com/rss/tecnologia--economia-digitale.xml",
        "https://www.ilsole24ore.com/rss/moda--economia-e-finanza.xml",
        "https://www.ilsole24ore.com/rss/radiocor--aziende-ed-istituzioni.xml",
    ]
    articles: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for feed_url in feed_urls:
        for row in parse_feed_articles(feed_url, session, start, end):
            row["source_discovery"] = "ilsole24ore_rss"
            if not matches_any_search_term(f"{row.get('title') or ''} {row.get('summary') or ''}", search_terms):
                continue
            article_url = clean_text(row.get("article_url"))
            if not article_url or article_url in seen_urls:
                continue
            seen_urls.add(article_url)
            articles.append(row)
    return articles


def run_explicit_media_adapter(
    entry: SourceEntry,
    session,
    start: datetime,
    end: datetime,
    adapter_configs: dict[str, dict[str, Any]],
    search_terms: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> ExplicitMediaRun | None:
    if matches_domain(entry.source_url, "ilsole24ore.com", "www.ilsole24ore.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: Il Sole 24 ORE RSS", collect_ilsole24ore_articles(session, start, end, search_terms), search_endpoint_status="rss")
    configured = adapter_config_for_url(entry.source_url, adapter_configs)
    if configured is not None:
        note = clean_text(configured.get("note")) or "configured site adapter"
        diagnostics: dict[str, Any] = {}
        try:
            articles = collect_configured_adapter_articles(
                entry,
                session,
                start,
                end,
                configured,
                search_terms,
                country_code,
                diagnostics=diagnostics,
            )
        except Exception as exc:
            return ExplicitMediaRun("error", f"{note}: {type(exc).__name__}: {exc}", [])
        return ExplicitMediaRun(
            "ok",
            note,
            articles,
            search_endpoint_status=clean_text(diagnostics.get("search_endpoint_status")),
            selector_match_count=int(diagnostics.get("selector_match_count") or 0),
            parsed_date_count=int(diagnostics.get("parsed_date_count") or 0),
        )
    if matches_domain(entry.source_url, "eczine.jp"):
        return ExplicitMediaRun("ok", "explicit site adapter: eczine search", collect_eczine_articles(session, start, end, search_terms))
    if matches_domain(entry.source_url, "netshop.impress.co.jp"):
        return ExplicitMediaRun("ok", "explicit site adapter: netshop search", collect_netshop_articles(session, start, end, search_terms))
    if matches_domain(entry.source_url, "netkeizai.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: netkeizai search", collect_netkeizai_articles(session, start, end, search_terms, country_code))
    if matches_domain(entry.source_url, "markezine.jp"):
        return ExplicitMediaRun("ok", "explicit site adapter: markezine search", collect_markezine_articles(session, start, end, search_terms))
    if matches_domain(entry.source_url, "nikkei.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: nikkei search", collect_nikkei_articles(session, start, end, search_terms, country_code))
    if matches_domain(entry.source_url, "news.yahoo.co.jp"):
        return ExplicitMediaRun("ok", "explicit site adapter: yahoo search", collect_yahoo_articles(session, start, end, search_terms, country_code))
    if matches_domain(entry.source_url, "prtimes.jp"):
        return ExplicitMediaRun("ok", "explicit site adapter: prtimes listing", collect_prtimes_articles(session, start, end, search_terms, country_code))
    if matches_domain(entry.source_url, "nbcnews.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: nbc listing", collect_nbc_articles(session, start, end, search_terms, country_code=country_code))
    if matches_domain(entry.source_url, "edition.cnn.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: edition cnn listing", collect_cnn_articles(session, start, end, search_terms, edition_base="https://edition.cnn.com", country_code=country_code))
    if matches_domain(entry.source_url, "cnn.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: cnn listing", collect_cnn_articles(session, start, end, search_terms, edition_base="https://www.cnn.com", country_code=country_code))
    if matches_domain(entry.source_url, "fashionsnap.com"):
        return ExplicitMediaRun("ok", "explicit site adapter: fashionsnap rss", collect_fashionsnap_articles(session, start, end, search_terms))
    if matches_domain(entry.source_url, "forbes.com"):
        return ExplicitMediaRun("blocked", "explicit search endpoint is blocked by anti-bot response", [])
    if matches_domain(entry.source_url, "reuters.com", "bloomberg.com", "wsj.com"):
        return ExplicitMediaRun("blocked", "site blocks scripted search requests with anti-bot or authentication", [])
    return ExplicitMediaRun("unsupported", "no explicit media adapter configured for this domain", [])


def run_official_side_adapter(
    entry: SourceEntry,
    session,
    start: datetime,
    end: datetime,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> ExplicitMediaRun | None:
    platform = clean_text(entry.platform)
    target = official_source_target(platform, country_code)
    target_label = target["label"]
    target_url = target["url"]
    if platform == "TikTok/TikTok Shop":
        articles = collect_tiktok_newsroom_articles(session, start, end, country_code)
        return ExplicitMediaRun(
            "ok",
            f"official public fallback: {target_label or 'TikTok Newsroom'}",
            articles,
            target_url,
        )
    if platform == "Amazon":
        articles = collect_aboutamazon_articles(session, start, end, country_code)
        return ExplicitMediaRun(
            "ok",
            f"official public fallback: {target_label or 'About Amazon'}",
            articles,
            target_url,
        )
    if platform == "Rakuten Ichiba":
        articles = collect_rakuten_press_articles(session, start, end, country_code)
        return ExplicitMediaRun(
            "ok",
            f"official public fallback: {target_label or 'Rakuten Press'}",
            articles,
            target_url,
        )
    if platform == "Qoo10":
        articles = collect_qoo10_feed_articles(session, start, end)
        return ExplicitMediaRun(
            "ok",
            "official public fallback: Qoo10 University Feed",
            articles,
            "https://article-university.qoo10.jp/feed",
        )
    if platform == "TEMU":
        articles = collect_temu_announcement_articles(session, start, end)
        return ExplicitMediaRun(
            "ok",
            "official public fallback: TEMU Announcements",
            articles,
            "https://www.temu.com/about_temu_home.html",
        )
    if platform == "Shein":
        articles = collect_shein_newsroom_articles(session, start, end)
        return ExplicitMediaRun(
            "ok",
            "official public fallback: SHEIN Group Newsroom",
            articles,
            "https://www.sheingroup.com/newsroom",
        )
    return None


def custom_site_articles(source_url: str, session, start: datetime, end: datetime) -> list[dict[str, Any]] | None:
    lowered = source_url.lower()
    if source_matches_official_platform(source_url, "TikTok/TikTok Shop"):
        country_code = country_for_official_source_url(source_url)
        return collect_tiktok_newsroom_articles(session, start, end, country_code)
    if source_matches_official_platform(source_url, "Amazon") or "aboutamazon.co.jp" in lowered:
        return collect_aboutamazon_articles(session, start, end, country_for_official_source_url(source_url))
    return None


def source_date_window_days(start: datetime, end: datetime) -> int:
    seconds = max(0.0, (end - start).total_seconds())
    return max(1, int((seconds + 86399) // 86400))


def should_use_source_shallow_mode(start: datetime, end: datetime, recall_mode: str = DEFAULT_RECALL_MODE) -> bool:
    return (
        clean_text(recall_mode).lower() == "balanced"
        and source_date_window_days(start, end) <= SOURCE_SHORT_WINDOW_DAYS
    )


def classify_zero_yield_reason(result: SourceResult) -> str:
    if result.recent_article_count > 0:
        return ""
    normalized = f"{result.status} {result.note}".lower()
    if "login" in normalized or result.status == "login_required":
        return "access_limited"
    if any(token in normalized for token in ["cloudflare", "403", "forbidden", "blocked", "anti bot", "anti-bot"]):
        return "blocked_or_anti_bot"
    if any(token in normalized for token in ["javascript", "client-side", "next.js", "hydration"]):
        return "client_side_rendering"
    if result.status == "error":
        return categorize_error_reason(result.note)
    endpoint_status = clean_text(result.search_endpoint_status).lower()
    if endpoint_status.startswith("http_4") or endpoint_status.startswith("http_5"):
        return "search_endpoint_invalid"
    if result.selector_match_count <= 0 and endpoint_status.startswith("http_2"):
        return "selector_no_match"
    if result.selector_match_count > 0 and result.parsed_date_count <= 0:
        return "date_not_parsed"
    if result.candidate_count <= 0:
        return "no_recent_candidates"
    if result.matched_brand_count <= 0:
        return "no_brand_match"
    return "no_recent_brand_articles"


def source_recommendation_for_reason(result: SourceResult) -> str:
    reason = result.zero_yield_reason or classify_zero_yield_reason(result)
    if not reason:
        return "keep"
    if reason in {"access_limited", "blocked_or_anti_bot", "client_side_rendering"}:
        return "skip_deep_crawl_or_disable"
    if reason in {"no_recent_candidates", "no_recent_brand_articles", "no_brand_match", "selector_no_match", "date_not_parsed"}:
        return "use_shallow_crawl_for_short_window"
    if reason == "search_endpoint_invalid":
        return "fix_or_replace_source_adapter"
    if reason in {"connect_timeout", "read_timeout", "proxy_error"}:
        return "retry_later_or_check_network"
    return "review_source_adapter"


def enrich_source_result(result: SourceResult, started_at: float) -> SourceResult:
    result.elapsed_seconds = round(max(0.0, time.monotonic() - started_at), 2)
    if result.candidate_count < 0:
        result.candidate_count = 0
    if result.matched_brand_count < 0:
        result.matched_brand_count = 0
    if not result.matched_brand_count and result.recent_article_count:
        result.matched_brand_count = result.recent_article_count
    result.recent_yield = "has_recent_articles" if result.recent_article_count > 0 else "zero_recent_yield"
    result.zero_yield_reason = classify_zero_yield_reason(result)
    result.source_recommendation = source_recommendation_for_reason(result)
    return result


def test_single_source(
    entry: SourceEntry,
    start: datetime,
    end: datetime,
    max_links: int,
    adapter_configs: dict[str, dict[str, Any]],
    site_credentials: dict[str, dict[str, Any]],
    media_search_terms: list[str] | None = None,
    media_targets: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
    recall_mode: str = DEFAULT_RECALL_MODE,
) -> tuple[SourceResult, list[dict[str, Any]]]:
    started_at = time.monotonic()
    session = build_session()
    fetch_url = canonicalize_source_url(entry.source_url)
    credentials_applied = apply_site_credentials(session, fetch_url, site_credentials)
    if should_use_source_shallow_mode(start, end, recall_mode):
        max_links = min(max_links, SOURCE_SHALLOW_MAX_LINKS)

    if entry.side != "media":
        try:
            explicit_run = run_official_side_adapter(entry, session, start, end, country_code)
        except Exception as exc:
            result = SourceResult(
                platform=entry.platform,
                side=entry.side,
                source_url=entry.source_url,
                fetch_url=fetch_url,
                source_site=source_site(entry.source_url),
                final_url=None,
                status="error",
                http_status=None,
                recent_article_count=0,
                note=f"official side adapter failed: {type(exc).__name__}: {exc}",
            )
            return enrich_source_result(result, started_at), []
        if explicit_run is not None:
            pseudo_response = SimpleNamespace(status_code=None, url=entry.source_url)
            result, articles = finalize_explicit_result(entry, pseudo_response, explicit_run, country_code=country_code)
            return enrich_source_result(result, started_at), articles

    try:
        response = fetch(session, fetch_url)
    except Exception as exc:
        result = SourceResult(
            platform=entry.platform,
            side=entry.side,
            source_url=entry.source_url,
            fetch_url=fetch_url,
            source_site=source_site(entry.source_url),
            final_url=None,
            status="error",
            http_status=None,
            recent_article_count=0,
            note=f"{type(exc).__name__}: {exc}" + ("; credentials were applied" if credentials_applied else ""),
        )
        return enrich_source_result(result, started_at), []

    if entry.side == "media":
        homepage_status, homepage_note = detect_source_status(
            entry.source_url,
            response.status_code,
            response.url,
            response.text[:50000],
        )
        if credentials_applied and homepage_status != "public":
            homepage_note = f"{homepage_note}; credentials were applied"
        configured_media_adapter = adapter_config_for_url(entry.source_url, adapter_configs)
        if should_use_source_shallow_mode(start, end, recall_mode) and homepage_status in {"login_required", "blocked"} and configured_media_adapter is None:
            result = SourceResult(
                platform=entry.platform,
                side=entry.side,
                source_url=entry.source_url,
                fetch_url=fetch_url,
                source_site=source_site(entry.source_url),
                final_url=response.url,
                status=homepage_status,
                http_status=response.status_code,
                recent_article_count=0,
                note=f"{homepage_note}; skipped deep crawl in short-window mode",
            )
            return enrich_source_result(result, started_at), []
        try:
            explicit_run = run_explicit_media_adapter(entry, session, start, end, adapter_configs, media_search_terms, country_code)
        except Exception as exc:
            result = SourceResult(
                platform=entry.platform,
                side=entry.side,
                source_url=entry.source_url,
                fetch_url=fetch_url,
                source_site=source_site(entry.source_url),
                final_url=response.url,
                status="error",
                http_status=response.status_code,
                recent_article_count=0,
                note=f"explicit media adapter failed: {type(exc).__name__}: {exc}",
            )
            return enrich_source_result(result, started_at), []
        result, articles = finalize_explicit_result(entry, response, explicit_run, media_targets, country_code)
        return enrich_source_result(result, started_at), articles

    status, note = detect_source_status(entry.source_url, response.status_code, response.url, response.text[:50000])
    if credentials_applied and status != "public":
        note = f"{note}; credentials were applied"
    if status != "public":
        result = SourceResult(
            platform=entry.platform,
            side=entry.side,
            source_url=entry.source_url,
            fetch_url=fetch_url,
            source_site=source_site(entry.source_url),
            final_url=response.url,
            status=status,
            http_status=response.status_code,
            recent_article_count=0,
            note=note,
        )
        return enrich_source_result(result, started_at), []

    custom_articles = custom_site_articles(response.url, session, start, end)
    if custom_articles is not None:
        articles = []
        for article in custom_articles:
            keep, matched_brands = should_keep_article(entry, article, media_targets, country_code)
            if keep:
                article["matched_brands"] = matched_brands
                articles.append(article)
        result = SourceResult(
            platform=entry.platform,
            side=entry.side,
            source_url=entry.source_url,
            fetch_url=fetch_url,
            source_site=source_site(entry.source_url),
            final_url=response.url,
            status="ok" if articles else "public_no_recent_articles",
            http_status=response.status_code,
            recent_article_count=len(articles),
            note="custom extractor" if articles else "custom extractor found no recent articles",
            candidate_count=len(custom_articles),
            matched_brand_count=len(articles),
        )
        return enrich_source_result(result, started_at), attach_source_context(entry, response.url, articles, country_code=country_code)

    soup = BeautifulSoup(response.text, "lxml")

    articles: list[dict[str, Any]] = []
    for feed_url in discover_feeds(soup, response.url):
        articles.extend(parse_feed_articles(feed_url, session, start, end))
        if articles:
            break

    if not articles:
        for candidate_url in discover_candidate_links(soup, response.url, max_links):
            article_meta = extract_article_metadata(candidate_url, session, start, end)
            if article_meta:
                articles.append(article_meta)

    deduped: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for article in articles:
        url = article["article_url"]
        if url in seen_urls:
            continue
        seen_urls.add(url)
        keep, matched_brands = should_keep_article(entry, article, media_targets, country_code)
        if not keep:
            continue
        article["matched_brands"] = matched_brands
        deduped.append(article)

    deduped.sort(key=lambda row: row["published_at"], reverse=True)
    result = SourceResult(
        platform=entry.platform,
        side=entry.side,
        source_url=entry.source_url,
        fetch_url=fetch_url,
        source_site=source_site(entry.source_url),
        final_url=response.url,
        status="ok" if deduped else "public_no_recent_articles",
        http_status=response.status_code,
        recent_article_count=len(deduped),
        note="generic discovery succeeded" if deduped else "public page fetched but no recent articles found",
        candidate_count=len(articles),
        matched_brand_count=len(deduped),
    )
    return enrich_source_result(result, started_at), attach_source_context(entry, response.url, deduped, country_code=country_code)


def attach_source_context(
    entry: SourceEntry,
    final_url: str,
    articles: list[dict[str, Any]],
    *,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for article in articles:
        row = dict(article)
        matched_platforms = row.get("matched_brands", [])
        resolved_platform = resolve_platform_for_article(entry, matched_platforms, country_code)
        display_platform_label = dedupe_utils.format_brand_labels(matched_platforms) if matched_platforms else resolved_platform
        row["platform"] = display_platform_label or resolved_platform
        row["platform_label"] = display_platform_label or resolved_platform
        row["source_platform"] = entry.platform
        row["side"] = entry.side
        row["source_url"] = entry.source_url
        row["source_final_url"] = final_url
        row["source_site"] = source_site(entry.source_url)
        row["source_country_guess"] = guess_country(article["article_url"])
        row["country_code"] = country_code
        enriched.append(row)
    return enriched


def finalize_explicit_result(
    entry: SourceEntry,
    response,
    explicit_run: ExplicitMediaRun,
    media_targets: list[str] | None = None,
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> tuple[SourceResult, list[dict[str, Any]]]:
    articles = []
    for article in explicit_run.articles:
        keep, matched_brands = should_keep_article(entry, article, media_targets, country_code)
        if keep:
            article["matched_brands"] = matched_brands
            articles.append(article)
    explicit_status = explicit_run.status
    if explicit_status == "ok":
        explicit_status = "ok" if articles else "public_no_recent_articles"
    final_url = explicit_run.final_url or response.url
    result = SourceResult(
        platform=entry.platform,
        side=entry.side,
        source_url=entry.source_url,
        fetch_url=canonicalize_source_url(entry.source_url),
        source_site=source_site(entry.source_url),
        final_url=final_url,
        status=explicit_status,
        http_status=response.status_code,
        recent_article_count=len(articles),
        note=explicit_run.note if articles or explicit_run.status != "ok" else f"{explicit_run.note}; no recent title matches",
        candidate_count=len(explicit_run.articles),
        matched_brand_count=len(articles),
        search_endpoint_status=explicit_run.search_endpoint_status,
        selector_match_count=explicit_run.selector_match_count,
        parsed_date_count=explicit_run.parsed_date_count,
    )
    return result, attach_source_context(entry, final_url, articles, country_code=country_code)


def build_side_label(sides: list[str]) -> str:
    ordered = [side for side in ["media", "buyer", "seller"] if side in sides]
    if not ordered:
        return "unknown"
    return "_".join(ordered)


def format_output_range_label(start: date, end: date) -> str:
    return f"range_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}"


def build_run_output_dir_name(country_slug: str, side_label: str, timestamp: str, start: date, end: date) -> str:
    return f"{country_slug}_xlsx_sources_{side_label}_run_{timestamp}_{format_output_range_label(start, end)}"


def emit_progress(
    callback,
    *,
    stage: str,
    total_sites: int,
    completed_sites: int,
    active_sites: int,
    current_site: str | None = None,
    last_completed_site: str | None = None,
    message: str | None = None,
    progress_percent: int | None = None,
) -> None:
    if callback is None:
        return
    callback(
        {
            "stage": stage,
            "total_sites": total_sites,
            "completed_sites": completed_sites,
            "active_sites": active_sites,
            "current_site": current_site or "",
            "last_completed_site": last_completed_site or "",
            "message": message or "",
            "progress_percent": progress_percent,
        }
    )


def categorize_error_reason(note: str | None) -> str:
    normalized = clean_text(note).lower()
    if not normalized:
        return "unknown"
    if "sslerror" in normalized or "certificate verify failed" in normalized or "self-signed certificate" in normalized:
        return "ssl_cert_verification"
    if "connecttimeout" in normalized or "connection to " in normalized and "timed out" in normalized:
        return "connect_timeout"
    if "readtimeout" in normalized or "read timed out" in normalized:
        return "read_timeout"
    if "proxyerror" in normalized or "proxy" in normalized:
        return "proxy_error"
    if "toomanyredirects" in normalized or "too many redirects" in normalized:
        return "too_many_redirects"
    if "login" in normalized or "signin" in normalized or "登录" in normalized:
        return "login_required"
    if "403" in normalized or "forbidden" in normalized or "blocked" in normalized:
        return "blocked"
    if "adapter failed" in normalized:
        return "adapter_failed"
    if "parse" in normalized or "selector" in normalized:
        return "parse_failure"
    return "other_network_or_parser_error"


def summarize_error_reasons(results: list[SourceResult]) -> Counter:
    return Counter(
        categorize_error_reason(result.note)
        for result in results
        if result.status == "error"
    )


def main(argv: list[str] | None = None, progress_callback=None) -> int:
    args = parse_args(argv)
    country_config = get_country_config(args.country)
    translator = TextTranslator(normalize_translation_target(args.translate_to))
    start, end = parse_date_range(args)
    survey_indicators = load_survey_indicators_from_xlsx(args.xlsx)
    survey_api_url, survey_api_key, survey_api_model = resolve_survey_api_credentials(args)
    survey_system_prompt = clean_text(args.survey_system_prompt) or default_survey_ai_system_prompt(args.country)

    selected_display_platforms, _, custom_platforms = resolve_requested_platforms(args.platforms, args.country)
    selected_platform_labels = selected_display_platforms + custom_platforms
    selected_search_terms = build_media_search_terms(args.platforms, args.country)
    selected_media_targets = media_match_targets(args.platforms, args.country)
    recall_mode = clean_text(getattr(args, "recall_mode", DEFAULT_RECALL_MODE)).lower()
    if recall_mode not in RECALL_MODE_CHOICES:
        recall_mode = DEFAULT_RECALL_MODE
    source_search_terms = selected_search_terms
    source_search_mode = "full"
    if should_use_source_shallow_mode(start, end, recall_mode):
        source_search_terms = build_source_search_terms(args.platforms, args.country)
        source_search_mode = "shallow_short_window"

    source_groups: list[list[SourceEntry]] = [load_extra_sources(args.extra_sources)]
    if bool(country_config.get("include_xlsx_sources")):
        source_groups.insert(0, load_sources_from_xlsx(args.xlsx))
    all_sources = merge_source_entries(*source_groups)
    sources = select_sources_for_run(
        all_sources,
        args.platforms,
        args.sides,
        args.country,
    )
    adapter_configs = load_adapter_configs(args.adapter_configs)
    site_credentials = load_site_credentials(args.site_credentials)
    results: list[SourceResult] = []
    all_articles: list[dict[str, Any]] = []
    total_sites = len(sources)
    completed_sites = 0

    emit_progress(
        progress_callback,
        stage="setup",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=min(args.workers, total_sites),
        current_site="准备分配抓取任务",
        message="已完成参数解析，准备开始抓取来源网站",
        progress_percent=NEWS_PROGRESS_SETUP,
    )

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                test_single_source,
                entry,
                start,
                end,
                args.max_links_per_source,
                adapter_configs,
                site_credentials,
                source_search_terms,
                selected_media_targets,
                args.country,
                recall_mode,
            ): entry
            for entry in sources
        }
        for future in as_completed(futures):
            entry = futures[future]
            result, articles = future.result()
            results.append(result)
            all_articles.extend(articles)
            completed_sites += 1
            remaining_sites = max(0, total_sites - completed_sites)
            crawl_progress_percent = interpolate_progress(
                NEWS_PROGRESS_CRAWL_START,
                NEWS_PROGRESS_CRAWL_END,
                completed_sites,
                total_sites,
            )
            emit_progress(
                progress_callback,
                stage="crawl_site",
                total_sites=total_sites,
                completed_sites=completed_sites,
                active_sites=min(args.workers, remaining_sites),
                current_site=f"并发抓取中（当前活跃 {min(args.workers, remaining_sites)} 个网站）" if remaining_sites else "抓取阶段即将完成",
                last_completed_site=entry.source_url,
                message=f"已完成 {completed_sites}/{total_sites} 个网站，最近完成：{entry.source_url}",
                progress_percent=crawl_progress_percent,
            )

    emit_progress(
        progress_callback,
        stage="translation",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="整理抓取结果",
        last_completed_site="",
        message="网站抓取完成，正在整理抓取结果",
        progress_percent=NEWS_PROGRESS_POST_CRAWL,
    )

    related_news_search_enabled = bool(getattr(args, "search_related_news", False))
    report_ranking_search_enabled = bool(getattr(args, "search_report_ranking", False))
    supplemental_search_enabled = related_news_search_enabled or report_ranking_search_enabled or bool(getattr(args, "promo_search", False))
    related_news_search_keywords = normalize_related_news_search_keywords_text(
        clean_text(getattr(args, "related_news_search_keywords", "")) or clean_text(getattr(args, "promo_search_keywords", "")),
        args.country,
        recall_mode,
    )
    report_search_keywords = normalize_report_search_keywords_text(getattr(args, "report_search_keywords", ""), args.country)
    promo_search_stats = {
        "task_count": 0,
        "article_count": 0,
        "related_news_task_count": 0,
        "report_task_count": 0,
        "related_news_article_count": 0,
        "report_article_count": 0,
    }
    if supplemental_search_enabled:
        emit_progress(
            progress_callback,
            stage="promo_search",
            total_sites=total_sites,
            completed_sites=completed_sites,
            active_sites=0,
            current_site="准备进行补充检索",
            last_completed_site="",
            message="来源网站抓取完成，开始执行相关新闻与数据/报告补充检索",
            progress_percent=NEWS_PROGRESS_POST_CRAWL,
        )
        promo_articles, promo_search_stats = collect_search_engine_promo_articles(
            args.platforms,
            engine_mode=args.promo_search_engine,
            keyword_blocks_text=related_news_search_keywords,
            survey_indicators=survey_indicators,
            related_news_enabled=related_news_search_enabled,
            related_news_keywords_text=related_news_search_keywords,
            report_ranking_enabled=report_ranking_search_enabled,
            report_keywords_text=report_search_keywords,
            country_code=args.country,
            recall_mode=recall_mode,
            start=start,
            end=end,
            progress_callback=progress_callback,
            total_sites=total_sites,
            completed_sites=completed_sites,
        )
        all_articles.extend(promo_articles)

    results.sort(key=lambda row: (row.platform, row.side, row.source_url))
    all_articles.sort(key=lambda row: (row["platform"], row["source_site"], row["published_at"]), reverse=True)
    raw_article_count = len(all_articles)
    articles_before_initial_dedupe = deepcopy(all_articles)
    initial_dedupe_stats: dict[str, Any] = {}
    all_articles = dedupe_utils.dedupe_articles(all_articles, stats=initial_dedupe_stats)
    deduped_article_count = len(all_articles)
    articles_after_initial_dedupe = deepcopy(all_articles)
    duplicate_article_count = max(0, raw_article_count - deduped_article_count)
    survey_filter_mode = clean_text(args.survey_filter_mode).lower() or DEFAULT_SURVEY_FILTER_MODE
    survey_filter_label = "AI 理解筛选" if survey_filter_mode == "ai" else "关键词筛选"

    emit_progress(
        progress_callback,
        stage="translation",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="翻译与整理文章内容",
        last_completed_site="",
        message="网站抓取完成，正在翻译与整理文章内容",
        progress_percent=NEWS_PROGRESS_TRANSLATION_START,
    )

    def translation_progress(payload: dict[str, Any]) -> None:
        completed = int(payload.get('completed') or 0)
        total = int(payload.get('total') or 0)
        retry_count = int(payload.get('retry_count') or 0)
        last_text = clean_text(payload.get('last_text') or '')
        if total <= 0:
            return
        progress_cap = NEWS_PROGRESS_TRANSLATION_AI_CAP if survey_filter_mode == 'ai' else NEWS_PROGRESS_TRANSLATION_KEYWORD_CAP
        progress_percent = min(
            progress_cap,
            interpolate_progress(NEWS_PROGRESS_TRANSLATION_START, progress_cap, completed, total),
        )
        emit_progress(
            progress_callback,
            stage="translation",
            total_sites=total_sites,
            completed_sites=completed_sites,
            active_sites=0,
            current_site=f"翻译与整理文章内容（已完成 {completed}/{total} 段文本）",
            last_completed_site=last_text[:120],
            message=(
                f"正在翻译与整理文章内容，已完成 {completed}/{total} 段文本"
                + (f"；重试 {retry_count} 段" if retry_count else "")
            ),
            progress_percent=progress_percent,
        )

    if survey_filter_mode != "ai":
        all_articles = add_translation_fields(all_articles, translator, progress_callback=translation_progress)
    articles_before_filter = deepcopy(all_articles)

    emit_progress(
        progress_callback,
        stage="survey_filter",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site=f"按指标进行{survey_filter_label}",
        last_completed_site="",
        message=f"正在按表格指标进行{survey_filter_label}，判断哪些新闻会影响问卷回答",
        progress_percent=NEWS_PROGRESS_FILTER_AI_START if survey_filter_mode == "ai" else NEWS_PROGRESS_FILTER_KEYWORD_START,
    )

    all_articles, survey_dimension_counts, original_article_count, survey_filter_stats = apply_survey_indicator_filter(
        all_articles,
        survey_indicators,
        mode=survey_filter_mode,
        api_url=survey_api_url,
        api_key=survey_api_key,
        api_model=survey_api_model,
        system_prompt=survey_system_prompt,
        ai_workers=args.survey_ai_workers,
        ai_batch_size=args.survey_ai_batch_size,
        country_code=args.country,
        recall_mode=recall_mode,
        progress_callback=progress_callback,
        total_sites=total_sites,
        completed_sites=completed_sites,
    )
    survey_filter_stats["low_volume_fill_added_count"] = 0
    survey_rows_before_final_dedupe = deepcopy(all_articles)

    emit_progress(
        progress_callback,
        stage="ai_dedupe",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="准备进行 AI 去重",
        last_completed_site="",
        message="指标筛选完成，正在准备进行最后一步 AI 去重",
        progress_percent=NEWS_PROGRESS_DEDUPE_PREP,
    )

    all_articles, ai_dedupe_stats = dedupe_utils.apply_final_ai_dedupe(
        all_articles,
        api_url=survey_api_url,
        api_key=survey_api_key,
        api_model=survey_api_model,
        call_filter_api=call_survey_filter_api,
        ai_workers=min(max(1, args.survey_ai_workers), 4),
        progress_callback=progress_callback,
        progress_emitter=emit_progress,
        total_sites=total_sites,
        completed_sites=completed_sites,
        progress_start=NEWS_PROGRESS_DEDUPE_PREP,
        progress_cap=NEWS_PROGRESS_DEDUPE_CAP,
    )

    translation_title_stats = {
        "failed_count": 0,
        "ai_fallback_count": 0,
        "ai_fallback_error_count": 0,
    }
    if survey_filter_mode == "ai":
        emit_progress(
            progress_callback,
            stage="translation",
            total_sites=total_sites,
            completed_sites=completed_sites,
            active_sites=0,
            current_site="翻译筛选后的新闻",
            last_completed_site="",
            message="AI 筛选完成，正在翻译最终保留的新闻",
            progress_percent=NEWS_PROGRESS_DEDUPE_CAP,
        )

        def final_translation_progress(payload: dict[str, Any]) -> None:
            completed = int(payload.get('completed') or 0)
            total = int(payload.get('total') or 0)
            retry_count = int(payload.get('retry_count') or 0)
            last_text = clean_text(payload.get('last_text') or '')
            if total <= 0:
                return
            progress_percent = min(
                NEWS_PROGRESS_OUTPUT_PREP,
                interpolate_progress(NEWS_PROGRESS_DEDUPE_CAP, NEWS_PROGRESS_OUTPUT_PREP, completed, total),
            )
            emit_progress(
                progress_callback,
                stage="translation",
                total_sites=total_sites,
                completed_sites=completed_sites,
                active_sites=0,
                current_site=f"翻译筛选后的新闻（已完成 {completed}/{total} 段文本）",
                last_completed_site=last_text[:120],
                message=(
                    f"正在翻译最终保留的新闻，已完成 {completed}/{total} 段文本"
                    + (f"；重试 {retry_count} 段" if retry_count else "")
                ),
                progress_percent=progress_percent,
            )

        all_articles = add_translation_fields(
            all_articles,
            translator,
            progress_callback=final_translation_progress,
            extra_text_fields=['survey_indicator_examples', 'survey_ai_reason_translated'],
        )
        all_articles, translation_title_stats = apply_ai_title_translation_fallback(
            all_articles,
            api_url=survey_api_url,
            api_key=survey_api_key,
            api_model=survey_api_model,
            target_language=translator.target_language,
        )
    else:
        translation_title_stats["failed_count"] = count_failed_title_translations(all_articles, translator.target_language)

    all_articles = [apply_industry_trend_fields(article) for article in all_articles]

    all_articles, briefing_sentiment_stats = add_briefing_sentiment_fields(
        all_articles,
        api_url=survey_api_url,
        api_key=survey_api_key,
        api_model=survey_api_model,
        country_code=args.country,
        batch_size=args.survey_ai_batch_size,
    )

    emit_progress(
        progress_callback,
        stage="finalize_output",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="准备输出目录与汇总信息",
        last_completed_site="",
        message="AI 去重完成，正在准备输出目录与汇总信息",
        progress_percent=NEWS_PROGRESS_OUTPUT_PREP,
    )

    side_label = build_side_label(args.sides)
    run_started_at = datetime.now()
    timestamp = run_started_at.strftime("%Y%m%d_%H%M%S")
    output_dir_name = build_run_output_dir_name(
        str(country_config["output_slug"]),
        side_label,
        timestamp,
        start.date(),
        end.date(),
    )
    output_dir = Path(args.output_dir) / output_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)

    before_file_translation_status = "raw_untranslated" if survey_filter_mode == "ai" else "translated"
    before_file_translation_note = (
        "AI mode filters first, so before.csv/before.json keep the pre-filter raw article pool."
        if survey_filter_mode == "ai"
        else "Keyword mode translates before filtering, so before.csv/before.json contain translated pre-filter articles."
    )
    brand_stage_funnel_summary = build_brand_stage_funnel_summary(
        selected_platform_labels,
        start=start,
        end=end,
        country_code=args.country,
        raw_rows=articles_before_initial_dedupe,
        initial_deduped_rows=articles_after_initial_dedupe,
        survey_rows=survey_rows_before_final_dedupe,
        final_rows=all_articles,
    )
    brand_stage_total_summary = build_brand_stage_total_summary(
        selected_platform_labels,
        brand_stage_funnel_summary,
    )
    industry_trend_rows = [row for row in all_articles if normalize_industry_trend_flag(row.get("industry_trend_flag"))]
    industry_trend_category_counts = Counter(
        normalize_industry_trend_category(row.get("industry_trend_category"))
        for row in industry_trend_rows
        if clean_text(row.get("industry_trend_category"))
    )
    industry_trend_impact_counts = Counter(
        normalize_industry_trend_impact(row.get("industry_trend_impact"))
        for row in industry_trend_rows
    )
    broad_entry_platform_summary = build_broad_entry_platform_summary(
        promo_search_stats,
        source_rows=articles_before_filter,
        survey_rows=survey_rows_before_final_dedupe,
        final_rows=all_articles,
        survey_filter_stats=survey_filter_stats,
    )
    platform_alias_summary = platform_alias_effective_summary(selected_platform_labels, args.country)
    low_volume_platform_diagnostics = build_low_volume_platform_diagnostics(
        brand_stage_total_summary,
        promo_search_stats,
        broad_entry_platform_summary,
        platforms=[platform for platform in ["TTS", "IG"] if platform in selected_platform_labels],
    )

    metadata = {
        "country_code": args.country,
        "country_label": country_config["label"],
        "country_market_label": country_config["market_label"],
        "generated_at": run_started_at.isoformat(),
        "side_label": side_label,
        "translation_target": translator.target_language,
        "translation_timing": "after_ai_filter" if survey_filter_mode == "ai" else "before_filter",
        "before_file_semantics": "pre_filter_article_pool",
        "before_file_translation_status": before_file_translation_status,
        "before_file_translation_note": before_file_translation_note,
        "after_file_semantics": "post_filter_final_articles",
        "after_file_translation_status": "translated",
        "range_start": start.isoformat(),
        "range_end": end.isoformat(),
        "selected_sides": args.sides,
        "selected_platforms": selected_platform_labels,
        "platform_alias_effective_summary": platform_alias_summary,
        "media_search_terms": selected_search_terms,
        "source_search_terms": source_search_terms,
        "source_search_mode": source_search_mode,
        "source_short_window_days": SOURCE_SHORT_WINDOW_DAYS,
        "source_elapsed_seconds_total": round(sum(float(result.elapsed_seconds or 0) for result in results), 2),
        "source_zero_yield_reason_counts": Counter(result.zero_yield_reason for result in results if result.zero_yield_reason),
        "source_recommendation_counts": Counter(result.source_recommendation for result in results if result.source_recommendation),
        "recall_mode": recall_mode,
        "brand_stage_funnel_summary": brand_stage_funnel_summary,
        "brand_stage_total_summary": brand_stage_total_summary,
        "promo_search_enabled": supplemental_search_enabled,
        "related_news_search_enabled": related_news_search_enabled,
        "report_ranking_search_enabled": report_ranking_search_enabled,
        "promo_search_engine": args.promo_search_engine,
        "related_news_search_keywords": related_news_search_keywords,
        "related_news_keyword_auto_split_count": keyword_auto_split_count(related_news_search_keywords),
        "related_news_search_keyword_block_count": len(normalize_related_news_search_keyword_blocks(related_news_search_keywords, args.country, recall_mode)),
        "report_search_keywords": report_search_keywords,
        "report_search_keyword_block_count": len(normalize_report_search_keyword_blocks(report_search_keywords, args.country)),
        "promo_search_task_count": int(promo_search_stats.get("task_count", 0) or 0),
        "promo_search_article_count": int(promo_search_stats.get("article_count", 0) or 0),
        "related_news_search_task_count": int(promo_search_stats.get("related_news_task_count", 0) or 0),
        "report_ranking_search_task_count": int(promo_search_stats.get("report_task_count", 0) or 0),
        "related_news_search_article_count": int(promo_search_stats.get("related_news_article_count", 0) or 0),
        "report_ranking_search_article_count": int(promo_search_stats.get("report_article_count", 0) or 0),
        "promo_search_platform_engine_summary": promo_search_stats.get("platform_engine_summary") or {},
        "promo_search_platform_query_summary": promo_search_stats.get("platform_query_summary") or {},
        "promo_search_broad_entry_platform_summary": broad_entry_platform_summary,
        "low_volume_platform_broad_query_summary": broad_entry_platform_summary,
        "low_volume_platform_diagnostics": low_volume_platform_diagnostics,
        "promo_search_task_diagnostics": promo_search_stats.get("task_diagnostics") or [],
        "extra_source_count": len(load_extra_sources(args.extra_sources)),
        "include_xlsx_sources": bool(country_config.get("include_xlsx_sources")),
        "adapter_config_count": len(adapter_configs),
        "site_credential_count": len(site_credentials),
        "source_count": len(sources),
        "source_status_counts": Counter(result.status for result in results),
        "error_reason_counts": summarize_error_reasons(results),
        "survey_indicator_filter_enabled": bool(survey_indicators),
        "survey_indicator_count": len(survey_indicators),
        "survey_filter_mode": survey_filter_mode,
        "survey_filter_label": survey_filter_label,
        "survey_filter_api_configured": bool(survey_api_url and survey_api_key and survey_api_model),
        "survey_filter_api_url": survey_api_url,
        "survey_filter_api_model": survey_api_model,
        "survey_filter_system_prompt": survey_system_prompt,
        "survey_filter_ai_workers": args.survey_ai_workers,
        "survey_filter_ai_batch_size": args.survey_ai_batch_size,
        "survey_filter_ai_request_count": int(survey_filter_stats.get("ai_request_count", 0) or 0),
        "survey_filter_hard_excluded_count": int(survey_filter_stats.get("hard_filtered_count", 0) or 0),
        "survey_filter_ai_evaluated_count": int(survey_filter_stats.get("ai_evaluated_count", 0) or 0),
        "survey_filter_api_error_count": int(survey_filter_stats.get("api_error_count", 0) or 0),
        "survey_filter_keyword_fallback_count": int(survey_filter_stats.get("keyword_fallback_count", 0) or 0),
        "survey_filter_ai_bridge_count": int(survey_filter_stats.get("ai_bridge_count", 0) or 0),
        "survey_filter_ai_excluded_reason_counts": survey_filter_stats.get("ai_excluded_reason_counts") or {},
        "survey_filter_low_volume_fill_count": int(survey_filter_stats.get("low_volume_fill_added_count", 0) or 0),
        "ai_dedupe_enabled": bool(ai_dedupe_stats.get("enabled")),
        "ai_dedupe_configured": bool(ai_dedupe_stats.get("configured")),
        "ai_dedupe_candidate_group_count": int(ai_dedupe_stats.get("candidate_group_count", 0) or 0),
        "ai_dedupe_request_count": int(ai_dedupe_stats.get("request_count", 0) or 0),
        "ai_dedupe_removed_count": int(ai_dedupe_stats.get("removed_count", 0) or 0),
        "ai_dedupe_api_error_count": int(ai_dedupe_stats.get("api_error_count", 0) or 0),
        "ai_dedupe_date_window_days": int(ai_dedupe_stats.get("date_window_days", 0) or 0),
        "translation_title_failed_count": int(translation_title_stats.get("failed_count", 0) or 0),
        "translation_title_ai_fallback_count": int(translation_title_stats.get("ai_fallback_count", 0) or 0),
        "translation_title_ai_fallback_error_count": int(translation_title_stats.get("ai_fallback_error_count", 0) or 0),
        "briefing_sentiment_enabled": bool(briefing_sentiment_stats.get("enabled")),
        "briefing_sentiment_configured": bool(briefing_sentiment_stats.get("configured")),
        "briefing_sentiment_request_count": int(briefing_sentiment_stats.get("request_count", 0) or 0),
        "briefing_sentiment_evaluated_count": int(briefing_sentiment_stats.get("evaluated_count", 0) or 0),
        "briefing_sentiment_error_count": int(briefing_sentiment_stats.get("api_error_count", 0) or 0),
        "briefing_sentiment_defaulted_count": int(briefing_sentiment_stats.get("defaulted_count", 0) or 0),
        "industry_trend_enabled": True,
        "industry_trend_article_count": len(industry_trend_rows),
        "industry_trend_category_counts": dict(industry_trend_category_counts),
        "industry_trend_impact_counts": dict(industry_trend_impact_counts),
        "survey_related_article_count_before_filter": original_article_count,
        "survey_related_article_count_after_filter": len(all_articles),
        "survey_dimension_article_counts": dict(survey_dimension_counts),
        "article_count_before_dedupe": raw_article_count,
        "duplicate_article_count_removed": duplicate_article_count,
        "initial_dedupe_removed_count": int(initial_dedupe_stats.get("removed_count", duplicate_article_count) or 0),
        "initial_dedupe_removed_by_reason": dict(initial_dedupe_stats.get("removed_by_reason") or {}),
        "initial_dedupe_missing_url_count": int(initial_dedupe_stats.get("missing_url_count", 0) or 0),
        "initial_dedupe_group_count": int(initial_dedupe_stats.get("group_count", 0) or 0),
        "initial_dedupe_removed_samples_json": "initial_dedupe_removed_samples.json" if initial_dedupe_stats.get("removed_samples") else "",
        "article_count": len(all_articles),
        "metadata_json": "metadata.json",
        "sources_json": "sources.json",
        "sources_csv": "sources.csv",
        "before_json": "before.json",
        "before_csv": "before.csv",
        "after_json": "after.json",
        "after_csv": "after.csv",
    }

    emit_progress(
        progress_callback,
        stage="finalize_output",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="写入 JSON 输出文件",
        last_completed_site="",
        message="正在写入 JSON 输出文件",
        progress_percent=NEWS_PROGRESS_JSON_WRITE,
    )
    output_writer.write_json(output_dir / "metadata.json", metadata)
    output_writer.write_json(output_dir / "sources.json", [asdict(result) for result in results])
    removed_samples = initial_dedupe_stats.get("removed_samples") or []
    if removed_samples:
        output_writer.write_json(output_dir / "initial_dedupe_removed_samples.json", removed_samples)
    output_writer.write_json(output_dir / "before.json", articles_before_filter)
    output_writer.write_json(output_dir / "after.json", all_articles)

    emit_progress(
        progress_callback,
        stage="finalize_output",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="写入 CSV 输出文件",
        last_completed_site="",
        message="正在写入 CSV 输出文件",
        progress_percent=NEWS_PROGRESS_CSV_WRITE,
    )
    output_writer.write_csv(output_dir / "sources.csv", [asdict(result) for result in results])
    output_writer.write_csv(output_dir / "before.csv", articles_before_filter)
    output_writer.write_csv(output_dir / "after.csv", all_articles)

    sqlite_write_error = ""
    try:
        db_store.save_run(output_dir, metadata, articles_before_filter, all_articles)
        metadata["sqlite_db_enabled"] = True
        metadata["sqlite_db_path"] = str(db_store.DEFAULT_DB_PATH)
        output_writer.write_json(output_dir / "metadata.json", metadata)
    except Exception as exc:
        sqlite_write_error = str(exc) or exc.__class__.__name__
        metadata["sqlite_db_enabled"] = False
        metadata["sqlite_write_error"] = sqlite_write_error
        try:
            output_writer.write_json(output_dir / "metadata.json", metadata)
        except Exception:
            pass

    emit_progress(
        progress_callback,
        stage="done",
        total_sites=total_sites,
        completed_sites=completed_sites,
        active_sites=0,
        current_site="",
        last_completed_site="",
        message="所有新闻抓取与输出写入已完成",
        progress_percent=100,
    )

    print(f"输出目录: {output_dir}")
    print(f"国家: {country_config['label']}")
    print(f"运行侧别: {side_label}")
    print(f"品牌范围: {', '.join(selected_platform_labels) if selected_platform_labels else '全部'}")
    print(f"召回模式: {recall_mode}")
    print(f"媒体侧搜索词: {', '.join(selected_search_terms) if selected_search_terms else '无'}")
    print(f"补充检索: {'开启' if supplemental_search_enabled else '关闭'}")
    print(f"相关新闻检索: {'开启' if related_news_search_enabled else '关闭'}")
    print(f"数据/报告检索: {'开启' if report_ranking_search_enabled else '关闭'}")
    if supplemental_search_enabled:
        print(f"全网补充搜索引擎: {args.promo_search_engine}")
        print(f"补充检索任务数: {int(promo_search_stats.get('task_count', 0) or 0)}")
        related_block_count = len(normalize_related_news_search_keyword_blocks(related_news_search_keywords, args.country, recall_mode))
        default_related_block_count = len(normalize_related_news_search_keyword_blocks('', args.country, recall_mode))
        related_auto_split_count = keyword_auto_split_count(related_news_search_keywords)
        print(f"相关新闻关键词块数: {related_block_count}")
        if related_auto_split_count:
            print(f"相关新闻关键词自动拆分数: {related_auto_split_count}")
        if related_block_count < max(1, default_related_block_count // 2):
            print("提示: 相关新闻关键词块偏少，可能导致筛选前新闻数量下降。")
        print(f"补充检索文章数: {int(promo_search_stats.get('article_count', 0) or 0)}")
        print(f"相关新闻任务数: {int(promo_search_stats.get('related_news_task_count', 0) or 0)}")
        print(f"数据/报告任务数: {int(promo_search_stats.get('report_task_count', 0) or 0)}")
        print(f"相关新闻文章数: {int(promo_search_stats.get('related_news_article_count', 0) or 0)}")
        print(f"数据/报告文章数: {int(promo_search_stats.get('report_article_count', 0) or 0)}")
    print(f"来源总数: {len(sources)}")
    print(f"抓取文章原始数: {raw_article_count}")
    print(f"去重移除重复新闻数: {duplicate_article_count}")
    print(f"指标筛选前新闻数: {len(articles_before_filter)}")
    print(f"去重后且指标筛选后新闻数: {len(all_articles)}")
    if before_file_translation_status == "raw_untranslated":
        print("指标筛选前文件: before.csv / before.json（原文未翻译；AI 模式先筛选后翻译）")
    else:
        print("指标筛选前文件: before.csv / before.json（已翻译）")
    print("指标筛选后文件: after.csv / after.json")
    if sqlite_write_error:
        print(f"SQLite 写入异常: {sqlite_write_error}")
    print(f"指标筛选方式: {survey_filter_label}")
    if survey_filter_mode == "ai":
        print(f"AI 模型: {survey_api_model or '未填写'}")
        print(f"AI 批量大小: {int(survey_filter_stats.get('ai_batch_size', args.survey_ai_batch_size) or args.survey_ai_batch_size)}")
        print(f"AI 请求批次数: {int(survey_filter_stats.get('ai_request_count', 0) or 0)}")
        print(f"AI 预先硬过滤数: {int(survey_filter_stats.get('hard_filtered_count', 0) or 0)}")
        print(f"AI 判定新闻数: {int(survey_filter_stats.get('ai_evaluated_count', 0) or 0)}")
        print(f"AI 异常批次数: {int(survey_filter_stats.get('api_error_count', 0) or 0)}")
        print(f"关键词回退保留数: {int(survey_filter_stats.get('keyword_fallback_count', 0) or 0)}")
        print(f"AI 桥接保留数: {int(survey_filter_stats.get('ai_bridge_count', 0) or 0)}")
        ai_excluded_reason_counts = survey_filter_stats.get('ai_excluded_reason_counts') or {}
        if ai_excluded_reason_counts:
            print("AI 筛选排除原因分布:")
            for reason, count in Counter(ai_excluded_reason_counts).most_common():
                print(f"  {reason}: {count}")
    print(f"最终 AI 去重: {'开启' if ai_dedupe_stats.get('enabled') else '关闭'}")
    if ai_dedupe_stats.get('configured'):
        print(f"最终 AI 去重候选组数: {int(ai_dedupe_stats.get('candidate_group_count', 0) or 0)}")
        print(f"最终 AI 去重请求数: {int(ai_dedupe_stats.get('request_count', 0) or 0)}")
        print(f"最终 AI 去重移除数: {int(ai_dedupe_stats.get('removed_count', 0) or 0)}")
        print(f"最终 AI 去重异常组数: {int(ai_dedupe_stats.get('api_error_count', 0) or 0)}")
    print(f"指标筛选前新闻数: {original_article_count}")
    print(f"指标筛选后新闻数: {len(all_articles)}")
    if brand_stage_total_summary:
        print("品牌阶段汇总概览:")
        for row in brand_stage_total_summary:
            print(
                f"  {row['brand']}: "
                f"入池候选 {row['raw_count']}，初步去重后 {row['initial_dedupe_count']}，"
                f"AI 筛选后 {row['survey_filter_count']}，最终保留 {row['final_count']}；"
                f"{row['diagnosis']}"
            )
    alias_warnings = {
        platform: row.get("warnings") or []
        for platform, row in platform_alias_summary.items()
        if isinstance(row, dict) and row.get("warnings")
    }
    if alias_warnings:
        print("品牌别名诊断:")
        for platform, warnings in alias_warnings.items():
            print(f"  {platform}: {'；'.join(str(item) for item in warnings)}")
    if low_volume_platform_diagnostics:
        print("TTS/IG 入池与筛选诊断:")
        for platform, row in low_volume_platform_diagnostics.items():
            top_skipped = row.get("top_skipped_reason_count") or {}
            skipped_label = "，".join(f"{reason} {count}" for reason, count in Counter(top_skipped).most_common(3)) or "无"
            print(
                f"  {platform}: 搜索 feed {row.get('search_feed_count', 0)}，"
                f"入池候选 {row.get('raw_count', 0)}，"
                f"AI 筛选后 {row.get('survey_filter_count', 0)}，"
                f"最终 {row.get('final_count', 0)}；"
                f"主要入池拦截：{skipped_label}"
            )
    if broad_entry_platform_summary:
        print("低量平台宽入池概览:")
        for platform, row in broad_entry_platform_summary.items():
            print(
                f"  {platform}: "
                f"本地候选 {row.get('target_market_kept_count', 0)}/{row.get('target_market_feed_count', 0)}，"
                f"欧洲/全球相关候选 {row.get('broad_kept_count', 0)}/{row.get('broad_feed_count', 0)}，"
                f"短宽查询 {row.get('broad_short_kept_count', 0)}/{row.get('broad_short_feed_count', 0)}，"
                f"宽入池 {row.get('broad_entry_before_filter_count', row.get('broad_entry_count', 0))}，"
                f"AI 保留 {row.get('broad_entry_ai_kept_count', 0)}，"
                f"最终 {row.get('broad_entry_final_count', 0)}"
            )
    if survey_dimension_counts:
        print('指标命中分布:')
        for dimension, count in survey_dimension_counts.most_common():
            print(f"  {dimension}: {count}")
    if industry_trend_rows:
        print("行业趋势新闻数:", len(industry_trend_rows))
        for category, count in industry_trend_category_counts.most_common():
            print(f"  {category}: {count}")
    for status, count in Counter(result.status for result in results).most_common():
        print(f"{status}: {count}")
    error_reason_counts = summarize_error_reasons(results)
    if error_reason_counts:
        print("error细分原因:")
        for reason, count in error_reason_counts.most_common():
            print(f"  {reason}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
