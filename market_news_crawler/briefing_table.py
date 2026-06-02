from __future__ import annotations

import json
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

import news_crawler
import xlsx_source_test
from country_config import DEFAULT_COUNTRY_CODE, get_country_config, normalize_country_code


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / 'outputs'
BRIEFING_TABLE_TEMPLATE_FILENAME = '?????.xlsx'
BRIEFING_TABLE_OUTPUT_DIR = OUTPUT_DIR / 'generated_briefing_tables'


def extract_api_error_detail(exc: Exception) -> str:
    response = getattr(exc, 'response', None)
    if response is None:
        message_parts: list[str] = []
        exception_chain = [exc]
        visited_ids = {id(exc)}
        current = exc
        while True:
            next_exc = getattr(current, '__cause__', None) or getattr(current, '__context__', None)
            if not isinstance(next_exc, BaseException) or id(next_exc) in visited_ids:
                break
            visited_ids.add(id(next_exc))
            exception_chain.append(next_exc)
            current = next_exc

        normalized_chain = ' | '.join(
            f'{type(item).__name__}: {str(item or "").strip()}'
            for item in exception_chain
            if str(item or '').strip()
        ).lower()
        if isinstance(exc, requests.exceptions.ProxyError) or 'proxyerror' in normalized_chain:
            message_parts.append('?????/VPN ???????? VPN ????? HTTPS ??????????')
        elif isinstance(exc, requests.exceptions.SSLError) or 'certificate verify failed' in normalized_chain or 'ssl' in normalized_chain:
            message_parts.append('??? SSL ???????VPN ????? HTTPS ??????????????')
        elif isinstance(exc, requests.exceptions.ConnectTimeout) or 'connecttimeout' in normalized_chain:
            message_parts.append('?????VPN ??????????DNS ??????')
        elif isinstance(exc, requests.exceptions.ReadTimeout) or 'readtimeout' in normalized_chain:
            message_parts.append('???????VPN ?????????????????')
        elif isinstance(exc, requests.exceptions.ConnectionError) or 'name resolution' in normalized_chain or 'nodename nor servname provided' in normalized_chain:
            message_parts.append('???????? VPN ??? DNS ?????????????')

        raw_message = ' | '.join(
            str(item or '').strip()
            for item in exception_chain
            if str(item or '').strip()
        )
        if raw_message:
            message_parts.append(raw_message[:500])
        return '?'.join(part for part in message_parts if part)
    try:
        body_text = str(response.text or '').strip()
    except Exception:
        body_text = ''
    if not body_text:
        return ''
    try:
        payload = json.loads(body_text)
    except Exception:
        return body_text[:500]
    if isinstance(payload, dict):
        error_payload = payload.get('error')
        if isinstance(error_payload, dict):
            message = str(error_payload.get('message') or '').strip()
            if message:
                return message
        message = str(payload.get('message') or '').strip()
        if message:
            return message
    return body_text[:500]


BRIEFING_TABLE_HEADERS = [
    'Tracking Week',
    'Source Country',
    'Platform',
    'NPS Overall Change',
    'Key User Segment Affected',
    'NPS Dimension',
    'News Title / Event',
    'News Type (Media / Official / Community)',
    'Source Name',
    'Publish Date',
    'News URL',
    'Core Summary',
    'Sentiment (Positive / Neutral / Negative)',
    'Affected Side (Buyer / Seller / Both)',
    'Impact Direction on NPS (Up / Down / Neutral)',
    'Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)',
    'Impact Level (High / Medium / Low)',
    'Short-term or Lagged Effect',
    'Quant Linkage Hypothesis',
    'Analyst Note',
    'Include in Client Readout (Y/N)',
    'Follow-up Needed (Y/N)',
]

BRIEFING_AI_HEADERS = [
    'Key User Segment Affected',
    'Core Summary',
    'Affected Side (Buyer / Seller / Both)',
    'Impact Direction on NPS (Up / Down / Neutral)',
    'Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)',
    'Impact Level (High / Medium / Low)',
    'Short-term or Lagged Effect',
    'Quant Linkage Hypothesis',
    'Analyst Note',
    'Include in Client Readout (Y/N)',
    'Follow-up Needed (Y/N)',
]

BRIEFING_COUNTRY_ZH_LABELS = {
    'japan': '日本',
    'france': '法国',
    'germany': '德国',
    'es': '西班牙',
    'spain': '西班牙',
    'italy': '意大利',
}

INVALID_BRIEFING_SEGMENT_VALUES = {
    '',
    'chinese user segment',
    'test segment',
    'user segment',
    'segment',
    '用户群体',
    '消费者',
    '用户',
    '平台用户',
    '普通消费者',
    '普通用户',
    '中国用户',
    '中文用户',
    '中国消费者',
    '中文消费者',
}


def find_briefing_table_template_path() -> Path:
    direct_path = BASE_DIR / BRIEFING_TABLE_TEMPLATE_FILENAME
    if direct_path.exists():
        return direct_path

    for candidate in BASE_DIR.glob('*.xlsx'):
        try:
            workbook = load_workbook(candidate, read_only=True, data_only=True)
            sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.active
            headers = [sheet.cell(1, col).value for col in range(1, min(sheet.max_column, len(BRIEFING_TABLE_HEADERS)) + 1)]
            workbook.close()
        except Exception:
            continue
        if headers == BRIEFING_TABLE_HEADERS[:len(headers)]:
            return candidate
    raise FileNotFoundError(f'找不到资讯表模板：{BRIEFING_TABLE_TEMPLATE_FILENAME}')


def parse_article_publish_dt(value: Any) -> datetime | None:
    parsed = news_crawler.parse_dt(str(value or '').strip())
    if parsed is not None:
        return parsed
    try:
        return datetime.fromisoformat(str(value or '').strip())
    except Exception:
        return None


def format_tracking_week(value: Any) -> str:
    parsed = parse_article_publish_dt(value) or datetime.now().astimezone()
    iso = parsed.isocalendar()
    return f'{iso.year} W{iso.week:02d}'


def format_briefing_publish_date(value: Any) -> str:
    parsed = parse_article_publish_dt(value)
    if parsed is None:
        return str(value or '').strip()
    return f'{parsed.month}/{parsed.day}/{parsed.year}'


def briefing_country_label(country_code: str) -> str:
    country_config = get_country_config(country_code)
    return str(country_config.get('google_news_gl') or country_code or '').strip().upper()


def briefing_country_zh_label(country_code: str) -> str:
    normalized = normalize_country_code(country_code)
    if normalized in BRIEFING_COUNTRY_ZH_LABELS:
        return BRIEFING_COUNTRY_ZH_LABELS[normalized]
    country_config = get_country_config(normalized)
    for key in ['label', 'market_label', 'output_slug']:
        value = str(country_config.get(key) or '').strip()
        if value:
            return value
    return normalized.upper()


def briefing_segment_options(country_code: str) -> list[str]:
    country = briefing_country_zh_label(country_code)
    return [
        f'{country}平台买家',
        f'{country}价格敏感型买家',
        f'{country}促销/优惠关注买家',
        f'{country}会员/忠诚度用户',
        f'{country} Prime 会员',
        f'偏好现金支付的{country}买家',
        f'{country}高频购物用户',
        f'{country}品质与合规敏感买家',
        f'{country}隐私与数据安全敏感买家',
        f'{country}售后与权益保障敏感买家',
        f'{country}内容/直播购物用户',
        f'{country}跨境购物用户',
        f'{country}新用户',
        f'{country}年轻购物用户',
        f'{country}平台卖家',
        f'{country}品牌商家',
    ]


def compact_cell_text(value: Any, max_length: int = 1200) -> str:
    text = news_crawler.clean_text(str(value or '')).strip()
    if len(text) > max_length:
        return text[:max_length - 1].rstrip() + '…'
    return text


def infer_briefing_news_type(row: dict[str, Any]) -> str:
    text = ' '.join(
        str(row.get(key) or '')
        for key in ['source_name', 'source_url', 'article_url', 'category']
    ).lower()
    if any(token in text for token in ['official', 'press', 'newsroom', 'aboutamazon', 'business.', 'company']):
        return 'Official'
    if any(token in text for token in ['reddit', 'forum', 'community', 'facebook', 'instagram', 'tiktok', 'x.com', 'twitter']):
        return 'Community'
    return 'Media'


def source_name_for_briefing(row: dict[str, Any]) -> str:
    source_name = compact_cell_text(row.get('source_name'), 180)
    if source_name:
        return source_name
    url = str(row.get('article_url') or row.get('source_url') or '').strip()
    try:
        return urlparse(url).netloc.replace('www.', '')
    except Exception:
        return ''


def nps_metric_from_dimensions(dimensions: str) -> str:
    normalized = str(dimensions or '').lower()
    if 'price' in normalized:
        return 'Price'
    if 'logistics' in normalized or 'fulfillment' in normalized or 'delivery' in normalized:
        return 'Fulfillment'
    if 'content' in normalized:
        return 'Content'
    if 'variety' in normalized or 'assortment' in normalized:
        return 'Assortment'
    if any(token in normalized for token in ['quality', 'customer', 'post-purchase', 'service', 'trust', 'seller']):
        return 'Trust'
    return 'Trust'


def direction_symbol(direction: str) -> str:
    normalized = str(direction or '').strip().lower()
    if normalized == 'up':
        return '↑'
    if normalized == 'down':
        return '↓'
    return '→'


def normalize_briefing_sentiment(value: Any) -> str:
    return xlsx_source_test.normalize_briefing_sentiment(value)


def normalize_briefing_segment_text(value: Any) -> str:
    return re.sub(r'\s+', ' ', news_crawler.clean_text(str(value or '')).strip()).strip()


def is_invalid_briefing_segment(value: Any, country_code: str) -> bool:
    segment = normalize_briefing_segment_text(value)
    if not segment:
        return True
    normalized = segment.lower()
    if normalized in INVALID_BRIEFING_SEGMENT_VALUES:
        return True
    country = briefing_country_zh_label(country_code)
    generic_country_values = {
        f'{country}用户',
        f'{country}消费者',
        f'{country}普通用户',
        f'{country}普通消费者',
        f'{country}用户群体',
    }
    if segment in generic_country_values:
        return True
    if 'chinese user segment' in normalized or '中国用户' in segment or '中文用户' in segment:
        return True
    if not news_crawler.contains_chinese_chars(segment):
        return True
    return False


def infer_briefing_segment(row: dict[str, Any], country_code: str) -> str:
    country = briefing_country_zh_label(country_code)
    text = ' '.join(
        str(row.get(key) or '')
        for key in [
            'platform_label',
            'title',
            'title_original',
            'title_display_zh',
            'summary',
            'survey_dimensions',
            'survey_question_ids',
            'survey_indicator_examples',
            'survey_ai_reason_raw',
            'survey_ai_reason_translated',
        ]
    ).lower()

    def has_any(tokens: list[str]) -> bool:
        return any(token.lower() in text for token in tokens)

    if has_any(['cash', 'contanti', '现金', 'payment', 'pagamenti', '支付']):
        return f'偏好现金支付的{country}买家'
    if has_any(['pfas', 'illegal', 'counterfeit', 'fake', 'authentic', 'compliance', 'regulator', 'investigation', '非法', '假货', '正品', '合规', '调查', '监管', '欧盟']):
        return f'{country}品质与合规敏感买家'
    if has_any(['prime']):
        return f'{country} Prime 会员'
    if has_any(['会员', '忠诚', 'loyalty', 'points', '积分', 'reward', '奖励']):
        return f'{country}会员/忠诚度用户'
    if has_any(['coupon', '优惠码', '折扣码', 'promo code', 'sconto', 'codice']):
        return f'{country}促销/优惠关注买家'
    if has_any(['discount', 'deal', 'deals', 'promo', 'promotion', 'sale', '折扣', '优惠', '促销', '大促', '特惠', '价格']):
        return f'{country}价格敏感型买家'
    if has_any(['brand merchant', '品牌商家', '品牌卖家']):
        return f'{country}品牌商家'
    if has_any(['seller', 'merchant', 'commission', '卖家', '商家', '佣金', '店铺政策', 'seller center']):
        return f'{country}平台卖家'
    if has_any(['privacy', 'data', 'security', 'personal information', '隐私', '数据', '安全', '个人信息']):
        return f'{country}隐私与数据安全敏感买家'
    if has_any(['refund', 'return', 'after-sales', 'post-purchase', 'service', 'rights', '退款', '退货', '售后', '权益', '客服']):
        return f'{country}售后与权益保障敏感买家'
    if has_any(['live', 'livestream', 'creator', 'content', 'recommendation', 'tiktok shop', '直播', '内容', '创作者', '推荐算法']):
        return f'{country}内容/直播购物用户'
    if has_any(['cross-border', 'global', 'import', 'export', '跨境', '海外']):
        return f'{country}跨境购物用户'
    if has_any(['new user', 'new customer', '新用户', '新客']):
        return f'{country}新用户'
    if has_any(['young', 'gen z', 'youth', '年轻', 'z世代']):
        return f'{country}年轻购物用户'
    if has_any(['heavy shopper', 'high-frequency', '高频', '重度购物']):
        return f'{country}高频购物用户'
    return f'{country}平台买家'


def normalize_briefing_segment(value: Any, row: dict[str, Any], country_code: str) -> str:
    segment = normalize_briefing_segment_text(value)
    options = briefing_segment_options(country_code)
    if segment in options and not is_invalid_briefing_segment(segment, country_code):
        return segment
    if is_invalid_briefing_segment(segment, country_code):
        return infer_briefing_segment(row, country_code)
    country = briefing_country_zh_label(country_code)
    if country not in segment:
        return infer_briefing_segment(row, country_code)
    return segment


def direct_briefing_fields(row: dict[str, Any], country_code: str) -> dict[str, str]:
    title_zh = compact_cell_text(row.get('title_display_zh') or row.get('title'), 500)
    title_original = compact_cell_text(row.get('title_original'), 500)
    title = title_zh or title_original
    return {
        'Tracking Week': format_tracking_week(row.get('published_at')),
        'Source Country': briefing_country_label(country_code),
        'Platform': compact_cell_text(row.get('platform_label'), 180),
        'NPS Dimension': compact_cell_text(row.get('survey_dimensions'), 180),
        'News Title / Event': title,
        'News Type (Media / Official / Community)': infer_briefing_news_type(row),
        'Source Name': source_name_for_briefing(row),
        'Publish Date': format_briefing_publish_date(row.get('published_at')),
        'News URL': str(row.get('article_url') or row.get('source_url') or '').strip(),
    }


def default_briefing_ai_fields(row: dict[str, Any], error_note: str = '', country_code: str = DEFAULT_COUNTRY_CODE) -> dict[str, str]:
    metric = nps_metric_from_dimensions(str(row.get('survey_dimensions') or ''))
    title = compact_cell_text(row.get('title_display_zh') or row.get('title') or row.get('title_original'), 400)
    summary = compact_cell_text(row.get('summary'), 600)
    analyst_note = compact_cell_text(row.get('survey_indicator_examples'), 800)
    trend_flag = xlsx_source_test.normalize_industry_trend_flag(row.get('industry_trend_flag'))
    trend_impact = xlsx_source_test.normalize_industry_trend_impact(row.get('industry_trend_impact'))
    trend_reason = compact_cell_text(row.get('industry_trend_reason'), 500)
    if trend_flag:
        trend_note = f"行业趋势标记：{trend_reason or '该新闻涉及品牌或平台整体层面的行业趋势。'}"
        analyst_note = f"{analyst_note} | {trend_note}" if analyst_note else trend_note
    if error_note:
        analyst_note = f'AI生成失败，需人工复核：{error_note[:180]}'
    direction = 'Neutral'
    if trend_flag and trend_impact == 'Positive':
        direction = 'Up'
    elif trend_flag and trend_impact == 'Negative':
        direction = 'Down'
    return {
        'Key User Segment Affected': infer_briefing_segment(row, country_code),
        'Core Summary': summary or title,
        'Sentiment (Positive / Neutral / Negative)': normalize_briefing_sentiment(row.get('briefing_sentiment')),
        'Affected Side (Buyer / Seller / Both)': 'Both',
        'Impact Direction on NPS (Up / Down / Neutral)': direction,
        'Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)': metric,
        'Impact Level (High / Medium / Low)': 'Medium',
        'Short-term or Lagged Effect': 'Immediate',
        'Quant Linkage Hypothesis': f'可能影响{metric}相关分项，并传导至总体 NPS。',
        'Analyst Note': analyst_note or 'AI生成失败，需人工复核',
        'Include in Client Readout (Y/N)': 'Y',
        'Follow-up Needed (Y/N)': 'Y',
    }


def build_briefing_ai_messages(rows: list[dict[str, Any]], country_code: str) -> list[dict[str, str]]:
    article_payload = []
    segment_options = briefing_segment_options(country_code)
    country_zh = briefing_country_zh_label(country_code)
    for index, row in enumerate(rows, start=1):
        article_payload.append(
            {
                'article_id': str(row.get('article_id') or index),
                'platform': row.get('platform_label') or '',
                'title_original': row.get('title_original') or '',
                'title_zh': row.get('title_display_zh') or row.get('title') or '',
                'summary': row.get('summary') or '',
                'published_at': row.get('published_at') or '',
                'source_name': row.get('source_name') or '',
                'source_url': row.get('source_url') or '',
                'article_url': row.get('article_url') or '',
                'nps_dimension': row.get('survey_dimensions') or '',
                'survey_question_ids': row.get('survey_question_ids') or '',
                'survey_indicator_examples': row.get('survey_indicator_examples') or '',
                'industry_trend_flag': bool(xlsx_source_test.normalize_industry_trend_flag(row.get('industry_trend_flag'))),
                'industry_trend_category': row.get('industry_trend_category') or '',
                'industry_trend_impact': row.get('industry_trend_impact') or '',
                'industry_trend_reason': row.get('industry_trend_reason') or '',
            }
        )
    schema = {
        'items': [
            {
                'article_id': 'same article_id from input',
                'Key User Segment Affected': 'must be one exact value from allowed_key_user_segments',
                'Core Summary': 'Chinese one-sentence event summary',
                'Affected Side (Buyer / Seller / Both)': 'Buyer|Seller|Both',
                'Impact Direction on NPS (Up / Down / Neutral)': 'Up|Down|Neutral',
                'Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)': 'Price|Trust|Fulfillment|Content|Assortment',
                'Impact Level (High / Medium / Low)': 'High|Medium|Low',
                'Short-term or Lagged Effect': 'Immediate|Lagged|Both',
                'Quant Linkage Hypothesis': 'Chinese hypothesis',
                'Analyst Note': 'Chinese analyst note',
                'Include in Client Readout (Y/N)': 'Y|N',
                'Follow-up Needed (Y/N)': 'Y|N',
            }
        ]
    }
    return [
        {
            'role': 'system',
            'content': (
                'You are an NPS market intelligence analyst. Return JSON only. '
                'Classify each news article for a client briefing table. '
                'Keep Chinese text for free-text fields. Use exactly the enum values shown. '
                'Do not omit any input article.'
            ),
        },
        {
            'role': 'user',
            'content': json.dumps(
                {
                    'country_code': country_code,
                    'country_zh': country_zh,
                    'allowed_key_user_segments': segment_options,
                    'output_schema': schema,
                    'rules': [
                        'Use the NPS dimension and survey linkage as primary evidence.',
                        'Chinese is only the output language; it does not mean China or Chinese users.',
                        'Key User Segment Affected must refer to the article market/country, not China.',
                        'Key User Segment Affected must be one exact value from allowed_key_user_segments.',
                        'Never output Chinese user segment, 中国用户, 中文用户, 消费者, 用户群体, or other generic placeholders.',
                        'Core Summary, Quant Linkage Hypothesis and Analyst Note must be concise Chinese.',
                        'If industry_trend_flag is true, reflect the trend/brand-level impact in Analyst Note and Quant Linkage Hypothesis.',
                        'If evidence is weak, choose Neutral, Medium, and mark Follow-up Needed as Y.',
                    ],
                    'articles': article_payload,
                },
                ensure_ascii=False,
                indent=2,
            ),
        },
    ]


def normalize_enum(value: Any, allowed: list[str], default: str) -> str:
    normalized = str(value or '').strip()
    if not normalized:
        return default
    for item in allowed:
        if normalized.lower() == item.lower():
            return item
    return default


def get_ai_item_value(item: dict[str, Any], header: str) -> Any:
    candidates = [
        header,
        header.split('(')[0].strip(),
        re.sub(r'[^0-9a-zA-Z]+', '_', header).strip('_').lower(),
    ]
    alias_map = {
        'Key User Segment Affected': ['key_user_segment_affected', 'segment'],
        'Core Summary': ['core_summary', 'summary'],
        'Sentiment (Positive / Neutral / Negative)': ['sentiment'],
        'Affected Side (Buyer / Seller / Both)': ['affected_side'],
        'Impact Direction on NPS (Up / Down / Neutral)': ['impact_direction_on_nps', 'impact_direction'],
        'Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)': ['impacted_nps_metric', 'metric'],
        'Impact Level (High / Medium / Low)': ['impact_level'],
        'Short-term or Lagged Effect': ['short_term_or_lagged_effect', 'short_or_lagged_effect', 'effect_timing'],
        'Quant Linkage Hypothesis': ['quant_linkage_hypothesis'],
        'Analyst Note': ['analyst_note'],
        'Include in Client Readout (Y/N)': ['include_in_client_readout', 'include'],
        'Follow-up Needed (Y/N)': ['follow_up_needed', 'followup_needed', 'follow_up'],
    }
    candidates.extend(alias_map.get(header, []))
    lower_item = {str(key).lower(): value for key, value in item.items()}
    for candidate in candidates:
        if candidate in item:
            return item.get(candidate)
        lowered = candidate.lower()
        if lowered in lower_item:
            return lower_item[lowered]
    return ''


def normalize_briefing_ai_item(
    item: dict[str, Any],
    row: dict[str, Any],
    country_code: str = DEFAULT_COUNTRY_CODE,
) -> dict[str, str]:
    values = default_briefing_ai_fields(row, country_code=country_code)
    for header in BRIEFING_AI_HEADERS:
        raw_value = get_ai_item_value(item, header)
        if raw_value not in (None, ''):
            values[header] = compact_cell_text(raw_value, 1200)

    values['Sentiment (Positive / Neutral / Negative)'] = normalize_enum(
        values.get('Sentiment (Positive / Neutral / Negative)'),
        ['Positive', 'Neutral', 'Negative'],
        'Neutral',
    )
    values['Affected Side (Buyer / Seller / Both)'] = normalize_enum(
        values.get('Affected Side (Buyer / Seller / Both)'),
        ['Buyer', 'Seller', 'Both'],
        'Both',
    )
    values['Impact Direction on NPS (Up / Down / Neutral)'] = normalize_enum(
        values.get('Impact Direction on NPS (Up / Down / Neutral)'),
        ['Up', 'Down', 'Neutral'],
        'Neutral',
    )
    values['Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)'] = normalize_enum(
        values.get('Impacted NPS Metric (Price / Trust / Fulfillment / Content / Assortment)'),
        ['Price', 'Trust', 'Fulfillment', 'Content', 'Assortment'],
        nps_metric_from_dimensions(str(row.get('survey_dimensions') or '')),
    )
    values['Impact Level (High / Medium / Low)'] = normalize_enum(
        values.get('Impact Level (High / Medium / Low)'),
        ['High', 'Medium', 'Low'],
        'Medium',
    )
    values['Include in Client Readout (Y/N)'] = normalize_enum(
        values.get('Include in Client Readout (Y/N)'),
        ['Y', 'N'],
        'Y',
    )
    values['Follow-up Needed (Y/N)'] = normalize_enum(
        values.get('Follow-up Needed (Y/N)'),
        ['Y', 'N'],
        'Y',
    )
    values['Key User Segment Affected'] = normalize_briefing_segment(
        values.get('Key User Segment Affected'),
        row,
        country_code,
    )
    return values


def briefing_payload_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ['items', 'rows', 'articles', 'decisions']:
        value = payload.get(key) if isinstance(payload, dict) else None
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def generate_briefing_ai_fields(
    rows: list[dict[str, Any]],
    *,
    country_code: str,
    api_settings: dict[str, str],
    batch_size: int = 10,
    stats: dict[str, int] | None = None,
) -> dict[str, dict[str, str]]:
    if stats is not None:
        stats.setdefault('batch_count', 0)
        stats.setdefault('failed_batch_count', 0)
        stats.setdefault('failed_row_count', 0)
        stats.setdefault('ai_completed_row_count', 0)

    output = {
        str(row.get('article_id') or index): default_briefing_ai_fields(row, country_code=country_code)
        for index, row in enumerate(rows, start=1)
    }
    row_by_id = {
        str(row.get('article_id') or index): row
        for index, row in enumerate(rows, start=1)
    }

    for offset in range(0, len(rows), batch_size):
        batch = rows[offset:offset + batch_size]
        if stats is not None:
            stats['batch_count'] += 1
        try:
            payload = xlsx_source_test.call_survey_filter_api(
                build_briefing_ai_messages(batch, country_code),
                api_settings['survey_api_url'],
                api_settings['survey_api_key'],
                api_settings['survey_api_model'],
            )
            for item in briefing_payload_items(payload):
                article_id = str(item.get('article_id') or '').strip()
                if article_id and article_id in row_by_id:
                    output[article_id] = normalize_briefing_ai_item(item, row_by_id[article_id], country_code)
                    if stats is not None:
                        stats['ai_completed_row_count'] += 1
        except Exception as exc:
            detail = extract_api_error_detail(exc)
            message = detail or str(exc) or exc.__class__.__name__
            if stats is not None:
                stats['failed_batch_count'] += 1
                stats['failed_row_count'] += len(batch)
            for index, row in enumerate(batch, start=offset + 1):
                article_id = str(row.get('article_id') or index)
                output[article_id] = default_briefing_ai_fields(row, message, country_code=country_code)
    return output


def build_article_briefing_table(
    rows: list[dict[str, Any]],
    *,
    country_code: str,
    api_settings: dict[str, str],
    ai_stats: dict[str, int] | None = None,
) -> Path:
    template_path = find_briefing_table_template_path()
    workbook = load_workbook(template_path)
    sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.active

    actual_headers = [sheet.cell(1, col).value for col in range(1, len(BRIEFING_TABLE_HEADERS) + 1)]
    if actual_headers != BRIEFING_TABLE_HEADERS:
        raise RuntimeError('资讯表模板 Sheet1 表头与预期不一致，请检查模板文件。')

    template_styles = [copy(sheet.cell(2, col)._style) for col in range(1, len(BRIEFING_TABLE_HEADERS) + 1)]
    template_height = sheet.row_dimensions[2].height
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    ai_fields = generate_briefing_ai_fields(rows, country_code=country_code, api_settings=api_settings, stats=ai_stats)
    for row_index, article in enumerate(rows, start=2):
        article_id = str(article.get('article_id') or (row_index - 1))
        direct_fields = direct_briefing_fields(article, country_code)
        ai_row = ai_fields.get(article_id) or default_briefing_ai_fields(article, country_code=country_code)
        direction = ai_row.get('Impact Direction on NPS (Up / Down / Neutral)', 'Neutral')
        row_values = {
            **direct_fields,
            **ai_row,
            'NPS Overall Change': direction_symbol(direction),
        }
        if template_height is not None:
            sheet.row_dimensions[row_index].height = template_height
        for col_index, header in enumerate(BRIEFING_TABLE_HEADERS, start=1):
            cell = sheet.cell(row_index, col_index)
            cell._style = copy(template_styles[col_index - 1])
            value = row_values.get(header, '')
            cell.value = value
            if header == 'News URL' and value:
                cell.hyperlink = str(value)
                cell.style = 'Hyperlink'

    BRIEFING_TABLE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"briefing_table_{country_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = BRIEFING_TABLE_OUTPUT_DIR / filename
    workbook.save(output_path)
    workbook.close()
    return output_path

